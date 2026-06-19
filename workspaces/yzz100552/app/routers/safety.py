import os
from datetime import date, datetime
from typing import Optional
from fastapi import APIRouter, Depends, Query
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.database import get_db
from app.services import get_safety_export

EXPORT_DIR = os.path.abspath("./exports")
os.makedirs(EXPORT_DIR, exist_ok=True)

router = APIRouter(prefix="/api/safety", tags=["安全员导出/环保管理"])


HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_header(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def _style_body(ws, start_row: int, end_row: int, cols: int):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = CENTER
            cell.border = BORDER


def _autosize(ws, cols: int):
    for c in range(1, cols + 1):
        max_len = 12
        for cell in ws[chr(64 + c)][1:]:
            try:
                if cell.value:
                    max_len = max(max_len, min(50, len(str(cell.value)) + 2))
            except Exception:
                pass
        ws.column_dimensions[chr(64 + c)].width = max_len


def _fmt_dt(v) -> str:
    return v.strftime("%Y-%m-%d %H:%M:%S") if isinstance(v, datetime) else (str(v) if v else "")


@router.get("/export", summary="安全员导出当天Excel（出场/被拦/返洗/环保异常）")
def api_safety_export(
    target_date: Optional[date] = Query(None, description="目标日期，默认今天"),
    db: Session = Depends(get_db),
):
    data = get_safety_export(db, target_date)
    d = data["date"]
    file_name = f"车辆安全记录_{d}.xlsx"
    file_path = os.path.join(EXPORT_DIR, file_name)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "当日出场车辆"
    headers1 = [
        "序号", "车牌号", "司机", "车辆类型", "所属工地", "装载货物", "重量(吨)",
        "进场时间", "装载完成", "洗轮完成", "检查时间", "出场时间",
        "是否返洗", "返洗次数", "被拦次数", "最后拦截原因", "检查员", "门岗"
    ]
    ws1.append([f"当日出场车辆统计 ({d})  共 {data['counts']['exited']} 辆"])
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers1))
    ws1.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws1.cell(row=1, column=1).alignment = CENTER
    ws1.append(headers1)
    _style_header(ws1, 2, len(headers1))
    for i, r in enumerate(data["exited_records"], 1):
        ws1.append([
            i, r.plate_number, r.driver_name or "", r.vehicle_type, r.construction_site or "",
            r.load_cargo or "", r.load_weight or 0,
            _fmt_dt(r.entry_time), _fmt_dt(r.load_complete_time), _fmt_dt(r.wash_complete_time),
            _fmt_dt(r.inspection_time), _fmt_dt(r.exit_time),
            "是" if r.is_rewashed else "否", r.rewash_count, r.block_count,
            r.last_block_reason or "", r.inspector or "", r.gate_operator or "",
        ])
    if data["exited_records"]:
        _style_body(ws1, 3, 2 + len(data["exited_records"]), len(headers1))
    _autosize(ws1, len(headers1))

    ws2 = wb.create_sheet("被拦记录")
    headers2 = [
        "序号", "车牌号", "拦截时间", "拦截类型", "拦截原因", "是否环保问题",
        "操作员", "是否已解决", "解决时间", "解决方式", "解决操作员"
    ]
    ws2.append([f"当日拦截记录 ({d})  共 {data['counts']['blocked']} 条"])
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers2))
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws2.cell(row=1, column=1).alignment = CENTER
    ws2.append(headers2)
    _style_header(ws2, 2, len(headers2))
    type_map = {
        "no_wash": "未洗轮", "no_tarp_photo": "无篷布照片",
        "inspection_failed": "检查不通过", "environmental": "环保异常"
    }
    for i, b in enumerate(data["block_records"], 1):
        ws2.append([
            i, b.plate_number, _fmt_dt(b.block_time), type_map.get(b.block_type, b.block_type),
            b.block_reason or "", "是" if b.is_environmental_issue else "否",
            b.block_operator or "", "是" if b.resolved else "否",
            _fmt_dt(b.resolve_time), b.resolve_method or "", b.resolve_operator or "",
        ])
    if data["block_records"]:
        _style_body(ws2, 3, 2 + len(data["block_records"]), len(headers2))
    _autosize(ws2, len(headers2))

    ws3 = wb.create_sheet("返洗记录")
    headers3 = [
        "序号", "车牌号", "司机", "进场时间", "出场时间", "返洗次数",
        "装载货物", "重量(吨)", "当前状态"
    ]
    ws3.append([f"当日返洗记录 ({d})  共 {data['counts']['rewashed']} 条"])
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers3))
    ws3.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws3.cell(row=1, column=1).alignment = CENTER
    ws3.append(headers3)
    _style_header(ws3, 2, len(headers3))
    status_map = {
        "entered": "已进场待装载", "loading": "装载中", "loaded": "装载完成待洗轮",
        "washing": "洗轮中", "washed": "洗轮完成待检查", "inspecting": "检查中",
        "blocked": "已拦截", "inspected": "检查通过待出场", "exited": "已出场"
    }
    for i, r in enumerate(data["rewash_records"], 1):
        ws3.append([
            i, r.plate_number, r.driver_name or "", _fmt_dt(r.entry_time),
            _fmt_dt(r.exit_time), r.rewash_count, r.load_cargo or "",
            r.load_weight or 0, status_map.get(r.status, r.status),
        ])
    if data["rewash_records"]:
        _style_body(ws3, 3, 2 + len(data["rewash_records"]), len(headers3))
    _autosize(ws3, len(headers3))

    ws4 = wb.create_sheet("环保异常")
    headers4 = [
        "序号", "车牌号", "异常时间", "异常类型", "异常原因",
        "操作员", "是否已解决", "解决方式"
    ]
    ws4.append([f"当日环保异常记录 ({d})  共 {data['counts']['environmental_issues']} 条"])
    ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers4))
    ws4.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws4.cell(row=1, column=1).alignment = CENTER
    ws4.append(headers4)
    _style_header(ws4, 2, len(headers4))
    for i, b in enumerate(data["environmental_issues"], 1):
        ws4.append([
            i, b.plate_number, _fmt_dt(b.block_time),
            type_map.get(b.block_type, b.block_type), b.block_reason or "",
            b.block_operator or "", "是" if b.resolved else "否", b.resolve_method or "",
        ])
    if data["environmental_issues"]:
        _style_body(ws4, 3, 2 + len(data["environmental_issues"]), len(headers4))
    _autosize(ws4, len(headers4))

    ws5 = wb.create_sheet("汇总统计")
    ws5["A1"] = f"当日汇总统计（{d}）"
    ws5["A1"].font = Font(bold=True, size=14)
    ws5.merge_cells("A1:B1")
    ws5.append(["指标", "数量"])
    _style_header(ws5, 2, 2)
    rows = [
        ("出场车辆", data["counts"]["exited"]),
        ("被拦次数", data["counts"]["blocked"]),
        ("返洗车辆", data["counts"]["rewashed"]),
        ("环保异常", data["counts"]["environmental_issues"]),
    ]
    for row in rows:
        ws5.append(list(row))
    _style_body(ws5, 3, 2 + len(rows), 2)
    ws5.column_dimensions["A"].width = 18
    ws5.column_dimensions["B"].width = 12

    wb.save(file_path)

    return FileResponse(
        file_path,
        filename=file_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/summary", summary="安全员当日汇总（非下载）")
def api_safety_summary(
    target_date: Optional[date] = Query(None),
    db: Session = Depends(get_db),
):
    data = get_safety_export(db, target_date)
    exited_data = []
    for r in data["exited_records"]:
        exited_data.append({
            "id": r.id, "plate_number": r.plate_number, "driver_name": r.driver_name,
            "entry_time": _fmt_dt(r.entry_time), "exit_time": _fmt_dt(r.exit_time),
            "is_rewashed": r.is_rewashed, "rewash_count": r.rewash_count,
            "block_count": r.block_count, "inspector": r.inspector,
        })
    blocks_data = []
    for b in data["block_records"]:
        blocks_data.append({
            "id": b.id, "plate_number": b.plate_number,
            "block_time": _fmt_dt(b.block_time), "block_type": b.block_type,
            "block_reason": b.block_reason, "is_environmental_issue": b.is_environmental_issue,
            "resolved": b.resolved, "resolve_method": b.resolve_method,
        })
    env_data = [b for b in blocks_data if b["is_environmental_issue"]]
    return {
        "date": data["date"],
        "counts": data["counts"],
        "exited_records": exited_data,
        "block_records": blocks_data,
        "environmental_issues": env_data,
    }
