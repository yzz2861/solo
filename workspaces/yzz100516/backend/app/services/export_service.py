from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from typing import List

from ..models.sample import Sample as SampleModel, SampleStatus, SamplePurpose


def get_status_text(status: SampleStatus) -> str:
    status_map = {
        SampleStatus.PENDING_APPROVAL: "待审批",
        SampleStatus.APPROVED: "已审批",
        SampleStatus.OUT: "已出区",
        SampleStatus.RETURNED: "已归还",
        SampleStatus.DESTROYED: "已销毁",
        SampleStatus.OVERDUE: "已超期",
    }
    return status_map.get(status, str(status))


def get_purpose_text(purpose: SamplePurpose) -> str:
    purpose_map = {
        SamplePurpose.RND: "研发",
        SamplePurpose.CUSTOMER: "客户展示",
        SamplePurpose.EXHIBITION: "展览",
        SamplePurpose.TESTING: "测试",
        SamplePurpose.OTHER: "其他",
    }
    return purpose_map.get(purpose, str(purpose))


def format_datetime(dt) -> str:
    if not dt:
        return ""
    if isinstance(dt, str):
        return dt
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def export_compliance_excel(data: dict) -> BytesIO:
    wb = Workbook()
    wb.remove(wb.active)

    headers = [
        "样品编号", "样品名称", "批次号", "用途", "申请人", "部门",
        "数量", "单位", "出区时间", "预计归还时间", "实际归还时间",
        "状态", "审批人", "审批时间", "备注"
    ]

    sheets = [
        ("出区样品清单", data.get("outbound_samples", [])),
        ("归还样品清单", data.get("returned_samples", [])),
        ("销毁样品清单", data.get("destroyed_samples", [])),
        ("缺资料清单", data.get("missing_docs_samples", [])),
    ]

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for sheet_name, samples in sheets:
        ws = wb.create_sheet(title=sheet_name)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        for row_idx, sample in enumerate(samples, 2):
            row_data = [
                sample.sample_no,
                sample.sample_name,
                sample.batch_number,
                get_purpose_text(sample.purpose),
                sample.applicant,
                sample.department or "",
                sample.quantity,
                sample.unit,
                format_datetime(sample.out_time),
                format_datetime(sample.expected_return_time),
                format_datetime(sample.actual_return_time),
                get_status_text(sample.status),
                sample.approver or "",
                format_datetime(sample.approval_time),
                sample.remark or "",
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        ws.row_dimensions[1].height = 25

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_release_order(sample: SampleModel) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "放行单"

    title_font = Font(bold=True, size=18)
    header_font = Font(bold=True, size=11)
    normal_font = Font(size=11)

    ws.merge_cells('A1:F1')
    title_cell = ws.cell(row=1, column=1, value="保税样品出区放行单")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A2:F2')
    no_cell = ws.cell(row=2, column=1, value=f"单号: {sample.sample_no}")
    no_cell.font = normal_font
    no_cell.alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[2].height = 25

    data = [
        ("样品名称", sample.sample_name, "批次号", sample.batch_number),
        ("用途", get_purpose_text(sample.purpose), "数量", f"{sample.quantity} {sample.unit}"),
        ("申请人", sample.applicant, "部门", sample.department or ""),
        ("出区时间", format_datetime(sample.out_time), "预计归还", format_datetime(sample.expected_return_time)),
        ("审批人", sample.approver or "", "审批时间", format_datetime(sample.approval_time)),
    ]

    start_row = 4
    for i, (label1, value1, label2, value2) in enumerate(data):
        row = start_row + i

        ws.cell(row=row, column=1, value=label1).font = header_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='right', vertical='center')

        ws.merge_cells(f'B{row}:C{row}')
        ws.cell(row=row, column=2, value=value1).font = normal_font
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center')

        ws.cell(row=row, column=4, value=label2).font = header_font
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='right', vertical='center')

        ws.merge_cells(f'E{row}:F{row}')
        ws.cell(row=row, column=5, value=value2).font = normal_font
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='left', vertical='center')

        ws.row_dimensions[row].height = 25

    remark_row = start_row + len(data) + 1
    ws.cell(row=remark_row, column=1, value="用途说明").font = header_font
    ws.cell(row=remark_row, column=1).alignment = Alignment(horizontal='right', vertical='top')
    ws.merge_cells(f'B{remark_row}:F{remark_row+2}')
    ws.cell(row=remark_row, column=2, value=sample.purpose_detail or "").font = normal_font
    ws.cell(row=remark_row, column=2).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    sign_row = remark_row + 4
    sign_data = [
        ("仓库交接:", "签字: __________  日期: __________"),
        ("申请人签收:", "签字: __________  日期: __________"),
    ]

    for i, (label, value) in enumerate(sign_data):
        row = sign_row + i
        ws.cell(row=row, column=1, value=label).font = header_font
        ws.merge_cells(f'B{row}:F{row}')
        ws.cell(row=row, column=2, value=value).font = normal_font
        ws.row_dimensions[row].height = 30

    footer_row = sign_row + len(sign_data) + 2
    ws.merge_cells(f'A{footer_row}:F{footer_row}')
    ws.cell(row=footer_row, column=1, value="本放行单一式两联，仓库和申请人各执一联").font = Font(size=9, italic=True)
    ws.cell(row=footer_row, column=1).alignment = Alignment(horizontal='center', vertical='center')

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 12

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
