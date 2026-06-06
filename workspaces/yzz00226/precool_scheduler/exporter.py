import os
import json
import csv
from typing import Dict, List, Any
from datetime import datetime


class ResultExporter:
    def __init__(self, output_dir: str = "./output"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def export(self, result: Dict, format: str = "csv") -> Dict[str, str]:
        format = format.lower()
        if format == "csv":
            return self._export_csv(result)
        elif format == "json":
            return self._export_json(result)
        elif format == "excel" or format == "xlsx":
            return self._export_excel(result)
        else:
            raise ValueError(f"不支持的导出格式: {format}，支持 csv/json/excel")

    def _export_csv(self, result: Dict) -> Dict[str, str]:
        batch_id = result["batch_id"]
        files = {}

        passed_path = os.path.join(self.output_dir, f"{batch_id}_passed.csv")
        self._write_csv(passed_path, result["passed"])
        files["passed"] = passed_path

        exception_path = os.path.join(self.output_dir, f"{batch_id}_exceptions.csv")
        self._write_csv(exception_path, result["exceptions"])
        files["exceptions"] = exception_path

        summary_path = os.path.join(self.output_dir, f"{batch_id}_summary.json")
        with open(summary_path, "w", encoding="utf-8") as f:
            json.dump(result["summary"], f, ensure_ascii=False, indent=2)
        files["summary"] = summary_path

        return files

    def _write_csv(self, path: str, records: List[Dict]):
        if not records:
            fieldnames = self._default_fieldnames()
        else:
            fieldnames = list(records[0].keys())
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for rec in records:
                writer.writerow(rec)

    def _default_fieldnames(self) -> List[str]:
        return [
            "record_id", "batch_id", "source_file", "row_number", "source_row_hash",
            "product_name", "product_type", "quantity", "unit",
            "inbound_date", "target_temp", "current_temp",
            "precool_hours", "precool_start", "precool_end", "precool_room",
            "status", "review_required", "review_reason", "errors",
        ]

    def _export_json(self, result: Dict) -> Dict[str, str]:
        batch_id = result["batch_id"]
        path = os.path.join(self.output_dir, f"{batch_id}_result.json")
        output = {
            "batch_id": batch_id,
            "idempotency_key": result.get("idempotency_key", ""),
            "is_cached": result.get("is_cached", False),
            "message": result.get("message", ""),
            "summary": result["summary"],
            "passed": result["passed"],
            "exceptions": result["exceptions"],
        }
        with open(path, "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=False, indent=2)
        return {"all": path}

    def _export_excel(self, result: Dict) -> Dict[str, str]:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("导出 Excel 需要 openpyxl 库，请运行: pip install openpyxl")

        batch_id = result["batch_id"]
        path = os.path.join(self.output_dir, f"{batch_id}_result.xlsx")

        wb = openpyxl.Workbook()

        ws_summary = wb.active
        ws_summary.title = "汇总摘要"
        summary = result["summary"]
        self._write_summary_sheet(ws_summary, summary)

        ws_passed = wb.create_sheet("通过清单")
        self._write_records_sheet(ws_passed, result["passed"])

        ws_exceptions = wb.create_sheet("异常清单")
        self._write_records_sheet(ws_exceptions, result["exceptions"])

        if result.get("is_cached"):
            ws_note = wb.create_sheet("说明")
            ws_note["A1"] = "本批次为幂等性缓存结果，与历史批次一致。"
            ws_note["A2"] = f"原批次号: {batch_id}"
            ws_note["A3"] = f"幂等键: {result.get('idempotency_key', '')}"

        wb.save(path)
        return {"all": path}

    def _write_summary_sheet(self, ws, summary: Dict):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")

        ws["A1"] = "冷库果蔬预冷排程 - 批次汇总"
        ws.merge_cells("A1:B1")
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        rows = [
            ("批次号", summary.get("batch_id", "")),
            ("生成时间", summary.get("generated_at", "")),
            ("来源文件", ", ".join(summary.get("source_files", []))),
            ("日期范围-开始", summary.get("date_range_start", "")),
            ("日期范围-结束", summary.get("date_range_end", "")),
            ("总记录数", summary.get("total_records", 0)),
            ("通过数量", summary.get("passed_records", 0)),
            ("异常数量", summary.get("exception_records", 0)),
            ("待复核数量", summary.get("review_records", 0)),
            ("预冷库位", ", ".join(summary.get("precool_rooms", []))),
            ("总预冷时长(小时)", summary.get("total_precool_hours", 0)),
            ("幂等性键", summary.get("idempotency_key", "")),
        ]
        for i, (k, v) in enumerate(rows, start=3):
            ws.cell(row=i, column=1, value=k).font = Font(bold=True)
            ws.cell(row=i, column=2, value=v)

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 60

    def _write_records_sheet(self, ws, records: List[Dict]):
        from openpyxl.styles import Font
        if not records:
            ws["A1"] = "无记录"
            return
        headers = list(records[0].keys())
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
        for row_idx, rec in enumerate(records, start=2):
            for col_idx, key in enumerate(headers, start=1):
                val = rec.get(key, "")
                if isinstance(val, list):
                    val = "; ".join(str(v) for v in val)
                ws.cell(row=row_idx, column=col_idx, value=val)
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col_idx)].width = 18
