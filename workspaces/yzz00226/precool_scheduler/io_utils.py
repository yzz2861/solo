import os
import csv
import json
from typing import List, Dict, Any
from .models import SourceRecord


def read_source_files(file_paths: List[str]) -> List[SourceRecord]:
    records = []
    for path in file_paths:
        if not os.path.exists(path):
            raise FileNotFoundError(f"文件不存在: {path}")
        if path.endswith(".csv"):
            records.extend(_read_csv(path))
        elif path.endswith(".json"):
            records.extend(_read_json(path))
        elif path.endswith(".xlsx") or path.endswith(".xls"):
            records.extend(_read_excel(path))
        else:
            raise ValueError(f"不支持的文件格式: {path}，支持 csv/json/excel")
    return records


def _read_csv(path: str) -> List[SourceRecord]:
    records = []
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader, start=2):
            records.append(SourceRecord(
                source_file=os.path.basename(path),
                row_number=i,
                raw_data=dict(row),
            ))
    return records


def _read_json(path: str) -> List[SourceRecord]:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, dict):
        data = [data]
    if not isinstance(data, list):
        raise ValueError(f"JSON 文件格式错误，应为列表或对象: {path}")
    records = []
    for i, item in enumerate(data, start=1):
        records.append(SourceRecord(
            source_file=os.path.basename(path),
            row_number=i,
            raw_data=item,
        ))
    return records


def _read_excel(path: str) -> List[SourceRecord]:
    try:
        import openpyxl
    except ImportError:
        raise ImportError("读取 Excel 需要 openpyxl 库，请运行: pip install openpyxl")
    records = []
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw = {}
        for col_idx, value in enumerate(row):
            key = headers[col_idx] if col_idx < len(headers) else f"col{col_idx}"
            raw[key] = value
        if any(v is not None and str(v).strip() != "" for v in raw.values()):
            records.append(SourceRecord(
                source_file=os.path.basename(path),
                row_number=row_idx,
                raw_data=raw,
            ))
    return records
