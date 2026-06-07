import os
import csv
import json
from typing import List, Dict, Any
from .models import ValveRecord, ProcessSummary, ExportFormat, RecordStatus


def export_results(
    passed_records: List[ValveRecord],
    exception_records: List[ValveRecord],
    bad_rows: List[Dict],
    summary: ProcessSummary,
    config,
) -> List[str]:
    output_dir = config.output_dir
    fmt = config.export_format
    batch_id = config.batch_id
    
    os.makedirs(output_dir, exist_ok=True)
    
    if fmt == ExportFormat.CSV:
        output_files = _build_csv_filenames(output_dir, batch_id)
    elif fmt == ExportFormat.JSON:
        output_files = [os.path.join(output_dir, f"{batch_id}_results.json")]
    elif fmt == ExportFormat.EXCEL:
        output_files = [os.path.join(output_dir, f"{batch_id}_results.xlsx")]
    else:
        output_files = []
    
    summary.output_files = output_files
    
    if fmt == ExportFormat.CSV:
        _write_csv_files(passed_records, exception_records, bad_rows, summary, output_dir, batch_id)
    elif fmt == ExportFormat.JSON:
        _write_json_file(passed_records, exception_records, bad_rows, summary, output_dir, batch_id)
    elif fmt == ExportFormat.EXCEL:
        _write_excel_file(passed_records, exception_records, bad_rows, summary, output_dir, batch_id)
    
    return output_files


def update_summary_output(summary: ProcessSummary, config):
    fmt = config.export_format
    output_dir = config.output_dir
    batch_id = config.batch_id
    
    if fmt == ExportFormat.CSV:
        summary_file = os.path.join(output_dir, f"{batch_id}_summary.json")
        if os.path.exists(summary_file):
            with open(summary_file, "w", encoding="utf-8") as f:
                json.dump(summary.to_dict(), f, ensure_ascii=False, indent=2)
    elif fmt == ExportFormat.JSON:
        json_file = os.path.join(output_dir, f"{batch_id}_results.json")
        if os.path.exists(json_file):
            with open(json_file, "r", encoding="utf-8") as f:
                data = json.load(f)
            data["summary"] = summary.to_dict()
            with open(json_file, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
    elif fmt == ExportFormat.EXCEL:
        try:
            import openpyxl
        except ImportError:
            return
        xlsx_file = os.path.join(output_dir, f"{batch_id}_results.xlsx")
        if os.path.exists(xlsx_file):
            wb = openpyxl.load_workbook(xlsx_file)
            if "汇总" in wb.sheetnames:
                ws = wb["汇总"]
                summary_dict = summary.to_dict()
                for row_idx, (k, v) in enumerate(summary_dict.items(), start=1):
                    val = v if not isinstance(v, (list, dict)) else json.dumps(v, ensure_ascii=False)
                    ws.cell(row=row_idx, column=2, value=str(val))
            wb.save(xlsx_file)


def _build_csv_filenames(output_dir, batch_id) -> List[str]:
    return [
        os.path.join(output_dir, f"{batch_id}_passed.csv"),
        os.path.join(output_dir, f"{batch_id}_exceptions.csv"),
        os.path.join(output_dir, f"{batch_id}_bad_rows.csv"),
        os.path.join(output_dir, f"{batch_id}_summary.json"),
    ]


def _write_csv_files(passed, exceptions, bad_rows, summary, output_dir, batch_id):
    files = _build_csv_filenames(output_dir, batch_id)
    
    _write_csv(files[0], [r.to_dict() for r in passed])
    _write_csv(files[1], [r.to_dict() for r in exceptions])
    
    if bad_rows:
        _write_csv(files[2], bad_rows)
    else:
        with open(files[2], "w", encoding="utf-8-sig", newline="") as f:
            f.write("")
    
    with open(files[3], "w", encoding="utf-8") as f:
        json.dump(summary.to_dict(), f, ensure_ascii=False, indent=2)


def _write_csv(filepath: str, rows: List[Dict[str, Any]]):
    if not rows:
        with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
            f.write("")
        return
    
    fieldnames = list(rows[0].keys())
    with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _write_json_file(passed, exceptions, bad_rows, summary, output_dir, batch_id):
    data = {
        "batch_id": batch_id,
        "summary": summary.to_dict(),
        "passed_records": [r.to_dict() for r in passed],
        "exception_records": [r.to_dict() for r in exceptions],
        "bad_rows": bad_rows,
    }
    
    out_file = os.path.join(output_dir, f"{batch_id}_results.json")
    with open(out_file, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _write_excel_file(passed, exceptions, bad_rows, summary, output_dir, batch_id):
    try:
        import openpyxl
    except ImportError:
        raise ImportError("导出Excel需要安装 openpyxl: pip install openpyxl")
    
    out_file = os.path.join(output_dir, f"{batch_id}_results.xlsx")
    
    wb = openpyxl.Workbook()
    
    ws_summary = wb.active
    ws_summary.title = "汇总"
    summary_dict = summary.to_dict()
    row_idx = 1
    for k, v in summary_dict.items():
        ws_summary.cell(row=row_idx, column=1, value=str(k))
        val = v if not isinstance(v, (list, dict)) else json.dumps(v, ensure_ascii=False)
        ws_summary.cell(row=row_idx, column=2, value=str(val))
        row_idx += 1
    
    if passed:
        ws_passed = wb.create_sheet("通过清单")
        _fill_excel_sheet(ws_passed, [r.to_dict() for r in passed])
    
    if exceptions:
        ws_exc = wb.create_sheet("异常清单")
        _fill_excel_sheet(ws_exc, [r.to_dict() for r in exceptions])
    
    if bad_rows:
        ws_bad = wb.create_sheet("坏行")
        _fill_excel_sheet(ws_bad, bad_rows)
    
    wb.save(out_file)


def _fill_excel_sheet(ws, rows: List[Dict]):
    if not rows:
        return
    headers = list(rows[0].keys())
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=str(header))
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            val = row.get(header, "")
            if isinstance(val, (list, dict)):
                val = json.dumps(val, ensure_ascii=False)
            ws.cell(row=row_idx, column=col_idx, value=str(val))


def preview_results(passed, exceptions, bad_rows, summary, max_show: int = 10) -> str:
    lines = []
    lines.append("=" * 60)
    lines.append("  处理结果预览 (DRY-RUN 模式)")
    lines.append("=" * 60)
    lines.append(f"批次号: {summary.batch_id}")
    lines.append(f"总记录数: {summary.total_records}")
    lines.append(f"通过: {summary.passed_count}")
    lines.append(f"异常: {summary.exception_count}")
    lines.append(f"坏行: {summary.bad_count}")
    lines.append("")
    
    if passed:
        lines.append(f"--- 通过清单 (显示前{min(max_show, len(passed))}条) ---")
        for r in passed[:max_show]:
            date_str = r.operation_date.strftime("%Y-%m-%d") if r.operation_date else "N/A"
            lines.append(f"  [{r.valve_id}] {r.valve_name} - {r.operation_type} - {date_str}")
        if len(passed) > max_show:
            lines.append(f"  ... 还有 {len(passed) - max_show} 条")
        lines.append("")
    
    if exceptions:
        lines.append(f"--- 异常清单 (显示前{min(max_show, len(exceptions))}条) ---")
        for r in exceptions[:max_show]:
            reasons = "; ".join(r.exception_reasons)
            exc_types = ",".join([e.value for e in r.exception_types])
            lines.append(f"  [{r.valve_id}] 异常类型: {exc_types}")
            lines.append(f"    原因: {reasons}")
        if len(exceptions) > max_show:
            lines.append(f"  ... 还有 {len(exceptions) - max_show} 条")
        lines.append("")
    
    if bad_rows:
        lines.append(f"--- 坏行清单 (显示前{min(max_show, len(bad_rows))}条) ---")
        for br in bad_rows[:max_show]:
            lines.append(f"  行 {br['row_number']} ({br['source_file']}): {br['error']}")
        if len(bad_rows) > max_show:
            lines.append(f"  ... 还有 {len(bad_rows) - max_show} 条")
        lines.append("")
    
    lines.append("--- 异常分类统计 ---")
    for k, v in summary.exception_breakdown.items():
        lines.append(f"  {k}: {v}")
    lines.append("")
    
    lines.append("=" * 60)
    lines.append("DRY-RUN 模式: 未生成任何输出文件")
    lines.append("=" * 60)
    
    return "\n".join(lines)
