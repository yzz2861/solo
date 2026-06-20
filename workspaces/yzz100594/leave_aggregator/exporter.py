import csv
import json
from datetime import date
from pathlib import Path
from typing import List, Dict, Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import LeaveRecord, LeaveType, Anomaly, SourceType, ClassSummary, SickAlert
from .statistics import StatisticsAnalyzer


class ExcelExporter:
    def __init__(
        self,
        all_records: List[LeaveRecord],
        merged_records: List[LeaveRecord],
        anomalies: List[Anomaly],
        analyzer: StatisticsAnalyzer,
        class_name: str = "未指定班级",
    ):
        self.all_records = all_records
        self.merged_records = merged_records
        self.anomalies = anomalies
        self.analyzer = analyzer
        self.class_name = class_name

        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.warn_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        self.error_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        self.info_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        self.sick_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        self.center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def export_all(self, output_dir: str) -> Dict[str, str]:
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)

        files = {}

        excel_path = output_path / f"{self.class_name}_请假汇总_{date.today().strftime('%Y%m%d')}.xlsx"
        self._export_excel(excel_path)
        files["excel"] = str(excel_path)

        class_list_path = output_path / f"{self.class_name}_班级清单_{date.today().strftime('%Y%m%d')}.csv"
        self._export_class_list_csv(class_list_path)
        files["class_list_csv"] = str(class_list_path)

        detail_path = output_path / f"{self.class_name}_明细记录_{date.today().strftime('%Y%m%d')}.csv"
        self._export_detail_csv(detail_path)
        files["detail_csv"] = str(detail_path)

        doctor_path = output_path / f"{self.class_name}_校医名单_{date.today().strftime('%Y%m%d')}.csv"
        self._export_doctor_csv(doctor_path)
        files["doctor_csv"] = str(doctor_path)

        anomaly_path = output_path / f"{self.class_name}_异常报告_{date.today().strftime('%Y%m%d')}.csv"
        self._export_anomaly_csv(anomaly_path)
        files["anomaly_csv"] = str(anomaly_path)

        summary_json = output_path / f"{self.class_name}_统计数据_{date.today().strftime('%Y%m%d')}.json"
        self._export_summary_json(summary_json)
        files["summary_json"] = str(summary_json)

        return files

    def _style_header(self, ws, row: int, max_col: int):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border

    def _style_data_range(self, ws, start_row: int, end_row: int, max_col: int):
        for row in range(start_row, end_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = self.thin_border

    def _autosize_columns(self, ws, max_col: int, min_width: int = 10, max_width: int = 40):
        for col in range(1, max_col + 1):
            max_len = min_width
            for cell in ws[get_column_letter(col)]:
                if cell.value:
                    cell_len = len(str(cell.value))
                    if cell_len > max_len:
                        max_len = min(cell_len, max_width)
            ws.column_dimensions[get_column_letter(col)].width = max_len + 2

    def _export_excel(self, excel_path: Path):
        wb = Workbook()
        wb.remove(wb.active)

        self._sheet_summary(wb)
        self._sheet_class_list(wb)
        self._sheet_detail(wb)
        self._sheet_raw_sources(wb)
        self._sheet_anomaly(wb)
        self._sheet_daily(wb)
        self._sheet_sick_doctor(wb)
        self._sheet_alerts(wb)

        wb.save(excel_path)

    def _sheet_summary(self, wb: Workbook):
        ws = wb.create_sheet("0-总览", 0)

        summary = self.analyzer.get_class_summary()
        type_stats = self.analyzer.get_type_stats()

        ws["A1"] = f"【{self.class_name}】学生请假汇总报告"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
        ws.merge_cells("A1:E1")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        ws["A3"] = "报告日期:"
        ws["B3"] = date.today().strftime("%Y年%m月%d日")
        ws["A4"] = "统计天数:"
        ws["B4"] = f"{summary.total_days} 天"
        ws["A5"] = "请假总人次:"
        ws["B5"] = summary.total_leave_count
        ws["A6"] = "涉及学生:"
        ws["B6"] = f"{len(summary.students_with_leave)} 人"
        ws["A7"] = "异常记录:"
        ws["B7"] = f"{summary.anomaly_count} 条"

        ws["A9"] = "按请假类型统计:"
        headers = ["类型", "次数", "人数", "占比"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=10, column=col, value=h)
        self._style_header(ws, 10, len(headers))

        row = 11
        total = max(summary.total_leave_count, 1)
        for t in [LeaveType.SICK.value, LeaveType.PERSONAL.value, LeaveType.PUBLIC.value, LeaveType.OTHER.value]:
            if t in type_stats:
                data = type_stats[t]
                ws.cell(row=row, column=1, value=t)
                ws.cell(row=row, column=2, value=data["count"])
                ws.cell(row=row, column=3, value=data["student_count"])
                ws.cell(row=row, column=4, value=f"{data['count'] / total * 100:.1f}%")
                if t == LeaveType.SICK.value:
                    for col in range(1, 5):
                        ws.cell(row=row, column=col).fill = self.sick_fill
                row += 1

        self._style_data_range(ws, 11, row - 1, len(headers))
        self._autosize_columns(ws, len(headers))

    def _sheet_class_list(self, wb: Workbook):
        ws = wb.create_sheet("1-班级清单", 1)

        headers = ["序号", "姓名", "请假天数(人次)", "病假", "事假", "公假", "其他", "最近一次请假", "备注"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        self._style_header(ws, 1, len(headers))

        student_stats = self.analyzer.get_student_stats()
        row = 2
        for idx, (student, data) in enumerate(student_stats.items(), 1):
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=student)
            ws.cell(row=row, column=3, value=data["total"])
            ws.cell(row=row, column=4, value=data[LeaveType.SICK.value])
            ws.cell(row=row, column=5, value=data[LeaveType.PERSONAL.value])
            ws.cell(row=row, column=6, value=data[LeaveType.PUBLIC.value])
            ws.cell(row=row, column=7, value=data[LeaveType.OTHER.value])

            if data["details"]:
                last_detail = data["details"][-1]
                ws.cell(
                    row=row,
                    column=8,
                    value=f"{last_detail[0].strftime('%m-%d')} {last_detail[1]} {last_detail[2]}",
                )

                has_anomaly = any(a for a in self.anomalies if any(r.student_name == student for r in a.related_records))
                if has_anomaly:
                    ws.cell(row=row, column=9, value="存在异常，请核对")
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row, column=col).fill = self.warn_fill

            row += 1

        self._style_data_range(ws, 2, row - 1, len(headers))
        self._autosize_columns(ws, len(headers))

    def _sheet_detail(self, wb: Workbook):
        ws = wb.create_sheet("2-合并明细", 2)

        headers = [
            "序号", "日期", "星期", "姓名", "节次", "请假类型",
            "请假原因", "数据来源", "原始来源数", "处理状态", "记录ID",
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        self._style_header(ws, 1, len(headers))

        merged_sorted = sorted(self.merged_records, key=lambda r: (r.record_date, r.student_name, r.period))
        weekday_map = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]

        row = 2
        for idx, r in enumerate(merged_sorted, 1):
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=r.record_date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=3, value=weekday_map[r.record_date.weekday()])
            ws.cell(row=row, column=4, value=r.student_name)
            ws.cell(row=row, column=5, value=r.period)
            ws.cell(row=row, column=6, value=r.leave_type.value)
            ws.cell(row=row, column=7, value=r.reason)
            ws.cell(row=row, column=8, value=r.source.value)

            key = r.merge_key()
            from collections import defaultdict
            if not hasattr(self, "_source_counts"):
                self._source_counts = defaultdict(set)
                for orig in self.all_records:
                    self._source_counts[orig.merge_key()].add(orig.source.value)
            source_count = len(self._source_counts.get(key, set()))
            ws.cell(row=row, column=9, value=source_count)

            status = "已合并" if source_count > 1 else "正常"
            ws.cell(row=row, column=10, value=status)
            ws.cell(row=row, column=11, value=r.record_id)

            if r.leave_type == LeaveType.SICK:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = self.sick_fill

            if source_count > 1:
                ws.cell(row=row, column=10).fill = self.info_fill

            row += 1

        self._style_data_range(ws, 2, row - 1, len(headers))
        self._autosize_columns(ws, len(headers))

    def _sheet_raw_sources(self, wb: Workbook):
        ws = wb.create_sheet("3-原始来源", 3)

        headers = [
            "序号", "记录ID", "数据来源", "日期", "姓名", "节次",
            "请假类型", "请假原因", "老师", "联系方式", "原始内容摘要",
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        self._style_header(ws, 1, len(headers))

        all_sorted = sorted(self.all_records, key=lambda r: (r.record_date, r.student_name, r.source.value))

        row = 2
        for idx, r in enumerate(all_sorted, 1):
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=r.record_id)
            ws.cell(row=row, column=3, value=r.source.value)
            ws.cell(row=row, column=4, value=r.record_date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=5, value=r.student_name)
            ws.cell(row=row, column=6, value=r.period)
            ws.cell(row=row, column=7, value=r.leave_type.value)
            ws.cell(row=row, column=8, value=r.reason)
            ws.cell(row=row, column=9, value=r.teacher or "")
            ws.cell(row=row, column=10, value=r.contact or "")
            ws.cell(row=row, column=11, value=r.raw_content[:80] if r.raw_content else "")

            if r.source == SourceType.SMS:
                ws.cell(row=row, column=3).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            elif r.source == SourceType.PAPER:
                ws.cell(row=row, column=3).fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            elif r.source == SourceType.ABSENCE:
                ws.cell(row=row, column=3).fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

            row += 1

        self._style_data_range(ws, 2, row - 1, len(headers))
        self._autosize_columns(ws, len(headers), max_width=50)

    def _sheet_anomaly(self, wb: Workbook):
        ws = wb.create_sheet("4-异常报告", 4)

        headers = [
            "序号", "异常类型", "严重级别", "异常描述",
            "涉及学生", "涉及记录数", "处理建议",
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        self._style_header(ws, 1, len(headers))

        severity_order = {"error": 1, "warning": 2, "info": 3}
        sorted_anomalies = sorted(self.anomalies, key=lambda a: (severity_order.get(a.severity, 9), a.anomaly_type.value))

        row = 2
        for idx, anomaly in enumerate(sorted_anomalies, 1):
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=anomaly.anomaly_type.value)
            ws.cell(row=row, column=3, value=anomaly.severity.upper())
            ws.cell(row=row, column=4, value=anomaly.description)
            ws.cell(row=row, column=5, value="、".join(sorted(set(r.student_name for r in anomaly.related_records))))
            ws.cell(row=row, column=6, value=len(anomaly.related_records))
            ws.cell(row=row, column=7, value=anomaly.suggestions)

            if anomaly.severity == "error":
                fill = self.error_fill
            elif anomaly.severity == "warning":
                fill = self.warn_fill
            else:
                fill = self.info_fill
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = fill

            row += 1

        self._style_data_range(ws, 2, row - 1, len(headers))
        self._autosize_columns(ws, len(headers), max_width=60)

    def _sheet_daily(self, wb: Workbook):
        ws = wb.create_sheet("5-每日统计", 5)

        headers = ["日期", "星期", "总人次", "病假", "事假", "公假", "其他", "涉及人数", "学生名单"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        self._style_header(ws, 1, len(headers))

        daily_stats = self.analyzer.get_daily_stats()
        weekday_map = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]

        row = 2
        for d, data in sorted(daily_stats.items()):
            ws.cell(row=row, column=1, value=d.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=2, value=weekday_map[d.weekday()])
            ws.cell(row=row, column=3, value=data["total"])
            ws.cell(row=row, column=4, value=data[LeaveType.SICK.value])
            ws.cell(row=row, column=5, value=data[LeaveType.PERSONAL.value])
            ws.cell(row=row, column=6, value=data[LeaveType.PUBLIC.value])
            ws.cell(row=row, column=7, value=data[LeaveType.OTHER.value])
            ws.cell(row=row, column=8, value=data["student_count"])
            ws.cell(row=row, column=9, value="、".join(data["student_names"]))

            if data[LeaveType.SICK.value] >= 5:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = self.warn_fill

            row += 1

        self._style_data_range(ws, 2, row - 1, len(headers))
        self._autosize_columns(ws, len(headers), max_width=80)

    def _sheet_sick_doctor(self, wb: Workbook):
        ws = wb.create_sheet("6-校医病假名单", 6)

        sick_data = self.analyzer.get_sick_list_for_doctor()

        headers = ["序号"] + list(sick_data[0].keys()) if sick_data else ["序号", "日期", "姓名", "节次", "症状/原因", "来源", "记录ID"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        self._style_header(ws, 1, len(headers))

        row = 2
        for idx, item in enumerate(sick_data, 1):
            ws.cell(row=row, column=1, value=idx)
            for col_idx, key in enumerate(list(item.keys()), 2):
                ws.cell(row=row, column=col_idx, value=item[key])
            row += 1

        if row > 2:
            self._style_data_range(ws, 2, row - 1, len(headers))
        self._autosize_columns(ws, len(headers))

    def _sheet_alerts(self, wb: Workbook):
        ws = wb.create_sheet("7-病假警报", 7)

        alerts = self.analyzer.check_sick_alerts()

        headers = ["序号", "警报日期", "当日病假人数", "阈值", "超出", "涉及学生", "警报信息"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        self._style_header(ws, 1, len(headers))

        row = 2
        for idx, alert in enumerate(alerts, 1):
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=alert.date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=3, value=alert.sick_count)
            ws.cell(row=row, column=4, value=alert.threshold)
            ws.cell(row=row, column=5, value=alert.sick_count - alert.threshold)
            ws.cell(row=row, column=6, value="、".join(alert.student_names))
            ws.cell(row=row, column=7, value=alert.message)

            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = self.error_fill

            row += 1

        if row == 2:
            ws["A2"] = "本周病假人数均在正常范围内，无警报。"
            ws["A2"].fill = self.info_fill
            ws.merge_cells("A2:G2")

        if row > 2:
            self._style_data_range(ws, 2, row - 1, len(headers))
        self._autosize_columns(ws, len(headers), max_width=60)

    def _export_class_list_csv(self, path: Path):
        student_stats = self.analyzer.get_student_stats()

        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["班级", "序号", "姓名", "请假总人次", "病假", "事假", "公假", "其他"])

            for idx, (student, data) in enumerate(student_stats.items(), 1):
                writer.writerow([
                    self.class_name,
                    idx,
                    student,
                    data["total"],
                    data[LeaveType.SICK.value],
                    data[LeaveType.PERSONAL.value],
                    data[LeaveType.PUBLIC.value],
                    data[LeaveType.OTHER.value],
                ])

    def _export_detail_csv(self, path: Path):
        all_sorted = sorted(self.all_records, key=lambda r: (r.record_date, r.student_name, r.source.value))

        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow([
                "记录ID", "班级", "数据来源", "日期", "姓名", "节次",
                "请假类型", "请假原因", "老师", "联系方式", "原始内容",
            ])

            for r in all_sorted:
                writer.writerow([
                    r.record_id,
                    self.class_name,
                    r.source.value,
                    r.record_date.strftime("%Y-%m-%d"),
                    r.student_name,
                    r.period,
                    r.leave_type.value,
                    r.reason,
                    r.teacher or "",
                    r.contact or "",
                    r.raw_content,
                ])

    def _export_doctor_csv(self, path: Path):
        sick_data = self.analyzer.get_sick_list_for_doctor()

        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["班级", "序号", "日期", "姓名", "节次", "症状/原因", "来源", "记录ID"])

            for idx, item in enumerate(sick_data, 1):
                writer.writerow([
                    self.class_name,
                    idx,
                    item["日期"],
                    item["姓名"],
                    item["节次"],
                    item["症状/原因"],
                    item["来源"],
                    item["记录ID"],
                ])

    def _export_anomaly_csv(self, path: Path):
        severity_order = {"error": 1, "warning": 2, "info": 3}
        sorted_anomalies = sorted(self.anomalies, key=lambda a: (severity_order.get(a.severity, 9), a.anomaly_type.value))

        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow([
                "班级", "序号", "异常类型", "严重级别", "异常描述",
                "涉及学生", "涉及记录数", "涉及记录ID", "处理建议",
            ])

            for idx, anomaly in enumerate(sorted_anomalies, 1):
                writer.writerow([
                    self.class_name,
                    idx,
                    anomaly.anomaly_type.value,
                    anomaly.severity.upper(),
                    anomaly.description,
                    "、".join(sorted(set(r.student_name for r in anomaly.related_records))),
                    len(anomaly.related_records),
                    "、".join(r.record_id for r in anomaly.related_records),
                    anomaly.suggestions,
                ])

    def _export_summary_json(self, path: Path):
        summary = self.analyzer.get_class_summary()
        type_stats = self.analyzer.get_type_stats()
        daily_stats = self.analyzer.get_daily_stats()
        alerts = self.analyzer.check_sick_alerts()

        data = {
            "班级": self.class_name,
            "生成日期": date.today().isoformat(),
            "汇总信息": {
                "统计天数": summary.total_days,
                "请假总人次": summary.total_leave_count,
                "病假人次": summary.sick_count,
                "事假人次": summary.personal_count,
                "公假人次": summary.public_count,
                "其他人次": summary.other_count,
                "异常记录数": summary.anomaly_count,
                "涉及学生数": len(summary.students_with_leave),
                "学生名单": summary.students_with_leave,
            },
            "按类型统计": {
                t: {
                    "人次": d["count"],
                    "人数": d["student_count"],
                    "学生名单": d["student_names"],
                }
                for t, d in type_stats.items()
            },
            "每日统计": {
                d.isoformat(): {
                    "总人次": data["total"],
                    "病假": data[LeaveType.SICK.value],
                    "事假": data[LeaveType.PERSONAL.value],
                    "公假": data[LeaveType.PUBLIC.value],
                    "其他": data[LeaveType.OTHER.value],
                    "人数": data["student_count"],
                    "学生名单": data["student_names"],
                }
                for d, data in daily_stats.items()
            },
            "异常报告": [
                {
                    "类型": a.anomaly_type.value,
                    "级别": a.severity,
                    "描述": a.description,
                    "涉及学生": list(set(r.student_name for r in a.related_records)),
                    "处理建议": a.suggestions,
                }
                for a in self.anomalies
            ],
            "病假警报": [
                {
                    "日期": a.date.isoformat(),
                    "人数": a.sick_count,
                    "阈值": a.threshold,
                    "学生名单": a.student_names,
                    "信息": a.message,
                }
                for a in alerts
            ],
        }

        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
