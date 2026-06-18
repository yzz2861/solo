"""联系人清单回写：按邮箱匹配，把退信分类结果写回 CSV/Excel 联系人表。

新增列：
- 退信分类
- 退信次数
- 最近退信时间
- 退信原因
- 错误码
- 是否建议清理
- 是否需客户经理联系
"""
from __future__ import annotations

import csv
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from .models import BounceRecord


EMAIL_HEADER_CANDIDATES = [
    "email", "e-mail", "mail", "邮箱", "电子邮件", "电子邮件地址",
    "联系人邮箱", "收件人", "收件邮箱", "工作邮箱", "联系邮箱", "客户邮箱",
]


def _detect_email_column(header: List[str]) -> Optional[int]:
    normalized = {h.strip().lower(): i for i, h in enumerate(header)}
    for cand in EMAIL_HEADER_CANDIDATES:
        if cand.lower() in normalized:
            return normalized[cand.lower()]
    regex = re.compile(r"e?.?mail|邮箱|邮\s*箱", re.IGNORECASE)
    for i, h in enumerate(header):
        if regex.search(h):
            return i
    for i, h in enumerate(header):
        sample = h.strip()
        if re.search(r"@[^\s]+\.[^\s]+", sample):
            return i
    return None


def _build_lookup(records: Iterable[BounceRecord]) -> Dict[str, BounceRecord]:
    return {r.recipient.lower(): r for r in records}


def _fmt_dt(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


def _record_to_fields(r: BounceRecord) -> Dict[str, str]:
    return {
        "退信分类": r.category.value,
        "退信次数": str(r.bounce_count),
        "最近退信时间": _fmt_dt(r.last_bounce_time),
        "退信原因": r.reason_text,
        "错误码": r.reason_code,
        "是否建议清理": "是" if r.needs_cleanup else "",
        "是否需客户经理联系": "是" if r.needs_contact_manager else "",
        "原活动": r.original_campaign,
    }


EXTRA_FIELDS = [
    "退信分类",
    "退信次数",
    "最近退信时间",
    "退信原因",
    "错误码",
    "是否建议清理",
    "是否需客户经理联系",
    "原活动",
]


def write_back_csv(
    contacts_csv: str,
    records: Iterable[BounceRecord],
    output_csv: Optional[str] = None,
) -> Tuple[str, Dict[str, int]]:
    """回写 CSV 联系人清单；返回 (输出路径, 匹配统计)。"""
    with open(contacts_csv, "r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        rows = list(reader)
    if not rows:
        raise ValueError("联系人 CSV 为空")
    header = rows[0]
    data_rows = rows[1:]

    email_idx = _detect_email_column(header)
    if email_idx is None:
        raise ValueError(
            f"未能从表头自动识别邮箱列，请手动确认。表头：{header}"
        )

    lookup = _build_lookup(records)
    matched = 0
    hard_or_black = 0
    manual = 0
    new_header = list(header) + [f for f in EXTRA_FIELDS if f not in header]
    extra_indices = {}
    for f in EXTRA_FIELDS:
        if f in new_header:
            extra_indices[f] = new_header.index(f)

    new_rows = [new_header]
    for row in data_rows:
        padded = row + [""] * (len(new_header) - len(row))
        email = (padded[email_idx] or "").strip().lower()
        if email and email in lookup:
            matched += 1
            r = lookup[email]
            fields = _record_to_fields(r)
            if r.needs_cleanup:
                hard_or_black += 1
            if r.needs_contact_manager:
                manual += 1
            for fname, fidx in extra_indices.items():
                if fidx < len(padded):
                    padded[fidx] = fields.get(fname, padded[fidx])
        new_rows.append(padded)

    out = output_csv or _add_suffix(contacts_csv, "_with_bounce")
    os.makedirs(os.path.dirname(os.path.abspath(out)) or ".", exist_ok=True)
    with open(out, "w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerows(new_rows)

    stats = {
        "联系人总数": len(data_rows),
        "匹配到退信": matched,
        "需清理(硬退+黑名单)": hard_or_black,
        "需客户经理联系": manual,
    }
    return out, stats


def write_back_xlsx(
    contacts_xlsx: str,
    records: Iterable[BounceRecord],
    output_xlsx: Optional[str] = None,
    sheet_name: Optional[str] = None,
) -> Tuple[str, Dict[str, int]]:
    try:
        from openpyxl import load_workbook, Workbook  # type: ignore
    except ImportError:
        raise ImportError("请安装 openpyxl 以支持 Excel: pip install openpyxl")

    wb = load_workbook(contacts_xlsx)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    data_rows: List[list] = []
    for row in rows_iter:
        data_rows.append(list(row))

    email_idx = _detect_email_column(header)
    if email_idx is None:
        raise ValueError(
            f"未能从表头自动识别邮箱列，请手动确认。表头：{header}"
        )

    lookup = _build_lookup(records)
    extra_start = ws.max_column + 1
    for offset, fname in enumerate(EXTRA_FIELDS):
        ws.cell(row=1, column=extra_start + offset, value=fname)

    matched = 0
    hard_or_black = 0
    manual = 0
    for r_idx, row in enumerate(data_rows, start=2):
        cell_val = ws.cell(row=r_idx, column=email_idx + 1).value
        email = str(cell_val or "").strip().lower()
        if email and email in lookup:
            matched += 1
            r = lookup[email]
            fields = _record_to_fields(r)
            if r.needs_cleanup:
                hard_or_black += 1
            if r.needs_contact_manager:
                manual += 1
            for offset, fname in enumerate(EXTRA_FIELDS):
                ws.cell(
                    row=r_idx,
                    column=extra_start + offset,
                    value=fields.get(fname, ""),
                )

    out = output_xlsx or _add_suffix(contacts_xlsx, "_with_bounce")
    os.makedirs(os.path.dirname(os.path.abspath(out)) or ".", exist_ok=True)
    wb.save(out)
    stats = {
        "联系人总数": len(data_rows),
        "匹配到退信": matched,
        "需清理(硬退+黑名单)": hard_or_black,
        "需客户经理联系": manual,
    }
    return out, stats


def _add_suffix(path: str, suffix: str) -> str:
    p = Path(path)
    return str(p.parent / f"{p.stem}{suffix}{p.suffix}")


def write_back(
    contacts_path: str,
    records: Iterable[BounceRecord],
    output_path: Optional[str] = None,
    sheet_name: Optional[str] = None,
) -> Tuple[str, Dict[str, int]]:
    """根据扩展名自动选择 CSV / Excel 回写。"""
    ext = Path(contacts_path).suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        return write_back_xlsx(contacts_path, records, output_path, sheet_name)
    return write_back_csv(contacts_path, records, output_path)
