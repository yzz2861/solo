"""位置对照表解析器 - 将原始文件名映射到项目、楼栋、楼层、部位信息"""

import csv
import os
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple
from pathlib import Path

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@dataclass
class LocationInfo:
    """位置信息数据类"""
    project: str = ""
    building: str = ""
    floor: str = ""
    position: str = ""
    original_keyword: str = ""

    def is_complete(self) -> bool:
        """检查位置信息是否完整（所有字段非空）"""
        return all([self.project, self.building, self.floor, self.position])

    def is_partial(self) -> bool:
        """检查是否有部分位置信息"""
        return any([self.project, self.building, self.floor, self.position])


@dataclass
class LocationMap:
    """位置映射表"""
    exact_matches: Dict[str, LocationInfo] = field(default_factory=dict)
    fuzzy_matches: List[Tuple[str, LocationInfo]] = field(default_factory=list)

    def lookup(self, filename: str) -> Optional[LocationInfo]:
        """
        根据文件名查找位置信息
        先精确匹配，再模糊匹配
        """
        base_name = os.path.splitext(filename)[0].upper()

        if base_name in self.exact_matches:
            return self.exact_matches[base_name]

        for keyword, info in self.fuzzy_matches:
            if keyword.upper() in base_name:
                return info

        return None

    def filter_by_building(self, building: str) -> "LocationMap":
        """按楼栋筛选位置映射"""
        filtered = LocationMap()
        building_upper = building.upper()

        for key, info in self.exact_matches.items():
            if info.building.upper() == building_upper:
                filtered.exact_matches[key] = info

        for keyword, info in self.fuzzy_matches:
            if info.building.upper() == building_upper:
                filtered.fuzzy_matches.append((keyword, info))

        return filtered

    def get_all_buildings(self) -> List[str]:
        """获取所有楼栋列表"""
        buildings = set()
        for info in self.exact_matches.values():
            if info.building:
                buildings.add(info.building)
        for _, info in self.fuzzy_matches:
            if info.building:
                buildings.add(info.building)
        return sorted(list(buildings))


class LocationMapper:
    """位置对照表解析器"""

    REQUIRED_COLUMNS = ["项目", "楼栋", "楼层", "部位"]
    KEYWORD_COLUMNS = ["原始文件名", "文件名", "关键词", "照片名"]

    def __init__(self, map_file: str):
        self.map_file = Path(map_file)
        if not self.map_file.exists():
            raise FileNotFoundError(f"位置对照表不存在: {map_file}")

    def parse(self) -> LocationMap:
        """解析位置对照表文件"""
        ext = self.map_file.suffix.lower()

        if ext in ['.xlsx', '.xls']:
            return self._parse_excel()
        elif ext == '.csv':
            return self._parse_csv()
        else:
            raise ValueError(f"不支持的文件格式: {ext}，请使用 .xlsx 或 .csv")

    def _parse_excel(self) -> LocationMap:
        """解析 Excel 格式的对照表"""
        if not HAS_OPENPYXL:
            raise ImportError("请安装 openpyxl: pip install openpyxl")

        wb = load_workbook(self.map_file, read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("对照表为空")

        headers = [str(h).strip() if h else "" for h in rows[0]]
        col_indices = self._validate_and_get_columns(headers)

        location_map = LocationMap()

        for row in rows[1:]:
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            try:
                info = self._row_to_location(row, col_indices)
                keyword_col_idx = col_indices.get('keyword')

                if keyword_col_idx is not None and row[keyword_col_idx]:
                    keyword = str(row[keyword_col_idx]).strip()
                    info.original_keyword = keyword

                    if '*' in keyword or '?' in keyword:
                        location_map.fuzzy_matches.append((keyword.replace('*', '').replace('?', ''), info))
                    else:
                        location_map.exact_matches[keyword.upper()] = info
            except (ValueError, IndexError) as e:
                continue

        wb.close()
        return location_map

    def _parse_csv(self) -> LocationMap:
        """解析 CSV 格式的对照表"""
        location_map = LocationMap()

        with open(self.map_file, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            headers = next(reader, None)

            if not headers:
                raise ValueError("对照表为空")

            headers = [h.strip() for h in headers]
            col_indices = self._validate_and_get_columns(headers)

            for row in reader:
                if not row or all(cell.strip() == "" for cell in row):
                    continue

                try:
                    info = self._row_to_location(row, col_indices)
                    keyword_col_idx = col_indices.get('keyword')

                    if keyword_col_idx is not None and row[keyword_col_idx]:
                        keyword = row[keyword_col_idx].strip()
                        info.original_keyword = keyword

                        if '*' in keyword or '?' in keyword:
                            location_map.fuzzy_matches.append((keyword.replace('*', '').replace('?', ''), info))
                        else:
                            location_map.exact_matches[keyword.upper()] = info
                except (ValueError, IndexError) as e:
                    continue

        return location_map

    def _validate_and_get_columns(self, headers: List[str]) -> Dict[str, int]:
        """验证表头并获取列索引"""
        col_indices = {}

        for col_name in self.REQUIRED_COLUMNS:
            for i, h in enumerate(headers):
                if col_name in h:
                    col_indices[col_name] = i
                    break
            if col_name not in col_indices:
                raise ValueError(f"对照表缺少必要列: {col_name}")

        for col_name in self.KEYWORD_COLUMNS:
            for i, h in enumerate(headers):
                if col_name in h:
                    col_indices['keyword'] = i
                    break
            if 'keyword' in col_indices:
                break

        if 'keyword' not in col_indices:
            raise ValueError(f"对照表缺少关键词列，请包含以下列名之一: {', '.join(self.KEYWORD_COLUMNS)}")

        return col_indices

    @staticmethod
    def _row_to_location(row: tuple, col_indices: Dict[str, int]) -> LocationInfo:
        """将行数据转换为 LocationInfo"""
        def get_val(idx: int) -> str:
            if idx >= len(row) or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        return LocationInfo(
            project=get_val(col_indices["项目"]),
            building=get_val(col_indices["楼栋"]),
            floor=get_val(col_indices["楼层"]),
            position=get_val(col_indices["部位"]),
        )
