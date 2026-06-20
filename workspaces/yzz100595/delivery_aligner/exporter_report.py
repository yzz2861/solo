import os
from typing import List, Optional
from datetime import date
import pandas as pd
from collections import defaultdict

from .models import MatchedOrder, OrderStatus
from .date_utils import format_date, describe_date_relative
from .matcher import (
    group_by_supplier,
    group_by_status,
    summarize,
    get_delayed_orders,
    get_no_promise_orders,
    get_promised_not_arrived,
)


def _build_summary_sheet(matched: List[MatchedOrder], today: date) -> pd.DataFrame:
    summary = summarize(matched)
    rows = [
        {"指标": "订单总数", "数值": summary["total_orders"], "说明": ""},
        {"指标": "总数量", "数值": summary["total_quantity"], "说明": ""},
        {"指标": "已到货数量", "数值": summary["delivered_quantity"], "说明": ""},
        {"指标": "未到货数量", "数值": summary["remaining_quantity"], "说明": ""},
        {"指标": "整体到货率(%)", "数值": summary["overall_delivery_rate"], "说明": ""},
        {"指标": "", "数值": "", "说明": ""},
    ]

    for status_name, count in summary["by_status"].items():
        rows.append({"指标": f"状态: {status_name}", "数值": count, "说明": ""})

    rows.extend([
        {"指标": "", "数值": "", "说明": ""},
        {"指标": "延期订单数", "数值": summary["delayed_count"], "说明": "重点关注"},
        {"指标": "无承诺订单数", "数值": summary["no_promise_count"], "说明": "需要采购跟进"},
        {"指标": "承诺未到订单数", "数值": summary["promised_pending_count"], "说明": "待交货"},
        {"指标": "延期总天数", "数值": summary["total_delay_days"], "说明": ""},
        {"指标": "平均延期天数", "数值": summary["avg_delay_days"], "说明": "仅统计延期订单"},
    ])

    rows.append({"指标": "", "数值": "", "说明": ""})
    rows.append({"指标": "按供应商统计(订单数)", "数值": "", "说明": ""})
    for supplier, count in sorted(summary["by_supplier_count"].items(), key=lambda x: -x[1]):
        rows.append({"指标": f"  {supplier}", "数值": count, "说明": ""})

    return pd.DataFrame(rows)


def _build_risk_sheet(matched: List[MatchedOrder], today: date) -> pd.DataFrame:
    rows = []
    risk_levels = []

    for m in matched:
        po = m.purchase_order
        if m.status == OrderStatus.FULLY_DELIVERED:
            continue

        risk_score = 0
        risk_factors = []

        if m.status == OrderStatus.DELAYED:
            risk_score += 50
            risk_factors.append(f"延期{m.delay_days}天")
        if not m.has_promise:
            risk_score += 30
            risk_factors.append("无供应商承诺")
        if m.remaining_quantity > 0 and po.quantity > 0:
            remaining_ratio = m.remaining_quantity / po.quantity
            if remaining_ratio >= 0.8:
                risk_score += 15
                risk_factors.append(f"剩余{round(remaining_ratio*100)}%未到")
            elif remaining_ratio >= 0.5:
                risk_score += 10
                risk_factors.append(f"剩余{round(remaining_ratio*100)}%未到")
        if po.plan_date and m.has_promise and m.latest_promise_date:
            if m.latest_promise_date > po.plan_date:
                diff = (m.latest_promise_date - po.plan_date).days
                if diff >= 7:
                    risk_score += 20
                    risk_factors.append(f"承诺晚于计划{diff}天")
                elif diff >= 3:
                    risk_score += 10
                    risk_factors.append(f"承诺晚于计划{diff}天")

        if risk_score >= 60:
            level = "高风险"
        elif risk_score >= 30:
            level = "中风险"
        else:
            level = "低风险"

        production_impact = _assess_production_impact(m, risk_score, today)

        rows.append({
            "风险等级": level,
            "风险评分": risk_score,
            "订单号": po.order_key,
            "物料编码": po.material_code,
            "物料名称": po.material_name,
            "供应商": po.supplier_short or po.supplier_full,
            "订单数量": po.quantity,
            "单位": po.unit,
            "已到货": m.delivered_quantity,
            "未到货": m.remaining_quantity,
            "缺口率%": round(m.remaining_quantity / po.quantity * 100, 1) if po.quantity > 0 else 0,
            "计划交期": format_date(po.plan_date),
            "计划交期倒计时": describe_date_relative(po.plan_date, today),
            "承诺交期": format_date(m.latest_promise_date),
            "承诺倒计时": describe_date_relative(m.latest_promise_date, today) if m.latest_promise_date else "无承诺",
            "实际状态": m.status.value,
            "延期天数": m.delay_days if m.delay_days > 0 else 0,
            "风险因素": "; ".join(risk_factors) if risk_factors else "",
            "生产影响评估": production_impact,
            "建议措施": _get_planning_action(m, level),
            "分批标记": "是" if m.is_partial_batch else "否",
            "拆分标记": "是" if po.is_split else "否",
            "备注": "; ".join(m.notes) if m.notes else "",
        })
        risk_levels.append(level)

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    sort_map = {"高风险": 0, "中风险": 1, "低风险": 2}
    df["_sort"] = df["风险等级"].map(sort_map)
    df = df.sort_values(by=["_sort", "风险评分"], ascending=[True, False])
    df = df.drop(columns=["_sort"])
    return df


def _assess_production_impact(m: MatchedOrder, risk_score: int, today: date) -> str:
    po = m.purchase_order
    impacts = []

    if po.plan_date:
        plan_diff = (po.plan_date - today).days
        if plan_diff < 0:
            impacts.append(f"已过计划交期，可能导致停产待料")
        elif plan_diff <= 3:
            impacts.append(f"计划交期临近({plan_diff}天内)，生产排程紧张")
        elif plan_diff <= 7:
            impacts.append(f"计划交期在一周内，需关注交付")

    if m.status == OrderStatus.DELAYED and m.delay_days >= 7:
        impacts.append("严重延期，可能影响订单交付")
    elif m.status == OrderStatus.DELAYED and m.delay_days >= 3:
        impacts.append("已延期，存在生产风险")

    if not m.has_promise and po.plan_date and (po.plan_date - today).days <= 14:
        impacts.append("无供应商承诺且计划交期在2周内，存在不确定性")

    if m.remaining_quantity == po.quantity and po.quantity > 0:
        impacts.append("完全未到货，直接影响投产")
    elif m.remaining_quantity > 0 and m.delivered_quantity > 0:
        ratio = m.remaining_quantity / po.quantity
        if ratio >= 0.5:
            impacts.append("超过半数未到货，可能需部分排产")

    return "; ".join(impacts) if impacts else "风险可控"


def _get_planning_action(m: MatchedOrder, risk_level: str) -> str:
    actions = []
    if risk_level == "高风险":
        actions.append("立即通知采购升级催办，评估备选供应商")
        if m.status != OrderStatus.FULLY_DELIVERED:
            actions.append("准备调整生产计划，考虑替代物料")
    elif risk_level == "中风险":
        actions.append("通知采购跟进，每日更新进度")
        actions.append("预留排程缓冲，准备调整方案")
    else:
        actions.append("正常跟进，按计划排产")

    if m.is_partial_batch:
        actions.append("根据分批到货安排分段排产")

    return "；".join(actions)


def _build_promise_gap_sheet(matched: List[MatchedOrder], today: date) -> pd.DataFrame:
    rows = []

    promised_pending = get_promised_not_arrived(matched)
    no_promise = get_no_promise_orders(matched)
    delayed = get_delayed_orders(matched)

    categories = [
        ("供应商承诺但未到货", [m for m in promised_pending if m.status != OrderStatus.DELAYED]),
        ("无承诺需要供应商确认", [m for m in no_promise if m.status != OrderStatus.DELAYED]),
        ("已承诺但已延期", [m for m in delayed if m.has_promise]),
        ("无承诺且已延期", [m for m in delayed if not m.has_promise]),
    ]

    for category, orders in categories:
        for m in orders:
            po = m.purchase_order
            gap_info = []

            if po.plan_date and m.latest_promise_date:
                gap_days = (m.latest_promise_date - po.plan_date).days
                if gap_days > 0:
                    gap_info.append(f"承诺比计划晚{gap_days}天")
                elif gap_days < 0:
                    gap_info.append(f"承诺比计划早{abs(gap_days)}天")

            if m.promises:
                date_changes = []
                sorted_promises = sorted(
                    [p for p in m.promises if p.promise_date],
                    key=lambda p: p.promise_date or date.min
                )
                if len(sorted_promises) >= 2:
                    for i in range(1, len(sorted_promises)):
                        prev = sorted_promises[i - 1].promise_date
                        curr = sorted_promises[i].promise_date
                        if prev and curr:
                            diff = (curr - prev).days
                            if diff != 0:
                                direction = "延后" if diff > 0 else "提前"
                                date_changes.append(f"{direction}{abs(diff)}天")
                    if date_changes:
                        gap_info.append(f"交期变更{len(date_changes)}次: " + ", ".join(date_changes))

            rows.append({
                "差异分类": category,
                "供应商": po.supplier_short or po.supplier_full,
                "订单号": po.order_key,
                "物料编码": po.material_code,
                "物料名称": po.material_name,
                "订单数量": po.quantity,
                "单位": po.unit,
                "已到货": m.delivered_quantity,
                "待到货": m.remaining_quantity,
                "计划交期": format_date(po.plan_date),
                "供应商承诺": format_date(m.latest_promise_date),
                "是否延期": "是" if m.status == OrderStatus.DELAYED else "否",
                "延期天数": m.delay_days if m.delay_days > 0 else 0,
                "承诺次数": len(m.promises),
                "差异说明": "; ".join(gap_info) if gap_info else "",
                "当前状态": m.status.value,
            })

    return pd.DataFrame(rows) if rows else pd.DataFrame()


def _build_supplier_scorecard(matched: List[MatchedOrder]) -> pd.DataFrame:
    by_supplier = group_by_supplier(matched)
    rows = []

    for supplier, orders in sorted(by_supplier.items(), key=lambda x: -len(x[1])):
        total = len(orders)
        on_time = sum(
            1 for m in orders
            if m.status in (OrderStatus.FULLY_DELIVERED, OrderStatus.PARTIALLY_DELIVERED)
            or (m.has_promise and m.latest_promise_date and m.status != OrderStatus.DELAYED)
        )
        delayed_count = sum(1 for m in orders if m.status == OrderStatus.DELAYED)
        no_promise_count = sum(
            1 for m in orders
            if not m.has_promise and m.status != OrderStatus.FULLY_DELIVERED
        )
        total_delay_days = sum(m.delay_days for m in orders if m.delay_days > 0)
        total_qty = sum(m.purchase_order.quantity for m in orders)
        delivered_qty = sum(m.delivered_quantity for m in orders)

        on_time_rate = round(on_time / total * 100, 1) if total > 0 else 0
        qty_rate = round(delivered_qty / total_qty * 100, 1) if total_qty > 0 else 0

        if on_time_rate >= 90 and delayed_count == 0 and no_promise_count == 0:
            grade = "A 优秀"
        elif on_time_rate >= 75 and delayed_count <= total * 0.1:
            grade = "B 良好"
        elif on_time_rate >= 60:
            grade = "C 一般"
        else:
            grade = "D 需改善"

        rows.append({
            "评级": grade,
            "供应商": supplier,
            "订单总数": total,
            "准时交付率%": on_time_rate,
            "数量完成率%": qty_rate,
            "延期单数": delayed_count,
            "无承诺单数": no_promise_count,
            "累计延期天数": total_delay_days,
            "平均延期/单": round(total_delay_days / delayed_count, 1) if delayed_count > 0 else 0,
            "处理建议": _get_supplier_action(grade, delayed_count, no_promise_count),
        })

    if rows:
        df = pd.DataFrame(rows)
        grade_map = {"A 优秀": 0, "B 良好": 1, "C 一般": 2, "D 需改善": 3}
        df["_sort"] = df["评级"].map(grade_map)
        df = df.sort_values(by=["_sort", "准时交付率%"], ascending=[True, False])
        df = df.drop(columns=["_sort"])
        return df
    return pd.DataFrame()


def _get_supplier_action(grade: str, delayed: int, no_promise: int) -> str:
    if grade == "A 优秀":
        return "保持合作，优先下单"
    elif grade == "B 良好":
        return "正常合作，定期复盘"
    elif grade == "C 一般":
        return "加强沟通，提升交期回复速度"
    else:
        actions = []
        if delayed > 0:
            actions.append(f"针对{delayed}个延期订单要求改善报告")
        if no_promise > 0:
            actions.append(f"敦促{no_promise}个订单尽快给出交期承诺")
        actions.append("评估备选供应商，降低依赖风险")
        return "；".join(actions)


def export_variance_report(
    matched: List[MatchedOrder],
    output_path: str,
    today: Optional[date] = None,
) -> str:
    today = today or date.today()

    summary_df = _build_summary_sheet(matched, today)
    risk_df = _build_risk_sheet(matched, today)
    gap_df = _build_promise_gap_sheet(matched, today)
    supplier_df = _build_supplier_scorecard(matched)

    ext = os.path.splitext(output_path)[1].lower()
    if ext not in (".xlsx", ".xls"):
        output_path = output_path + ".xlsx"

    output_dir = os.path.dirname(os.path.abspath(output_path))
    os.makedirs(output_dir, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="总览", index=False)
        if not risk_df.empty:
            risk_df.to_excel(writer, sheet_name="风险评估", index=False)
            _style_risk_sheet(writer, "风险评估", risk_df)
        if not gap_df.empty:
            gap_df.to_excel(writer, sheet_name="承诺差异", index=False)
        if not supplier_df.empty:
            supplier_df.to_excel(writer, sheet_name="供应商评级", index=False)
            _style_supplier_sheet(writer, "供应商评级", supplier_df)

        _style_summary_sheet(writer, "总览", summary_df)

    return output_path


def _style_summary_sheet(writer, sheet_name: str, df: pd.DataFrame) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = writer.sheets[sheet_name]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx in range(2, len(df) + 2):
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = left_align if col_idx == 1 or col_idx == 3 else center_align

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 25


def _style_risk_sheet(writer, sheet_name: str, df: pd.DataFrame) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = writer.sheets[sheet_name]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    high_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    mid_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    low_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx in range(2, len(df) + 2):
        level_cell = ws.cell(row=row_idx, column=1)
        level_val = str(level_cell.value or "")
        if "高风险" in level_val:
            row_fill = high_fill
        elif "中风险" in level_val:
            row_fill = mid_fill
        else:
            row_fill = low_fill

        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.fill = row_fill
            cell.alignment = left_align if col_idx >= 17 else center_align

    col_widths = {
        1: 10, 2: 10, 3: 16, 4: 14, 5: 20, 6: 12, 7: 10, 8: 8, 9: 10,
        10: 10, 11: 10, 12: 18, 13: 16, 14: 18, 15: 16, 16: 10,
        17: 10, 18: 30, 19: 35, 20: 30, 21: 10, 22: 10, 23: 30,
    }
    for col_idx, width in col_widths.items():
        if col_idx <= len(df.columns):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _style_supplier_sheet(writer, sheet_name: str, df: pd.DataFrame) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = writer.sheets[sheet_name]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    grade_fills = {
        "A 优秀": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "B 良好": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "C 一般": PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),
        "D 需改善": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx in range(2, len(df) + 2):
        grade_cell = ws.cell(row=row_idx, column=1)
        grade_val = str(grade_cell.value or "")
        row_fill = grade_fills.get(grade_val, PatternFill())

        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.fill = row_fill
            cell.alignment = left_align if col_idx in (2, 10) else center_align

    col_widths = {1: 12, 2: 16, 3: 12, 4: 14, 5: 14, 6: 12, 7: 14, 8: 14, 9: 14, 10: 40}
    for col_idx, width in col_widths.items():
        if col_idx <= len(df.columns):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
