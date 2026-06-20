import os
from typing import List, Optional
from datetime import date
import pandas as pd

from .models import MatchedOrder, OrderStatus
from .date_utils import format_date, describe_date_relative
from .matcher import (
    get_delayed_orders,
    get_no_promise_orders,
    get_promised_not_arrived,
    group_by_supplier,
)


def _build_urgent_list_rows(
    matched: List[MatchedOrder],
    today: date,
) -> List[dict]:
    rows = []

    delayed = get_delayed_orders(matched)
    no_promise = get_no_promise_orders(matched)
    promised_pending = get_promised_not_arrived(matched)

    sections = [
        ("严重延期", delayed),
        ("承诺未到货", [m for m in promised_pending if m.status != OrderStatus.DELAYED]),
        ("无承诺待跟进", [m for m in no_promise if m.status != OrderStatus.DELAYED]),
        ("部分到货待补", [m for m in matched if m.status == OrderStatus.PARTIALLY_DELIVERED]),
    ]

    for section_name, orders in sections:
        by_supplier = group_by_supplier(orders)
        for supplier, supplier_orders in sorted(by_supplier.items()):
            for m in sorted(supplier_orders, key=lambda x: -(x.delay_days or 0)):
                po = m.purchase_order
                priority = _get_priority(m)
                action_suggestion = _get_action_suggestion(m, today)
                contact_note = _get_contact_note(m)

                rows.append({
                    "优先级": priority,
                    "分类": section_name,
                    "供应商": supplier,
                    "供应商全称": po.supplier_full,
                    "订单号": po.order_key,
                    "物料编码": po.material_code,
                    "物料名称": po.material_name,
                    "订单数量": po.quantity,
                    "单位": po.unit,
                    "已到货": m.delivered_quantity,
                    "未到货": m.remaining_quantity,
                    "完成率%": m.delivery_rate,
                    "计划交期": format_date(po.plan_date),
                    "承诺交期": format_date(m.latest_promise_date),
                    "最近到货": format_date(m.latest_arrival_date),
                    "延期天数": m.delay_days if m.delay_days > 0 else "",
                    "承诺状态": "有承诺" if m.has_promise else "无承诺",
                    "分批到货": "是" if m.is_partial_batch else "否",
                    "催办建议": action_suggestion,
                    "联系方式/邮件要点": contact_note,
                    "备注": "; ".join(m.notes) if m.notes else "",
                })

    return rows


def _get_priority(m: MatchedOrder) -> str:
    if m.status == OrderStatus.DELAYED and m.delay_days >= 7:
        return "紧急 P0"
    elif m.status == OrderStatus.DELAYED:
        return "高 P1"
    elif not m.has_promise and m.purchase_order.plan_date:
        return "高 P1"
    elif m.status == OrderStatus.PARTIALLY_DELIVERED and m.remaining_quantity > m.delivered_quantity:
        return "中 P2"
    elif m.has_promise and not m.has_arrival:
        return "中 P2"
    else:
        return "低 P3"


def _get_action_suggestion(m: MatchedOrder, today: date) -> str:
    if m.status == OrderStatus.DELAYED:
        base = f"立即联系供应商确认最新交期，已延期{m.delay_days}天"
        if m.remaining_quantity > 0:
            base += f"，剩余{m.remaining_quantity}{m.purchase_order.unit}未到"
        return base
    elif not m.has_promise:
        if m.purchase_order.plan_date:
            plan_desc = describe_date_relative(m.purchase_order.plan_date, today)
            return f"催要交期承诺，计划交期{format_date(m.purchase_order.plan_date)}（{plan_desc}）"
        else:
            return "催要交期承诺，无计划交期也无供应商承诺"
    elif m.status == OrderStatus.PARTIALLY_DELIVERED:
        promise_desc = describe_date_relative(m.latest_promise_date, today) if m.latest_promise_date else ""
        return f"催促剩余{m.remaining_quantity}{m.purchase_order.unit}交货，承诺交期{format_date(m.latest_promise_date)}（{promise_desc}）"
    elif m.has_promise and not m.has_arrival:
        promise_desc = describe_date_relative(m.latest_promise_date, today) if m.latest_promise_date else ""
        return f"确认备货进度，承诺交期{format_date(m.latest_promise_date)}（{promise_desc}）"
    else:
        return "持续跟进"


def _get_contact_note(m: MatchedOrder) -> str:
    parts = []
    if m.promises:
        sources = set(p.source for p in m.promises if p.source)
        if sources:
            parts.append(f"最近承诺来源: {', '.join(sources)}")
        batch_info = []
        for i, p in enumerate(m.promises, 1):
            if p.promise_date:
                qty = f"{p.promise_quantity}" if p.promise_quantity else "全部"
                batch_info.append(f"第{i}次承诺{qty}@{format_date(p.promise_date)}")
        if batch_info:
            parts.append("历史承诺: " + "; ".join(batch_info))
    if m.arrivals:
        parts.append(f"已到{len(m.arrivals)}批，最近到货{format_date(m.latest_arrival_date)}")
    if parts:
        return " | ".join(parts)
    return "首次联系，请建立承诺记录"


def export_urgent_list(
    matched: List[MatchedOrder],
    output_path: str,
    today: Optional[date] = None,
) -> str:
    today = today or date.today()
    rows = _build_urgent_list_rows(matched, today)

    if not rows:
        return "暂无需要催办的订单"

    output_dir = os.path.dirname(os.path.abspath(output_path))
    os.makedirs(output_dir, exist_ok=True)

    df = pd.DataFrame(rows)

    ext = os.path.splitext(output_path)[1].lower()
    if ext in (".xlsx", ".xls"):
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="催货清单", index=False)
            _style_excel(writer, "催货清单", df)
    elif ext == ".csv":
        df.to_csv(output_path, index=False, encoding="utf-8-sig")
    else:
        output_path = output_path + ".xlsx"
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="催货清单", index=False)
            _style_excel(writer, "催货清单", df)

    return output_path


def _style_excel(writer, sheet_name: str, df: pd.DataFrame) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = writer.sheets[sheet_name]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    p0_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    p1_fill = PatternFill(start_color="FFA94D", end_color="FFA94D", fill_type="solid")
    p2_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
    p3_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx in range(2, len(df) + 2):
        priority_cell = ws.cell(row=row_idx, column=1)
        priority_val = str(priority_cell.value or "")
        if "P0" in priority_val:
            row_fill = p0_fill
        elif "P1" in priority_val:
            row_fill = p1_fill
        elif "P2" in priority_val:
            row_fill = p2_fill
        else:
            row_fill = p3_fill

        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = left_align if col_idx >= 16 else center_align
            if row_fill != p3_fill:
                cell.fill = row_fill

    col_widths = {
        1: 10, 2: 12, 3: 12, 4: 22, 5: 16, 6: 14, 7: 20,
        8: 10, 9: 8, 10: 10, 11: 10, 12: 10,
        13: 18, 14: 18, 15: 18, 16: 10, 17: 10, 18: 10,
        19: 35, 20: 40, 21: 30,
    }
    for col_idx, width in col_widths.items():
        if col_idx <= len(df.columns):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
