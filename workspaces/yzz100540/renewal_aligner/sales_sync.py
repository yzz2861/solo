from __future__ import annotations

import csv
import json
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

import pandas as pd

from .models import (
    CustomerRenewalRecord,
    ForecastCategory,
    RiskLevel,
    RiskType,
)
from .risk_engine import RISK_ORDER


RISK_LEVEL_ORDER = {lvl: i for i, lvl in enumerate(RISK_ORDER)}


@dataclass
class SalesSyncItem:
    canonical_customer_name: str
    risk_level: RiskLevel
    contract_value: float
    days_to_expiry: Optional[int]
    csm_owner: str
    sales_owner: str
    customer_success_concerns: list[str] = field(default_factory=list)
    needs_sales_input: list[str] = field(default_factory=list)
    meeting_agenda_items: list[str] = field(default_factory=list)
    current_forecast: str = ""
    expected_forecast: str = ""
    forecast_gap_notes: str = ""
    next_steps_confirmed: bool = False
    next_step_owner: str = ""
    next_step_deadline: Optional[date] = None
    notes_for_sales: str = ""
    raw_record: Optional[CustomerRenewalRecord] = None


@dataclass
class SalesPreMeetingReport:
    report_date: date
    total_high_risk_customers: int
    by_sales_owner: dict[str, list[SalesSyncItem]] = field(default_factory=dict)
    urgent_items: list[SalesSyncItem] = field(default_factory=list)
    forecast_mismatch_items: list[SalesSyncItem] = field(default_factory=list)
    items_needing_confirmation: list[SalesSyncItem] = field(default_factory=list)


def _needs_sales_escalation(record: CustomerRenewalRecord) -> tuple[bool, list[str], list[str]]:
    escalate = False
    csm_concerns: list[str] = []
    ask_sales: list[str] = []
    types = {r.risk_type for r in record.risks}
    levels = {r.risk_type: r.risk_level for r in record.risks}

    if RiskType.FORECAST_MISSING in types:
        escalate = True
        csm_concerns.append("合同即将到期但系统中缺少对应销售商机，客户已表达过续签意向但销售尚未录入")
        ask_sales.append("请确认是否已在CRM创建续费商机？若没有请本周内完成录入")

    if RiskType.FORECAST_MISMATCH in types:
        escalate = True
        for r in record.risks:
            if r.risk_type == RiskType.FORECAST_MISMATCH:
                csm_concerns.append(r.message)
        ask_sales.append("请与销售管理层确认预测金额/阶段是否准确，并同步客户真实续费意向")

    if RiskType.EXPIRING_SOON in types:
        level = levels.get(RiskType.EXPIRING_SOON, RiskLevel.NONE)
        if level in (RiskLevel.HIGH, RiskLevel.CRITICAL):
            escalate = True
            for r in record.risks:
                if r.risk_type == RiskType.EXPIRING_SOON:
                    csm_concerns.append(r.message)
            ask_sales.append("请确认是否已启动商务谈判？当前客户决策链是否完整？")

    usage_high_risk = any(
        r.risk_type == RiskType.LOW_USAGE and r.risk_level in (RiskLevel.HIGH, RiskLevel.CRITICAL)
        for r in record.risks
    ) or RiskType.ZERO_USAGE_WITH_PILOT in types
    if usage_high_risk:
        escalate = True
        for r in record.risks:
            if r.risk_type in (RiskType.LOW_USAGE, RiskType.ZERO_USAGE_WITH_PILOT):
                csm_concerns.append(r.message)
        ask_sales.append("客户使用情况异常，建议销售与客户决策层安排一次战略沟通，了解真实使用评价与未来规划")

    ticket_high_risk = any(
        r.risk_type in (RiskType.HIGH_TICKETS, RiskType.TICKET_REOPENED)
        and r.risk_level in (RiskLevel.HIGH, RiskLevel.CRITICAL)
        for r in record.risks
    )
    if ticket_high_risk:
        escalate = True
        for r in record.risks:
            if r.risk_type in (RiskType.HIGH_TICKETS, RiskType.TICKET_REOPENED):
                csm_concerns.append(r.message)
        ask_sales.append("请协助向客户表达我方对当前积压问题的重视，必要时安排高管拜访")

    return escalate, csm_concerns, ask_sales


def build_sales_sync_list(
    records: dict[str, CustomerRenewalRecord],
    min_risk_level: RiskLevel = RiskLevel.HIGH,
    baseline: Optional[date] = None,
) -> list[SalesSyncItem]:
    base = baseline or date.today()
    items: list[SalesSyncItem] = []
    for name, rec in records.items():
        if RISK_LEVEL_ORDER.get(rec.highest_risk_level, -1) < RISK_LEVEL_ORDER.get(min_risk_level, 99):
            escalate, _, _ = _needs_sales_escalation(rec)
            if not escalate:
                continue
        escalate, concerns, asks = _needs_sales_escalation(rec)
        days_to = None
        c = rec.primary_contract
        if c and c.end_date:
            days_to = (c.end_date - base).days
        current_fmt = "-"
        expected_fmt = "-"
        gap_notes = ""
        if rec.has_forecast:
            committed = rec.committed_forecast_amount
            total = sum(f.amount for f in rec.forecasts)
            cats = sorted({f.category.value for f in rec.forecasts})
            current_fmt = f"￥{total:,.0f}（其中承诺￥{committed:,.0f}）阶段：{'/'.join(cats)}"
        if rec.contract_value > 0:
            expected_fmt = f"￥{rec.contract_value:,.0f}（与合同额对齐）"
            if rec.has_forecast and rec.committed_forecast_amount > 0:
                delta = rec.committed_forecast_amount - rec.contract_value
                if abs(delta) / rec.contract_value >= 0.10:
                    if delta < 0:
                        gap_notes = f"承诺预测比合同低￥{abs(delta):,.0f}（{abs(delta)/rec.contract_value*100:.0f}%），可能缩量或流失"
                    else:
                        gap_notes = f"承诺预测比合同高￥{delta:,.0f}，可能为增购，需销售确认"
        agenda = []
        agenda.extend(concerns)
        agenda.extend(asks)
        if rec.next_action:
            agenda.append(f"（客成建议）{rec.next_action}")
        item = SalesSyncItem(
            canonical_customer_name=name,
            risk_level=rec.highest_risk_level,
            contract_value=rec.contract_value,
            days_to_expiry=days_to,
            csm_owner=rec.csm_owner or "-",
            sales_owner=rec.sales_owner or "待分配",
            customer_success_concerns=concerns,
            needs_sales_input=asks,
            meeting_agenda_items=agenda,
            current_forecast=current_fmt,
            expected_forecast=expected_fmt,
            forecast_gap_notes=gap_notes,
            next_step_deadline=rec.follow_up_date,
            next_step_owner=rec.csm_owner or rec.sales_owner or "-",
            notes_for_sales="\n".join(concerns) if concerns else "",
            raw_record=rec,
        )
        items.append(item)
    items.sort(key=lambda it: (
        -RISK_LEVEL_ORDER.get(it.risk_level, 0),
        (9999 if it.days_to_expiry is None else (
            -9999 if it.days_to_expiry < 0 else -it.days_to_expiry
        )),
        -it.contract_value,
    ))
    return items


def build_pre_meeting_report(
    items: list[SalesSyncItem],
    baseline: Optional[date] = None,
) -> SalesPreMeetingReport:
    base = baseline or date.today()
    report = SalesPreMeetingReport(
        report_date=base,
        total_high_risk_customers=len(items),
    )
    for it in items:
        report.by_sales_owner.setdefault(it.sales_owner, []).append(it)
        if it.days_to_expiry is not None and 0 <= it.days_to_expiry <= 30:
            report.urgent_items.append(it)
        if it.forecast_gap_notes:
            report.forecast_mismatch_items.append(it)
        if it.needs_sales_input:
            report.items_needing_confirmation.append(it)
    return report


def export_sales_sync_excel(items: list[SalesSyncItem], output_path: Path) -> Path:
    rows: list[dict[str, Any]] = []
    for idx, it in enumerate(items):
        expiry = "-"
        if it.days_to_expiry is not None:
            expiry = f"已过期{abs(it.days_to_expiry)}天" if it.days_to_expiry < 0 else f"{it.days_to_expiry}天"
        rows.append({
            "序号": idx + 1,
            "风险等级": it.risk_level.value.upper(),
            "客户名称": it.canonical_customer_name,
            "合同额(元)": it.contract_value,
            "到期剩余": expiry,
            "客成经理": it.csm_owner,
            "销售负责人": it.sales_owner,
            "客成关注要点": "\n".join(f"• {c}" for c in it.customer_success_concerns) or "无",
            "需销售确认事项": "\n".join(f"• {a}" for a in it.needs_sales_input) or "无",
            "当前销售预测": it.current_forecast,
            "预测偏差说明": it.forecast_gap_notes or "无偏差",
            "会议讨论议题": "\n".join(f"• {a}" for a in it.meeting_agenda_items) or "待定",
            "会后下一步责任人": it.next_step_owner,
            "建议完成截止": it.next_step_deadline.isoformat() if it.next_step_deadline else "-",
            "□ 销售已确认": "",
        })
    df = pd.DataFrame(rows)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="高风险客户同步")
        ws = writer.sheets["高风险客户同步"]
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        level_col = 2
        risk_colors = {
            "critical": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
            "high": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        }
        for row_idx in range(2, len(rows) + 2):
            lv = (ws.cell(row=row_idx, column=level_col).value or "").lower()
            for k, fill in risk_colors.items():
                if k in lv:
                    for col in range(1, len(rows[0]) + 1):
                        ws.cell(row=row_idx, column=col).fill = fill
            for col in range(1, len(rows[0]) + 1):
                ws.cell(row=row_idx, column=col).alignment = Alignment(vertical="center", wrap_text=True)
        widths = {
            1: 6, 2: 10, 3: 24, 4: 12, 5: 10, 6: 12, 7: 14,
            8: 40, 9: 40, 10: 28, 11: 28, 12: 45, 13: 16, 14: 14, 15: 12,
        }
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = "D2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(rows[0]))}{len(rows) + 1}"

        owner_rows: dict[str, list[dict[str, Any]]] = {}
        for it in items:
            owner_rows.setdefault(it.sales_owner, []).append({
                "客户": it.canonical_customer_name,
                "风险等级": it.risk_level.value.upper(),
                "合同额": f"￥{it.contract_value:,.0f}",
                "到期剩余": f"{it.days_to_expiry}天" if it.days_to_expiry is not None else "-",
                "需确认": "；".join(it.needs_sales_input) or "无",
                "□ 已同步": "",
            })
        if owner_rows:
            all_rows = []
            for owner, custs in owner_rows.items():
                all_rows.append({"分组": f"【销售：{owner} - 共{len(custs)}位】", "": "", "  ": "", "    ": "", "      ": "", "□ 已同步": ""})
                all_rows.extend(custs)
                all_rows.append({"分组": "", "": "", "  ": "", "    ": "", "      ": "", "□ 已同步": ""})
            df2 = pd.DataFrame(all_rows)
            df2.to_excel(writer, index=False, sheet_name="按销售分组清单")
            ws2 = writer.sheets["按销售分组清单"]
            for cell in ws2[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws2.freeze_panes = "A2"
    return output_path


def export_sales_pre_meeting_md(
    report: SalesPreMeetingReport,
    output_path: Path,
) -> Path:
    lines: list[str] = []
    lines.append(f"# 续费会前确认 - 高风险客户同步清单")
    lines.append("")
    lines.append(f"会议日期：{report.report_date.isoformat()}")
    lines.append(f"需要销售对齐的高风险客户数：**{report.total_high_risk_customers}**")
    lines.append("")
    lines.append("## 📊 总体概览")
    lines.append("")
    lines.append(f"- 🔴 极度紧急（30天内到期或极高风险）：**{len(report.urgent_items)}** 位")
    lines.append(f"- ⚠️ 预测金额/阶段与合同不匹配：**{len(report.forecast_mismatch_items)}** 位")
    lines.append(f"- ❓ 有明确事项需销售确认：**{len(report.items_needing_confirmation)}** 位")
    lines.append("")
    lines.append("## 👥 按销售分组")
    lines.append("")
    for owner in sorted(report.by_sales_owner.keys()):
        its = report.by_sales_owner[owner]
        total_value = sum(i.contract_value for i in its)
        lines.append(f"### {owner}（{len(its)} 位，合同总额 ￥{total_value:,.0f}）")
        lines.append("")
        for i, it in enumerate(its):
            expiry = f"{it.days_to_expiry}天" if it.days_to_expiry is not None and it.days_to_expiry >= 0 else (
                f"已过期{abs(it.days_to_expiry)}天" if it.days_to_expiry is not None else "-"
            )
            lvl_mark = "🔴" if it.risk_level == RiskLevel.CRITICAL else "🟠"
            lines.append(f"{i+1}. **{it.canonical_customer_name}** {lvl_mark} 合同￥{it.contract_value:,.0f} / 到期剩{expiry}")
            if it.notes_for_sales:
                lines.append(f"   - 🚨 客成关注：{it.notes_for_sales}")
            if it.forecast_gap_notes:
                lines.append(f"   - ⚠️ 预测偏差：{it.forecast_gap_notes}")
            lines.append(f"   - 📋 需销售确认：{'；'.join(it.needs_sales_input) if it.needs_sales_input else '无'}")
            lines.append(f"   - ➡️ 建议下一步：截止{it.next_step_deadline.isoformat() if it.next_step_deadline else '-'}，责任人{it.next_step_owner}")
            lines.append("")
        lines.append("")
    if report.urgent_items:
        lines.append("## 🚨 极度紧急（需本周内完成对齐）")
        lines.append("")
        for it in report.urgent_items:
            lines.append(f"- **{it.canonical_customer_name}**（销售：{it.sales_owner}）")
            lines.append(f"  - 合同 ￥{it.contract_value:,.0f}，到期剩 {it.days_to_expiry} 天")
            for a in it.meeting_agenda_items:
                lines.append(f"  - {a}")
            lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## ✅ 会后确认栏（销售填写）")
    lines.append("")
    lines.append("| 客户 | 销售 | □ 已同步风险 | □ 已确认预测 | □ 已明确下一步 | 备注 |")
    lines.append("|------|------|:-----------:|:-----------:|:-------------:|------|")
    for owner in sorted(report.by_sales_owner.keys()):
        for it in report.by_sales_owner[owner]:
            lines.append(f"| {it.canonical_customer_name} | {owner} | | | | |")
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines), encoding="utf-8")
    return output_path
