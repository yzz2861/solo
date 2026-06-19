from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path
from typing import Any, Iterable, Optional

import pandas as pd

from .history import DiffResult, CustomerChange
from .models import (
    CustomerRenewalRecord,
    RiskLevel,
)
from .risk_engine import RISK_ORDER


RISK_LEVEL_ORDER = {lvl: i for i, lvl in enumerate(RISK_ORDER)}
RISK_EMOJI = {
    RiskLevel.CRITICAL: "🔴",
    RiskLevel.HIGH: "🟠",
    RiskLevel.MEDIUM: "🟡",
    RiskLevel.LOW: "🟢",
    RiskLevel.NONE: "⚪",
}


@dataclass
class FollowUpItem:
    priority_rank: int
    canonical_customer_name: str
    all_names: str
    risk_level: RiskLevel
    risk_summary: str
    risk_changes: str
    contract_value: float
    days_to_expiry: Optional[int]
    csm_owner: str
    sales_owner: str
    next_action: str
    follow_up_date: Optional[date]
    special_notes: str
    ticket_summary: str
    usage_summary: str
    forecast_summary: str
    raw_record: CustomerRenewalRecord


def _format_risk_summary(record: CustomerRenewalRecord) -> str:
    seen = set()
    parts = []
    for r in record.risks:
        tag = r.risk_type.value
        if tag in seen:
            continue
        seen.add(tag)
        lvl_mark = RISK_EMOJI.get(r.risk_level, "")
        change_mark = ""
        if r.change_since_last:
            if r.change_since_last == "新增":
                change_mark = "[新]"
            elif "升级" in r.change_since_last:
                change_mark = "[↑]"
            elif "已解决" in r.change_since_last:
                change_mark = "[✓]"
            elif r.change_since_last == "持续":
                change_mark = "[→]"
        parts.append(f"{lvl_mark}{change_mark}{tag}")
    return " ".join(parts) if parts else "无风险"


def _format_risk_changes(record: CustomerRenewalRecord, diff: Optional[DiffResult]) -> str:
    if not diff:
        return "首次运行"
    all_changes = diff.worsened + diff.improved + diff.unchanged_high_risk
    for ch in all_changes:
        if ch.canonical_customer_name == record.canonical_customer_name:
            parts = []
            if ch.is_new_customer:
                parts.append("新进入名单")
            if ch.level_delta > 0:
                parts.append(f"风险升级 {ch.previous_level.value}→{ch.current_level.value}")
            elif ch.level_delta < 0:
                parts.append(f"风险下降 {ch.previous_level.value}→{ch.current_level.value}")
            if ch.added_risks:
                parts.append(f"新增{len(ch.added_risks)}项风险")
            if ch.removed_risks:
                parts.append(f"解决{len(ch.removed_risks)}项风险")
            if ch.escalated_risks:
                parts.append(f"{len(ch.escalated_risks)}项风险加重")
            return "；".join(parts) if parts else "无变化"
    if record.canonical_customer_name in diff.new_customers:
        return "新进入名单"
    return "无变化"


def _format_special_notes(record: CustomerRenewalRecord) -> str:
    notes: list[str] = []
    if record.was_renamed:
        old = "、".join((record.rename_details or {}).get("previous_names") or [])
        notes.append(f"客户已改名（曾用名：{old}）")
    c = record.primary_contract
    if c and c.is_renewal_in_progress:
        notes.append("续签进行中")
    u = record.latest_usage
    if u and u.is_zero_usage and u.has_pilot:
        notes.append("零使用但有试点项目")
    if record.reopened_ticket_count > 0:
        notes.append(f"有{record.reopened_ticket_count}个重开工单")
    return "；".join(notes)


def _format_usage(record: CustomerRenewalRecord) -> str:
    u = record.latest_usage
    if not u:
        return "无数据"
    if u.is_zero_usage:
        return f"零使用（{u.total_licenses}席位{'，含试点' if u.has_pilot else ''}）"
    rate = u.utilization_rate * 100
    return f"使用率{rate:.0f}%（{u.active_users}/{u.total_licenses}活跃），近30天登录{u.login_last_30_days}人"


def _format_tickets(record: CustomerRenewalRecord) -> str:
    if not record.tickets:
        return "无工单"
    open_c = record.open_ticket_count
    reopen_c = record.reopened_ticket_count
    total = len(record.tickets)
    parts = [f"未关闭{open_c}/{total}个"]
    if reopen_c:
        parts.append(f"重开{reopen_c}个")
    return "，".join(parts)


def _format_forecast(record: CustomerRenewalRecord) -> str:
    if not record.has_forecast:
        c = record.primary_contract
        if c and c.end_date and (c.end_date - date.today()).days <= 90:
            return "⚠️ 缺少对应销售预测"
        return "无预测"
    committed = record.committed_forecast_amount
    total = sum(f.amount for f in record.forecasts)
    cats = {f.category.value for f in record.forecasts}
    return f"预测额￥{total:,.0f}（承诺￥{committed:,.0f}），阶段：{'/'.join(sorted(cats))}"


def build_follow_up_list(
    records: dict[str, CustomerRenewalRecord],
    diff: Optional[DiffResult] = None,
    min_risk_level: RiskLevel = RiskLevel.LOW,
    include_next_days: int = 30,
    owner_filter: Optional[str] = None,
    baseline: Optional[date] = None,
) -> list[FollowUpItem]:
    base = baseline or date.today()
    items: list[FollowUpItem] = []
    for name, rec in records.items():
        if owner_filter:
            owner = (rec.csm_owner or "").lower()
            if owner_filter.lower() not in owner:
                continue
        level_ok = RISK_LEVEL_ORDER.get(rec.highest_risk_level, -1) >= RISK_LEVEL_ORDER.get(min_risk_level, 99)
        soon = False
        c = rec.primary_contract
        if c and c.end_date:
            d = (c.end_date - base).days
            if 0 <= d <= include_next_days:
                soon = True
        if not level_ok and not soon:
            continue
        days_to = None
        if c and c.end_date:
            days_to = (c.end_date - base).days
        item = FollowUpItem(
            priority_rank=0,
            canonical_customer_name=name,
            all_names=" / ".join(rec.all_names) if rec.all_names else name,
            risk_level=rec.highest_risk_level,
            risk_summary=_format_risk_summary(rec),
            risk_changes=_format_risk_changes(rec, diff),
            contract_value=rec.contract_value,
            days_to_expiry=days_to,
            csm_owner=rec.csm_owner or "-",
            sales_owner=rec.sales_owner or "-",
            next_action=rec.next_action or "-",
            follow_up_date=rec.follow_up_date,
            special_notes=_format_special_notes(rec),
            ticket_summary=_format_tickets(rec),
            usage_summary=_format_usage(rec),
            forecast_summary=_format_forecast(rec),
            raw_record=rec,
        )
        items.append(item)
    items.sort(key=_sort_key, reverse=True)
    for i, it in enumerate(items):
        it.priority_rank = i + 1
    return items


def _sort_key(item: FollowUpItem) -> tuple:
    lvl_score = RISK_LEVEL_ORDER.get(item.risk_level, 0) * 10000
    value_score = min(item.contract_value, 9999999) / 100
    urgency = 0
    if item.days_to_expiry is not None:
        days = item.days_to_expiry
        if days <= 0:
            urgency = 100000
        elif days <= 30:
            urgency = (30 - days) * 1000
        elif days <= 90:
            urgency = (90 - days) * 100
    has_new = 1 if "[新]" in item.risk_summary or "新进入" in item.risk_changes else 0
    has_escalation = 1 if "[↑]" in item.risk_summary or "风险升级" in item.risk_changes else 0
    return (lvl_score + urgency + has_new * 500 + has_escalation * 300, value_score, -item.priority_rank)


def export_follow_up_excel(items: list[FollowUpItem], output_path: Path) -> Path:
    rows: list[dict[str, Any]] = []
    for it in items:
        expiry_str = "-"
        if it.days_to_expiry is not None:
            if it.days_to_expiry < 0:
                expiry_str = f"已过期{abs(it.days_to_expiry)}天"
            elif it.days_to_expiry == 0:
                expiry_str = "今日到期"
            else:
                expiry_str = f"{it.days_to_expiry}天"
        rows.append({
            "优先级": f"#{it.priority_rank}",
            "风险等级": f"{RISK_EMOJI.get(it.risk_level, '')} {it.risk_level.value}",
            "客户名称": it.canonical_customer_name,
            "所有别名/曾用名": it.all_names,
            "合同额(元)": it.contract_value,
            "到期剩余": expiry_str,
            "客成经理": it.csm_owner,
            "销售负责人": it.sales_owner,
            "风险类型": it.risk_summary,
            "较上次变化": it.risk_changes,
            "特殊情况说明": it.special_notes or "-",
            "使用情况": it.usage_summary,
            "工单情况": it.ticket_summary,
            "销售预测": it.forecast_summary,
            "建议下一步动作": it.next_action,
            "建议跟进日期": it.follow_up_date.isoformat() if it.follow_up_date else "-",
        })
    df = pd.DataFrame(rows)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="下周重点跟进")
        ws = writer.sheets["下周重点跟进"]
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        level_col = 2
        risk_colors = {
            "critical": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
            "high": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            "medium": PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid"),
            "low": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        }
        for row_idx in range(2, len(rows) + 2):
            lv_cell = ws.cell(row=row_idx, column=level_col)
            lv_val = (lv_cell.value or "").lower()
            for key, fill in risk_colors.items():
                if key in lv_val:
                    for col in range(1, len(rows[0]) + 1):
                        ws.cell(row=row_idx, column=col).fill = fill
                    break
            for col in range(1, len(rows[0]) + 1):
                ws.cell(row=row_idx, column=col).alignment = Alignment(vertical="center", wrap_text=True)
        widths = {
            1: 8, 2: 12, 3: 24, 4: 30, 5: 12, 6: 10, 7: 12, 8: 14,
            9: 36, 10: 20, 11: 28, 12: 30, 13: 16, 14: 30, 15: 50, 16: 14,
        }
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(rows[0]))}{len(rows) + 1}"
    return output_path


def export_follow_up_markdown(items: list[FollowUpItem], output_path: Path, title: str = "下周重点跟进清单") -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    lines: list[str] = []
    today = date.today()
    lines.append(f"# {title}")
    lines.append("")
    lines.append(f"生成日期: {today.isoformat()}")
    lines.append(f"共 {len(items)} 位客户需要重点跟进")
    lines.append("")
    by_owner: dict[str, list[FollowUpItem]] = {}
    for it in items:
        by_owner.setdefault(it.csm_owner or "未分配", []).append(it)
    lines.append("## 概览")
    lines.append("")
    lines.append("| 客成经理 | 待跟进客户数 | CRITICAL | HIGH | MEDIUM | LOW |")
    lines.append("|----------|-------------:|---------:|-----:|-------:|----:|")
    for owner in sorted(by_owner.keys()):
        own_items = by_owner[owner]
        counts = {RiskLevel.CRITICAL: 0, RiskLevel.HIGH: 0, RiskLevel.MEDIUM: 0, RiskLevel.LOW: 0}
        for it in own_items:
            if it.risk_level in counts:
                counts[it.risk_level] += 1
        lines.append(f"| {owner} | {len(own_items)} | {counts[RiskLevel.CRITICAL]} | {counts[RiskLevel.HIGH]} | {counts[RiskLevel.MEDIUM]} | {counts[RiskLevel.LOW]} |")
    lines.append("")
    lines.append("## 详细跟进列表")
    lines.append("")
    for it in items:
        expiry = "-"
        if it.days_to_expiry is not None:
            if it.days_to_expiry < 0:
                expiry = f"已过期 {abs(it.days_to_expiry)} 天 ⚠️"
            else:
                expiry = f"{it.days_to_expiry} 天"
        emoji = RISK_EMOJI.get(it.risk_level, "")
        lines.append(f"### {it.priority_rank}. {emoji} {it.canonical_customer_name}（{it.risk_level.value.upper()}）")
        lines.append("")
        if it.all_names and it.all_names != it.canonical_customer_name:
            lines.append(f"- **别名/曾用名**: {it.all_names}")
        lines.append(f"- **合同额**: ￥{it.contract_value:,.0f}")
        lines.append(f"- **距到期**: {expiry}")
        lines.append(f"- **客成经理**: {it.csm_owner}　|　**销售**: {it.sales_owner}")
        lines.append(f"- **风险标签**: {it.risk_summary}")
        if it.risk_changes and it.risk_changes not in ("无变化", "首次运行"):
            lines.append(f"- **变化情况**: {it.risk_changes}")
        if it.special_notes:
            lines.append(f"- **特别说明**: {it.special_notes}")
        lines.append(f"- **使用**: {it.usage_summary}")
        lines.append(f"- **工单**: {it.ticket_summary}")
        lines.append(f"- **预测**: {it.forecast_summary}")
        lines.append(f"- **下一步**: {it.next_action}")
        lines.append(f"- **建议跟进日**: {it.follow_up_date.isoformat() if it.follow_up_date else '-'}")
        lines.append("")
    output_path.write_text("\n".join(lines), encoding="utf-8")
    return output_path
