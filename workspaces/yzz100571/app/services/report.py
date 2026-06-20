from datetime import datetime, date
from typing import List, Tuple, Optional
from io import BytesIO
from sqlalchemy.orm import Session
from sqlalchemy import and_, func
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..models import (
    Cabinet,
    RackRequest,
    RackRequestStatus,
    DeviceModel,
    AuditLog,
    AuditAction,
    User,
)
from ..schemas import (
    QuarterlyReport,
    CabinetUtilizationReport,
    AuditLogResponse,
)


def get_cabinet_utilization(
    db: Session,
    cabinet_id: Optional[int] = None
) -> List[CabinetUtilizationReport]:
    query = db.query(Cabinet)
    if cabinet_id:
        query = query.filter(Cabinet.id == cabinet_id)

    cabinets = query.all()
    reports = []

    for cabinet in cabinets:
        active_requests = db.query(RackRequest).filter(
            RackRequest.cabinet_id == cabinet.id,
            RackRequest.status.in_([
                RackRequestStatus.APPROVED,
                RackRequestStatus.IN_PROGRESS,
                RackRequestStatus.COMPLETED,
            ])
        ).join(DeviceModel, RackRequest.device_model_id == DeviceModel.id).all()

        used_u = 0
        used_power = 0

        for req in active_requests:
            u_start = req.actual_u_start or req.planned_u_start
            u_end = req.actual_u_end or req.planned_u_end
            used_u += (u_end - u_start + 1)
            used_power += req.power_draw_watts

        utilization_rate = (used_u / cabinet.total_u * 100) if cabinet.total_u > 0 else 0
        power_utilization_rate = (used_power / cabinet.max_power_watts * 100) if cabinet.max_power_watts > 0 else 0

        reports.append(CabinetUtilizationReport(
            cabinet_id=cabinet.id,
            cabinet_name=cabinet.name,
            location=cabinet.location,
            total_u=cabinet.total_u,
            used_u=used_u,
            utilization_rate=round(utilization_rate, 2),
            max_power_watts=cabinet.max_power_watts,
            used_power_watts=used_power,
            power_utilization_rate=round(power_utilization_rate, 2),
            device_count=len(active_requests)
        ))

    return reports


def get_quarter_range(quarter: Optional[str] = None) -> Tuple[str, datetime, datetime]:
    today = datetime.utcnow()

    if quarter:
        year = int(quarter[:4])
        q = int(quarter[-1])
        if q == 1:
            start = datetime(year, 1, 1)
            end = datetime(year, 3, 31, 23, 59, 59)
        elif q == 2:
            start = datetime(year, 4, 1)
            end = datetime(year, 6, 30, 23, 59, 59)
        elif q == 3:
            start = datetime(year, 7, 1)
            end = datetime(year, 9, 30, 23, 59, 59)
        elif q == 4:
            start = datetime(year, 10, 1)
            end = datetime(year, 12, 31, 23, 59, 59)
        else:
            raise ValueError("季度必须是 1-4")
    else:
        year = today.year
        q = (today.month - 1) // 3 + 1
        if q == 1:
            start = datetime(year, 1, 1)
            end = datetime(year, 3, 31, 23, 59, 59)
        elif q == 2:
            start = datetime(year, 4, 1)
            end = datetime(year, 6, 30, 23, 59, 59)
        elif q == 3:
            start = datetime(year, 7, 1)
            end = datetime(year, 9, 30, 23, 59, 59)
        else:
            start = datetime(year, 10, 1)
            end = datetime(year, 12, 31, 23, 59, 59)

        quarter = f"{year}Q{q}"

    return quarter, start, end


def generate_quarterly_report(
    db: Session,
    quarter: Optional[str] = None
) -> QuarterlyReport:
    quarter_str, start_date, end_date = get_quarter_range(quarter)

    cabinet_utilization = get_cabinet_utilization(db)

    total_requests = db.query(RackRequest).filter(
        RackRequest.created_at >= start_date,
        RackRequest.created_at <= end_date
    ).count()

    approved_requests = db.query(RackRequest).filter(
        RackRequest.approved_at >= start_date,
        RackRequest.approved_at <= end_date,
        RackRequest.status.in_([
            RackRequestStatus.APPROVED,
            RackRequestStatus.IN_PROGRESS,
            RackRequestStatus.COMPLETED,
            RackRequestStatus.DECOMMISSIONED,
        ])
    ).count()

    rejected_requests = db.query(RackRequest).filter(
        RackRequest.status == RackRequestStatus.REJECTED,
        func.date(RackRequest.created_at) >= start_date.date(),
        func.date(RackRequest.created_at) <= end_date.date()
    ).count()

    rejection_rate = (rejected_requests / total_requests * 100) if total_requests > 0 else 0

    abnormal_releases_query = db.query(AuditLog).filter(
        AuditLog.created_at >= start_date,
        AuditLog.created_at <= end_date,
        AuditLog.is_abnormal_release == True
    ).order_by(AuditLog.created_at.desc()).all()

    abnormal_releases = []
    for log in abnormal_releases_query:
        user = db.query(User).filter(User.id == log.user_id).first()
        abnormal_releases.append(AuditLogResponse(
            id=log.id,
            rack_request_id=log.rack_request_id,
            user_id=log.user_id,
            user_name=user.full_name if user else None,
            action=log.action,
            old_value=log.old_value,
            new_value=log.new_value,
            remark=log.remark,
            created_at=log.created_at,
            is_abnormal_release=log.is_abnormal_release
        ))

    avg_utilization = (
        sum(r.utilization_rate for r in cabinet_utilization) / len(cabinet_utilization)
        if cabinet_utilization else 0
    )

    return QuarterlyReport(
        quarter=quarter_str,
        start_date=start_date,
        end_date=end_date,
        cabinet_utilization=cabinet_utilization,
        total_requests=total_requests,
        approved_requests=approved_requests,
        rejected_requests=rejected_requests,
        rejection_rate=round(rejection_rate, 2),
        abnormal_releases=abnormal_releases,
        average_utilization_rate=round(avg_utilization, 2)
    )


def export_quarterly_report_to_excel(
    db: Session,
    quarter: Optional[str] = None
) -> BytesIO:
    report = generate_quarterly_report(db, quarter)

    output = BytesIO()
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    ws1 = wb.active
    ws1.title = "概览"

    ws1["A1"] = f"机房机柜季度报告 - {report.quarter}"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1.merge_cells("A1:F1")

    ws1["A3"] = "统计周期"
    ws1["B3"] = f"{report.start_date.strftime('%Y-%m-%d')} 至 {report.end_date.strftime('%Y-%m-%d')}"
    ws1["A4"] = "总申请单数"
    ws1["B4"] = report.total_requests
    ws1["A5"] = "已批准数"
    ws1["B5"] = report.approved_requests
    ws1["A6"] = "已驳回数"
    ws1["B6"] = report.rejected_requests
    ws1["A7"] = "驳回率"
    ws1["B7"] = f"{report.rejection_rate}%"
    ws1["A8"] = "平均机柜利用率"
    ws1["B8"] = f"{report.average_utilization_rate}%"

    ws2 = wb.create_sheet("机柜利用率")
    headers2 = ["机柜ID", "机柜名称", "位置", "总U位", "已用U位", "U位利用率%", "最大功率(W)", "已用功率(W)", "功率利用率%", "设备数量"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for row, data in enumerate(report.cabinet_utilization, 2):
        ws2.cell(row=row, column=1, value=data.cabinet_id)
        ws2.cell(row=row, column=2, value=data.cabinet_name)
        ws2.cell(row=row, column=3, value=data.location)
        ws2.cell(row=row, column=4, value=data.total_u)
        ws2.cell(row=row, column=5, value=data.used_u)
        ws2.cell(row=row, column=6, value=data.utilization_rate)
        ws2.cell(row=row, column=7, value=data.max_power_watts)
        ws2.cell(row=row, column=8, value=data.used_power_watts)
        ws2.cell(row=row, column=9, value=data.power_utilization_rate)
        ws2.cell(row=row, column=10, value=data.device_count)

    for col in range(1, 11):
        ws2.column_dimensions[chr(64 + col)].width = 15

    ws3 = wb.create_sheet("被驳回申请")
    headers3 = ["申请单号", "设备名称", "机柜", "申请人", "申请日期", "驳回原因"]
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    rejected_requests = db.query(RackRequest).filter(
        RackRequest.status == RackRequestStatus.REJECTED,
        RackRequest.created_at >= report.start_date,
        RackRequest.created_at <= report.end_date
    ).all()

    for row, req in enumerate(rejected_requests, 2):
        creator = db.query(User).filter(User.id == req.created_by).first()
        ws3.cell(row=row, column=1, value=req.request_no)
        ws3.cell(row=row, column=2, value=req.device_name)
        ws3.cell(row=row, column=3, value=req.cabinet.name if req.cabinet else "")
        ws3.cell(row=row, column=4, value=creator.full_name if creator else "")
        ws3.cell(row=row, column=5, value=req.created_at.strftime("%Y-%m-%d %H:%M:%S"))
        ws3.cell(row=row, column=6, value=req.reject_reason)

    for col in range(1, 7):
        ws3.column_dimensions[chr(64 + col)].width = 20

    ws4 = wb.create_sheet("异常释放记录")
    headers4 = ["ID", "关联申请单", "操作人", "操作类型", "备注", "操作时间", "是否异常"]
    for col, header in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for row, log in enumerate(report.abnormal_releases, 2):
        ws4.cell(row=row, column=1, value=log.id)
        ws4.cell(row=row, column=2, value=log.rack_request_id or "")
        ws4.cell(row=row, column=3, value=log.user_name or "")
        ws4.cell(row=row, column=4, value=log.action.value)
        ws4.cell(row=row, column=5, value=log.remark or "")
        ws4.cell(row=row, column=6, value=log.created_at.strftime("%Y-%m-%d %H:%M:%S"))
        ws4.cell(row=row, column=7, value="是" if log.is_abnormal_release else "否")

    for col in range(1, 8):
        ws4.column_dimensions[chr(64 + col)].width = 20

    wb.save(output)
    output.seek(0)
    return output
