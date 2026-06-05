#!/usr/bin/env python3
"""
Run one idea from idea_history.md in Trae CN and save the copied final output.

The runner deliberately prefers stable automation surfaces in this order:
1. Trae's `trae-cn chat` CLI to create/start the chat.
2. macOS keyboard shortcuts to focus and nudge Trae.
3. Accessibility UI text/button lookup for "任务完成" and "复制全部".
4. Screenshots plus trace logs for manual recovery when UI automation is blocked.
"""

from __future__ import annotations

import argparse
import contextlib
import csv
import fcntl
import json
import os
import re
import shlex
import shutil
import subprocess
import sys
import tempfile
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import requests

from runtime_compat import bootstrap_script_auto_runtime, existing_legacy_paths


SCRIPT_DIR = Path(__file__).resolve().parent


def discover_solo_root() -> Path:
    env_root = os.environ.get("SOLO_ROOT")
    if env_root:
        return Path(env_root).expanduser().resolve()
    cwd = Path.cwd().resolve()
    if (cwd / "idea_history.md").exists():
        return cwd
    parent = SCRIPT_DIR.parent.resolve()
    if (parent / "idea_history.md").exists():
        return parent
    return cwd


ROOT = discover_solo_root()
RUN_LOG_ROOT = Path(os.environ.get("SOLO_RUN_LOG_ROOT", ROOT / "run_log")).expanduser().resolve()
SCRIPT_AUTO_RUN_LOG_DIR = RUN_LOG_ROOT / "script_auto"
bootstrap_script_auto_runtime(ROOT, SCRIPT_DIR, RUN_LOG_ROOT)
LOCAL_FEISHU_ENV_PATH = SCRIPT_AUTO_RUN_LOG_DIR / ".feishu_solo.env"


def load_local_env(path: Path) -> None:
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue
        key, value = stripped.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


load_local_env(LOCAL_FEISHU_ENV_PATH)
if os.environ.get("SOLO_EXTRA_PATHS"):
    os.environ["PATH"] = os.environ["SOLO_EXTRA_PATHS"] + os.pathsep + os.environ.get("PATH", "")

IDEA_HISTORY = ROOT / "idea_history.md"
RUNS_DIR = SCRIPT_AUTO_RUN_LOG_DIR / "runs"
STATE_PATH = SCRIPT_AUTO_RUN_LOG_DIR / ".trae_idea_runner_state.json"
RUNNER_LOCK_PATH = SCRIPT_AUTO_RUN_LOG_DIR / ".trae_idea_runner.lock"
RUNNER_ACTIVE_PATH = SCRIPT_AUTO_RUN_LOG_DIR / ".trae_idea_runner_active.json"
RUNNER_ACTIVE_DIR = SCRIPT_AUTO_RUN_LOG_DIR / ".trae_idea_runner_active.d"
SHARED_WRITE_LOCK_PATH = SCRIPT_AUTO_RUN_LOG_DIR / ".trae_idea_shared_write.lock"
TRAE_WINDOW_LOCK_PATH = SCRIPT_AUTO_RUN_LOG_DIR / ".trae_idea_window.lock"
TRAE_WINDOW_LOCK_TIMEOUT_SECONDS = 120
PUBLISH_LOCK_PATH = SCRIPT_AUTO_RUN_LOG_DIR / ".trae_idea_publish.lock"
LEGACY_RUNNER_LOCK_PATHS = existing_legacy_paths(SCRIPT_DIR / ".trae_idea_runner.lock")
LEGACY_SHARED_WRITE_LOCK_PATHS = existing_legacy_paths(SCRIPT_DIR / ".trae_idea_shared_write.lock")
LEGACY_TRAE_WINDOW_LOCK_PATHS = existing_legacy_paths(SCRIPT_DIR / ".trae_idea_window.lock")
LEGACY_PUBLISH_LOCK_PATHS = existing_legacy_paths(SCRIPT_DIR / ".trae_idea_publish.lock")
RESULT_CSV_PATH = SCRIPT_AUTO_RUN_LOG_DIR / "solo_runs.csv"
RESULT_XLSX_PATH = SCRIPT_AUTO_RUN_LOG_DIR / "solo_runs.xlsx"
RESULT_DATA_DIR = ROOT / "result"
RESULT_SCREENSHOT_DIR = RESULT_DATA_DIR / "截图"
QC_MARKS_PATH = SCRIPT_AUTO_RUN_LOG_DIR / "qc_failed_projects.json"
SOLO_HELPER = SCRIPT_DIR / "solo.py"
PUBLISH_BRANCH = "pro2"
PUBLISH_FIRST_ROUND_BRANCH = "pro3"
PUBLISH_MULTI_ROUND_BRANCH = "pro2"
PUBLISH_REPO_WEB_URL = "https://github.com/yzz2861/trae_eval"
PUBLISH_REPO_GIT_URL = os.environ.get("SOLO_PUBLISH_REPO_GIT_URL", "git@github.com:yzz2861/trae_eval.git")
SKIP_GIT_INTERACTION = os.environ.get("SOLO_SKIP_GIT_INTERACTION", "1") != "0"
SERVICE_EXCEPTION_MARKERS = (
    "服务端异常",
    "failed to recv azure server response",
    "code_comp_fail",
)
STOPPED_RUN_MARKERS = (
    "手动终止输出",
    "输出已终止",
    "手动停止输出",
    "Trae 已停止运行",
    "trae 已停止运行",
)
DELETE_CONFIRMATION_MARKERS = (
    "检测到高风险命令",
    "高风险命令",
    "运行命令可能会带来严重后果",
    "是否仍要",
    "确认删除",
    "删除确认",
    "delete confirmation",
    "are you sure",
)
DELETE_COMMAND_PATTERNS = (
    r"\brm\b",
    r"\brmdir\b",
    r"\bunlink\b",
    r"\bshred\b",
    r"\bpkill\b",
    r"\bdel\b",
    r"\berase\b",
    r"\brd\b",
    r"\bRemove-Item\b",
    r"\bfind\b.*\b-delete\b",
    r"\bgit\b.*\b(clean|restore|checkout|stash\s+clear)\b",
)
DELETE_SKIP_BUTTON_NEEDLES = ["跳过", "Skip"]
CHMOD_COMMAND_PATTERNS = (r"\bchmod\b",)
SERVICE_EXCEPTION_CONTINUE_BUTTON_NEEDLES = ["继续", "Continue"]
STOPPED_RETRY_BUTTON_NEEDLES = ["重试", "Retry"]
CONFIRMATION_CANCEL_BUTTON_NEEDLES = ["取消", "Cancel"]
HIGH_RISK_RUN_BUTTON_NEEDLES = ["仍要运行", "确认运行", "运行", "Run", "Continue", "继续"]
RUNNER_SERVICE_EXCEPTION_EXIT_CODE = 3
RUNNER_STOPPED_EXIT_CODE = 4
CODEX_QA_MODEL = "gpt-5.5"
CODEX_QA_REASONING_EFFORT = "high"
CODEX_PUSH_RECOVERY_REPORT = "codex_push_recovery_report.md"
DEFAULT_TASK_TYPE = "0-1代码生成"
TASK_TYPE_OPTIONS = ("Bug修复", "0-1代码生成", "Feature迭代", "代码理解", "代码重构", "代码测试")
DEFAULT_MODIFICATION_SCOPE = "跨系统多模块"
MODIFICATION_SCOPE_OPTIONS = ("单文件", "模块内多文件", "跨模块多文件", "跨系统多模块")
MAX_TABLE_TEXT_CHARS = 30000
SOLO_QC_HEADERS = [
    "session_id",
    "轮次",
    "提示词",
    "任务类型",
    "业务领域",
    "修改范围",
    "任务是否完成",
    "产物及过程是否满意",
    "不满意原因",
    "远端Github地址",
    "分支文件夹",
    "截图",
    "日志轨迹",
]
DEFAULT_FEISHU_APP_ID = os.environ.get("FEISHU_APP_ID", "cli_a96c06c136789bd2")
DEFAULT_FEISHU_APP_SECRET = os.environ.get("FEISHU_APP_SECRET", "iMaUswb6BKfnE8lxOmQEvdK0ANpgv601")
DEFAULT_FEISHU_BASE_TOKEN = os.environ.get("FEISHU_BASE_TOKEN") or os.environ.get("BASE_TOKEN", "NVLabI2piaNiVHsn6GYchYlmnRt")
DEFAULT_FEISHU_TABLE_ID = os.environ.get("FEISHU_TABLE_ID") or os.environ.get("TABLE_ID", "tblONK81uEWGM2jF")
TRAE_APP_NAME = os.environ.get("TRAE_APP_NAME", "Trae CN")
TRAE_APP_EXECUTABLE_MARKER = os.environ.get("TRAE_APP_EXECUTABLE_MARKER", f"{TRAE_APP_NAME}.app/Contents/MacOS/Electron")
SCREENSHOT_OVERLAY_APPS_TO_HIDE = ("Feishu", "Lark", "飞书")
SCREENSHOT_WINDOW_WIDTH = 1440
SCREENSHOT_WINDOW_HEIGHT = 900
SCREENSHOT_WINDOW_X = 35
SCREENSHOT_WINDOW_Y = 50
PUBLISH_EXCLUDED_NAMES = {
    ".DS_Store",
    ".env",
    ".env.local",
    ".git",
    ".mypy_cache",
    ".pytest_cache",
    ".ruff_cache",
    ".venv",
    "__pycache__",
    "node_modules",
    "venv",
}
PUBLISH_EXCLUDED_SUFFIXES = (".pyc", ".pyo", ".db", ".sqlite", ".sqlite3")


def trae_cli_command(*args: str) -> list[str]:
    env_cli = os.environ.get("TRAE_CLI")
    found_cli = shutil.which("trae-cn")
    if env_cli:
        return [env_cli, *args]
    if found_cli:
        return [found_cli, *args]
    return ["open", "-na", TRAE_APP_NAME, "--args", *args]


@dataclass
class Idea:
    idea_id: str
    title: str
    fields: dict[str, str]
    raw: str


class Trace:
    def __init__(self, run_dir: Path) -> None:
        self.run_dir = run_dir
        self.path = run_dir / "trace.jsonl"

    def write(self, event: str, **data: Any) -> None:
        row = {
            "ts": datetime.now().isoformat(timespec="seconds"),
            "event": event,
            **data,
        }
        with self.path.open("a", encoding="utf-8") as f:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")
        print(f"[{row['ts']}] {event}: {data}", flush=True)


def lock_paths(paths: list[Path], *, nonblocking: bool) -> list[Any]:
    handles = []
    try:
        for path in paths:
            path.parent.mkdir(parents=True, exist_ok=True)
            handle = path.open("a+", encoding="utf-8")
            fcntl.flock(handle.fileno(), fcntl.LOCK_EX | (fcntl.LOCK_NB if nonblocking else 0))
            handles.append(handle)
    except Exception:
        for handle in reversed(handles):
            try:
                fcntl.flock(handle.fileno(), fcntl.LOCK_UN)
            finally:
                handle.close()
        raise
    return handles


def unlock_paths(handles: list[Any]) -> None:
    for handle in reversed(handles):
        try:
            fcntl.flock(handle.fileno(), fcntl.LOCK_UN)
        finally:
            handle.close()


class RunnerLock:
    def __init__(self, path: Path, timeout_seconds: int, legacy_paths: list[Path] | None = None) -> None:
        self.path = path
        self.timeout_seconds = timeout_seconds
        self.paths = [path, *(legacy_paths or [])]
        self.handles: list[Any] = []

    def __enter__(self) -> "RunnerLock":
        deadline = time.time() + self.timeout_seconds
        while True:
            try:
                self.handles = lock_paths(self.paths, nonblocking=True)
                payload = json.dumps({
                    "pid": os.getpid(),
                    "started_at": datetime.now().isoformat(timespec="seconds"),
                }, ensure_ascii=False) + "\n"
                for handle in self.handles:
                    handle.seek(0)
                    handle.truncate()
                    handle.write(payload)
                    handle.flush()
                return self
            except BlockingIOError:
                unlock_paths(self.handles)
                self.handles = []
                if self.timeout_seconds <= 0 or time.time() >= deadline:
                    raise SystemExit(f"已有 trae_idea_runner 在运行，跳过以避免混用 Trae 前台窗口: {self.path}")
                time.sleep(1)

    def __exit__(self, exc_type: object, exc: object, tb: object) -> None:
        unlock_paths(self.handles)
        self.handles = []


@contextlib.contextmanager
def shared_write_lock() -> Any:
    handles = lock_paths([SHARED_WRITE_LOCK_PATH, *LEGACY_SHARED_WRITE_LOCK_PATHS], nonblocking=False)
    try:
        yield
    finally:
        unlock_paths(handles)


@contextlib.contextmanager
def traced_file_lock(
    path: Path,
    trace: "Trace | None",
    purpose: str,
    timeout_seconds: int | None = None,
    legacy_paths: list[Path] | None = None,
) -> Any:
    paths = [path, *(legacy_paths or [])]
    started = time.time()
    if trace:
        trace.write(
            f"{purpose}_lock_wait",
            path=str(path),
            legacy_paths=[str(item) for item in paths[1:]],
            timeout_seconds=timeout_seconds,
        )
    handles: list[Any] = []
    acquired = False
    deadline = time.time() + timeout_seconds if timeout_seconds and timeout_seconds > 0 else None
    try:
        while True:
            try:
                if deadline is None:
                    handles = lock_paths(paths, nonblocking=False)
                else:
                    handles = lock_paths(paths, nonblocking=True)
                break
            except BlockingIOError:
                unlock_paths(handles)
                handles = []
                waited = round(time.time() - started, 3)
                if deadline is not None and time.time() >= deadline:
                    if trace:
                        trace.write(f"{purpose}_lock_timeout", path=str(path), waited_seconds=waited)
                    raise TimeoutError(f"等待锁超时 {timeout_seconds}s: {path}")
                time.sleep(1)
        waited = round(time.time() - started, 3)
        payload = json.dumps({
            "pid": os.getpid(),
            "purpose": purpose,
            "acquired_at": datetime.now().isoformat(timespec="seconds"),
        }, ensure_ascii=False) + "\n"
        for handle in handles:
            handle.seek(0)
            handle.truncate()
            handle.write(payload)
            handle.flush()
        acquired = True
        if trace:
            trace.write(f"{purpose}_lock_acquired", path=str(path), waited_seconds=waited)
        yield
    finally:
        if acquired and trace:
            trace.write(f"{purpose}_lock_released", path=str(path))
        unlock_paths(handles)


def trae_window_lock(trace: "Trace | None", purpose: str = "trae_window") -> Any:
    return traced_file_lock(
        TRAE_WINDOW_LOCK_PATH,
        trace,
        f"trae_window_{purpose}",
        TRAE_WINDOW_LOCK_TIMEOUT_SECONDS,
        LEGACY_TRAE_WINDOW_LOCK_PATHS,
    )


def publish_lock_path(branch_name: str) -> Path:
    safe_branch = re.sub(r"[^A-Za-z0-9_.-]+", "_", branch_name or PUBLISH_BRANCH)
    return SCRIPT_AUTO_RUN_LOG_DIR / f".trae_idea_publish_{safe_branch}.lock"


def publish_lock(trace: "Trace | None", branch_name: str = PUBLISH_BRANCH) -> Any:
    lock_path = publish_lock_path(branch_name)
    return traced_file_lock(lock_path, trace, f"publish_{branch_name}", legacy_paths=[])


def skipped_publish_result(reason: str = "git_interaction_disabled") -> dict[str, str]:
    return {
        "status": "skipped",
        "commit": "",
        "remote": "",
        "branch": "",
        "reason": reason,
    }


def run(cmd: list[str], *, cwd: Path | None = None, check: bool = True) -> subprocess.CompletedProcess[str]:
    return subprocess.run(cmd, cwd=str(cwd) if cwd else None, text=True, capture_output=True, check=check)


def osascript(script: str, *, check: bool = False, timeout_seconds: float | None = 15) -> str:
    try:
        if "\n" in script.strip():
            proc = subprocess.run(
                ["osascript"],
                input=script,
                text=True,
                capture_output=True,
                check=check,
                timeout=timeout_seconds,
            )
        else:
            proc = subprocess.run(
                ["osascript", "-e", script],
                text=True,
                capture_output=True,
                check=check,
                timeout=timeout_seconds,
            )
    except subprocess.TimeoutExpired as exc:
        if check:
            raise
        return ((exc.stdout or "") + "\n" + (exc.stderr or "") + "\n[osascript timeout]").strip()
    return (proc.stdout or proc.stderr).strip()


def shell_quote(text: str) -> str:
    return shlex.quote(text)


def red_text(text: str) -> str:
    return f"\033[31m{text}\033[0m"


def is_github_push_failure(error: BaseException | str) -> bool:
    text = str(error)
    markers = (
        "Git push",
        "push_failed",
        "clone 发布仓库失败",
        "推送发布分支失败",
        "远端 push 校验失败",
    )
    return any(marker in text for marker in markers)


def tail_text(path: Path, limit: int = 6000) -> str:
    if not path.exists():
        return ""
    text = path.read_text(encoding="utf-8", errors="ignore")
    return text[-limit:]


def set_clipboard(text: str) -> None:
    p = subprocess.Popen(["pbcopy"], stdin=subprocess.PIPE, text=True)
    p.communicate(text)
    if p.returncode:
        raise RuntimeError("pbcopy failed")


def get_clipboard() -> str:
    return run(["pbpaste"], check=False).stdout


def key_code(key: str, modifiers: list[str] | None = None) -> None:
    using = ""
    if modifiers:
        using = " using {" + ", ".join(modifiers) + " down" + "}" if len(modifiers) == 1 else (
            " using {" + ", ".join(f"{m} down" for m in modifiers) + "}"
        )
    osascript(f'tell application "System Events" to key code {key}{using}', check=False)


def click_at(x: int, y: int) -> None:
    osascript(f'tell application "System Events" to click at {{{x}, {y}}}', check=False)


def double_click_at(x: int, y: int) -> None:
    click_at(x, y)
    time.sleep(0.08)
    click_at(x, y)


def post_jxa_mouse_click(x: int, y: int) -> None:
    script = f'''
ObjC.import('CoreGraphics')
ObjC.import('Foundation')
const point = $.CGPointMake({x}, {y})
function post(type) {{
  const event = $.CGEventCreateMouseEvent(null, type, point, $.kCGMouseButtonLeft)
  $.CGEventPost($.kCGHIDEventTap, event)
}}
post($.kCGEventMouseMoved)
$.NSThread.sleepForTimeInterval(0.05)
post($.kCGEventLeftMouseDown)
$.NSThread.sleepForTimeInterval(0.04)
post($.kCGEventLeftMouseUp)
'''
    subprocess.run(["osascript", "-l", "JavaScript"], input=script, text=True, capture_output=True, check=False)


def post_jxa_mouse_double_click(x: int, y: int) -> None:
    script = f'''
ObjC.import('CoreGraphics')
ObjC.import('Foundation')
const point = $.CGPointMake({x}, {y})
function post(type) {{
  const event = $.CGEventCreateMouseEvent(null, type, point, $.kCGMouseButtonLeft)
  $.CGEventPost($.kCGHIDEventTap, event)
}}
post($.kCGEventMouseMoved)
$.NSThread.sleepForTimeInterval(0.05)
post($.kCGEventLeftMouseDown)
post($.kCGEventLeftMouseUp)
$.NSThread.sleepForTimeInterval(0.08)
post($.kCGEventLeftMouseDown)
post($.kCGEventLeftMouseUp)
'''
    subprocess.run(["osascript", "-l", "JavaScript"], input=script, text=True, capture_output=True, check=False)


def post_jxa_scroll(x: int, y: int, delta: int, repeats: int) -> None:
    script = f'''
ObjC.import('CoreGraphics')
ObjC.import('Foundation')
const point = $.CGPointMake({x}, {y})
const move = $.CGEventCreateMouseEvent(null, $.kCGEventMouseMoved, point, $.kCGMouseButtonLeft)
$.CGEventPost($.kCGHIDEventTap, move)
for (let i = 0; i < {repeats}; i++) {{
  const event = $.CGEventCreateScrollWheelEvent(null, $.kCGScrollEventUnitPixel, 1, {delta})
  $.CGEventPost($.kCGHIDEventTap, event)
  $.NSThread.sleepForTimeInterval(0.02)
}}
'''
    subprocess.run(["osascript", "-l", "JavaScript"], input=script, text=True, capture_output=True, check=False)


def keystroke(text: str, modifiers: list[str] | None = None) -> None:
    using = ""
    if modifiers:
        using = " using {" + ", ".join(f"{m} down" for m in modifiers) + "}"
    osascript(f'tell application "System Events" to keystroke {json.dumps(text)}{using}', check=False)


def activate_trae() -> None:
    osascript(f'tell application "{TRAE_APP_NAME}" to activate', check=False)
    time.sleep(1.0)


def quit_trae() -> None:
    osascript(f'tell application "{TRAE_APP_NAME}" to quit', check=False)


def trae_main_pid() -> int | None:
    proc = run(["ps", "-axo", "pid=,args="], check=False)
    for line in proc.stdout.splitlines():
        if TRAE_APP_EXECUTABLE_MARKER not in line:
            continue
        try:
            return int(line.strip().split(None, 1)[0])
        except (IndexError, ValueError):
            continue
    return None


def hide_overlay_apps_for_screenshot() -> None:
    app_names = ", ".join(json.dumps(name) for name in SCREENSHOT_OVERLAY_APPS_TO_HIDE)
    script = f'''
tell application "System Events"
  repeat with appName in {{{app_names}}}
    try
      if exists application process (appName as text) then
        set visible of application process (appName as text) to false
      end if
    end try
  end repeat
end tell
'''
    osascript(script, check=False)
    time.sleep(0.2)


def normalize_trae_window_for_screenshot() -> None:
    hide_overlay_apps_for_screenshot()
    activate_trae()
    pid = trae_main_pid()
    if pid:
        process_selector = f'first application process whose unix id is {pid}'
    else:
        process_selector = 'first application process whose name is "Electron"'
    script = f'''
tell application "System Events"
  set traeProcess to {process_selector}
  set frontmost of traeProcess to true
  tell traeProcess
    if (count of windows) is 0 then return
    set w to missing value
    repeat with candidate in windows
      try
        if (value of attribute "AXMain" of candidate) is true then
          set w to candidate
          exit repeat
        end if
      end try
    end repeat
    if w is missing value then set w to window 1
    try
      set value of attribute "AXFullScreen" of w to false
      delay 0.3
    end try
    set position of w to {{{SCREENSHOT_WINDOW_X}, {SCREENSHOT_WINDOW_Y}}}
    set size of w to {{{SCREENSHOT_WINDOW_WIDTH}, {SCREENSHOT_WINDOW_HEIGHT}}}
  end tell
end tell
'''
    osascript(script, check=False)
    time.sleep(0.5)


def trae_window_rect() -> tuple[int, int, int, int] | None:
    normalize_trae_window_for_screenshot()
    pid = trae_main_pid()
    if pid:
        process_selector = f'first application process whose unix id is {pid}'
    else:
        process_selector = 'first application process whose name is "Electron"'
    script = f'''
tell application "System Events"
  set traeProcess to {process_selector}
  tell traeProcess
    if (count of windows) is 0 then return ""
    set w to missing value
    repeat with candidate in windows
      try
        if (value of attribute "AXMain" of candidate) is true then
          set w to candidate
          exit repeat
        end if
      end try
    end repeat
    if w is missing value then set w to window 1
    set {{x, y}} to position of w
    set {{ww, hh}} to size of w
    return (x as text) & "," & (y as text) & "," & (ww as text) & "," & (hh as text)
  end tell
end tell
'''
    raw = osascript(script, check=False).strip()
    parts = raw.split(",")
    if len(parts) != 4:
        return None
    try:
        x, y, width, height = [int(float(p)) for p in parts]
    except ValueError:
        return None
    if width <= 0 or height <= 0:
        return None
    return x, y, width, height


def trae_window_titles() -> list[str]:
    pid = trae_main_pid()
    if pid:
        process_selector = f'first application process whose unix id is {pid}'
    else:
        process_selector = 'first application process whose name is "Electron"'
    script = f'''
tell application "System Events"
  tell {process_selector}
    if (count of windows) is 0 then return ""
    return name of every window
  end tell
end tell
'''
    raw = osascript(script, check=False).strip()
    if not raw:
        return []
    return [part.strip() for part in raw.split(",") if part.strip()]


def trae_main_window_title() -> str:
    pid = trae_main_pid()
    if pid:
        process_selector = f'first application process whose unix id is {pid}'
    else:
        process_selector = 'first application process whose name is "Electron"'
    script = f'''
tell application "System Events"
  try
    set traeProcess to {process_selector}
  on error
    return ""
  end try
  tell traeProcess
    if (count of windows) is 0 then return ""
    repeat with candidate in windows
      try
        if (value of attribute "AXMain" of candidate) is true then return name of candidate as text
      end try
    end repeat
    return name of window 1 as text
  end tell
end tell
'''
    return osascript(script, check=False).strip()


def normalize_title_parts(title_parts: Any) -> list[str]:
    if isinstance(title_parts, (str, Path)):
        raw_parts = [str(title_parts)]
    else:
        try:
            raw_parts = [str(part) for part in title_parts]
        except TypeError:
            raw_parts = [str(title_parts)]
    result: list[str] = []
    seen: set[str] = set()
    for part in raw_parts:
        part = part.strip()
        if not part or part in seen:
            continue
        seen.add(part)
        result.append(part)
    return result


def raise_trae_window_by_titles(title_parts: Any, trace: Trace | None = None) -> bool:
    parts = normalize_title_parts(title_parts)
    if not parts:
        return False
    apple_parts = ", ".join(json.dumps(part) for part in parts)
    script = f'''
tell application "System Events"
  set titleParts to {{{apple_parts}}}
  repeat with p in application processes
    try
      if (name of p as text) is "Electron" or (name of p as text) is "{TRAE_APP_NAME}" then
        repeat with i from 1 to (count of windows of p)
          try
            set windowName to name of window i of p as text
            repeat with titlePart in titleParts
              if windowName contains (titlePart as text) then
                set frontmost of p to true
                try
                  set value of attribute "AXMain" of window i of p to true
                end try
                try
                  set value of attribute "AXFocused" of window i of p to true
                end try
                perform action "AXRaise" of window i of p
                return (titlePart as text) & linefeed & windowName
              end if
            end repeat
          end try
        end repeat
      end if
    end try
  end repeat
  return ""
end tell
'''
    raw = osascript(script, check=False).strip()
    raw_lines = raw.splitlines()
    matched_part = raw_lines[0].strip() if raw_lines else ""
    raised = raw_lines[1].strip() if len(raw_lines) > 1 else raw
    time.sleep(0.2)
    current = trae_main_window_title()
    verified_part = next((part for part in parts if current and part in current), "")
    verified = bool(verified_part)
    if not verified and raised:
        titles = trae_window_titles()
        attempts = min(8, max(3, len(titles) + 1))
        activate_trae()
        for attempt in range(attempts):
            current = trae_main_window_title()
            verified_part = next((part for part in parts if current and part in current), "")
            if verified_part:
                verified = True
                if trace:
                    trace.write("trae_window_cycle_matched", title_part=verified_part, title=current, attempt=attempt)
                break
            key_code("50", ["command"])  # Command+` cycles windows in the active app.
            time.sleep(0.35)
    if trace:
        trace.write(
            "trae_window_raise_by_title",
            title_parts=parts,
            title_part=matched_part or (parts[0] if parts else ""),
            matched_part=matched_part,
            raised=raised,
            current=current,
            verified=verified,
            verified_part=verified_part,
        )
    time.sleep(0.2)
    return verified


def raise_trae_window_by_title(title_part: str, trace: Trace | None = None) -> bool:
    return raise_trae_window_by_titles([title_part], trace)


def close_trae_window_by_title(title_part: str, trace: Trace | None = None) -> bool:
    if not title_part:
        return False
    if not raise_trae_window_by_title(title_part, trace):
        if trace:
            trace.write("trae_window_close_by_title", title_part=title_part, closed="", verified=False)
        return False
    script = f'''
tell application "System Events"
  set p to first process whose frontmost is true
  if (count of windows of p) is 0 then return ""
  set w to missing value
  repeat with candidate in windows of p
    try
      if (value of attribute "AXMain" of candidate) is true then
        set w to candidate
        exit repeat
      end if
    end try
  end repeat
  if w is missing value then set w to window 1 of p
  set windowName to name of w as text
  if windowName does not contain {json.dumps(title_part)} then return ""
  try
    perform action "AXPress" of button 1 of w
  on error
    try
      click button 1 of w
    on error
      keystroke "w" using command down
    end try
  end try
  return windowName
end tell
'''
    closed = osascript(script, check=False).strip()
    if trace:
        trace.write("trae_window_close_by_title", title_part=title_part, closed=closed, verified=bool(closed))
    time.sleep(0.6)
    return bool(closed)


def screenshot(path: Path) -> None:
    rect = trae_window_rect()
    if rect:
        x, y, width, height = rect
        result = run(["screencapture", "-x", "-R", f"{x},{y},{width},{height}", str(path)], check=False)
        if result.returncode == 0 and path.exists() and path.stat().st_size > 0:
            return
    run(["screencapture", "-x", str(path)], check=False)


def read_json_file(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return default


def write_active_runner(idea_id: str, run_dir: Path, workspace: Path, mode: str) -> None:
    payload = {
        "pid": os.getpid(),
        "idea_id": idea_id,
        "run_dir": str(run_dir),
        "workspace": str(workspace),
        "mode": mode,
        "ts": datetime.now().isoformat(timespec="seconds"),
    }
    RUNNER_ACTIVE_DIR.mkdir(parents=True, exist_ok=True)
    (RUNNER_ACTIVE_DIR / f"{os.getpid()}.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    RUNNER_ACTIVE_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def clear_active_runner() -> None:
    pid = os.getpid()
    try:
        (RUNNER_ACTIVE_DIR / f"{pid}.json").unlink()
    except FileNotFoundError:
        pass
    active = read_json_file(RUNNER_ACTIVE_PATH, {})
    if isinstance(active, dict) and active.get("pid") == pid:
        try:
            RUNNER_ACTIVE_PATH.unlink()
        except FileNotFoundError:
            pass


def idea_session_filename(idea_id: str) -> str:
    digits = "".join(re.findall(r"\d+", idea_id))
    suffix = digits or idea_id
    return f"sessionId_{suffix}.txt"


def parse_session_identifier(value: str) -> tuple[str, str] | None:
    m = re.search(r":([0-9a-f]+)_([0-9a-f]+)\.", value)
    if not m:
        return None
    return m.group(1), m.group(2)


def find_trae_log_dir_for_session(run_dir: Path) -> Path | None:
    base = Path.home() / "Library/Application Support/Trae CN/logs"
    if not base.exists():
        return None
    for session_file in sorted(run_dir.glob("sessionId_*.txt"), key=lambda p: p.stat().st_mtime, reverse=True):
        parsed = parse_session_identifier(session_file.read_text(encoding="utf-8", errors="ignore"))
        if not parsed:
            continue
        trace_id, session_id = parsed
        for log_dir in sorted((p for p in base.iterdir() if p.is_dir()), key=lambda p: p.stat().st_mtime, reverse=True):
            modular_dir = log_dir / "Modular"
            if not modular_dir.exists():
                continue
            for stdout_log in modular_dir.glob("ai-agent_*_stdout.log"):
                text = stdout_log.read_text(encoding="utf-8", errors="ignore")
                if trace_id in text and session_id in text:
                    return log_dir
    return None


def load_state() -> dict[str, Any]:
    if not STATE_PATH.exists():
        return {"used": []}
    try:
        return json.loads(STATE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {"used": []}


def save_state(state: dict[str, Any]) -> None:
    STATE_PATH.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def mark_idea_used(idea_id: str, run_dir: Path, trace: Trace) -> None:
    with shared_write_lock():
        state = load_state()
        used = state.setdefault("used", [])
        if idea_id not in used:
            used.append(idea_id)
        state["last_run"] = {"idea_id": idea_id, "run_dir": str(run_dir), "ts": datetime.now().isoformat(timespec="seconds")}
        save_state(state)
    trace.write("state_updated", path=str(STATE_PATH))


def mark_idea_failed(idea_id: str, run_dir: Path, reason: str, trace: Trace) -> None:
    with shared_write_lock():
        state = load_state()
        start_failed = state.setdefault("start_failed", {})
        if not isinstance(start_failed, dict):
            start_failed = {}
            state["start_failed"] = start_failed
        start_failed[idea_id] = {
            "run_dir": str(run_dir),
            "reason": reason,
            "ts": datetime.now().isoformat(timespec="seconds"),
        }
        started = state.get("started", [])
        if isinstance(started, list):
            state["started"] = [item for item in started if item != idea_id]
        state["last_failed"] = {"idea_id": idea_id, "run_dir": str(run_dir), "reason": reason, "ts": datetime.now().isoformat(timespec="seconds")}
        save_state(state)
    trace.write("state_failed_updated", path=str(STATE_PATH), reason=reason)


def infer_idea_id_from_run_dir(run_dir: Path) -> str | None:
    idea_json = run_dir / "idea.json"
    if idea_json.exists():
        try:
            data = json.loads(idea_json.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            data = {}
        if data.get("idea_id"):
            return str(data["idea_id"])
    m = re.search(r"-([A-Za-z]+[0-9A-Za-z_-]*)$", run_dir.name)
    return m.group(1) if m else None


def workspace_from_run_dir(run_dir: Path, fallback: Path) -> Path:
    idea_json = run_dir / "idea.json"
    if idea_json.exists():
        try:
            data = json.loads(idea_json.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            data = {}
        if data.get("workspace"):
            return Path(str(data["workspace"]))
    return fallback


def resolve_finish_run_dir(idea_id: str | None, run_dir_arg: str | None) -> Path:
    if run_dir_arg:
        run_dir = Path(run_dir_arg).expanduser()
        if not run_dir.exists():
            raise SystemExit(f"指定的 run 目录不存在: {run_dir}")
        return run_dir

    if idea_id:
        matches = sorted(
            (p for p in RUNS_DIR.glob(f"*-{idea_id}") if not (p / ".dry_run").exists()),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        if matches:
            return matches[0]
        raise SystemExit(f"找不到 {idea_id} 对应的 run 目录，可以补 --run-dir 指定。")

    last_run = load_state().get("last_run", {}).get("run_dir")
    if last_run and Path(last_run).exists():
        return Path(last_run)

    matches = sorted(
        (p for p in RUNS_DIR.iterdir() if p.is_dir() and not (p / ".dry_run").exists()),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if matches:
        return matches[0]
    raise SystemExit("找不到可收尾的 run 目录。")


def parse_ideas(path: Path) -> list[Idea]:
    text = path.read_text(encoding="utf-8")
    matches = list(re.finditer(r"^##\s+([A-Za-z0-9_-]+)\s+-\s+(.+?)\s*$", text, re.M))
    ideas: list[Idea] = []
    for i, m in enumerate(matches):
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        block = text[start:end].strip()
        fields: dict[str, str] = {}
        for line in block.splitlines():
            fm = re.match(r"^-\s*([^：:]+)[：:]\s*(.*)$", line.strip())
            if fm:
                fields[fm.group(1).strip()] = fm.group(2).strip()
        ideas.append(Idea(m.group(1).strip(), m.group(2).strip(), fields, block))
    return ideas


def choose_idea(ideas: list[Idea], idea_id: str | None, order: str) -> Idea:
    if idea_id:
        for idea in ideas:
            if idea.idea_id == idea_id:
                return idea
        raise SystemExit(f"找不到点子 ID: {idea_id}")

    used = set(load_state().get("used", []))
    candidates = list(reversed(ideas)) if order == "newest" else list(ideas)
    for idea in candidates:
        if idea.idea_id not in used:
            return idea
    raise SystemExit("没有未运行的点子了。可以传 --idea-id 指定重跑某条。")


def unquote_prompt_value(value: str) -> str:
    value = value.strip()
    quote_pairs = {
        "'": "'",
        '"': '"',
        "“": "”",
        "‘": "’",
    }
    while len(value) >= 2 and value[0] in quote_pairs and value[-1] == quote_pairs[value[0]]:
        value = value[1:-1].strip()
    return value


def prompt_from_idea(idea: Idea) -> tuple[str, str]:
    inline_prompt = idea.fields.get("提示词")
    if inline_prompt:
        return unquote_prompt_value(inline_prompt), "idea-history:提示词"

    prompt_file = idea.fields.get("提示词文件")
    if prompt_file:
        prompt_file_value = unquote_prompt_value(prompt_file)
        if Path(prompt_file_value).exists():
            return Path(prompt_file_value).read_text(encoding="utf-8"), prompt_file_value
        return prompt_file_value, "idea-history:提示词文件"

    lines = [
        f"请基于这个点子从 0 到 1 实现一个可运行项目：{idea.title}。",
    ]
    for key in ("核心场景", "关键机制", "形态", "业务领域"):
        value = idea.fields.get(key)
        if value:
            lines.append(f"{key}：{value}")
    lines.append("请补齐必要的前后端/脚本/数据/README，保证我能按 README 运行并看到核心流程。")
    return "\n".join(lines), "synthesized-from-idea-history"


def latest_renderer_log() -> Path | None:
    base = Path.home() / "Library/Application Support/Trae CN/logs"
    logs = sorted(base.glob("*/window*/renderer.log"), key=lambda p: p.stat().st_mtime, reverse=True)
    return logs[0] if logs else None


def read_recent_renderer_log(lines: int = 500) -> str:
    path = latest_renderer_log()
    if not path:
        return ""
    data = path.read_text(encoding="utf-8", errors="ignore").splitlines()
    return "\n".join(data[-lines:])


def read_recent_renderer_log_for_workspace(workspace: Path, lines: int = 500) -> tuple[str, Path | None, Path | None]:
    log_dir = find_trae_log_dir_for_workspace(workspace)
    if not log_dir:
        return "", None, None
    renderer_log = find_renderer_log_for_workspace_activity(log_dir, workspace)
    renderer_logs = sorted(log_dir.glob("window*/renderer.log"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not renderer_log and renderer_logs:
        renderer_log = renderer_logs[0]
    if not renderer_log:
        return "", log_dir, None
    data = renderer_log.read_text(encoding="utf-8", errors="ignore").splitlines()
    return "\n".join(data[-lines:]), log_dir, renderer_log


def read_recent_text(path: Path, lines: int = 500) -> str:
    try:
        data = path.read_text(encoding="utf-8", errors="ignore").splitlines()
    except OSError:
        return ""
    return "\n".join(data[-lines:])


def find_trae_log_dir_for_workspace(workspace: Path) -> Path | None:
    base = Path.home() / "Library/Application Support/Trae CN/logs"
    if not base.exists():
        return None
    needle = str(workspace)
    for log_dir in sorted((p for p in base.iterdir() if p.is_dir()), key=lambda p: p.stat().st_mtime, reverse=True):
        main_log = log_dir / "main.log"
        if main_log.exists():
            try:
                if needle in main_log.read_text(encoding="utf-8", errors="ignore"):
                    return log_dir
            except OSError:
                pass
        for renderer_log in sorted(log_dir.glob("window*/renderer.log"), key=lambda p: p.stat().st_mtime, reverse=True):
            try:
                if needle in renderer_log.read_text(encoding="utf-8", errors="ignore"):
                    return log_dir
            except OSError:
                continue
    return None


def renderer_workspace_score(renderer_log: Path, workspace: Path) -> tuple[int, int, int]:
    try:
        lines = renderer_log.read_text(encoding="utf-8", errors="ignore").splitlines()
    except OSError:
        return (0, 0, 0)
    workspace_text = str(workspace)
    idea_id = workspace.name
    file_write_refs = 0
    workspace_refs = 0
    session_refs = 0
    for line in lines:
        has_workspace = workspace_text in line
        if has_workspace:
            workspace_refs += 1
            if "ChatSnapshotService._writeAndSaveFile" in line or '"filePath"' in line:
                file_write_refs += 1
        if idea_id and idea_id in line:
            workspace_refs += 1
        if "session_id" in line and "event:" in line:
            session_refs += 1
    return (file_write_refs, workspace_refs, session_refs)


def find_renderer_log_for_workspace_activity(log_dir: Path, workspace: Path) -> Path | None:
    best: tuple[tuple[int, int, int], float, Path] | None = None
    for renderer_log in sorted(log_dir.glob("window*/renderer.log"), key=lambda p: p.stat().st_mtime, reverse=True):
        score = renderer_workspace_score(renderer_log, workspace)
        if score[0] <= 0 and score[1] <= 0:
            continue
        item = (score, renderer_log.stat().st_mtime, renderer_log)
        if best is None or item[:2] > best[:2]:
            best = item
    return best[2] if best else None


def parse_json_after(line: str, marker: str) -> Any | None:
    idx = line.find(marker)
    if idx < 0:
        return None
    payload = line[idx + len(marker):].strip()
    try:
        return json.loads(payload)
    except json.JSONDecodeError:
        return None


def extract_event_params(line: str) -> tuple[str, dict[str, Any]] | None:
    m = re.search(r"event:\s+([A-Za-z0-9_]+)\s+;\s+params:\s+(\{.*\})\s*$", line)
    if not m:
        return None
    try:
        return m.group(1), json.loads(m.group(2))
    except json.JSONDecodeError:
        return None


def unescape_rust_debug_string(value: str) -> str | None:
    try:
        return json.loads('"' + value + '"')
    except json.JSONDecodeError:
        return None


def normalize_tool_name(name: str | None) -> str:
    return {
        "LS": "view_folder",
        "RunCommand": "run_command",
        "TodoWrite": "todo_write",
        "Edit": "edit_file_search_replace",
    }.get(name or "", name or "unknown")


def content_text(message: dict[str, Any]) -> str:
    parts: list[str] = []
    for item in message.get("content") or []:
        if isinstance(item, dict) and item.get("type") == "text":
            parts.append(item.get("text") or "")
    return "".join(parts)


def parse_tool_result(text: str) -> tuple[str, str]:
    status_match = re.search(r"<toolcall_status>(.*?)</toolcall_status>", text, re.S)
    result_match = re.search(r"<toolcall_result>(.*?)</toolcall_result>", text, re.S)
    status = (status_match.group(1).strip() if status_match else "").lower()
    result = result_match.group(1).strip() if result_match else text.strip()
    if status == "done":
        status = "success"
    return status or "unknown", result


def extract_server_history_transcript(log_dir: Path, run_dir: Path, metadata: dict[str, Any], trace: Trace) -> None:
    stdout_logs = sorted((log_dir / "Modular").glob("ai-agent_*_stdout.log"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not stdout_logs:
        trace.write("full_transcript_stdout_not_found", log_dir=str(log_dir))
        return

    session_id = metadata.get("session_id")
    raw_messages: list[dict[str, Any]] = []
    for stdout_log in stdout_logs:
        for line in stdout_log.read_text(encoding="utf-8", errors="ignore").splitlines():
            if "History: HistoryEvent" not in line or "messages: " not in line:
                continue
            if session_id and f"conversation_id: \"{session_id}\"" not in line:
                continue
            m = re.search(r'messages: "((?:\\.|[^"\\])*)", token_usage:', line)
            if not m:
                continue
            decoded = unescape_rust_debug_string(m.group(1))
            if not decoded:
                continue
            try:
                data = json.loads(decoded)
            except json.JSONDecodeError:
                continue
            raw_messages.extend(data.get("raw_messages") or [])

    if not raw_messages:
        trace.write("full_transcript_empty", log_dir=str(log_dir), session_id=session_id)
        return

    call_names: dict[str, str] = {}
    call_args: dict[str, str] = {}
    transcript_items: list[dict[str, Any]] = []
    md: list[str] = ["# Trae Full Transcript", ""]

    for msg in raw_messages:
        role = msg.get("role")
        if role == "user":
            text = content_text(msg)
            user_match = re.search(r"<user_input>\s*(.*?)\s*</user_input>", text, re.S)
            if user_match:
                prompt = user_match.group(1).strip()
                transcript_items.append({"role": "user", "text": prompt})
                md.extend(["## User", "", prompt, ""])
            continue

        if role == "assistant":
            text = content_text(msg).strip()
            if text:
                transcript_items.append({"role": "assistant", "text": text})
                md.extend([text, ""])

            for tool_call in msg.get("tool_calls") or []:
                call_id = tool_call.get("id")
                fc = tool_call.get("function_call") or {}
                tool_name = normalize_tool_name(fc.get("name"))
                args = fc.get("arguments") or ""
                if call_id:
                    call_names[call_id] = tool_name
                    call_args[call_id] = args
                item = {"role": "tool_call", "toolName": tool_name, "tool_call_id": call_id, "arguments": args}
                transcript_items.append(item)
                md.extend(["```text", f"toolName: {tool_name}", "status: running"])
                try:
                    parsed_args = json.loads(args)
                except json.JSONDecodeError:
                    parsed_args = {}
                if isinstance(parsed_args, dict):
                    for key in ("path", "file_path", "filePath", "command"):
                        if parsed_args.get(key):
                            label = "filePath" if key in {"file_path", "filePath"} else key
                            md.append(f"{label}: {parsed_args[key]}")
                elif args:
                    md.append(f"arguments: {args}")
                md.extend(["```", ""])
            continue

        if role == "tool":
            call_id = msg.get("tool_call_id")
            tool_name = call_names.get(call_id, "unknown")
            status, result = parse_tool_result(content_text(msg))
            item = {"role": "tool_result", "toolName": tool_name, "tool_call_id": call_id, "status": status, "result": result}
            transcript_items.append(item)
            md.extend(["```text", f"toolName: {tool_name}", f"status: {status}"])
            args = call_args.get(call_id, "")
            try:
                parsed_args = json.loads(args)
            except json.JSONDecodeError:
                parsed_args = {}
            if isinstance(parsed_args, dict):
                for key in ("path", "file_path", "filePath", "command"):
                    if parsed_args.get(key):
                        label = "filePath" if key in {"file_path", "filePath"} else key
                        md.append(f"{label}: {parsed_args[key]}")
            if result:
                md.append(result)
            md.extend(["```", ""])

    json_path = run_dir / "trae_full_transcript.json"
    md_path = run_dir / "trae_full_transcript.md"
    json_path.write_text(json.dumps({
        "metadata": metadata,
        "source_logs": [str(p) for p in stdout_logs],
        "items": transcript_items,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    md_path.write_text("\n".join(md).rstrip() + "\n", encoding="utf-8")
    trace.write("full_transcript_saved", markdown_path=str(md_path), json_path=str(json_path), items=len(transcript_items))


def extract_trae_trajectory(run_dir: Path, workspace: Path, trace: Trace) -> None:
    session_metadata = load_trae_session_metadata(run_dir)
    metadata_log_dir = Path(session_metadata.get("log_dir", "")) if session_metadata.get("log_dir") else None
    log_dir = (
        metadata_log_dir if metadata_log_dir and metadata_log_dir.exists() else None
    ) or find_trae_log_dir_for_session(run_dir) or find_trae_log_dir_for_workspace(workspace)
    if not log_dir:
        trace.write("trajectory_log_dir_not_found", workspace=str(workspace))
        return

    metadata_log_file = Path(session_metadata.get("log_file", "")) if session_metadata.get("log_file") else None
    workspace_renderer_log = find_renderer_log_for_workspace_activity(log_dir, workspace)
    renderer_log: Path | None = None
    if workspace_renderer_log and metadata_log_file and metadata_log_file.exists():
        workspace_score = renderer_workspace_score(workspace_renderer_log, workspace)
        metadata_score = renderer_workspace_score(metadata_log_file, workspace)
        renderer_log = workspace_renderer_log if workspace_score > metadata_score else metadata_log_file
        if renderer_log != metadata_log_file:
            trace.write(
                "trajectory_renderer_log_corrected_by_workspace",
                recorded_log=str(metadata_log_file),
                workspace_log=str(workspace_renderer_log),
                recorded_score=metadata_score,
                workspace_score=workspace_score,
            )
    elif workspace_renderer_log:
        renderer_log = workspace_renderer_log
    elif metadata_log_file and metadata_log_file.exists():
        renderer_log = metadata_log_file
    else:
        renderer_logs = sorted(log_dir.glob("window*/renderer.log"), key=lambda p: p.stat().st_mtime, reverse=True)
        renderer_log = renderer_logs[0] if renderer_logs else None
    if not renderer_log:
        trace.write("trajectory_renderer_log_not_found", log_dir=str(log_dir))
        return

    lines = renderer_log.read_text(encoding="utf-8", errors="ignore").splitlines()
    metadata: dict[str, Any] = {"log_dir": str(log_dir), "renderer_log": str(renderer_log), "workspace": str(workspace)}
    tool_events: list[dict[str, Any]] = []
    seen_tool_events: set[tuple[str, str]] = set()
    command_by_tool_key: dict[str, dict[str, Any]] = {}
    file_events: list[dict[str, Any]] = []
    copied_events: list[dict[str, Any]] = []
    completion_events: list[dict[str, Any]] = []

    for line in lines:
        event = extract_event_params(line)
        if event:
            event_name, params = event
            if not metadata.get("session_id") and params.get("session_id"):
                metadata.update({
                    "session_id": params.get("session_id"),
                    "message_id": params.get("message_id"),
                    "chat_model": params.get("chat_model"),
                    "agent_name": params.get("agent_name"),
                    "agent_id": params.get("agent_id"),
                })
            tool_type = params.get("tool_type")
            tool_id = params.get("tool_id")
            if event_name in {"tool_call_show", "file_tool_show", "run_script_show", "run_script_success"} and tool_type:
                key = (str(tool_id), event_name)
                if key not in seen_tool_events:
                    seen_tool_events.add(key)
                    tool_events.append({
                        "ts": line.split(" [", 1)[0],
                        "event": event_name,
                        "toolName": tool_type,
                        "tool_id": tool_id,
                        "block_id": params.get("block_id"),
                        "status": "success" if event_name == "run_script_success" else "shown",
                        "runtime_duration": params.get("runtime_duration"),
                    })
            if event_name == "code_comp_complete_shown":
                completion_events.append({
                    "ts": line.split(" [", 1)[0],
                    "event": event_name,
                    "tool_count": params.get("tool_count"),
                    "request_round_count": params.get("request_round_count"),
                    "duration": params.get("duration"),
                    "is_interrupted": params.get("is_interrupted"),
                })
            if event_name == "code_comp_copy_click":
                copied_events.append({"ts": line.split(" [", 1)[0], "event": event_name, "result_type": params.get("result_type")})

        terminal_trace = parse_json_after(line, "[ToolingTerminalTrace]toolcall_run_command_tracing ")
        if isinstance(terminal_trace, dict):
            cat = terminal_trace.get("categories") or {}
            key = cat.get("tool_call_key")
            if key:
                command_by_tool_key[key] = {
                    "tool_call_key": key,
                    "command": cat.get("command"),
                    "exitCode": cat.get("exitCode"),
                    "terminal_type": cat.get("terminal_type"),
                    "blocking": cat.get("blocking"),
                    "metrics": terminal_trace.get("metrics") or {},
                }

        m_result = re.search(r"commandResult=\s+(\[.*\])", line)
        if m_result:
            try:
                result_items = json.loads(m_result.group(1))
            except json.JSONDecodeError:
                result_items = []
            for item in result_items:
                key = item.get("serverCallId")
                if key:
                    command_by_tool_key.setdefault(key, {"tool_call_key": key})
                    command_by_tool_key[key].update({
                        "command": item.get("command") or command_by_tool_key[key].get("command"),
                        "exitCode": item.get("exitCode"),
                        "terminalId": item.get("terminalId"),
                        "chatSessionId": item.get("chatSessionId"),
                        "logs": item.get("logs", []),
                    })

        if "ChatSnapshotService._writeAndSaveFile " in line:
            file_path = line.split("ChatSnapshotService._writeAndSaveFile ", 1)[1].strip()
            file_events.append({"ts": line.split(" [", 1)[0], "toolName": "Write", "status": "success", "filePath": file_path})
        elif "notifyApplyStateChange" in line and '"filePath"' in line and '"confirming"' in line:
            payload = line[line.find("{"):]
            try:
                data = json.loads(payload)
            except json.JSONDecodeError:
                data = {}
            if data.get("filePath"):
                item = {"ts": line.split(" [", 1)[0], "toolName": "Write", "status": "success", "filePath": data["filePath"]}
                if item not in file_events:
                    file_events.append(item)

    if metadata.get("session_id") and str(session_metadata.get("log_file") or "") != str(renderer_log):
        corrected_metadata = dict(session_metadata)
        corrected_metadata.update({
            "session_id": str(metadata.get("session_id") or ""),
            "message_id": str(metadata.get("message_id") or ""),
            "log_file": str(renderer_log),
            "log_dir": str(log_dir),
        })
        (run_dir / "trae_session.json").write_text(json.dumps(corrected_metadata, ensure_ascii=False, indent=2), encoding="utf-8")
        trace.write("trae_session_metadata_corrected_by_workspace", session_id=corrected_metadata["session_id"], log_file=str(renderer_log))

    trajectory = {
        "metadata": metadata,
        "completion_events": completion_events,
        "copy_events": copied_events,
        "tool_events": tool_events,
        "command_events": list(command_by_tool_key.values()),
        "file_events": file_events,
    }

    events_path = run_dir / "trae_trajectory_events.json"
    events_path.write_text(json.dumps(trajectory, ensure_ascii=False, indent=2), encoding="utf-8")

    md_lines = [
        "# Trae Trajectory",
        "",
        f"- workspace: `{workspace}`",
        f"- log_dir: `{log_dir}`",
        f"- renderer_log: `{renderer_log}`",
    ]
    for key in ("session_id", "message_id", "chat_model", "agent_name", "agent_id"):
        if metadata.get(key):
            md_lines.append(f"- {key}: `{metadata[key]}`")
    md_lines.append("")
    md_lines.append("## Completion")
    if completion_events:
        for item in completion_events:
            md_lines.append(
                f"- {item['ts']} complete: tool_count={item.get('tool_count')}, "
                f"rounds={item.get('request_round_count')}, duration_ms={item.get('duration')}, interrupted={item.get('is_interrupted')}"
            )
    else:
        md_lines.append("- no completion event found")

    md_lines.append("")
    md_lines.append("## Commands")
    for item in command_by_tool_key.values():
        if item.get("command"):
            md_lines.append("")
            md_lines.append(f"toolName: run_command")
            md_lines.append(f"status: {'success' if str(item.get('exitCode')) == '0' else 'exit_' + str(item.get('exitCode'))}")
            md_lines.append(f"tool_call_key: {item.get('tool_call_key')}")
            md_lines.append(f"command: {item.get('command')}")
            logs = item.get("logs") or []
            if logs:
                snippet = "\n".join(logs)
                if len(snippet) > 3000:
                    snippet = snippet[:3000] + "\n...<truncated>..."
                md_lines.append("")
                md_lines.append("```text")
                md_lines.append(snippet.rstrip())
                md_lines.append("```")

    md_lines.append("")
    md_lines.append("## File Writes")
    seen_paths: set[str] = set()
    for item in file_events:
        file_path = item.get("filePath")
        if file_path and file_path not in seen_paths:
            seen_paths.add(file_path)
            md_lines.append("")
            md_lines.append("toolName: Write")
            md_lines.append("status: success")
            md_lines.append(f"filePath: {file_path}")

    md_lines.append("")
    md_lines.append("## Tool Timeline")
    for item in tool_events:
        md_lines.append(f"- {item['ts']} {item['event']} {item['toolName']} {item.get('status')} tool_id={item.get('tool_id')}")

    md_path = run_dir / "trae_trajectory.md"
    md_path.write_text("\n".join(md_lines) + "\n", encoding="utf-8")

    (run_dir / "trae_log_dir.txt").write_text(str(log_dir) + "\n", encoding="utf-8")
    extract_server_history_transcript(log_dir, run_dir, metadata, trace)
    trace.write(
        "trajectory_saved",
        log_dir=str(log_dir),
        events_path=str(events_path),
        markdown_path=str(md_path),
        tool_events=len(tool_events),
        command_events=len(command_by_tool_key),
        file_events=len(file_events),
    )


def format_trae_time(ts: str) -> str:
    try:
        dt = datetime.fromisoformat(ts)
    except ValueError:
        return ts
    return f"{dt.year}/{dt.month}/{dt.day} {dt.hour}:{dt.minute:02d}:{dt.second:02d}"


def derive_session_identifier_from_logs(workspace: Path, trace: Trace, run_dir: Path | None = None) -> str | None:
    log_dir: Path | None = None
    if run_dir:
        log_dir = find_trae_log_dir_for_session(run_dir)
        recorded_log_dir = run_dir / "trae_log_dir.txt"
        if not log_dir and recorded_log_dir.exists():
            candidate = Path(recorded_log_dir.read_text(encoding="utf-8").strip())
            if candidate.exists():
                log_dir = candidate
    if not log_dir:
        log_dir = find_trae_log_dir_for_workspace(workspace)
    if not log_dir:
        trace.write("session_id_log_dir_not_found", workspace=str(workspace))
        return None

    user_id = ""
    main_log = log_dir / "main.log"
    if main_log.exists():
        m = re.search(r'"userId":"([^"]+)"', main_log.read_text(encoding="utf-8", errors="ignore"))
        if m:
            user_id = m.group(1)

    session_id = ""
    trace_id = ""
    first_message_id = ""
    started_at = ""
    stdout_logs = sorted((log_dir / "Modular").glob("ai-agent_*_stdout.log"), key=lambda p: p.stat().st_mtime)
    for stdout_log in stdout_logs:
        for line in stdout_log.read_text(encoding="utf-8", errors="ignore").splitlines():
            if "generate start" not in line or "trace_id=" not in line or "message_id=" not in line:
                continue
            m = re.search(
                r"^([0-9T:\-.+]+).*?trace_id=\"([0-9a-f]+)\".*?session_id=([0-9a-f]+).*?message_id=([0-9a-f]+)",
                line,
            )
            if m:
                started_at, trace_id, session_id, first_message_id = m.groups()
                break
        if trace_id:
            break

    final_message_id = ""
    renderer_logs = sorted(log_dir.glob("window*/renderer.log"), key=lambda p: p.stat().st_mtime, reverse=True)
    for renderer_log in renderer_logs:
        for line in renderer_log.read_text(encoding="utf-8", errors="ignore").splitlines():
            event = extract_event_params(line)
            if not event:
                continue
            _event_name, params = event
            if params.get("session_id") == session_id and params.get("message_id"):
                final_message_id = params["message_id"]
                break
        if final_message_id:
            break

    missing = [
        name for name, value in {
            "user_id": user_id,
            "trace_id": trace_id,
            "session_id": session_id,
            "first_message_id": first_message_id,
            "final_message_id": final_message_id,
            "started_at": started_at,
        }.items()
        if not value
    ]
    if missing:
        trace.write("session_id_log_derive_failed", log_dir=str(log_dir), missing=missing)
        return None

    return f".{user_id}:{trace_id}_{session_id}.{first_message_id}.{final_message_id}:Trae CN.T({format_trae_time(started_at)})"


def save_session_identifier(run_dir: Path, idea_id: str, value: str, trace: Trace, source: str) -> Path:
    session_path = run_dir / idea_session_filename(idea_id)
    session_path.write_text(value.strip() + "\n", encoding="utf-8")
    trace.write("session_id_saved", path=str(session_path), source=source, chars=len(value.strip()))
    return session_path


def accessibility_text_snapshot() -> str:
    activate_trae()
    script = r'''
tell application "System Events"
  set p to first process whose frontmost is true
  set outText to ""
  try
    set elems to entire contents of p
    repeat with e in elems
      try
        set outText to outText & " " & ((name of e) as text)
      end try
      try
        set outText to outText & " " & ((description of e) as text)
      end try
    end repeat
  end try
  return outText
end tell
'''
    return osascript(script, check=False, timeout_seconds=4)


def accessibility_text_contains(needles: list[str]) -> bool:
    text = accessibility_text_snapshot()
    return any(n in text for n in needles)


def click_accessibility_button(needles: list[str], *, require_button: bool = False) -> bool:
    activate_trae()
    needle_list = "{" + ",".join(json.dumps(n) for n in needles) + "}"
    require_button_value = "true" if require_button else "false"
    script = f'''
set needles to {needle_list}
set requireButton to {require_button_value}

on matchesNeedle(t, needles)
  repeat with n in needles
    if t contains (n as text) then return true
  end repeat
  return false
end matchesNeedle

on clickMatch(theElement, needles, requireButton, depthLevel)
  tell application "System Events"
    if depthLevel > 18 then return false
    set labelText to ""
    try
      set labelText to labelText & " " & ((name of theElement) as text)
    end try
    try
      set labelText to labelText & " " & ((description of theElement) as text)
    end try
    try
      set labelText to labelText & " " & ((help of theElement) as text)
    end try
    set roleText to ""
    try
      set roleText to ((role of theElement) as text)
    end try
    if my matchesNeedle(labelText, needles) and ((requireButton is false) or roleText contains "button") then
      try
        perform action "AXPress" of theElement
        return true
      end try
      try
        click theElement
        return true
      end try
    end if
    try
      set childElements to UI elements of theElement
      repeat with childElement in childElements
        if my clickMatch(childElement, needles, requireButton, depthLevel + 1) then return true
      end repeat
    end try
    return false
  end tell
end clickMatch

tell application "System Events"
  set p to first process whose frontmost is true
  try
    repeat with e in entire contents of p
      set label to ""
      try
        set label to label & " " & (name of e as text)
      end try
      try
        set label to label & " " & (description of e as text)
      end try
      try
        set label to label & " " & (help of e as text)
      end try
      set roleText to ""
      try
        set roleText to (role of e as text)
      end try
      if my matchesNeedle(label, needles) and ((requireButton is false) or roleText contains "button") then
        try
          perform action "AXPress" of e
          return true
        end try
        try
          click e
          return true
        end try
      end if
    end repeat
  end try
  return my clickMatch(p, needles, requireButton, 0)
end tell
'''
    return osascript(script, check=False, timeout_seconds=6).lower() == "true"


def delete_command_confirmation_visible(text: str) -> bool:
    if not text:
        return False
    lowered = text.lower()
    has_confirmation = any(marker.lower() in lowered for marker in DELETE_CONFIRMATION_MARKERS)
    has_delete_command = any(re.search(pattern, text, flags=re.IGNORECASE) for pattern in DELETE_COMMAND_PATTERNS)
    return has_confirmation and has_delete_command


def chmod_command_confirmation_visible(text: str) -> bool:
    if not text:
        return False
    lowered = text.lower()
    has_confirmation = any(marker.lower() in lowered for marker in DELETE_CONFIRMATION_MARKERS)
    has_chmod_command = any(re.search(pattern, text, flags=re.IGNORECASE) for pattern in CHMOD_COMMAND_PATTERNS)
    return has_confirmation and has_chmod_command


def skip_high_risk_delete_command_if_present(
    trace: Trace,
    run_dir: Path,
    workspace: Path,
    idea_id: str,
    *,
    acquire_lock: bool = True,
) -> bool:
    def run_locked() -> bool:
        raised = raise_trae_window_by_title(idea_id, trace) or raise_trae_window_by_title(workspace.name, trace)
        if not raised:
            trace.write("delete_command_skip_window_not_found", idea_id=idea_id, workspace=str(workspace))
            return False

        text = accessibility_text_snapshot()
        if not delete_command_confirmation_visible(text):
            trace.write("delete_command_skip_not_visible", idea_id=idea_id)
            return False

        before = run_dir / "delete_command_skip_before.png"
        screenshot(before)
        clicked = click_accessibility_button(DELETE_SKIP_BUTTON_NEEDLES, require_button=True)
        if clicked:
            time.sleep(1.2)
            screenshot(run_dir / "delete_command_skip_after.png")
        trace.write(
            "delete_command_skip_done",
            idea_id=idea_id,
            clicked=clicked,
            screenshot=str(before),
        )
        return clicked

    if acquire_lock:
        with trae_window_lock(trace, "delete_command_skip"):
            return run_locked()
    return run_locked()


def skip_high_risk_chmod_command_if_present(
    trace: Trace,
    run_dir: Path,
    workspace: Path,
    idea_id: str,
    *,
    acquire_lock: bool = True,
) -> bool:
    def run_locked() -> bool:
        raised = raise_trae_window_by_title(idea_id, trace) or raise_trae_window_by_title(workspace.name, trace)
        if not raised:
            trace.write("chmod_command_skip_window_not_found", idea_id=idea_id, workspace=str(workspace))
            return False

        before = run_dir / "chmod_command_skip_before.png"
        screenshot(before)
        cancelled = click_accessibility_button(CONFIRMATION_CANCEL_BUTTON_NEEDLES, require_button=True)
        if cancelled:
            time.sleep(0.8)
        clicked = click_accessibility_button(DELETE_SKIP_BUTTON_NEEDLES, require_button=True)
        if not clicked:
            trace.write("chmod_command_skip_not_visible", idea_id=idea_id, screenshot=str(before))
            return False
        if clicked:
            time.sleep(1.2)
            screenshot(run_dir / "chmod_command_skip_after.png")
        trace.write(
            "chmod_command_skip_done",
            idea_id=idea_id,
            clicked=clicked,
            screenshot=str(before),
        )
        return clicked

    if acquire_lock:
        with trae_window_lock(trace, "chmod_command_skip"):
            return run_locked()
    return run_locked()


def run_high_risk_chmod_command_if_present(
    trace: Trace,
    run_dir: Path,
    workspace: Path,
    idea_id: str,
    *,
    acquire_lock: bool = True,
) -> bool:
    trace.write("chmod_command_run_redirected_to_skip", idea_id=idea_id)
    return skip_high_risk_chmod_command_if_present(trace, run_dir, workspace, idea_id, acquire_lock=acquire_lock)


def send_shortcut_sequence(trace: Trace, prompt: str, enabled: bool) -> None:
    if not enabled:
        return
    activate_trae()

    # Trae custom keybinding on this machine: toggles/focuses the unified sidebar.
    trace.write("shortcut", keys="option+cmd+s", purpose="focus Trae unified sidebar/chat area")
    key_code("1", ["option", "command"])  # key code 1 = s
    time.sleep(0.7)

    trace.write("shortcut", keys="cmd+v", purpose="paste prompt")
    set_clipboard(prompt)
    keystroke("v", ["command"])
    time.sleep(0.5)

    trace.write("shortcut", keys="return", purpose="send prompt")
    key_code("36")  # return


def send_prompt_with_solo_helper_once(trace: Trace, prompt: str, timeout_seconds: int, attempt: int, workspace: Path, reuse_task: bool = False) -> dict[str, Any]:
    cmd = [
        sys.executable,
        str(SOLO_HELPER),
        "send-prompt-gui",
        "--prompt",
        prompt,
        "--timeout",
        str(timeout_seconds),
    ]
    if reuse_task:
        cmd.append("--reuse-task")
    reuse_part = " --reuse-task" if reuse_task else ""
    trace.write("send_prompt_gui_start", attempt=attempt, reuse_task=reuse_task, cmd=f"{sys.executable} {SOLO_HELPER} send-prompt-gui --prompt <prompt> --timeout {timeout_seconds}{reuse_part}")
    proc = subprocess.Popen(cmd, cwd=str(ROOT), text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    visible_failure = False
    # `send-prompt-gui` may fall back to `trae-cn chat` when macOS
    # Accessibility/AppleScript is unavailable, then still needs time to verify
    # the fresh `code_comp_trigger` event from Trae logs.
    deadline = time.time() + timeout_seconds + 90
    while proc.poll() is None and time.time() < deadline:
        time.sleep(1)
        if accessibility_text_contains(["请求失败", "异常打断", "SOLO Agent 请求失败"]):
            visible_failure = True
            trace.write("send_prompt_gui_visible_failure", attempt=attempt)
            proc.terminate()
            try:
                proc.wait(timeout=5)
            except subprocess.TimeoutExpired:
                proc.kill()
            break
    if proc.poll() is None:
        trace.write("send_prompt_gui_timeout", attempt=attempt)
        proc.kill()
    stdout, stderr = proc.communicate()
    trace.write(
        "send_prompt_gui_exit",
        attempt=attempt,
        returncode=proc.returncode,
        visible_failure=visible_failure,
        stdout_chars=len(stdout or ""),
        stderr=(stderr or "").strip()[-2000:],
    )
    payload: dict[str, Any] = {}
    if stdout.strip():
        try:
            payload = json.loads(stdout)
        except json.JSONDecodeError:
            trace.write("send_prompt_gui_json_parse_failed", attempt=attempt, stdout=(stdout or "").strip()[-2000:])
    trace.write("send_prompt_gui_payload", attempt=attempt, payload=payload)
    best_event = payload.get("best_session_event") if isinstance(payload, dict) else None
    if visible_failure or proc.returncode != 0 or not isinstance(best_event, dict) or not best_event.get("profile_match"):
        raise RuntimeError(send_prompt_failure_message(payload, proc.returncode, visible_failure))
    return payload


def compact_error_detail(value: object, max_chars: int = 900) -> str:
    text = str(value or "").strip()
    if len(text) <= max_chars:
        return text
    return text[: max_chars - 3] + "..."


def send_prompt_failure_message(payload: dict[str, Any], returncode: int | None, visible_failure: bool) -> str:
    preflight = payload.get("preflight") if isinstance(payload, dict) else None
    if not isinstance(preflight, dict):
        preflight = {}
    failure_kind = str(payload.get("failure_kind") or preflight.get("failure_kind") or "")
    failure_summary = str(payload.get("failure_summary") or preflight.get("failure_summary") or "")
    prepare_stderr = str(preflight.get("prepare_stderr") or "")
    gui_stderr = str(payload.get("gui_send_stderr") or "")

    if failure_kind == "macos_accessibility_denied":
        return (
            "macOS 辅助访问权限被拒，脚本没有权限通过 osascript/System Events 操作 Trae，"
            "因此没有真正完成粘贴/发送，也抓不到 Seed-Code-DogFooding + SOLO Agent 会话。"
            "请到 系统设置 -> 隐私与安全性 -> 辅助功能，给启动脚本的入口"
            "（Terminal/iTerm/Codex/Python）授权；如果列表里出现 /usr/bin/osascript，也一起打开。"
            f" 原始错误: {compact_error_detail(prepare_stderr or gui_stderr or failure_summary)}"
        )

    if visible_failure:
        return "Trae 界面出现请求失败/异常打断，send-prompt-gui 已停止等待，请检查当前 Trae 会话状态后重试。"

    best_error = payload.get("best_error_event") if isinstance(payload, dict) else None
    error_detail = ""
    if isinstance(best_error, dict):
        error_detail = str(best_error.get("error") or best_error.get("line_text") or "")
    detail = compact_error_detail(failure_summary or prepare_stderr or gui_stderr or error_detail)
    if returncode not in (0, None):
        return (
            f"send-prompt-gui 退出码 {returncode}，未抓到 required Seed-Code-DogFooding + SOLO Agent 会话。"
            + (f" 关键错误: {detail}" if detail else "")
        )
    return (
        "send-prompt-gui 已执行，但没有在 Trae 日志中抓到新的 Seed-Code-DogFooding + SOLO Agent "
        "`code_comp_trigger` 事件。"
        + (f" 关键日志: {detail}" if detail else "")
    )


def send_prompt_with_solo_helper(trace: Trace, prompt: str, workspace: Path, timeout_seconds: int = 45, max_attempts: int = 2, reuse_task: bool = False) -> dict[str, Any]:
    last_error: str | None = None
    for attempt in range(1, max_attempts + 1):
        try:
            return send_prompt_with_solo_helper_once(trace, prompt, timeout_seconds, attempt, workspace, reuse_task=reuse_task)
        except RuntimeError as exc:
            last_error = str(exc)
            trace.write("send_prompt_gui_attempt_failed", attempt=attempt, max_attempts=max_attempts, error=last_error)
            if "macOS 辅助访问权限被拒" in last_error:
                break
            if attempt < max_attempts:
                time.sleep(3)
    raise RuntimeError(last_error or "send-prompt-gui failed")


def save_trae_session_metadata(trace: Trace, payload: dict[str, Any]) -> None:
    best_event = payload.get("best_session_event") if isinstance(payload, dict) else None
    if not isinstance(best_event, dict):
        return
    log_file = str(best_event.get("log_file") or "")
    metadata = {
        "session_id": str(best_event.get("session_id") or ""),
        "message_id": str(best_event.get("message_id") or ""),
        "trace_id": str(best_event.get("trace_id") or ""),
        "log_file": log_file,
        "log_dir": str(Path(log_file).parents[1]) if log_file else "",
        "timestamp": str(best_event.get("timestamp") or ""),
    }
    path = trace.run_dir / "trae_session.json"
    path.write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    trace.write("trae_session_metadata_saved", path=str(path), session_id=metadata["session_id"], log_file=metadata["log_file"])


def load_trae_session_metadata(run_dir: Path) -> dict[str, str]:
    path = run_dir / "trae_session.json"
    data = read_json_file(path, {})
    if isinstance(data, dict) and data.get("log_file"):
        return {str(k): str(v or "") for k, v in data.items()}

    trace_path = run_dir / "trace.jsonl"
    if not trace_path.exists():
        return {}
    metadata: dict[str, str] = {}
    for line in trace_path.read_text(encoding="utf-8", errors="ignore").splitlines():
        try:
            row = json.loads(line)
        except json.JSONDecodeError:
            continue
        if row.get("event") != "send_prompt_gui_payload":
            continue
        payload = row.get("payload")
        best_event = payload.get("best_session_event") if isinstance(payload, dict) else None
        if not isinstance(best_event, dict):
            continue
        log_file = str(best_event.get("log_file") or "")
        metadata = {
            "session_id": str(best_event.get("session_id") or ""),
            "message_id": str(best_event.get("message_id") or ""),
            "trace_id": str(best_event.get("trace_id") or ""),
            "log_file": log_file,
            "log_dir": str(Path(log_file).parents[1]) if log_file else "",
            "timestamp": str(best_event.get("timestamp") or ""),
        }
    return metadata


def open_trae_project(trace: Trace, workspace: Path, timeout_seconds: int = 30) -> None:
    cmd = trae_cli_command("-n", str(workspace))
    trace.write("open_trae_project", workspace=str(workspace), cmd=" ".join(shell_quote(part) for part in cmd))
    proc = subprocess.run(
        cmd,
        cwd=str(workspace),
        text=True,
        capture_output=True,
        check=False,
        timeout=30,
    )
    trace.write(
        "open_trae_project_exit",
        returncode=proc.returncode,
        stdout=(proc.stdout or "").strip()[-1000:],
        stderr=(proc.stderr or "").strip()[-1000:],
    )
    if proc.returncode != 0:
        raise RuntimeError("Trae project window did not open successfully")

    deadline = time.time() + timeout_seconds
    while time.time() < deadline:
        titles = trae_window_titles()
        if any(workspace.name in title or str(workspace) in title for title in titles):
            trace.write("open_trae_project_verified", workspace=str(workspace), method="window_title", titles=titles)
            return
        if accessibility_text_contains([workspace.name, str(workspace)]):
            trace.write("open_trae_project_verified", workspace=str(workspace), method="accessibility_text")
            return
        time.sleep(1)
    screenshot_path = trace.run_dir / "open_project_failed.png"
    screenshot(screenshot_path)
    titles = trae_window_titles()
    trace.write("open_trae_project_verify_failed", workspace=str(workspace), screenshot=str(screenshot_path), titles=titles)
    raise RuntimeError(f"Trae window did not attach to workspace: {workspace}")


def start_trae_chat(trace: Trace, workspace: Path, prompt: str, use_cli: bool, reuse_task: bool = False) -> subprocess.Popen[str] | None:
    open_trae_project(trace, workspace)

    if use_cli:
        payload = send_prompt_with_solo_helper(trace, prompt, workspace, reuse_task=reuse_task)
        save_trae_session_metadata(trace, payload)
        return None

    trace.write("trae_project_opened_without_cli", cwd=str(workspace))
    return None


def open_trae_workspace_for_finish(trace: Trace, workspace: Path, title_parts: Any | None = None, timeout_seconds: float = 4.0) -> bool:
    candidates = normalize_title_parts([*(normalize_title_parts(title_parts) if title_parts else []), workspace.name, str(workspace)])
    trace.write("open_trae_workspace_for_finish", workspace=str(workspace), title_parts=candidates)
    if raise_trae_window_by_titles(candidates, trace):
        trace.write("finish_workspace_existing_window_reused", workspace=str(workspace))
        return True
    subprocess.Popen(trae_cli_command("--new-window", str(workspace)))
    deadline = time.time() + timeout_seconds
    while time.time() < deadline:
        titles = trae_window_titles()
        if any(part in title for title in titles for part in candidates):
            if raise_trae_window_by_titles(candidates, trace):
                trace.write("finish_workspace_new_window_ready", workspace=str(workspace), titles=titles)
                return True
        time.sleep(0.35)
    activate_trae()
    raised = raise_trae_window_by_titles(candidates, trace)
    trace.write("finish_workspace_new_window_poll_timeout", workspace=str(workspace), raised=raised)
    return raised


def latest_ask_user_question_pending(text: str, session_id: str) -> bool:
    pending = False
    waiting = False
    for line in text.splitlines():
        event = extract_event_params(line)
        if not event:
            continue
        event_name, params = event
        if session_id and params.get("session_id") != session_id:
            continue
        tool_type = params.get("tool_type")
        if event_name in {"tool_call_show", "file_tool_show"} and tool_type == "AskUserQuestion":
            pending = True
        elif event_name == "tool_call_click" and tool_type == "AskUserQuestion" and params.get("action") == "submit":
            pending = False
        elif event_name == "icube_task_monitor":
            try:
                waiting = int(params.get("waiting_session_cnt") or 0) > 0
            except (TypeError, ValueError):
                waiting = False
    return pending and waiting


def service_exception_seen(text: str, session_id: str, workspace: Path, idea_id: str) -> bool:
    if not text:
        return False
    workspace_markers = {idea_id, workspace.name, str(workspace)}
    has_workspace_context = any(marker and marker in text for marker in workspace_markers)
    if not has_workspace_context:
        return False

    session_failed = not session_id
    for line in text.splitlines():
        event = extract_event_params(line)
        if event:
            event_name, params = event
            if session_id and params.get("session_id") != session_id:
                continue
            if event_name == "code_comp_complete_shown" and str(params.get("is_interrupted") or "0") in {"", "0", "False", "false"}:
                return False
            if event_name == "code_comp_fail":
                error_message = str(params.get("error_message") or "")
                if any(marker in error_message for marker in SERVICE_EXCEPTION_MARKERS):
                    return True
                session_failed = True
                continue

        if any(marker in line for marker in SERVICE_EXCEPTION_MARKERS):
            if not session_id or session_id in line or session_failed:
                return True
    return False


def stopped_run_seen(text: str, session_id: str, workspace: Path, idea_id: str) -> bool:
    if not text:
        return False
    workspace_markers = {idea_id, workspace.name, str(workspace)}
    has_workspace_context = any(marker and marker in text for marker in workspace_markers)
    if not has_workspace_context and not session_id:
        return False

    for line in text.splitlines():
        event = extract_event_params(line)
        if event:
            event_name, params = event
            if session_id and params.get("session_id") != session_id:
                continue
            if event_name == "code_comp_complete_shown" and str(params.get("is_interrupted") or "0") not in {"", "0", "False", "false"}:
                return True
            if event_name == "icube_task_monitor":
                try:
                    running = int(params.get("runining_session_cnt") or 0)
                    waiting = int(params.get("waiting_session_cnt") or 0)
                    completed = int(params.get("completed_session_cnt") or 0)
                    interrupted = int(params.get("interupted_session_cnt") or 0)
                except (TypeError, ValueError):
                    running = waiting = completed = interrupted = 0
                if running == 0 and waiting == 0 and completed == 0 and interrupted > 0:
                    return True

        if any(marker in line for marker in STOPPED_RUN_MARKERS):
            if not session_id or session_id in line or has_workspace_context:
                return True
    return False


def write_abort_marker(run_dir: Path, reason: str, detail: str) -> None:
    marker_path = run_dir / "aborted.json"
    marker_path.write_text(json.dumps({
        "reason": reason,
        "detail": detail,
        "ts": datetime.now().isoformat(timespec="seconds"),
    }, ensure_ascii=False, indent=2), encoding="utf-8")


def close_project_window_for_service_exception(
    trace: Trace,
    run_dir: Path,
    workspace: Path,
    idea_id: str,
) -> bool:
    with trae_window_lock(trace, "server_exception_close"):
        raise_trae_window_by_title(idea_id, trace) or raise_trae_window_by_title(workspace.name, trace)
        screenshot(run_dir / "server_exception.png")
        closed = close_trae_window_by_title(idea_id, trace) or close_trae_window_by_title(workspace.name, trace)
    write_abort_marker(run_dir, "server_exception", "Trae 显示服务端异常，runner 已关闭对应项目窗口。")
    mark_idea_failed(idea_id, run_dir, "server_exception", trace)
    trace.write("service_exception_window_closed", idea_id=idea_id, closed=closed)
    return closed


def continue_service_exception_if_present(
    trace: Trace,
    run_dir: Path,
    workspace: Path,
    idea_id: str,
    *,
    acquire_lock: bool = True,
) -> bool:
    def run_locked() -> bool:
        raised = raise_trae_window_by_title(idea_id, trace) or raise_trae_window_by_title(workspace.name, trace)
        if not raised:
            trace.write("service_exception_continue_window_not_found", idea_id=idea_id, workspace=str(workspace))
            return False
        scroll_left_conversation_to_bottom(trace)
        before = run_dir / "service_exception_continue_before.png"
        screenshot(before)
        clicked = click_accessibility_button(SERVICE_EXCEPTION_CONTINUE_BUTTON_NEEDLES, require_button=True)
        if clicked:
            time.sleep(1.5)
            screenshot(run_dir / "service_exception_continue_after.png")
        trace.write(
            "service_exception_continue_done",
            idea_id=idea_id,
            clicked=clicked,
            screenshot=str(before),
        )
        return clicked

    if acquire_lock:
        with trae_window_lock(trace, "service_exception_continue"):
            return run_locked()
    return run_locked()


def retry_stopped_run_if_present(
    trace: Trace,
    run_dir: Path,
    workspace: Path,
    idea_id: str,
    *,
    acquire_lock: bool = True,
) -> bool:
    def run_locked() -> bool:
        raised = raise_trae_window_by_title(idea_id, trace) or raise_trae_window_by_title(workspace.name, trace)
        if not raised:
            trace.write("stopped_run_retry_window_not_found", idea_id=idea_id, workspace=str(workspace))
            return False
        scroll_left_conversation_to_bottom(trace)
        before = run_dir / "stopped_run_retry_before.png"
        screenshot(before)
        clicked = click_accessibility_button(STOPPED_RETRY_BUTTON_NEEDLES, require_button=True)
        if clicked:
            time.sleep(1.5)
            screenshot(run_dir / "stopped_run_retry_after.png")
        trace.write("stopped_run_retry_done", idea_id=idea_id, clicked=clicked, screenshot=str(before))
        return clicked

    if acquire_lock:
        with trae_window_lock(trace, "stopped_run_retry"):
            return run_locked()
    return run_locked()


def close_project_window_for_stopped_run(
    trace: Trace,
    run_dir: Path,
    workspace: Path,
    idea_id: str,
    reason: str,
    detail: str,
) -> bool:
    with trae_window_lock(trace, "stopped_run_close"):
        raise_trae_window_by_title(idea_id, trace) or raise_trae_window_by_title(workspace.name, trace)
        screenshot(run_dir / f"{reason}.png")
        closed = close_trae_window_by_title(idea_id, trace) or close_trae_window_by_title(workspace.name, trace)
    write_abort_marker(run_dir, reason, detail)
    mark_idea_failed(idea_id, run_dir, reason, trace)
    trace.write("stopped_run_window_closed", idea_id=idea_id, reason=reason, closed=closed)
    return closed


def service_exception_detected_for_run(
    trace: Trace,
    run_dir: Path,
    workspace: Path,
    idea_id: str,
    *,
    check_ui: bool,
) -> bool:
    session_metadata = load_trae_session_metadata(run_dir)
    session_id = str(session_metadata.get("session_id") or "")
    sources: list[tuple[str, str, str]] = []
    recent_log, _log_dir, renderer_log = read_recent_renderer_log_for_workspace(workspace, 2500)
    if recent_log:
        sources.append(("workspace_renderer.log", recent_log, str(renderer_log or "")))
    session_log = Path(str(session_metadata.get("log_file") or ""))
    if session_log.exists():
        sources.append(("recorded_session_renderer.log", read_recent_text(session_log, 2500), str(session_log)))

    session_candidates = [session_id] if session_id else [""]
    for source, text, path in sources:
        if not text:
            continue
        for candidate_session_id in dict.fromkeys(session_candidates):
            if service_exception_seen(text, candidate_session_id, workspace, idea_id):
                trace.write(
                    "service_exception_detected",
                    source=source,
                    path=path,
                    session_id=candidate_session_id,
                )
                return True

    if not check_ui:
        return False

    with trae_window_lock(trace, "server_exception_probe"):
        raised = raise_trae_window_by_title(idea_id, trace) or raise_trae_window_by_title(workspace.name, trace)
        if not raised:
            trace.write("service_exception_probe_window_not_found", idea_id=idea_id, workspace=str(workspace))
            return False
        scroll_left_conversation_to_bottom(trace)
        probe_path = run_dir / "service_exception_probe.png"
        screenshot(probe_path)
        if accessibility_text_contains(list(SERVICE_EXCEPTION_MARKERS)):
            trace.write("service_exception_detected", source="accessibility", screenshot=str(probe_path))
            return True
    return False


def stopped_run_detected_for_run(
    trace: Trace,
    run_dir: Path,
    workspace: Path,
    idea_id: str,
    *,
    check_ui: bool,
) -> tuple[bool, str]:
    if not trae_main_pid():
        trace.write("stopped_run_detected", source="process", reason="trae_not_running")
        return True, "trae_not_running"

    session_metadata = load_trae_session_metadata(run_dir)
    session_id = str(session_metadata.get("session_id") or "")
    sources: list[tuple[str, str, str]] = []
    recent_log, _log_dir, renderer_log = read_recent_renderer_log_for_workspace(workspace, 2500)
    if recent_log:
        sources.append(("workspace_renderer.log", recent_log, str(renderer_log or "")))
    session_log = Path(str(session_metadata.get("log_file") or ""))
    if session_log.exists():
        sources.append(("recorded_session_renderer.log", read_recent_text(session_log, 2500), str(session_log)))

    session_candidates = [session_id] if session_id else []
    session_candidates.append("")
    for source, text, path in sources:
        if not text:
            continue
        for candidate_session_id in dict.fromkeys(session_candidates):
            if stopped_run_seen(text, candidate_session_id, workspace, idea_id):
                trace.write("stopped_run_detected", source=source, path=path, session_id=candidate_session_id, reason="stopped_or_interrupted")
                return True, "stopped_or_interrupted"

    if not check_ui:
        return False, ""

    with trae_window_lock(trace, "stopped_run_probe"):
        raised = raise_trae_window_by_title(idea_id, trace) or raise_trae_window_by_title(workspace.name, trace)
        if not raised:
            trace.write("stopped_run_probe_window_not_found", idea_id=idea_id, workspace=str(workspace))
            return False, ""
        scroll_left_conversation_to_bottom(trace)
        probe_path = run_dir / "stopped_run_probe.png"
        screenshot(probe_path)
        if accessibility_text_contains(list(STOPPED_RUN_MARKERS)):
            trace.write("stopped_run_detected", source="accessibility", screenshot=str(probe_path), reason="manual_stop_output")
            return True, "manual_stop_output"
    return False, ""


def find_visual_next_button(image_path: Path) -> tuple[float, float, tuple[int, int, int, int, int]] | None:
    try:
        from PIL import Image
    except Exception:
        return None

    image = Image.open(image_path).convert("RGB")
    width, height = image.size
    light_pixels: set[tuple[int, int]] = set()
    for y in range(int(height * 0.35), int(height * 0.82)):
        for x in range(0, int(width * 0.36)):
            r, g, b = image.getpixel((x, y))
            if r > 175 and g > 175 and b > 175 and abs(r - g) < 45 and abs(g - b) < 45:
                light_pixels.add((x, y))

    seen: set[tuple[int, int]] = set()
    best: tuple[int, int, int, int, int] | None = None
    for pixel in list(light_pixels):
        if pixel in seen:
            continue
        stack = [pixel]
        seen.add(pixel)
        xs: list[int] = []
        ys: list[int] = []
        while stack:
            x, y = stack.pop()
            xs.append(x)
            ys.append(y)
            for neighbor in ((x + 1, y), (x - 1, y), (x, y + 1), (x, y - 1)):
                if neighbor in light_pixels and neighbor not in seen:
                    seen.add(neighbor)
                    stack.append(neighbor)
        if not xs:
            continue
        x1, y1, x2, y2 = min(xs), min(ys), max(xs), max(ys)
        component_width = x2 - x1 + 1
        component_height = y2 - y1 + 1
        if (
            70 <= component_width <= 190
            and 30 <= component_height <= 75
            and y1 > height * 0.55
            and x1 > width * 0.18
        ):
            component = (len(xs), x1, y1, x2, y2)
            if best is None or component[0] > best[0]:
                best = component

    if best is None:
        return None
    _, x1, y1, x2, y2 = best
    return (x1 + x2) / 2, (y1 + y2) / 2, best


def auto_answer_question_cards(
    trace: Trace,
    run_dir: Path,
    workspace: Path,
    idea_id: str,
    *,
    max_steps: int = 8,
    acquire_lock: bool = True,
) -> int:
    def run_locked() -> int:
        raised = raise_trae_window_by_title(idea_id, trace) or raise_trae_window_by_title(workspace.name, trace)
        if not raised:
            trace.write("auto_question_window_not_found", idea_id=idea_id, workspace=str(workspace))
            return 0

        clicked = 0
        for step in range(max_steps):
            shot_path = run_dir / f"auto_question_step_{step}.png"
            screenshot(shot_path)
            next_button = find_visual_next_button(shot_path)
            if not next_button:
                trace.write("auto_question_next_not_found", step=step, screenshot=str(shot_path))
                break
            point_x, point_y, component = next_button
            rect = trae_window_rect()
            if not rect:
                trace.write("auto_question_window_rect_missing", step=step)
                break
            x, y, width, height = rect
            try:
                from PIL import Image

                image_width, image_height = Image.open(shot_path).size
            except Exception:
                image_width, image_height = width, height
            scale_x = image_width / max(1, width)
            scale_y = image_height / max(1, height)
            screen_x = int(x + point_x / scale_x)
            screen_y = int(y + point_y / scale_y)
            trace.write(
                "auto_question_click_next",
                step=step,
                screenshot=str(shot_path),
                component=component,
                screen_x=screen_x,
                screen_y=screen_y,
            )
            click_at(screen_x, screen_y)
            clicked += 1
            time.sleep(1.4)
        if clicked:
            screenshot(run_dir / "auto_question_after.png")
        trace.write("auto_question_done", clicked=clicked, idea_id=idea_id)
        return clicked

    if acquire_lock:
        with trae_window_lock(trace, "auto_question"):
            return run_locked()
    return run_locked()


def find_visual_bottom_light_button(image_path: Path) -> tuple[float, float, tuple[int, int, int, int, int]] | None:
    try:
        from PIL import Image
    except Exception:
        return None

    image = Image.open(image_path).convert("RGB")
    width, height = image.size
    light_pixels: set[tuple[int, int]] = set()
    for y in range(int(height * 0.68), int(height * 0.90)):
        for x in range(int(width * 0.16), int(width * 0.45)):
            r, g, b = image.getpixel((x, y))
            if r > 175 and g > 175 and b > 175 and abs(r - g) < 45 and abs(g - b) < 45:
                light_pixels.add((x, y))

    seen: set[tuple[int, int]] = set()
    candidates: list[tuple[int, int, int, int, int]] = []
    for pixel in list(light_pixels):
        if pixel in seen:
            continue
        stack = [pixel]
        seen.add(pixel)
        xs: list[int] = []
        ys: list[int] = []
        while stack:
            x, y = stack.pop()
            xs.append(x)
            ys.append(y)
            for neighbor in ((x + 1, y), (x - 1, y), (x, y + 1), (x, y - 1)):
                if neighbor in light_pixels and neighbor not in seen:
                    seen.add(neighbor)
                    stack.append(neighbor)
        if not xs:
            continue
        x1, y1, x2, y2 = min(xs), min(ys), max(xs), max(ys)
        component_width = x2 - x1 + 1
        component_height = y2 - y1 + 1
        if 70 <= component_width <= 220 and 28 <= component_height <= 80:
            candidates.append((len(xs), x1, y1, x2, y2))

    if not candidates:
        return None
    best = sorted(candidates, key=lambda c: (c[2], c[0]), reverse=True)[0]
    _, x1, y1, x2, y2 = best
    return (x1 + x2) / 2, (y1 + y2) / 2, best


def find_visual_pending_review_accept_button(image_path: Path) -> tuple[float, float, tuple[int, int, int, int, int]] | None:
    try:
        from PIL import Image
    except Exception:
        return None

    image = Image.open(image_path).convert("RGB")
    width, height = image.size
    light_pixels: set[tuple[int, int]] = set()
    for y in range(int(height * 0.70), int(height * 0.84)):
        for x in range(int(width * 0.20), int(width * 0.30)):
            r, g, b = image.getpixel((x, y))
            if r > 155 and g > 155 and b > 155 and abs(r - g) < 45 and abs(g - b) < 45:
                light_pixels.add((x, y))

    seen: set[tuple[int, int]] = set()
    candidates: list[tuple[int, int, int, int, int]] = []
    for pixel in list(light_pixels):
        if pixel in seen:
            continue
        stack = [pixel]
        seen.add(pixel)
        xs: list[int] = []
        ys: list[int] = []
        while stack:
            x, y = stack.pop()
            xs.append(x)
            ys.append(y)
            for neighbor in ((x + 1, y), (x - 1, y), (x, y + 1), (x, y - 1)):
                if neighbor in light_pixels and neighbor not in seen:
                    seen.add(neighbor)
                    stack.append(neighbor)
        if not xs:
            continue
        x1, y1, x2, y2 = min(xs), min(ys), max(xs), max(ys)
        component_width = x2 - x1 + 1
        component_height = y2 - y1 + 1
        if 24 <= component_width <= 90 and 24 <= component_height <= 90 and len(xs) >= 200:
            candidates.append((len(xs), x1, y1, x2, y2))

    if not candidates:
        return None
    _, x1, y1, x2, y2 = sorted(candidates, key=lambda c: (c[2], c[3], c[0]), reverse=True)[0]
    return (x1 + x2) / 2, (y1 + y2) / 2, (len(candidates), x1, y1, x2, y2)


def visual_completed_status_icon_present(image_path: Path) -> bool:
    try:
        from PIL import Image
    except Exception:
        return False

    image = Image.open(image_path).convert("RGB")
    width, height = image.size
    green_pixels: set[tuple[int, int]] = set()
    for y in range(int(height * 0.72), int(height * 0.86)):
        for x in range(0, int(width * 0.06)):
            r, g, b = image.getpixel((x, y))
            if g > 130 and r < 120 and b > 80 and g > r * 1.4 and g > b * 0.8:
                green_pixels.add((x, y))

    seen: set[tuple[int, int]] = set()
    for pixel in list(green_pixels):
        if pixel in seen:
            continue
        stack = [pixel]
        seen.add(pixel)
        xs: list[int] = []
        ys: list[int] = []
        while stack:
            x, y = stack.pop()
            xs.append(x)
            ys.append(y)
            for neighbor in ((x + 1, y), (x - 1, y), (x, y + 1), (x, y - 1)):
                if neighbor in green_pixels and neighbor not in seen:
                    seen.add(neighbor)
                    stack.append(neighbor)
        if not xs:
            continue
        x1, y1, x2, y2 = min(xs), min(ys), max(xs), max(ys)
        if len(xs) >= 180 and 12 <= x2 - x1 + 1 <= 42 and 12 <= y2 - y1 + 1 <= 42:
            return True
    return False


def accept_all_changes_if_present(trace: Trace, run_dir: Path, workspace: Path, idea_id: str) -> bool:
    raised = raise_trae_window_by_title(idea_id, trace) or raise_trae_window_by_title(workspace.name, trace)
    if not raised:
        trace.write("accept_all_window_not_found", idea_id=idea_id, workspace=str(workspace))
        return False
    scroll_left_conversation_to_bottom(trace)

    before = run_dir / "accept_all_before.png"
    screenshot(before)
    clicked = click_accessibility_button(["全部保留", "保留全部", "Keep All", "Accept All"])
    method = "accessibility" if clicked else ""

    if not clicked:
        button = find_visual_bottom_light_button(before)
        if button:
            point_x, point_y, component = button
            rect = trae_window_rect()
            if rect:
                x, y, width, height = rect
                try:
                    from PIL import Image

                    image_width, image_height = Image.open(before).size
                except Exception:
                    image_width, image_height = width, height
                scale_x = image_width / max(1, width)
                scale_y = image_height / max(1, height)
                screen_x = int(x + point_x / scale_x)
                screen_y = int(y + point_y / scale_y)
                trace.write(
                    "accept_all_visual_click",
                    screenshot=str(before),
                    component=component,
                    screen_x=screen_x,
                    screen_y=screen_y,
                )
                post_jxa_mouse_click(screen_x, screen_y)
                time.sleep(0.25)
                post_jxa_mouse_click(screen_x, screen_y)
                clicked = True
                method = "visual"

    if clicked:
        time.sleep(2.0)
        screenshot(run_dir / "accept_all_after.png")
    trace.write("accept_all_done", clicked=clicked, method=method, idea_id=idea_id)
    return clicked


def accept_visible_pending_review_if_present(trace: Trace, run_dir: Path, workspace: Path, idea_id: str) -> bool:
    raised = raise_trae_window_by_title(idea_id, trace) or raise_trae_window_by_title(workspace.name, trace)
    if not raised:
        trace.write("accept_pending_review_window_not_found", idea_id=idea_id, workspace=str(workspace))
        return False
    before = run_dir / "accept_pending_review_before.png"
    screenshot(before)
    accessibility_visible = accessibility_text_contains(["文件待审查", "Files to review", "待审查"])
    button = find_visual_pending_review_accept_button(before)
    if not accessibility_visible and not button:
        trace.write("accept_pending_review_not_visible", idea_id=idea_id, screenshot=str(before))
        return False
    if not button:
        trace.write(
            "accept_pending_review_button_not_found",
            screenshot=str(before),
            idea_id=idea_id,
            accessibility_visible=accessibility_visible,
        )
        return False

    point_x, point_y, component = button
    rect = trae_window_rect()
    if not rect:
        trace.write("accept_pending_review_window_rect_missing", screenshot=str(before), idea_id=idea_id)
        return False

    x, y, width, height = rect
    try:
        from PIL import Image

        image_width, image_height = Image.open(before).size
    except Exception:
        image_width, image_height = width, height
    scale_x = image_width / max(1, width)
    scale_y = image_height / max(1, height)
    screen_x = int(x + point_x / scale_x)
    screen_y = int(y + point_y / scale_y)
    trace.write(
        "accept_pending_review_visual_click",
        screenshot=str(before),
        component=component,
        screen_x=screen_x,
        screen_y=screen_y,
        idea_id=idea_id,
    )
    post_jxa_mouse_click(screen_x, screen_y)
    time.sleep(0.15)
    click_at(screen_x, screen_y)
    time.sleep(1.8)
    after = run_dir / "accept_pending_review_after.png"
    screenshot(after)
    still_visible = find_visual_pending_review_accept_button(after) is not None
    trace.write(
        "accept_pending_review_done",
        clicked=not still_visible,
        method="visual+jxa+system_events",
        idea_id=idea_id,
        still_visible=still_visible,
        screenshot=str(after),
    )
    if still_visible:
        raise RuntimeError(f"Trae 待审查文件确认按钮点击后仍可见，停止发送以避免提示词卡住: {idea_id}")
    return not still_visible


def project_completion_visible(trace: Trace, run_dir: Path, workspace: Path, idea_id: str) -> bool:
    raised = raise_trae_window_by_title(idea_id, trace) or raise_trae_window_by_title(workspace.name, trace)
    if not raised:
        trace.write("completion_probe_window_not_found", idea_id=idea_id, workspace=str(workspace))
        return False
    scroll_left_conversation_to_bottom(trace)
    probe_path = run_dir / "completion_probe_bottom.png"
    screenshot(probe_path)
    accessibility_visible = accessibility_text_contains(["任务完成", "Task completed"])
    visual_accept_button = find_visual_bottom_light_button(probe_path)
    visual_done_icon = visual_completed_status_icon_present(probe_path)
    visible = accessibility_visible or bool(visual_accept_button) or visual_done_icon
    trace.write(
        "completion_probe_bottom_status",
        visible=visible,
        source=(
            "accessibility" if accessibility_visible
            else "visual_accept_button" if visual_accept_button
            else "visual_done_icon" if visual_done_icon
            else "none"
        ),
        visual_accept_button=visual_accept_button[2] if visual_accept_button else None,
        visual_done_icon=visual_done_icon,
        idea_id=idea_id,
    )
    return visible


def wait_for_completion(trace: Trace, timeout_seconds: int, poll_seconds: int, run_dir: Path, workspace: Path) -> bool:
    deadline = time.time() + timeout_seconds
    trace.write("wait_completion_start", timeout_seconds=timeout_seconds, workspace=str(workspace))
    seen_log_dir: Path | None = None
    session_metadata = load_trae_session_metadata(run_dir)
    session_id = str(session_metadata.get("session_id") or "")
    session_log = Path(str(session_metadata.get("log_file") or ""))
    session_log_seen = False
    idea_id = infer_idea_id_from_run_dir(run_dir) or workspace.name
    next_ui_probe_at = 0.0
    next_auto_question_at = 0.0
    while time.time() < deadline:
        if not trae_main_pid():
            close_project_window_for_stopped_run(
                trace,
                run_dir,
                workspace,
                idea_id,
                "trae_not_running",
                "Trae 已停止运行，runner 已放弃当前项目并释放调度位。",
            )
            return False
        recent_log, log_dir, renderer_log = read_recent_renderer_log_for_workspace(workspace, 350)
        if log_dir and log_dir != seen_log_dir:
            trace.write("monitor_workspace_log", log_dir=str(log_dir), renderer_log=str(renderer_log) if renderer_log else None)
            seen_log_dir = log_dir
        log_sources = [("workspace_renderer.log", recent_log, log_dir)]
        if session_log.exists():
            if not session_log_seen:
                trace.write("monitor_session_log", renderer_log=str(session_log), session_id=session_id)
                session_log_seen = True
            log_sources.append(("session_renderer.log", read_recent_text(session_log, 5000), session_log.parent.parent))
        for source, text, source_log_dir in log_sources:
            if not text:
                continue
            if service_exception_seen(text, session_id, workspace, idea_id):
                trace.write("service_exception_detected", source=source, session_id=session_id)
                close_project_window_for_service_exception(trace, run_dir, workspace, idea_id)
                return False
            if stopped_run_seen(text, session_id, workspace, idea_id):
                trace.write("stopped_run_detected", source=source, session_id=session_id)
                close_project_window_for_stopped_run(
                    trace,
                    run_dir,
                    workspace,
                    idea_id,
                    "stopped_or_interrupted",
                    "Trae 显示手动终止输出或会话已中断，runner 已关闭对应项目窗口并释放调度位。",
                )
                return False
            if latest_ask_user_question_pending(text, session_id) and time.time() >= next_auto_question_at:
                trace.write("auto_question_pending_detected", source=source, session_id=session_id)
                clicked = auto_answer_question_cards(trace, run_dir, workspace, idea_id, acquire_lock=True)
                next_auto_question_at = time.time() + (10 if clicked else 60)
                if clicked:
                    next_ui_probe_at = time.time() + max(120, poll_seconds)
                    break
            session_ok = not session_id or session_id in text
            completion_seen = (
                '"last_message_is_finish":"1"' in text
                or '"last_message_is_finish": "1"' in text
                or (session_ok and "code_comp_complete_shown" in text)
                or (session_ok and '"completed_session_cnt":1' in text and '"runining_session_cnt":0' in text)
                or (session_ok and '"completed_session_cnt": 1' in text and '"runining_session_cnt": 0' in text)
            )
            if completion_seen:
                trace.write("completion_seen", source=source, text="trae completion marker", log_dir=str(source_log_dir) if source_log_dir else "")
                with trae_window_lock(trace, "completion_screenshot"):
                    accept_all_changes_if_present(trace, run_dir, workspace, idea_id)
                    screenshot(run_dir / "completion.png")
                return True

        now_ts = time.time()
        if now_ts >= next_ui_probe_at:
            service_exception_visible = False
            stopped_run_visible = False
            with trae_window_lock(trace, "completion_probe"):
                if raise_trae_window_by_title(idea_id, trace) or raise_trae_window_by_title(workspace.name, trace):
                    scroll_left_conversation_to_bottom(trace)
                    if accessibility_text_contains(list(SERVICE_EXCEPTION_MARKERS)):
                        trace.write("service_exception_detected", source="accessibility_probe")
                        service_exception_visible = True
                    if accessibility_text_contains(list(STOPPED_RUN_MARKERS)):
                        trace.write("stopped_run_detected", source="accessibility_probe", reason="manual_stop_output")
                        stopped_run_visible = True
                if not service_exception_visible and not stopped_run_visible and project_completion_visible(trace, run_dir, workspace, idea_id):
                    trace.write("completion_seen", source="accessibility", text="任务完成")
                    accept_all_changes_if_present(trace, run_dir, workspace, idea_id)
                    screenshot(run_dir / "completion.png")
                    return True
                screenshot(run_dir / "latest.png")
            if service_exception_visible:
                close_project_window_for_service_exception(trace, run_dir, workspace, idea_id)
                return False
            if stopped_run_visible:
                close_project_window_for_stopped_run(
                    trace,
                    run_dir,
                    workspace,
                    idea_id,
                    "manual_stop_output",
                    "Trae 显示“手动终止输出”，runner 已关闭对应项目窗口并释放调度位。",
                )
                return False
            next_ui_probe_at = now_ts + max(120, poll_seconds)
        trace.write("completion_poll", remaining_seconds=max(0, int(deadline - time.time())))
        time.sleep(poll_seconds)

    trace.write("completion_timeout")
    return False


def copy_final_output(trace: Trace, run_dir: Path, workspace: Path, idea_id: str) -> str:
    raised = raise_trae_window_by_title(idea_id, trace) or raise_trae_window_by_title(workspace.name, trace)
    if not raised:
        screenshot(run_dir / "copy_window_not_found.png")
        raise RuntimeError("Trae 目标窗口不可见，无法自动复制最终回答")
    scroll_left_conversation_to_bottom(trace)
    activate_trae()
    before = get_clipboard()

    if click_accessibility_button(["复制全部", "Copy All", "copy_button_output", "复制"]):
        trace.write("copy_button_clicked", method="accessibility")
    else:
        # Keyboard fallback: focus chat area/sidebar, then copy. This is less exact than the button,
        # but it often captures the selected/focused final answer if Trae keeps it active.
        trace.write("copy_button_not_found", fallback="option+cmd+s then cmd+a/c")
        key_code("1", ["option", "command"])  # s
        time.sleep(0.3)
        keystroke("a", ["command"])
        time.sleep(0.2)
        keystroke("c", ["command"])

    time.sleep(1.0)
    after = get_clipboard()
    if after == before or not after.strip():
        screenshot(run_dir / "copy_failed.png")
        trace.write("copy_failed_or_unchanged", clipboard_chars=len(after))
        raise RuntimeError("没有拿到新的剪贴板内容，已保存 copy_failed.png")

    trace.write("copied_output", clipboard_chars=len(after))
    return after


def find_accessibility_element_center(needles: list[str]) -> tuple[int, int] | None:
    activate_trae()
    needle_list = "{" + ",".join(json.dumps(n) for n in needles) + "}"
    script = f'''
set needles to {needle_list}

on matchesNeedle(t, needles)
  repeat with n in needles
    if t contains (n as text) then return true
  end repeat
  return false
end matchesNeedle

on findMatch(e, needles, depth)
  if depth > 10 then return ""
  set label to ""
  try
    set label to label & " " & (name of e as text)
  end try
  try
    set label to label & " " & (description of e as text)
  end try
  try
    set label to label & " " & (value of e as text)
  end try
  try
    set label to label & " " & (help of e as text)
  end try
  if my matchesNeedle(label, needles) then
    try
      set {{x, y}} to position of e
      set {{ww, hh}} to size of e
      return ((x + (ww / 2)) as integer as text) & "," & ((y + (hh / 2)) as integer as text)
    end try
  end if
  try
    repeat with c in UI elements of e
      set found to my findMatch(c, needles, depth + 1)
      if found is not "" then return found
    end repeat
  end try
  return ""
end findMatch

tell application "System Events"
  set p to first process whose frontmost is true
  return my findMatch(p, needles, 0)
end tell
'''
    raw = osascript(script, check=False).strip()
    parts = raw.split(",")
    if len(parts) != 2:
        return None
    try:
        return int(float(parts[0])), int(float(parts[1]))
    except ValueError:
        return None


def scroll_left_conversation_to_top(trace: Trace) -> bool:
    rect = trae_window_rect()
    if not rect:
        trace.write("left_chat_scroll_failed", reason="window_rect_not_found")
        return False

    x, y, width, height = rect
    focus_x = x + min(360, max(120, width // 5))
    focus_y = y + max(160, height // 2)
    post_jxa_mouse_double_click(focus_x, focus_y)
    time.sleep(0.1)
    post_jxa_scroll(focus_x, focus_y, delta=500, repeats=28)

    # Keep the key fallback: it is harmless when the wheel already reached the top.
    for _ in range(4):
        key_code("115")  # Home
        time.sleep(0.12)

    trace.write("left_chat_scrolled_to_top", focus_x=focus_x, focus_y=focus_y)
    return True


def scroll_left_conversation_to_bottom(trace: Trace) -> bool:
    rect = trae_window_rect()
    if not rect:
        trace.write("left_chat_scroll_failed", reason="window_rect_not_found", direction="bottom")
        return False

    x, y, width, height = rect
    focus_x = x + min(360, max(120, width // 5))
    focus_y = y + max(160, height // 2)
    post_jxa_mouse_double_click(focus_x, focus_y)
    time.sleep(0.1)
    post_jxa_scroll(focus_x, focus_y, delta=-500, repeats=36)

    for _ in range(4):
        key_code("119")  # End
        time.sleep(0.12)

    trace.write("left_chat_scrolled_to_bottom", focus_x=focus_x, focus_y=focus_y)
    return True


def copy_session_id_from_thinking_button(trace: Trace, run_dir: Path, idea_id: str, workspace: Path | None = None) -> str | None:
    before = get_clipboard()
    if not scroll_left_conversation_to_bottom(trace):
        return None

    rect = trae_window_rect()
    if not rect:
        trace.write("session_id_copy_failed", reason="window_rect_not_found")
        return None

    x, y, _width, _height = rect
    center = (x + 55, y + 280)
    post_jxa_mouse_double_click(*center)
    time.sleep(0.8)
    after = get_clipboard().strip()
    if not after or after == before.strip():
        screenshot(run_dir / "session_id_copy_failed.png")
        trace.write(
            "session_id_copy_failed_or_unchanged",
            center_x=center[0],
            center_y=center[1],
            clipboard_chars=len(after),
            screenshot=str(run_dir / "session_id_copy_failed.png"),
        )
        return None

    save_session_identifier(run_dir, idea_id, after, trace, source="clipboard")
    return after


def copy_thinking_process(trace: Trace, run_dir: Path, workspace: Path) -> Path | None:
    before = get_clipboard()
    if not scroll_left_conversation_to_bottom(trace):
        return None

    center = find_accessibility_element_center(["思考过程", "Thought", "thinking", "Thinking"])
    if center:
        post_jxa_mouse_double_click(*center)
        time.sleep(1.0)
    else:
        rect = trae_window_rect()
        if rect:
            x, y, _width, height = rect
            post_jxa_mouse_double_click(x + 110, y + max(260, height - 360))
            time.sleep(1.0)

    after = get_clipboard().strip()
    thinking_path = run_dir / "thinking_process.md"
    if after and after != before.strip() and len(after) > 20:
        thinking_path.write_text(after + "\n", encoding="utf-8")
        trace.write("thinking_process_saved", path=str(thinking_path), source="clipboard", chars=len(after))
        return thinking_path

    transcript_path = run_dir / "trae_full_transcript.md"
    if transcript_path.exists():
        text = transcript_path.read_text(encoding="utf-8", errors="ignore")
        thinking_path.write_text(text, encoding="utf-8")
        trace.write("thinking_process_saved", path=str(thinking_path), source="full_transcript_fallback", chars=len(text))
        return thinking_path

    trace.write("thinking_process_not_found", workspace=str(workspace))
    return None


def first_line(text: str, limit: int = 300) -> str:
    compact = " ".join(text.strip().split())
    return compact[:limit]


def table_text(text: Any, limit: int = MAX_TABLE_TEXT_CHARS) -> str:
    value = str(text or "")
    value = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", value)
    if len(value) <= limit:
        return value
    return value[:limit] + "\n...[truncated]..."


def normalize_modification_scope(value: str) -> str:
    value = (value or "").strip()
    for option in MODIFICATION_SCOPE_OPTIONS:
        if value == option or value.startswith(f"{option}：") or value.startswith(f"{option}:"):
            return option
    return value or DEFAULT_MODIFICATION_SCOPE


def normalize_business_domain(value: Any) -> str:
    text = re.sub(r"\s+", "", str(value or ""))
    return text or "Web前端"


def is_runner_fallback_trajectory(text: Any) -> bool:
    value = str(text or "").strip()
    return value.startswith("Trae 轨迹疑似串窗") or "\n## Runner trace\n" in value[:3000]


def trajectory_content_from_row(row: dict[str, str]) -> str:
    if row.get("日志轨迹") and not is_runner_fallback_trajectory(row.get("日志轨迹")):
        return table_text(row["日志轨迹"])
    if trajectory_workspace_mismatch_note(row):
        return ""
    for key in ("manual_trajectory_path", "full_transcript_path", "trajectory_path", "thinking_path", "output_path", "trace_path"):
        path = row.get(key, "")
        if path and Path(path).exists():
            return table_text(Path(path).read_text(encoding="utf-8", errors="ignore"))
    return ""


def trajectory_workspace_mismatch_note(row: dict[str, str]) -> str:
    workspace = row.get("workspace", "")
    run_dir = row.get("run_dir", "")
    if not workspace or not run_dir:
        return ""
    events_path = Path(run_dir) / "trae_trajectory_events.json"
    if not events_path.exists():
        return ""
    data = read_json_file(events_path, {})
    if not isinstance(data, dict):
        return ""
    file_paths = [
        str(item.get("filePath") or "")
        for item in data.get("file_events", [])
        if isinstance(item, dict) and item.get("filePath")
    ]
    if not file_paths:
        return ""
    if any(path.startswith(workspace) for path in file_paths):
        return ""
    sample_paths = ", ".join(sorted({path for path in file_paths[:6]})[:3])
    return (
        "Trae 轨迹疑似串窗：采集到的 Trae file_events 不属于当前工作区 "
        f"{workspace}。样例路径：{sample_paths}。为避免把别的项目轨迹写进本表，"
        "这里改用 runner 本地 trace 作为审计轨迹。"
    )


def backfill_submission_fields(row: dict[str, str]) -> dict[str, str]:
    updated = dict(row)

    if not updated.get("提示词"):
        prompt_path = updated.get("prompt_path", "")
        if prompt_path and Path(prompt_path).exists():
            updated["提示词"] = Path(prompt_path).read_text(encoding="utf-8", errors="ignore").strip()

    if not updated.get("轮次"):
        updated["轮次"] = "第一轮"

    idea_data: dict[str, Any] = {}
    run_dir = updated.get("run_dir", "")
    idea_json = Path(run_dir) / "idea.json" if run_dir else None
    if idea_json and idea_json.exists():
        idea_data = read_json_file(idea_json, {})

    if not updated.get("业务领域"):
        updated["业务领域"] = idea_field(idea_data, "业务领域", "business_domain") or "Web前端"
    updated["业务领域"] = normalize_business_domain(updated.get("业务领域"))
    if not updated.get("修改范围"):
        updated["修改范围"] = normalize_modification_scope(idea_field(idea_data, "修改范围", "modification_scope"))
    if not updated.get("截图"):
        updated["截图"] = updated.get("screenshot_path", "")

    qa_report_path = updated.get("qa_report_path", "")
    if qa_report_path and Path(qa_report_path).exists():
        verdict = qa_verdict_from_report(Path(qa_report_path), updated.get("qa_status", ""))
        if not updated.get("任务是否完成"):
            updated["任务是否完成"] = verdict["task_completed"]
        if not updated.get("产物及过程是否满意"):
            updated["产物及过程是否满意"] = verdict["satisfied"]
        if not updated.get("不满意原因"):
            updated["不满意原因"] = verdict["reason"]
        if verdict.get("task_type") and (not updated.get("任务类型") or updated.get("轮次") == "第二轮"):
            updated["任务类型"] = verdict["task_type"]

    if not updated.get("任务类型"):
        updated["任务类型"] = DEFAULT_TASK_TYPE

    if not updated.get("远端Github地址"):
        updated["远端Github地址"] = updated.get("git_remote", "")
    if not updated.get("分支文件夹"):
        updated["分支文件夹"] = updated.get("git_branch", "")

    return updated


def extract_report_field(text: str, field: str) -> str:
    pattern = rf"^\s*[-*]?\s*{re.escape(field)}\s*[:：]\s*(.*?)\s*$"
    match = re.search(pattern, text, re.M)
    return match.group(1).strip() if match else ""


def normalize_task_type(value: str) -> str:
    text = value.strip()
    if not text:
        return ""
    for option in TASK_TYPE_OPTIONS:
        if text == option or option in text:
            return option
    return text[:40]


def relativeize_report_paths(text: str) -> str:
    root_pattern = re.escape(str(ROOT))
    workspace_prefix = rf"{root_pattern}/workspaces/[^/]+/"

    def replace_workspace_link(match: re.Match[str]) -> str:
        label = match.group(1).strip()
        target = match.group(2).strip()
        if not label or label == target or label in target:
            return target
        return f"{label} ({target})"

    text = re.sub(rf"\[([^\]]+)\]\({workspace_prefix}([^)]+)\)", replace_workspace_link, text)
    text = re.sub(rf"{workspace_prefix}", "", text)
    text = re.sub(rf"{root_pattern}/", "", text)
    return text


def qa_verdict_from_report(report_path: Path | None, qa_status: str) -> dict[str, str]:
    text = report_path.read_text(encoding="utf-8", errors="ignore") if report_path and report_path.exists() else ""
    verdict = extract_report_field(text, "结论")
    task_type = normalize_task_type(extract_report_field(text, "任务类型"))
    task_completed = extract_report_field(text, "任务是否完成")
    reason = extract_report_field(text, "未完成原因") or extract_report_field(text, "不满意原因")

    passed = qa_status == "success" and bool(re.search(r"结论\s*[:：]\s*通过", text))
    failed = bool(re.search(r"结论\s*[:：]\s*不通过", text)) or qa_status != "success"
    if passed and not failed:
        return {
            "discard_status": "discarded",
            "task_type": task_type,
            "task_completed": task_completed or "完成了任务",
            "satisfied": "满意",
            "reason": "",
        }

    if not reason:
        summary = first_line(text, 260) if text else f"Codex 质检未正常生成报告，qa_status={qa_status}"
        reason = f"产物不满意：{summary}"
    elif not reason.startswith("产物不满意："):
        reason = f"产物不满意：{reason}"
    reason = relativeize_report_paths(reason)
    return {
        "discard_status": "not_discarded",
        "task_type": task_type,
        "task_completed": task_completed if task_completed in {"完成了任务", "未完成任务"} else "未完成任务",
        "satisfied": "不满意",
        "reason": reason,
    }


def idea_field(idea_data: dict[str, Any], *names: str) -> str:
    fields = idea_data.get("fields") if isinstance(idea_data.get("fields"), dict) else {}
    for name in names:
        value = fields.get(name) if isinstance(fields, dict) else None
        if value:
            return str(value).strip()
    return ""


def write_solo_qc_workbook(row: dict[str, str], trace: Trace, append: bool = False) -> Path:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception as exc:
        trace.write("solo_qc_workbook_skipped", reason=type(exc).__name__)
        raise

    row = backfill_submission_fields(row)
    if not row.get("日志轨迹") or is_runner_fallback_trajectory(row.get("日志轨迹")):
        row["日志轨迹"] = trajectory_content_from_row(row)

    idea_id = row["idea_id"]
    output_path = RESULT_DATA_DIR / f"{idea_id}.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="1F2937")
    widths = {
        "session_id": 42,
        "轮次": 12,
        "提示词": 42,
        "任务类型": 16,
        "业务领域": 18,
        "修改范围": 18,
        "任务是否完成": 16,
        "产物及过程是否满意": 20,
        "不满意原因": 58,
        "日志轨迹": 82,
        "远端Github地址": 42,
        "分支文件夹": 18,
        "截图": 42,
    }
    if append and output_path.exists():
        wb = load_workbook(output_path)
        ws = wb[wb.sheetnames[0]]
        for col_idx in range(ws.max_column, 0, -1):
            if ws.cell(1, col_idx).value == "任务难度":
                ws.delete_cols(col_idx)
                trace.write("solo_qc_workbook_removed_obsolete_column", path=str(output_path), header="任务难度")
        existing_headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        if existing_headers != SOLO_QC_HEADERS:
            wb.close()
            raise RuntimeError(f"solo qc workbook header mismatch before append: {existing_headers}")
        write_row_idx = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "solo质检"
        for col_idx, header in enumerate(SOLO_QC_HEADERS, start=1):
            header_cell = ws.cell(1, col_idx)
            header_cell.value = header
            header_cell.fill = header_fill
            header_cell.font = header_font
            header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[header_cell.column_letter].width = widths.get(header, 18)
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = "A2"
        write_row_idx = 2

    for col_idx, header in enumerate(SOLO_QC_HEADERS, start=1):
        ws.cell(write_row_idx, col_idx).value = table_text(row.get(header, ""))
        ws.cell(write_row_idx, col_idx).alignment = Alignment(vertical="top", wrap_text=True)
        if not ws.column_dimensions[ws.cell(1, col_idx).column_letter].width:
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = widths.get(header, 18)
    ws.row_dimensions[write_row_idx].height = 180
    wb.save(output_path)
    wb.close()

    check_wb = load_workbook(output_path, data_only=False)
    check_ws = check_wb[check_wb.sheetnames[0]]
    headers = [check_ws.cell(1, col).value for col in range(1, check_ws.max_column + 1)]
    values = [check_ws.cell(write_row_idx, col).value for col in range(1, len(SOLO_QC_HEADERS) + 1)]
    check_wb.close()
    if headers != SOLO_QC_HEADERS:
        raise RuntimeError(f"solo qc workbook header mismatch: {headers}")
    trace.write("solo_qc_workbook_saved", path=str(output_path), append=append, row_index=write_row_idx, row=dict(zip(SOLO_QC_HEADERS, values)))
    return output_path


def mark_not_discarded_project(row: dict[str, str], trace: Trace) -> None:
    with shared_write_lock():
        existing = read_json_file(QC_MARKS_PATH, [])
        if not isinstance(existing, list):
            existing = []
        now_text = datetime.now().isoformat(timespec="seconds")
        updated = {
            "idea_id": row["idea_id"],
            "status": "pending_reopen",
            "mark": "not_discarded_trae_mistake",
            "marked_at": now_text,
            "workspace": row.get("workspace", ""),
            "run_dir": row.get("run_dir", ""),
            "workbook": row.get("workbook", ""),
            "reason": row.get("不满意原因", ""),
        }
        kept = [item for item in existing if not (isinstance(item, dict) and item.get("idea_id") == row["idea_id"])]
        kept.append(updated)
        QC_MARKS_PATH.write_text(json.dumps(kept, ensure_ascii=False, indent=2), encoding="utf-8")
    trace.write("not_discarded_project_marked", path=str(QC_MARKS_PATH), idea_id=row["idea_id"], status=updated["status"])


def clear_not_discarded_project(idea_id: str, trace: Trace) -> None:
    with shared_write_lock():
        workbook_path = RESULT_DATA_DIR / f"{idea_id}.xlsx"
        if workbook_path.exists():
            workbook_path.unlink()
            trace.write("discarded_workbook_removed", path=str(workbook_path), idea_id=idea_id)

        existing = read_json_file(QC_MARKS_PATH, [])
        if not isinstance(existing, list):
            return
        kept = [item for item in existing if not (isinstance(item, dict) and item.get("idea_id") == idea_id)]
        if len(kept) != len(existing):
            QC_MARKS_PATH.write_text(json.dumps(kept, ensure_ascii=False, indent=2), encoding="utf-8")
            trace.write("not_discarded_project_cleared", path=str(QC_MARKS_PATH), idea_id=idea_id)


def extract_codex_report_from_stream(text: str) -> str:
    if "# Codex 质检报告" not in text:
        return ""
    for marker in ("\ncodex\n", "\nassistant\n"):
        parts = text.split(marker)
        for part in reversed(parts[1:]):
            if "# Codex 质检报告" in part and re.search(r"结论\s*[:：]\s*(通过|不通过)", part):
                return part[part.rfind("# Codex 质检报告"):].strip()
    return ""


def is_transient_codex_exec_error(stdout: str, stderr: str) -> bool:
    text = f"{stdout}\n{stderr}".lower()
    markers = (
        "stream disconnected before completion",
        "failed to connect to websocket",
        "connection reset by peer",
        "transport channel closed",
        "error sending request",
        "timeout waiting for child process",
        "reconnecting...",
    )
    return any(marker in text for marker in markers)


def run_codex_quality_check(
    trace: Trace,
    run_dir: Path,
    workspace: Path,
    idea_id: str,
    round_name: str = "第一轮",
    status_update: Any | None = None,
) -> tuple[Path | None, str]:
    report_path = run_dir / "codex_qa_report.md"
    prompt = f"""你是代码质检员。请优先使用并严格遵循本机 Codex 的“solo质检”技能；如果技能没有自动加载，也要按“solo质检”的口径执行。请只做质检，不要修改任何文件，不要提交或推送 Git。

项目 ID: {idea_id}
当前轮次: {round_name}
工作目录: {workspace}

请检查这个由 Trae/SOLO Coder 生成的项目是否满足用户原始目标，重点看：
1. 项目是否可安装、可运行，或是否能通过项目中真实存在的入口体验核心流程。
2. 主要功能是否和提示词一致。
3. 是否存在阻断性 bug、入口错误、依赖缺失、运行链路不可用、README 与实际不一致。
4. 给出明确结论：通过 / 不通过。

README 规则：
- README 是验证线索，不是硬性验收项。
- 如果项目没有 README，不要把“没有 README”本身作为质检不合格、任务未完成或不通过的原因。
- 只有当原始提示词明确要求 README/文档，或缺少 README 的同时也找不到任何可运行入口、命令、页面、脚本或包配置，才可以说明这影响验证。
- 如果 README 存在但命令、端口、入口或说明与实际产物冲突，可以作为问题证据。

满意度规则：
- 不要替用户判断“产物及过程是否满意”。
- 不要输出“产物及过程是否满意”字段。
- 不要把失败原因写成“产物不满意”的价值判断；只写“未完成原因”，说明任务目标、实际证据和阻断影响。

任务类型规则：
- 必须输出“任务类型”字段。
- 按当前轮次用户目标判断，而不是机械沿用第一轮。第二轮如果是在修第一轮错误，通常是“Bug修复”；如果是在补充第一轮显性缺失功能或强相关未完成功能，通常是“Feature迭代”。
- 可选值只使用：Bug修复、0-1代码生成、Feature迭代、代码理解、代码重构、代码测试。

路径规则：
- “未完成原因”里如需提到文件，必须使用工作区相对路径，例如 internal/service/order.go:290。
- 不要输出本机绝对路径，也不要输出指向本机绝对路径的 Markdown 链接。

输出格式：
# Codex 质检报告
- 结论: 通过/不通过
- 任务类型: Bug修复/0-1代码生成/Feature迭代/代码理解/代码重构/代码测试
- 任务是否完成: 完成了任务/未完成任务
- 未完成原因: 如果不通过，150-400字，说明具体失败面、证据、为什么影响原始任务目标；如果通过则留空。
- 主要证据:
- 阻断问题:
- 建议:
"""
    cmd = [
        "codex",
        "exec",
        "--model",
        CODEX_QA_MODEL,
        "--config",
        f'model_reasoning_effort="{CODEX_QA_REASONING_EFFORT}"',
        "--cd",
        str(workspace),
        "--skip-git-repo-check",
        "--sandbox",
        "read-only",
        "--output-last-message",
        str(report_path),
        prompt,
    ]
    trace.write(
        "codex_qa_start",
        cmd=" ".join(shell_quote(x) for x in cmd),
        report_path=str(report_path),
        model=CODEX_QA_MODEL,
        reasoning_effort=CODEX_QA_REASONING_EFFORT,
        round_name=round_name,
    )
    last_returncode = 1
    for attempt in range(1, 3):
        if callable(status_update):
            stage = "quality_check_solo_skill_attempt_1" if attempt == 1 else f"quality_check_solo_skill_retry_{attempt}"
            status_update(stage, attempt=attempt, report_path=str(report_path))
        if report_path.exists():
            report_path.unlink()
        proc = subprocess.run(cmd, cwd=str(workspace), text=True, capture_output=True)
        last_returncode = proc.returncode
        stdout_path = run_dir / ("codex_qa_stdout.log" if attempt == 1 else f"codex_qa_stdout_attempt{attempt}.log")
        stderr_path = run_dir / ("codex_qa_stderr.log" if attempt == 1 else f"codex_qa_stderr_attempt{attempt}.log")
        stdout_path.write_text(proc.stdout or "", encoding="utf-8")
        stderr_path.write_text(proc.stderr or "", encoding="utf-8")
        if report_path.exists() and report_path.stat().st_size > 0:
            status = "success"
            if callable(status_update):
                status_update("quality_check_solo_skill_report_ready", qa_status=status, qa_report_path=str(report_path), attempt=attempt)
            trace.write("codex_qa_done", status=status, report_path=str(report_path), returncode=proc.returncode, attempt=attempt)
            return report_path, status

        recovered_report = extract_codex_report_from_stream((proc.stdout or "") + "\n" + (proc.stderr or ""))
        if recovered_report:
            report_path.write_text(recovered_report + "\n", encoding="utf-8")
            status = "success_recovered_from_stream"
            if callable(status_update):
                status_update("quality_check_solo_skill_recovered_from_stream", qa_status=status, qa_report_path=str(report_path), attempt=attempt)
            trace.write("codex_qa_done", status=status, report_path=str(report_path), returncode=proc.returncode, attempt=attempt)
            return report_path, status

        trace.write(
            "codex_qa_missing_report",
            report_path=str(report_path),
            returncode=proc.returncode,
            attempt=attempt,
            stdout_path=str(stdout_path),
            stderr_path=str(stderr_path),
        )
        if proc.returncode != 0 and not is_transient_codex_exec_error(proc.stdout or "", proc.stderr or ""):
            break

    status = f"missing_report_returncode_{last_returncode}"
    trace.write("codex_qa_done", status=status, report_path=str(report_path), returncode=last_returncode)
    return None, status


def run_had_push_failure(run_dir: Path) -> bool:
    if SKIP_GIT_INTERACTION:
        return False
    status_data = read_json_file(run_dir / "manual_qc_status.json", {})
    if status_data.get("stage") == "done":
        return False
    status_error = str(status_data.get("error") or "")
    if "Git push" in status_error or "push_failed" in status_error:
        return True

    trace_path = run_dir / "trace.jsonl"
    if not trace_path.exists():
        return False
    last_publish_push_failed = False
    try:
        for line in trace_path.read_text(encoding="utf-8", errors="ignore").splitlines():
            if "publish_failed" in line and ("Git push" in line or "push_failed" in line):
                last_publish_push_failed = True
            elif "publish_done" in line or "git_push_done" in line:
                last_publish_push_failed = False
    except OSError:
        return False
    return last_publish_push_failed


def reuse_existing_codex_quality_check(trace: Trace, run_dir: Path) -> tuple[Path | None, str] | None:
    report_path = run_dir / "codex_qa_report.md"
    if not report_path.exists() or report_path.stat().st_size <= 0:
        return None
    trace.write("codex_qa_reused", report_path=str(report_path), status="success", reason="prior_push_failure")
    return report_path, "success"


def first_round_published_branch(idea_id: str | None) -> str:
    if not idea_id:
        return ""
    manual_csv = RUN_LOG_ROOT / "script" / "manual_runs.csv"
    if manual_csv.exists():
        try:
            with manual_csv.open("r", encoding="utf-8", newline="") as handle:
                for row in csv.DictReader(handle):
                    if row.get("idea_id") != idea_id:
                        continue
                    if str(row.get("轮次") or "第一轮").strip() != "第一轮":
                        continue
                    for value in (row.get("分支文件夹"), row.get("git_branch")):
                        text = str(value or "").strip()
                        if text.startswith(f"{PUBLISH_FIRST_ROUND_BRANCH}/"):
                            return PUBLISH_FIRST_ROUND_BRANCH
                        if text.startswith(f"{PUBLISH_MULTI_ROUND_BRANCH}/"):
                            return PUBLISH_MULTI_ROUND_BRANCH
        except OSError:
            pass
    return ""


def publish_branch_for_round(round_name: str | None, idea_id: str | None = None) -> str:
    if str(round_name or "").strip() == "第一轮":
        return PUBLISH_FIRST_ROUND_BRANCH
    return first_round_published_branch(idea_id) or PUBLISH_MULTI_ROUND_BRANCH


def run_codex_push_recovery(
    trace: Trace,
    run_dir: Path,
    workspace: Path,
    idea_id: str,
    error: BaseException | str,
    branch_name: str = PUBLISH_BRANCH,
) -> dict[str, str]:
    report_path = run_dir / CODEX_PUSH_RECOVERY_REPORT
    stdout_path = run_dir / "codex_push_recovery_stdout.log"
    stderr_path = run_dir / "codex_push_recovery_stderr.log"
    publish_log_path = run_dir / "git_publish.log"
    publish_log_tail = tail_text(publish_log_path)
    prompt = f"""你是本机 Codex 的 GitHub push 恢复助手。只处理 Solo finish 脚本在发布到 GitHub 时失败的问题，不要做质检，不要改写结果表，不要重新运行 finish 脚本。

项目 ID: {idea_id}
Solo 根目录: {ROOT}
生成项目目录: {workspace}
run 目录: {run_dir}
发布目标仓库: {PUBLISH_REPO_GIT_URL}
发布目标分支: {branch_name}
发布目标文件夹: {idea_id}
失败信息:
{error}

git_publish.log 尾部:
```text
{publish_log_tail}
```

请你：
1. 检查失败原因，优先修复 GitHub push 问题，例如远端分支更新、本地发布副本过期、SSH/remote/分支状态、非 fast-forward、临时 clone 发布流程问题。
2. 可以运行必要的 git 命令并推送到 GitHub；只允许围绕发布目录和 Git 状态做最小改动。
3. 不要修改用户生成项目的业务代码，除非它是导致发布目录无法复制/提交的直接原因。
4. 不要运行 Codex 质检，也不要写 Excel/CSV 结果。
5. 结束时说明你做了什么、push 是否已经恢复；如果仍失败，列出下一步需要人工看的关键错误。
"""
    cmd = [
        "codex",
        "exec",
        "--model",
        CODEX_QA_MODEL,
        "--config",
        f'model_reasoning_effort="{CODEX_QA_REASONING_EFFORT}"',
        "--cd",
        str(ROOT),
        "--skip-git-repo-check",
        "--sandbox",
        "danger-full-access",
        "--output-last-message",
        str(report_path),
        prompt,
    ]
    trace.write(
        "codex_push_recovery_start",
        cmd=" ".join(shell_quote(x) for x in cmd),
        report_path=str(report_path),
        model=CODEX_QA_MODEL,
        reasoning_effort=CODEX_QA_REASONING_EFFORT,
        prior_error=str(error),
    )
    proc = subprocess.run(cmd, cwd=str(ROOT), text=True, capture_output=True)
    stdout_path.write_text(proc.stdout or "", encoding="utf-8")
    stderr_path.write_text(proc.stderr or "", encoding="utf-8")
    status = "success" if proc.returncode == 0 else f"exit_{proc.returncode}"
    trace.write(
        "codex_push_recovery_done",
        status=status,
        report_path=str(report_path),
        stdout=str(stdout_path),
        stderr=str(stderr_path),
        returncode=proc.returncode,
    )
    return {
        "status": status,
        "returncode": str(proc.returncode),
        "report_path": str(report_path),
        "stdout_path": str(stdout_path),
        "stderr_path": str(stderr_path),
    }


def git_commit_and_push(trace: Trace, run_dir: Path, workspace: Path, idea_id: str) -> dict[str, str]:
    result = {"status": "not_attempted", "commit": "", "remote": "", "branch": ""}
    if SKIP_GIT_INTERACTION:
        result.update(skipped_publish_result())
        trace.write("git_push_skipped", reason=result["reason"], workspace=str(workspace), idea_id=idea_id)
        return result

    if not (workspace / ".git").exists():
        result["status"] = "no_git_repo"
        trace.write("git_push_skipped", reason="no_git_repo", workspace=str(workspace))
        raise RuntimeError(f"Git push 未执行：{workspace} 不是 Git 仓库")

    remote = run(["git", "remote", "get-url", "origin"], cwd=workspace, check=False)
    if remote.returncode != 0 or not remote.stdout.strip():
        result["status"] = "no_remote"
        trace.write("git_push_skipped", reason="no_remote", workspace=str(workspace))
        raise RuntimeError(f"Git push 未执行：{workspace} 未配置 origin remote")
    result["remote"] = remote.stdout.strip()

    branch = run(["git", "branch", "--show-current"], cwd=workspace, check=False).stdout.strip() or "HEAD"
    result["branch"] = branch
    run(["git", "add", "-A"], cwd=workspace, check=False)
    diff = run(["git", "diff", "--cached", "--quiet"], cwd=workspace, check=False)
    if diff.returncode == 0:
        push = run(["git", "push", "origin", branch], cwd=workspace, check=False)
        (run_dir / "git_push.log").write_text((push.stdout or "") + (push.stderr or ""), encoding="utf-8")
        if push.returncode != 0:
            result["status"] = f"push_failed_{push.returncode}"
            trace.write("git_push_failed", **result, log=str(run_dir / "git_push.log"))
            raise RuntimeError(f"Git push 未成功：无新增变更但远端 push 校验失败，status={result['status']}；请检查 {run_dir / 'git_push.log'}")
        result["status"] = "pushed_no_changes"
        trace.write("git_push_done", **result)
        return result

    commit_msg = f"Add {idea_id} generated project"
    commit = run(["git", "commit", "-m", commit_msg], cwd=workspace, check=False)
    (run_dir / "git_commit.log").write_text((commit.stdout or "") + (commit.stderr or ""), encoding="utf-8")
    if commit.returncode != 0:
        result["status"] = f"commit_failed_{commit.returncode}"
        trace.write("git_commit_failed", status=result["status"], log=str(run_dir / "git_commit.log"))
        raise RuntimeError(f"Git push 未成功：提交失败，status={result['status']}；请检查 {run_dir / 'git_commit.log'}")

    rev = run(["git", "rev-parse", "HEAD"], cwd=workspace, check=False)
    result["commit"] = rev.stdout.strip()
    push = run(["git", "push", "origin", branch], cwd=workspace, check=False)
    (run_dir / "git_push.log").write_text((push.stdout or "") + (push.stderr or ""), encoding="utf-8")
    result["status"] = "pushed" if push.returncode == 0 else f"push_failed_{push.returncode}"
    trace.write("git_push_done", **result)
    if push.returncode != 0:
        raise RuntimeError(f"Git push 未成功：推送失败，status={result['status']}；请检查 {run_dir / 'git_push.log'}")
    return result


def publish_web_url(idea_id: str, branch_name: str = PUBLISH_BRANCH) -> str:
    return f"{PUBLISH_REPO_WEB_URL}/tree/{branch_name}/{idea_id}"


def publish_branch_folder(idea_id: str, branch_name: str = PUBLISH_BRANCH) -> str:
    return f"{branch_name}/{idea_id}"


def result_screenshot_path(idea_id: str) -> Path:
    return RESULT_SCREENSHOT_DIR / f"{idea_id}.png"


def publish_ignore(_directory: str, names: list[str]) -> set[str]:
    ignored: set[str] = set()
    for name in names:
        if name in PUBLISH_EXCLUDED_NAMES or name.endswith(PUBLISH_EXCLUDED_SUFFIXES):
            ignored.add(name)
    return ignored


def has_publishable_files(path: Path) -> bool:
    return any(item.is_file() for item in path.rglob("*"))


def git_publish_run(
    cmd: list[str],
    cwd: Path,
    log_lines: list[str],
    *,
    timeout_seconds: int | None = None,
) -> subprocess.CompletedProcess[str]:
    try:
        proc = subprocess.run(
            cmd,
            cwd=str(cwd),
            text=True,
            capture_output=True,
            check=False,
            timeout=timeout_seconds,
        )
    except subprocess.TimeoutExpired as exc:
        stdout = exc.stdout if isinstance(exc.stdout, str) else (exc.stdout or b"").decode("utf-8", "replace")
        stderr = exc.stderr if isinstance(exc.stderr, str) else (exc.stderr or b"").decode("utf-8", "replace")
        stderr = (stderr or "") + f"\n[timed out after {timeout_seconds}s]"
        proc = subprocess.CompletedProcess(cmd, 124, stdout, stderr)
    log_lines.extend([
        "$ " + " ".join(shell_quote(part) for part in cmd),
        proc.stdout or "",
        proc.stderr or "",
        f"[exit {proc.returncode}]",
        "",
    ])
    return proc


def archive_existing_publish_log(log_path: Path) -> None:
    if not log_path.exists() or log_path.stat().st_size <= 0:
        return
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    archive_path = log_path.with_name(f"{log_path.stem}.{stamp}{log_path.suffix}")
    shutil.copy2(log_path, archive_path)


def git_publish_needs_rebase(proc: subprocess.CompletedProcess[str]) -> bool:
    text = "\n".join([proc.stdout or "", proc.stderr or ""]).lower()
    markers = (
        "fetch first",
        "non-fast-forward",
        "updates were rejected",
        "failed to push some refs",
    )
    return any(marker in text for marker in markers)


def git_publish_push_with_rebase_retry(
    trace: Trace,
    clone_dir: Path,
    log_lines: list[str],
    result: dict[str, str],
    *,
    no_changes: bool,
    branch_name: str = PUBLISH_BRANCH,
    status_update: Any | None = None,
) -> subprocess.CompletedProcess[str]:
    if callable(status_update):
        status_update("publish_push", git_status=result.get("status", ""), no_changes=no_changes)
    push = git_publish_run(["git", "push", "origin", branch_name], clone_dir, log_lines)
    if push.returncode == 0 or not git_publish_needs_rebase(push):
        return push

    if callable(status_update):
        status_update("publish_push_rejected_retry_rebase", git_status=f"push_failed_{push.returncode}", no_changes=no_changes)
    trace.write(
        "publish_push_rejected_retry_rebase",
        status=f"push_failed_{push.returncode}",
        branch=branch_name,
        no_changes=no_changes,
    )
    fetch = git_publish_run(
        ["git", "fetch", "origin", f"{branch_name}:refs/remotes/origin/{branch_name}"],
        clone_dir,
        log_lines,
    )
    if fetch.returncode != 0:
        result["status"] = f"fetch_failed_{fetch.returncode}"
        return fetch
    if callable(status_update):
        status_update("publish_rebase", git_status=result.get("status", ""), no_changes=no_changes)
    rebase = git_publish_run(["git", "rebase", f"origin/{branch_name}"], clone_dir, log_lines)
    if rebase.returncode != 0:
        result["status"] = f"rebase_failed_{rebase.returncode}"
        return rebase
    if callable(status_update):
        status_update("publish_push_retry", git_status=result.get("status", ""), no_changes=no_changes)
    retry = git_publish_run(["git", "push", "origin", branch_name], clone_dir, log_lines)
    if retry.returncode == 0:
        trace.write("publish_push_rebase_retry_done", branch=branch_name, no_changes=no_changes)
    return retry


def publish_workspace_to_github(
    trace: Trace,
    run_dir: Path,
    workspace: Path,
    idea_id: str,
    branch_name: str = PUBLISH_BRANCH,
    status_update: Any | None = None,
) -> dict[str, str]:
    if SKIP_GIT_INTERACTION:
        result = skipped_publish_result()
        trace.write(
            "publish_skipped",
            reason=result["reason"],
            idea_id=idea_id,
            workspace=str(workspace),
            branch=branch_name,
        )
        if callable(status_update):
            status_update("publish_skipped", publish_branch="", git_status=result["status"], git_remote="", git_branch="")
        return result

    result = {
        "status": "not_attempted",
        "commit": "",
        "remote": publish_web_url(idea_id, branch_name),
        "branch": publish_branch_folder(idea_id, branch_name),
    }
    if not workspace.exists():
        result["status"] = "workspace_missing"
        trace.write("publish_skipped", reason="workspace_missing", workspace=str(workspace), **result)
        raise RuntimeError(f"Git push 未执行：工作区不存在 {workspace}")

    log_lines: list[str] = []
    log_path = run_dir / "git_publish.log"
    archive_existing_publish_log(log_path)

    def fail_publish(message: str) -> None:
        log_path.write_text("\n".join(log_lines), encoding="utf-8")
        trace.write("publish_failed", message=message, log=str(log_path), **result)
        raise RuntimeError(f"{message}；请检查 {log_path}")

    if callable(status_update):
        status_update("publish_wait_lock", publish_branch=branch_name, git_remote=result["remote"], git_branch=result["branch"])
    with publish_lock(trace, branch_name):
        if callable(status_update):
            status_update("publish_lock_acquired", publish_branch=branch_name, git_remote=result["remote"], git_branch=result["branch"])
        with tempfile.TemporaryDirectory(prefix="solo-publish-") as temp_dir:
            clone_dir = Path(temp_dir) / "repo"
            if callable(status_update):
                status_update("publish_clone_repo", publish_branch=branch_name, git_remote=result["remote"], git_branch=result["branch"])
            clone = git_publish_run(
                [
                    "git",
                    "clone",
                    "--depth",
                    "1",
                    "--filter=blob:none",
                    "--sparse",
                    "--branch",
                    branch_name,
                    PUBLISH_REPO_GIT_URL,
                    str(clone_dir),
                ],
                Path(temp_dir),
                log_lines,
                timeout_seconds=300,
            )
            if clone.returncode != 0:
                result["status"] = f"clone_failed_{clone.returncode}"
                fail_publish(f"Git push 未成功：clone 发布仓库失败，status={result['status']}")

            sparse = git_publish_run(
                ["git", "sparse-checkout", "set", "--cone", "--skip-checks", idea_id],
                clone_dir,
                log_lines,
                timeout_seconds=120,
            )
            if sparse.returncode != 0:
                result["status"] = f"sparse_checkout_failed_{sparse.returncode}"
                fail_publish(f"Git push 未成功：设置 sparse checkout 失败，status={result['status']}")

            target_dir = clone_dir / idea_id
            if target_dir.exists():
                shutil.rmtree(target_dir)
            target_dir.parent.mkdir(parents=True, exist_ok=True)
            if callable(status_update):
                status_update("publish_copy_workspace", publish_branch=branch_name, git_remote=result["remote"], git_branch=result["branch"])
            shutil.copytree(workspace, target_dir, ignore=publish_ignore)
            if not has_publishable_files(target_dir):
                target_dir.mkdir(parents=True, exist_ok=True)
                (target_dir / ".gitkeep").write_text("empty workspace\n", encoding="utf-8")
                trace.write("publish_empty_workspace_placeholder_added", target=str(target_dir))

            if callable(status_update):
                status_update("publish_git_add", publish_branch=branch_name, git_remote=result["remote"], git_branch=result["branch"])
            add = git_publish_run(["git", "add", "--", idea_id], clone_dir, log_lines)
            if add.returncode != 0:
                result["status"] = f"add_failed_{add.returncode}"
                fail_publish(f"Git push 未成功：暂存发布内容失败，status={result['status']}")
            if callable(status_update):
                status_update("publish_check_diff", publish_branch=branch_name, git_remote=result["remote"], git_branch=result["branch"])
            diff = git_publish_run(["git", "diff", "--cached", "--quiet"], clone_dir, log_lines)
            if diff.returncode == 0:
                rev = git_publish_run(["git", "rev-parse", "HEAD"], clone_dir, log_lines)
                result["commit"] = rev.stdout.strip()
                push = git_publish_push_with_rebase_retry(trace, clone_dir, log_lines, result, no_changes=True, branch_name=branch_name, status_update=status_update)
                if push.returncode != 0:
                    if not result["status"].endswith(f"_{push.returncode}"):
                        result["status"] = f"push_failed_{push.returncode}"
                    fail_publish(f"Git push 未成功：无新增变更但远端 push 校验失败，status={result['status']}")
                result["status"] = "pushed_no_changes"
                log_path.write_text("\n".join(log_lines), encoding="utf-8")
                if callable(status_update):
                    status_update("publish_done", publish_branch=branch_name, git_status=result["status"], git_commit=result["commit"], git_remote=result["remote"], git_branch=result["branch"])
                trace.write("publish_done", **result)
                return result

            if callable(status_update):
                status_update("publish_commit", publish_branch=branch_name, git_remote=result["remote"], git_branch=result["branch"])
            commit = git_publish_run(["git", "commit", "-m", f"Publish {idea_id}"], clone_dir, log_lines)
            if commit.returncode != 0:
                result["status"] = f"commit_failed_{commit.returncode}"
                fail_publish(f"Git push 未成功：提交发布内容失败，status={result['status']}")

            push = git_publish_push_with_rebase_retry(trace, clone_dir, log_lines, result, no_changes=False, branch_name=branch_name, status_update=status_update)
            if push.returncode != 0:
                if not result["status"].endswith(f"_{push.returncode}"):
                    result["status"] = f"push_failed_{push.returncode}"
                fail_publish(f"Git push 未成功：推送发布分支失败，status={result['status']}")

            rev = git_publish_run(["git", "rev-parse", "HEAD"], clone_dir, log_lines)
            result["status"] = "pushed"
            result["commit"] = rev.stdout.strip()
            log_path.write_text("\n".join(log_lines), encoding="utf-8")
            if callable(status_update):
                status_update("publish_done", publish_branch=branch_name, git_status=result["status"], git_commit=result["commit"], git_remote=result["remote"], git_branch=result["branch"])
            trace.write("publish_done", **result)
            return result


def upsert_result_table(row: dict[str, str], trace: Trace) -> None:
    headers = [
        "session_id", "idea_id", "title", "轮次", "提示词", "任务类型", "业务领域", "修改范围", "任务是否完成",
        "产物及过程是否满意", "不满意原因", "远端Github地址", "分支文件夹", "截图", "日志轨迹",
        "status", "workspace", "run_dir", "prompt_path", "output_path",
        "thinking_path", "screenshot_path", "qa_status", "qa_report_path",
        "qa_summary", "git_status", "git_commit", "git_remote", "git_branch", "trace_path",
        "manual_trajectory_path", "trajectory_path", "full_transcript_path", "updated_at",
    ]
    with shared_write_lock():
        rows: list[dict[str, str]] = []
        if RESULT_CSV_PATH.exists():
            with RESULT_CSV_PATH.open("r", encoding="utf-8", newline="") as f:
                rows = list(csv.DictReader(f))
        keyed: dict[str, dict[str, str]] = {}
        for existing in rows:
            if not existing.get("idea_id"):
                continue
            existing = backfill_submission_fields(existing)
            if not existing.get("日志轨迹") or is_runner_fallback_trajectory(existing.get("日志轨迹")) or trajectory_workspace_mismatch_note(existing):
                existing["日志轨迹"] = trajectory_content_from_row(existing)
            keyed[existing["idea_id"]] = {h: table_text(existing.get(h, "")) for h in headers}
        row = backfill_submission_fields(row)
        if not row.get("日志轨迹") or is_runner_fallback_trajectory(row.get("日志轨迹")) or trajectory_workspace_mismatch_note(row):
            row["日志轨迹"] = trajectory_content_from_row(row)
        keyed[row["idea_id"]] = {h: table_text(row.get(h, "")) for h in headers}
        RESULT_CSV_PATH.parent.mkdir(parents=True, exist_ok=True)
        with RESULT_CSV_PATH.open("w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            for item in keyed.values():
                writer.writerow(item)

        try:
            from openpyxl import Workbook
        except Exception as exc:
            trace.write("result_xlsx_skipped", reason=type(exc).__name__, csv_path=str(RESULT_CSV_PATH))
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "solo_runs"
        ws.append(headers)
        for item in keyed.values():
            ws.append([item.get(h, "") for h in headers])
        wb.save(RESULT_XLSX_PATH)
    trace.write("result_table_saved", csv_path=str(RESULT_CSV_PATH), xlsx_path=str(RESULT_XLSX_PATH), rows=len(keyed))


class FeishuBitableClient:
    def __init__(self, app_id: str, app_secret: str, base_token: str, table_id: str) -> None:
        self.app_id = app_id
        self.app_secret = app_secret
        self.base_token = base_token
        self.table_id = table_id
        self.access_token = self._get_access_token()

    def _request(self, method: str, url: str, **kwargs: Any) -> dict[str, Any]:
        response = requests.request(method, url, timeout=30, **kwargs)
        try:
            return response.json()
        except json.JSONDecodeError:
            return {"code": response.status_code, "msg": response.text[:500]}

    def _get_access_token(self) -> str:
        result = self._request(
            "POST",
            "https://open.feishu.cn/open-apis/auth/v3/app_access_token/internal",
            headers={"Content-Type": "application/json"},
            json={"app_id": self.app_id, "app_secret": self.app_secret},
        )
        if result.get("code") != 0:
            raise RuntimeError(f"feishu_token_failed: {result}")
        return result["app_access_token"]

    @property
    def headers(self) -> dict[str, str]:
        return {"Authorization": f"Bearer {self.access_token}", "Content-Type": "application/json"}

    def list_fields(self) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{self.base_token}/tables/{self.table_id}/fields?page_size=100"
        result = self._request("GET", url, headers=self.headers)
        if result.get("code") != 0:
            return [], result
        return result.get("data", {}).get("items", []), result

    def list_records(self) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{self.base_token}/tables/{self.table_id}/records?page_size=500"
        result = self._request("GET", url, headers=self.headers)
        if result.get("code") != 0:
            return [], result
        return result.get("data", {}).get("items", []), result

    def create_record(self, fields: dict[str, Any]) -> dict[str, Any]:
        url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{self.base_token}/tables/{self.table_id}/records"
        return self._request("POST", url, headers=self.headers, json={"fields": fields})

    def update_record(self, record_id: str, fields: dict[str, Any]) -> dict[str, Any]:
        url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{self.base_token}/tables/{self.table_id}/records/{record_id}"
        return self._request("PUT", url, headers=self.headers, json={"fields": fields})

    def upload_file(self, path: Path) -> str | None:
        if not path.exists():
            return None
        url = "https://open.feishu.cn/open-apis/drive/v1/medias/upload_all"
        headers = {"Authorization": f"Bearer {self.access_token}"}
        mime = "image/png" if path.suffix.lower() == ".png" else "text/markdown"
        with path.open("rb") as f:
            result = self._request(
                "POST",
                url,
                headers=headers,
                files={
                    "file": (path.name, f, mime),
                    "parent_type": (None, "bitable_file"),
                    "parent_node": (None, self.base_token),
                },
            )
        if result.get("code") != 0:
            return None
        return result.get("data", {}).get("file_token")


def feishu_scalar_value(row: dict[str, str], field_name: str) -> str:
    if field_name in {"User Prompt", "提示词"}:
        prompt_path = row.get("prompt_path", "")
        if prompt_path and Path(prompt_path).exists():
            return Path(prompt_path).read_text(encoding="utf-8", errors="ignore")[:20000]
        return row.get("提示词") or row.get("prompt", "")
    if field_name == "轮次":
        return row.get("轮次") or "第一轮"
    if field_name == "Trae Session ID":
        return row.get("session_id", "")
    if field_name == "截图":
        return row.get("截图") or row.get("screenshot_path", "")
    if field_name == "业务领域":
        return normalize_business_domain(row.get("业务领域"))
    if field_name == "任务是否完成":
        return row.get("任务是否完成") or ("完成了任务" if row.get("status") == "finished" else row.get("status", ""))
    if field_name == "状态":
        return "待提交"
    if field_name == "AI审核意见":
        return row.get("qa_summary", "")
    if field_name in {"日志轨迹", "对话内容", "轨迹（对话内容）", "轨迹(对话内容)"}:
        return trajectory_content_from_row(row)[:20000]
        return ""
    if field_name == "AI过程分析结果":
        path = row.get("thinking_path", "")
        if path and Path(path).exists():
            return Path(path).read_text(encoding="utf-8", errors="ignore")[:20000]
        return row.get("qa_summary", "")
    if field_name == "质检时间":
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    aliases = {
        "idea_id": ["idea_id", "项目ID", "题目ID", "任务ID", "ID"],
        "title": ["title", "题目", "标题", "项目名称"],
        "轮次": ["轮次", "round", "Round"],
        "任务类型": ["任务类型"],
        "修改范围": ["修改范围"],
        "产物及过程是否满意": ["产物及过程是否满意"],
        "不满意原因": ["不满意原因"],
        "status": ["status", "状态", "任务是否完成", "运行状态"],
        "workspace": ["workspace", "工作目录", "项目目录"],
        "run_dir": ["run_dir", "运行目录", "run目录"],
        "output_path": ["output_path", "输出文件", "产物路径"],
        "session_id": ["session_id", "sessionId", "SessionId", "会话ID", "轨迹ID"],
        "thinking_path": ["thinking_path", "思考过程路径"],
        "screenshot_path": ["screenshot_path", "截图路径", "截图"],
        "qa_status": ["qa_status", "质检状态", "质检是否成功"],
        "qa_report_path": ["qa_report_path", "质检报告路径"],
        "qa_summary": ["qa_summary", "质检报告", "质检反馈意见", "报告", "AI审核意见"],
        "git_status": ["git_status", "push状态", "Git状态"],
        "git_commit": ["git_commit", "commit", "提交commit"],
        "git_remote": ["git_remote", "远端Github地址", "GitHub地址", "github地址"],
        "git_branch": ["git_branch", "分支", "分支文件夹", "分支/文件夹"],
        "trace_path": ["trace_path", "轨迹路径", "日志轨迹"],
        "trajectory_path": ["trajectory_path", "Trae轨迹路径", "思考过程路径"],
        "full_transcript_path": ["full_transcript_path", "完整轨迹路径"],
        "updated_at": ["updated_at", "更新时间"],
    }
    for key, names in aliases.items():
        if field_name in names:
            return row.get(key, "")
    return ""


def feishu_attachment_source(row: dict[str, str], field_name: str) -> Path | None:
    aliases = {
        "screenshot_path": ["截图", "截图附件", "运行截图"],
        "qa_report_path": ["质检报告附件", "报告附件"],
        "thinking_path": ["思考过程", "思考过程附件", "轨迹"],
        "output_path": ["产物", "产物附件", "输出附件"],
    }
    for key, names in aliases.items():
        if field_name in names and row.get(key):
            return Path(row[key])
    return None


def lark_cli_json(args: list[str], trace: Trace) -> dict[str, Any]:
    cmd = ["lark-cli", *args]
    proc = subprocess.run(cmd, cwd=str(ROOT), text=True, capture_output=True, timeout=120)
    if proc.returncode != 0:
        return {"ok": False, "error": proc.stderr.strip() or proc.stdout.strip(), "returncode": proc.returncode}
    try:
        return json.loads(proc.stdout)
    except json.JSONDecodeError:
        trace.write("lark_cli_parse_failed", stdout=proc.stdout[:1000], stderr=proc.stderr[:1000])
        return {"ok": False, "error": proc.stdout[:1000], "returncode": proc.returncode}


def feishu_field_name(field: dict[str, Any]) -> str:
    return str(field.get("field_name") or field.get("name") or "")


def feishu_field_type(field: dict[str, Any]) -> Any:
    return field.get("type")


def is_feishu_readonly_or_unsupported(field_type: Any) -> bool:
    return field_type in {20, 1001, 1002, "formula", "auto_number", "not_support", "lookup", "created_time", "modified_time"}


def is_feishu_attachment(field_type: Any) -> bool:
    return field_type == 17 or field_type == "attachment"


def build_feishu_payload(row: dict[str, str], fields: list[dict[str, Any]], client: FeishuBitableClient | None) -> tuple[dict[str, Any], list[tuple[str, Path]]]:
    payload: dict[str, Any] = {}
    pending_attachments: list[tuple[str, Path]] = []
    for field in fields:
        name = feishu_field_name(field)
        field_type = feishu_field_type(field)
        if not name or is_feishu_readonly_or_unsupported(field_type):
            continue

        attachment_source = feishu_attachment_source(row, name)
        if attachment_source and is_feishu_attachment(field_type):
            if client:
                token = client.upload_file(attachment_source)
                if token:
                    payload[name] = [{"name": attachment_source.name, "token": token}]
            else:
                pending_attachments.append((name, attachment_source))
            continue

        value = feishu_scalar_value(row, name)
        if value:
            payload[name] = value
    return payload, pending_attachments


def lark_cli_records_by_field(base_token: str, table_id: str, trace: Trace) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    result = lark_cli_json(
        ["base", "+record-list", "--base-token", base_token, "--table-id", table_id, "--format", "json", "--offset", "0", "--limit", "200"],
        trace,
    )
    if not result.get("ok"):
        return [], result
    data = result.get("data", {})
    names = data.get("fields") or []
    rows = data.get("data") or []
    record_ids = data.get("record_id_list") or []
    records: list[dict[str, Any]] = []
    for record_id, values in zip(record_ids, rows):
        records.append({"record_id": record_id, "fields": dict(zip(names, values))})
    return records, result


def field_value_matches_idea(value: Any, idea_id: str) -> bool:
    candidates = {idea_id}
    if idea_id.startswith("zy") and idea_id[2:].isdigit():
        candidates.add(idea_id[2:])
    if isinstance(value, list):
        return any(field_value_matches_idea(item, idea_id) for item in value)
    if isinstance(value, dict):
        return any(field_value_matches_idea(item, idea_id) for item in value.values())
    return str(value or "").strip() in candidates


def find_feishu_record_id(records: list[dict[str, Any]], idea_id: str) -> str:
    match_fields = ("idea_id", "项目ID", "题目ID", "任务ID", "ID", "UID", "uid")
    for record in records:
        record_fields = record.get("fields") or {}
        if any(field_value_matches_idea(record_fields.get(name), idea_id) for name in match_fields):
            return record.get("record_id") or ""
    return ""


def upsert_feishu_result_with_lark_cli(row: dict[str, str], trace: Trace, base_token: str, table_id: str) -> bool:
    fields_result = lark_cli_json(
        ["base", "+field-list", "--base-token", base_token, "--table-id", table_id, "--offset", "0", "--limit", "100"],
        trace,
    )
    field_data = fields_result.get("data", {})
    fields = field_data.get("items") or field_data.get("fields") or []
    if not fields:
        trace.write("feishu_lark_cli_failed", action="field-list", result=fields_result)
        return False

    payload, pending_attachments = build_feishu_payload(row, fields, None)
    if not payload and not pending_attachments:
        trace.write("feishu_lark_cli_skipped", reason="no_matching_fields", field_names=[feishu_field_name(f) for f in fields])
        return False

    records, records_result = lark_cli_records_by_field(base_token, table_id, trace)
    if not records and not records_result.get("ok"):
        trace.write("feishu_lark_cli_failed", action="record-list", result=records_result)
        return False

    record_id = find_feishu_record_id(records, row["idea_id"])
    args = ["base", "+record-upsert", "--base-token", base_token, "--table-id", table_id, "--json", json.dumps(payload, ensure_ascii=False)]
    action = "create"
    if record_id:
        args[4:4] = ["--record-id", record_id]
        action = "update"

    result = lark_cli_json(args, trace)
    if not result.get("ok"):
        trace.write("feishu_lark_cli_failed", action=action, result=result, fields=list(payload))
        return False

    record_id = record_id or result.get("data", {}).get("record", {}).get("record_id", "")
    uploaded: list[str] = []
    if record_id:
        for field_name, path in pending_attachments:
            upload_result = lark_cli_json(
                [
                    "base",
                    "+record-upload-attachment",
                    "--base-token",
                    base_token,
                    "--table-id",
                    table_id,
                    "--record-id",
                    record_id,
                    "--field-id",
                    field_name,
                    "--file",
                    str(path),
                    "--name",
                    path.name,
                ],
                trace,
            )
            if upload_result.get("ok"):
                uploaded.append(field_name)
            else:
                trace.write("feishu_lark_cli_attachment_failed", field=field_name, path=str(path), result=upload_result)

    trace.write("feishu_lark_cli_success", action=action, record_id=record_id, fields=list(payload), attachments=uploaded)
    return True


def upsert_feishu_result(row: dict[str, str], trace: Trace, base_token: str, table_id: str) -> None:
    try:
        client = FeishuBitableClient(DEFAULT_FEISHU_APP_ID, DEFAULT_FEISHU_APP_SECRET, base_token, table_id)
    except Exception as exc:
        trace.write("feishu_token_failed", error=str(exc))
        return

    fields, field_result = client.list_fields()
    if not fields:
        trace.write("feishu_permission_failed", action="list_fields", code=field_result.get("code"), msg=field_result.get("msg"))
        if str(field_result.get("code")) == "91403":
            upsert_feishu_result_with_lark_cli(row, trace, base_token, table_id)
        return

    payload, _pending_attachments = build_feishu_payload(row, fields, client)

    if not payload:
        trace.write("feishu_write_skipped", reason="no_matching_fields", field_names=[feishu_field_name(f) for f in fields])
        return

    records, records_result = client.list_records()
    if records_result.get("code") != 0:
        trace.write("feishu_permission_failed", action="list_records", code=records_result.get("code"), msg=records_result.get("msg"))
        if str(records_result.get("code")) == "91403":
            upsert_feishu_result_with_lark_cli(row, trace, base_token, table_id)
        return

    record_id = find_feishu_record_id(records, row["idea_id"])

    if record_id:
        result = client.update_record(record_id, payload)
        action = "update"
    else:
        result = client.create_record(payload)
        action = "create"

    if result.get("code") == 0:
        trace.write("feishu_write_success", action=action, record_id=record_id or result.get("data", {}).get("record", {}).get("record_id"), fields=list(payload))
    else:
        trace.write("feishu_write_failed", action=action, code=result.get("code"), msg=result.get("msg"), fields=list(payload))


def fallback_output_from_workspace(run_dir: Path, workspace: Path, idea_id: str, error: RuntimeError) -> str:
    prompt_path = run_dir / "prompt.txt"
    prompt = prompt_path.read_text(encoding="utf-8", errors="ignore").strip() if prompt_path.exists() else ""
    files: list[str] = []
    if workspace.exists():
        for path in sorted(p for p in workspace.rglob("*") if p.is_file()):
            if any(part in {".git", "node_modules", "__pycache__"} for part in path.parts):
                continue
            try:
                rel = path.relative_to(workspace)
            except ValueError:
                rel = path
            files.append(str(rel))
            if len(files) >= 80:
                break
    return "\n".join([
        f"# Trae Output Fallback - {idea_id}",
        "",
        "Trae 已出现完成标记，但自动复制最终回答失败，runner 已继续用工作区产物做后续质检。",
        "",
        f"- copy_error: {error}",
        f"- workspace: {workspace}",
        f"- prompt: {prompt}",
        "",
        "## Workspace Files",
        *(f"- {item}" for item in files),
        "",
    ])


def run_delivery_postprocess(trace: Trace, run_dir: Path, workspace: Path, idea_id: str, *, close_before_qa: bool) -> bool:
    if service_exception_detected_for_run(trace, run_dir, workspace, idea_id, check_ui=True):
        trace.write("delivery_skipped_service_exception", idea_id=idea_id)
        close_project_window_for_service_exception(trace, run_dir, workspace, idea_id)
        return False

    idea_data = read_json_file(run_dir / "idea.json", {})
    output_path = run_dir / "trae_copied_output.md"
    final_screenshot = result_screenshot_path(idea_id)
    thinking_path = None
    with trae_window_lock(trace, "delivery_capture"):
        raised = raise_trae_window_by_title(idea_id, trace) or raise_trae_window_by_title(workspace.name, trace)
        if raised:
            final_screenshot.parent.mkdir(parents=True, exist_ok=True)
            screenshot(final_screenshot)
            trace.write("final_screenshot_saved", path=str(final_screenshot))
            thinking_path = copy_thinking_process(trace, run_dir, workspace)
            if close_before_qa:
                closed = close_trae_window_by_title(idea_id, trace) or close_trae_window_by_title(workspace.name, trace)
                trace.write("trae_window_closed_before_qa", idea_id=idea_id, closed=closed)
        else:
            trace.write("delivery_capture_window_not_found", idea_id=idea_id, workspace=str(workspace))
    prompt_text = (run_dir / "prompt.txt").read_text(encoding="utf-8", errors="ignore") if (run_dir / "prompt.txt").exists() else ""
    business_domain = normalize_business_domain(idea_field(idea_data, "业务领域", "business_domain") or "Web前端")
    modification_scope = normalize_modification_scope(idea_field(idea_data, "修改范围", "modification_scope"))
    reuse_qa_after_publish = run_had_push_failure(run_dir)
    publish_result = {
        "status": "skipped",
        "commit": "",
        "remote": publish_web_url(idea_id, PUBLISH_FIRST_ROUND_BRANCH),
        "branch": publish_branch_folder(idea_id, PUBLISH_FIRST_ROUND_BRANCH),
    }
    trace.write("publish_skipped", reason="disabled", idea_id=idea_id, workspace=str(workspace))
    reused_qa = reuse_existing_codex_quality_check(trace, run_dir) if reuse_qa_after_publish else None
    if reused_qa:
        qa_report_path, qa_status = reused_qa
    else:
        qa_report_path, qa_status = run_codex_quality_check(trace, run_dir, workspace, idea_id)
    verdict = qa_verdict_from_report(qa_report_path, qa_status)
    qa_summary = ""
    if qa_report_path and qa_report_path.exists():
        qa_summary = first_line(qa_report_path.read_text(encoding="utf-8", errors="ignore"), 500)

    row = {
        "idea_id": idea_id,
        "title": str(idea_data.get("title", "")),
        "status": verdict["discard_status"],
        "workspace": str(workspace),
        "run_dir": str(run_dir),
        "prompt_path": str(run_dir / "prompt.txt"),
        "output_path": str(output_path) if output_path.exists() else "",
        "session_id": "",
        "thinking_path": str(thinking_path) if thinking_path else "",
        "screenshot_path": str(final_screenshot),
        "qa_status": qa_status,
        "qa_report_path": str(qa_report_path) if qa_report_path else "",
        "qa_summary": qa_summary,
        "git_status": publish_result["status"],
        "git_commit": publish_result["commit"],
        "git_remote": publish_result["remote"],
        "git_branch": publish_result["branch"],
        "trace_path": str(run_dir / "trace.jsonl"),
        "trajectory_path": str(run_dir / "trae_trajectory.md"),
        "full_transcript_path": str(run_dir / "trae_full_transcript.md"),
        "日志轨迹": "",
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "轮次": "第一轮",
        "提示词": prompt_text.strip(),
        "任务类型": verdict.get("task_type") or DEFAULT_TASK_TYPE,
        "业务领域": business_domain,
        "修改范围": modification_scope,
        "任务是否完成": verdict["task_completed"],
        "产物及过程是否满意": verdict["satisfied"],
        "不满意原因": verdict["reason"],
        "远端Github地址": publish_result["remote"],
        "分支文件夹": publish_result["branch"],
        "截图": str(final_screenshot),
    }
    upsert_result_table(row, trace)
    if verdict["discard_status"] == "not_discarded":
        workbook_path = write_solo_qc_workbook(row, trace)
        row["workbook"] = str(workbook_path)
        mark_not_discarded_project(row, trace)
    else:
        clear_not_discarded_project(idea_id, trace)
        trace.write("project_discarded_after_qc", idea_id=idea_id, status="作废了", qa_report_path=str(qa_report_path) if qa_report_path else "")
    return True


def run_finish_steps(
    *,
    trace: Trace,
    run_dir: Path,
    workspace: Path,
    idea_id: str,
    timeout_minutes: int,
    poll_seconds: int,
    keep_trae_open: bool,
    update_state: bool,
) -> int:
    output_path = run_dir / "trae_copied_output.md"
    if output_path.exists():
        trace.write("existing_output_found", path=str(output_path), action="skip_completion_wait")
    elif (run_dir / "completion_adopted.json").exists():
        trace.write("completion_adopted_marker_found", path=str(run_dir / "completion_adopted.json"), action="skip_completion_wait")
    else:
        ok = wait_for_completion(trace, timeout_minutes * 60, poll_seconds, run_dir, workspace)
        if not ok:
            abort_marker = run_dir / "aborted.json"
            reason = "timeout"
            if abort_marker.exists():
                data = read_json_file(abort_marker, {})
                reason = str(data.get("reason") or "aborted") if isinstance(data, dict) else "aborted"
            trace.write("stop_without_copy", reason=reason)
            if reason == "server_exception":
                return RUNNER_SERVICE_EXCEPTION_EXIT_CODE
            if reason in {"manual_stop_output", "stopped_or_interrupted", "trae_not_running"}:
                return RUNNER_STOPPED_EXIT_CODE
            return 2

    if service_exception_detected_for_run(trace, run_dir, workspace, idea_id, check_ui=True):
        trace.write("stop_before_final_copy", reason="server_exception")
        close_project_window_for_service_exception(trace, run_dir, workspace, idea_id)
        return RUNNER_SERVICE_EXCEPTION_EXIT_CODE

    stopped_detected, stopped_reason = stopped_run_detected_for_run(trace, run_dir, workspace, idea_id, check_ui=True)
    if stopped_detected:
        reason = stopped_reason or "stopped_or_interrupted"
        detail = (
            "Trae 已停止运行，runner 已放弃当前项目并释放调度位。"
            if reason == "trae_not_running"
            else "Trae 显示手动终止输出或会话已中断，runner 已关闭对应项目窗口并释放调度位。"
        )
        trace.write("stop_before_final_copy", reason=reason)
        close_project_window_for_stopped_run(trace, run_dir, workspace, idea_id, reason, detail)
        return RUNNER_STOPPED_EXIT_CODE

    with trae_window_lock(trace, "final_copy"):
        try:
            output = copy_final_output(trace, run_dir, workspace, idea_id)
            output_path.write_text(output, encoding="utf-8")
            trace.write("output_saved", path=str(output_path))
        except RuntimeError as exc:
            if not output_path.exists():
                fallback = fallback_output_from_workspace(run_dir, workspace, idea_id, exc)
                output_path.write_text(fallback, encoding="utf-8")
                trace.write("copy_failed_using_workspace_fallback", path=str(output_path), error=str(exc))
            else:
                trace.write("copy_failed_using_existing_output", path=str(output_path), error=str(exc))

    extract_trae_trajectory(run_dir, workspace, trace)
    delivered = run_delivery_postprocess(trace, run_dir, workspace, idea_id, close_before_qa=not keep_trae_open)
    if not delivered:
        return RUNNER_SERVICE_EXCEPTION_EXIT_CODE

    if update_state:
        mark_idea_used(idea_id, run_dir, trace)

    if not keep_trae_open:
        trace.write("trae_window_close_policy", policy="target_window_closed_before_qa")

    print(str(output_path))
    return 0


def run_main(args: argparse.Namespace) -> int:
    if args.finish_only:
        run_dir = resolve_finish_run_dir(args.idea_id, args.run_dir)
        idea_id = args.idea_id or infer_idea_id_from_run_dir(run_dir)
        if not idea_id:
            raise SystemExit(f"无法从 run 目录推断 idea_id，请补 --idea-id: {run_dir}")
        workspace = workspace_from_run_dir(run_dir, Path(args.workspace_root) / idea_id)
        trace = Trace(run_dir)
        write_active_runner(idea_id, run_dir, workspace, mode="finish_only")
        trace.write("finish_only_start", idea_id=idea_id, run_dir=str(run_dir), workspace=str(workspace))
        with trae_window_lock(trace, "finish_open_workspace"):
            open_trae_workspace_for_finish(trace, workspace)
        return run_finish_steps(
            trace=trace,
            run_dir=run_dir,
            workspace=workspace,
            idea_id=idea_id,
            timeout_minutes=args.timeout_minutes,
            poll_seconds=args.poll_seconds,
            keep_trae_open=args.keep_trae_open,
            update_state=True,
        )

    ideas = parse_ideas(IDEA_HISTORY)
    idea = choose_idea(ideas, args.idea_id, args.order)
    prompt, prompt_source = prompt_from_idea(idea)

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    run_dir = RUNS_DIR / f"{timestamp}-{idea.idea_id}"
    run_dir.mkdir(parents=True, exist_ok=True)
    trace = Trace(run_dir)

    workspace = Path(args.workspace_root) / idea.idea_id
    workspace.mkdir(parents=True, exist_ok=True)
    write_active_runner(idea.idea_id, run_dir, workspace, mode="run")
    (run_dir / "prompt.txt").write_text(prompt, encoding="utf-8")
    (run_dir / "idea.json").write_text(json.dumps({
        "idea_id": idea.idea_id,
        "title": idea.title,
        "fields": idea.fields,
        "prompt_source": prompt_source,
        "workspace": str(workspace),
    }, ensure_ascii=False, indent=2), encoding="utf-8")

    trace.write("selected_idea", idea_id=idea.idea_id, title=idea.title, prompt_source=prompt_source)
    trace.write("workspace_ready", workspace=str(workspace))

    if args.dry_run:
        (run_dir / ".dry_run").write_text(datetime.now().isoformat(timespec="seconds") + "\n", encoding="utf-8")
        trace.write("dry_run_done", run_dir=str(run_dir), prompt_chars=len(prompt))
        print(json.dumps({
            "idea_id": idea.idea_id,
            "title": idea.title,
            "prompt_source": prompt_source,
            "workspace": str(workspace),
            "run_dir": str(run_dir),
            "prompt_path": str(run_dir / "prompt.txt"),
        }, ensure_ascii=False, indent=2))
        return 0

    with trae_window_lock(trace, "prompt_send"):
        proc = start_trae_chat(trace, workspace, prompt, use_cli=not args.no_cli)
        time.sleep(4.0)
        if args.no_cli:
            send_shortcut_sequence(trace, prompt, enabled=not args.no_shortcuts)
    trace.write("prompt_handoff_ready", idea_id=idea.idea_id, run_dir=str(run_dir), workspace=str(workspace))

    if args.start_only:
        trace.write("start_only_done", idea_id=idea.idea_id, run_dir=str(run_dir), workspace=str(workspace))
        print(str(run_dir))
        return 0

    status = run_finish_steps(
        trace=trace,
        run_dir=run_dir,
        workspace=workspace,
        idea_id=idea.idea_id,
        timeout_minutes=args.timeout_minutes,
        poll_seconds=args.poll_seconds,
        keep_trae_open=args.keep_trae_open,
        update_state=True,
    )

    if proc and proc.poll() is None:
        trace.write("cli_process_still_running", pid=proc.pid)

    return status


def main() -> int:
    parser = argparse.ArgumentParser(description="Run one idea from idea_history.md in Trae CN.")
    parser.add_argument("--idea-id", help="指定点子 ID；不传则取未运行的一条。")
    parser.add_argument("--order", choices=["newest", "oldest"], default="newest", help="自动取点子的顺序。默认 newest。")
    parser.add_argument("--timeout-minutes", type=int, default=90, help="等待任务完成的最长时间。")
    parser.add_argument("--poll-seconds", type=int, default=30, help="轮询完成状态的间隔。")
    parser.add_argument("--no-cli", action="store_true", help="不用 trae-cn chat，只打开 Trae 窗口后走快捷键粘贴发送。")
    parser.add_argument("--no-shortcuts", action="store_true", help="不执行快捷键粘贴/发送。通常只在 CLI 已能直接发送时使用。")
    parser.add_argument("--keep-trae-open", action="store_true", help="完成后不关闭 Trae。")
    parser.add_argument("--workspace-root", default=str(ROOT / "workspaces"), help="每条点子的工作目录根路径。")
    parser.add_argument("--run-dir", help="指定已有 run 目录，常用于 --finish-only 重跑收尾。")
    parser.add_argument("--finish-only", action="store_true", help="不启动新任务，只对已有 Trae 窗口/run 目录执行收尾验证。")
    parser.add_argument("--start-only", action="store_true", help="只打开工作区并发送提示词，不等待完成、不复制结果、不做收尾。")
    parser.add_argument("--lock-timeout", type=int, default=0, help="等待 runner 全局锁的秒数；默认 0 表示已有运行就立即退出。")
    parser.add_argument("--allow-parallel", action="store_true", help="允许多个 runner 并行；由 controller 负责并发上限和启动节流。")
    parser.add_argument("--dry-run", action="store_true", help="只解析点子和生成提示词，不启动 Trae。")
    args = parser.parse_args()

    try:
        if args.allow_parallel:
            return run_main(args)
        with RunnerLock(RUNNER_LOCK_PATH, args.lock_timeout, LEGACY_RUNNER_LOCK_PATHS):
            return run_main(args)
    finally:
        clear_active_runner()


if __name__ == "__main__":
    raise SystemExit(main())
