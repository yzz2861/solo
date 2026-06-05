#!/usr/bin/env python3
"""Helpers for the Solo Coder workflow on this machine."""

from __future__ import annotations

import argparse
import base64
import datetime as dt
import json
import os
import re
import shlex
import shutil
import subprocess
import sys
import tempfile
import textwrap
import time
import unicodedata
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Iterable
from urllib import error as urlerror
from urllib import parse as urlparse
from urllib import request as urlrequest

from openpyxl import load_workbook


PROJECT_ROOT = Path("~/pro/solocoder").expanduser()
DEFAULT_IDEAS_ROOT = PROJECT_ROOT / "ideas"
DEFAULT_START_ROOT = PROJECT_ROOT / "pro"
DEFAULT_REPO_URL = "git@github.com:yzz2861/trae_eval.git"
DEFAULT_TEMPLATE_WORKBOOK = PROJECT_ROOT / "【中强智联】Solo Coder_数据表_表格.xlsx"
SKIP_GIT_PUSH = os.environ.get("SOLO_SKIP_GIT_INTERACTION", os.environ.get("SOLO_SKIP_GIT_PUSH", "1")) != "0"
TRAE_APP = "Trae CN"
TRAE_CLI = Path("/Applications/Trae CN.app/Contents/Resources/app/bin/trae-cn")
TRAE_LOG_ROOT = Path.home() / "Library/Application Support/Trae CN/logs"
REQUIRED_MODEL = "Seed-Code-DogFooding"
REQUIRED_AGENT_ID = "solo_agent"
REQUIRED_AGENT_TYPE = "solo_agent"
REQUIRED_AGENT_NAME = "SOLO Agent"
BUSINESS_DOMAINS = [
    "纯后端API服务",
    "Web前端",
    "全栈Web应用",
    "游戏开发",
    "数据分析与可视化",
    "3D/交互可视化",
    "AI/ML应用",
    "科学计算",
    "命令行工具",
    "桌面应用(含GUI）",
    "自动化与工具脚本",
]
TASK_TYPES = [
    "Bug修复",
    "0-1代码生成",
    "Feature迭代",
    "代码理解",
    "代码重构",
    "工程化",
    "代码测试",
]
MODIFICATION_SCOPES = [
    "单文件",
    "模块内多文件",
    "跨模块多文件",
    "跨系统多模块",
]
TASK_COMPLETION_STATUSES = [
    "未完成任务",
    "完成了任务",
]
SATISFACTION_STATUSES = [
    "满意",
    "不满意",
]
PRODUCT_REASON_TEMPLATES = [
    "只解决了部分需求",
    "根因定位错误",
    "违反用户约束",
    "引入新问题或回归",
]
PROCESS_REASON_TEMPLATES = [
    "缺少可验证证据",
    "缺少自查与验证",
    "过程规划混乱",
    "重复让用户兜底",
]

SESSION_PATTERNS = [
    re.compile(r'"session_id":"(?P<session_id>[^"]+)"(?:,"message_id":"(?P<message_id>[^"]+)")?'),
    re.compile(r'"sessionId":"(?P<session_id>[^"]+)"'),
    re.compile(r"sessionId:\s*(?P<session_id>[a-zA-Z0-9-]+)"),
]
TRACE_PATTERN = re.compile(r'"traceId":"(?P<trace_id>[^"]+)"')
CHAT_MODEL_PATTERN = re.compile(r'"chat_model":"(?P<chat_model>[^"]+)"')
AGENT_ID_PATTERN = re.compile(r'"agent_id":"(?P<agent_id>[^"]+)"')
AGENT_TYPE_PATTERN = re.compile(r'"agent_type":"(?P<agent_type>[^"]+)"')
AGENT_NAME_PATTERN = re.compile(r'"agent_name":"(?P<agent_name>[^"]+)"')
ERROR_PATTERNS = [
    re.compile(r"command 'workbench\.action\.chat\.open' not found"),
    re.compile(r"\[error\]\s+(?P<error>.+)"),
]
ROUND_ORDER = {
    "第一轮": 1,
    "第二轮": 2,
    "第三轮": 3,
    "第四轮": 4,
    "第五轮": 5,
}
ROUND_NAMES = list(ROUND_ORDER.keys())


@dataclass
class SessionEvent:
    session_id: str | None = None
    message_id: str | None = None
    trace_id: str | None = None
    connect_session_id: str | None = None
    chat_model: str | None = None
    agent_id: str | None = None
    agent_type: str | None = None
    agent_name: str | None = None
    error: str | None = None
    log_file: str | None = None
    line_no: int | None = None
    line_text: str | None = None
    timestamp: str | None = None
    profile_match: bool = False


def parse_timestamp(line: str) -> str | None:
    match = re.match(r"(?P<ts>\d{4}-\d{2}-\d{2}T[0-9:.+-]+)", line)
    return match.group("ts") if match else None


def maybe_asdict(event: SessionEvent | None) -> dict[str, object] | None:
    return asdict(event) if event else None


def run_command(
    cmd: list[str],
    cwd: Path | None = None,
    timeout: int | None = None,
    input_text: str | None = None,
) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        cmd,
        cwd=str(cwd) if cwd else None,
        capture_output=True,
        text=True,
        input=input_text,
        timeout=timeout,
        check=False,
    )


def slugify(text: str, fallback: str = "idea") -> str:
    normalized = unicodedata.normalize("NFKD", text)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", ascii_text).strip("-").lower()
    slug = re.sub(r"-{2,}", "-", slug)
    return slug[:48] if slug else fallback


def extract_short_title(text: str) -> str:
    compact = re.sub(r"\s+", " ", text).strip()
    return compact[:60] if compact else "solo-coder-idea"


def infer_task_type(idea: str) -> str:
    lowered = idea.lower()
    if any(token in lowered for token in ["bug", "报错", "修复", "排查", "异常"]):
        return "Bug修复"
    if any(token in lowered for token in ["重构", "优化结构", "解耦"]):
        return "代码重构"
    if any(token in lowered for token in ["测试", "单测", "集成测试", "test"]):
        return "代码测试"
    if any(token in lowered for token in ["配置", "部署", "ci", "cd", "脚手架", "工程化"]):
        return "工程化"
    if any(token in lowered for token in ["理解", "解释代码", "阅读代码", "分析代码"]):
        return "代码理解"
    if any(token in lowered for token in ["迭代", "新增功能", "扩展", "升级"]):
        return "Feature迭代"
    return "0-1代码生成"


def infer_business_domain(idea: str) -> str:
    lowered = idea.lower()
    if any(token in lowered for token in ["ai", "ml", "llm", "agent", "rag", "embedding", "模型"]):
        return "AI/ML应用"
    if any(token in lowered for token in ["3d", "three", "webgl", "交互可视化"]):
        return "3D/交互可视化"
    if any(token in lowered for token in ["数据分析", "图表", "可视化", "dashboard", "streamlit"]):
        return "数据分析与可视化"
    if any(token in lowered for token in ["api", "接口", "后端", "backend", "server"]):
        return "纯后端API服务"
    if any(token in lowered for token in ["桌面", "gui", "electron", "tauri"]):
        return "桌面应用(含GUI）"
    if any(token in lowered for token in ["脚本", "自动化", "crawler", "爬虫", "workflow"]):
        return "自动化与工具脚本"
    if any(token in lowered for token in ["命令行", "cli", "terminal"]):
        return "命令行工具"
    if any(token in lowered for token in ["全栈", "fullstack", "前后端", "登录", "数据库"]):
        return "全栈Web应用"
    return "Web前端"


def infer_modification_scope(task_type: str) -> str:
    if task_type == "0-1代码生成":
        return "跨模块多文件"
    if task_type in {"Feature迭代", "工程化"}:
        return "跨模块多文件"
    if task_type == "Bug修复":
        return "模块内多文件"
    return "单文件"


def round_index(round_name: str) -> int:
    return ROUND_ORDER.get(round_name, 1)


def build_branch_name(idea_id: str, round_name: str) -> str:
    return f"idea/{idea_id}/r{round_index(round_name)}"


def build_base_branch_name(idea_id: str) -> str:
    return f"idea/{idea_id}/base"


def build_pr_title(idea_id: str, round_name: str, title: str) -> str:
    return f"[{idea_id}][{round_name}] {title[:50]}"


def build_folder_name(idea_id: str, idea: str) -> str:
    return f"{idea_id}-{slugify(extract_short_title(idea), fallback='solo-coder-idea')}"


def build_optimized_prompt(idea_id: str, round_name: str, folder_name: str, idea: str) -> str:
    return textwrap.dedent(
        f"""\
        你现在在一个真实开发任务里，请直接产出可运行代码，而不是只给方案。

        项目编号：{idea_id}
        本轮：{round_name}
        模型要求：必须使用 {REQUIRED_MODEL}
        智能体要求：必须使用 {REQUIRED_AGENT_NAME}
        代码目录要求：在仓库根目录创建并只修改 `{folder_name}/`

        原始想法：
        {idea}

        开发目标：
        - 将上面的想法落成一个不简单、可运行、可演示的项目原型
        - 所有代码集中在 `{folder_name}/`
        - 如果仓库当前为空，请自行初始化最小但完整的项目结构
        - 优先做成具备明显功能闭环的版本，不要只搭空壳

        工作方式要求：
        - 先快速审视当前仓库是否为空、已有文件结构和运行方式
        - 再给出极简执行计划，然后立即开始写代码
        - 不要把大量时间花在空泛分析或模板化总结上
        - 不要偏离需求去做无关的大重构
        - 如需新增依赖，只增加必要依赖，并说明原因

        交付要求：
        - 必须给出完整代码
        - 必须给出运行步骤
        - 必须说明关键文件和实现点
        - 如有未完成项，明确列出

        验收标准：
        - 项目可以启动或运行
        - 页面或功能不是纯占位
        - 至少包含 3 个以上互相关联的功能点或模块
        - 代码结构清晰，文件划分合理
        - 不要只生成单文件 demo

        输出方式：
        - 先简短说明你将创建什么
        - 然后直接开始改文件
        - 最后用简短清单说明运行命令、关键文件、已完成项、未完成项
        """
    ).strip() + "\n"


def build_pr_template(
    idea_id: str,
    round_name: str,
    prompt: str,
    task_type: str,
    business_domain: str,
    modification_scope: str,
) -> str:
    return textwrap.dedent(
        f"""\
        ## Summary

        - Idea ID: `{idea_id}`
        - Round: `{round_name}`
        - Task Type: `{task_type}`
        - Business Domain: `{business_domain}`
        - Modification Scope: `{modification_scope}`

        ## User Prompt

        ```text
        {prompt.strip()}
        ```

        ## Trae Session

        - Session ID: `<to-fill>`
        - Model: `{REQUIRED_MODEL}`
        - Agent: `{REQUIRED_AGENT_NAME}`

        ## Review

        - Task Completed: `<to-fill>`
        - Satisfied: `<to-fill>`
        - Dissatisfaction Notes: `<to-fill>`
        """
    )


def detect_remote_repo_state(repo_url: str) -> dict[str, object]:
    refs = run_command(["git", "ls-remote", repo_url], timeout=20)
    has_refs = bool(refs.stdout.strip())
    symref = run_command(["git", "ls-remote", "--symref", repo_url, "HEAD"], timeout=20)
    default_branch = None
    for line in symref.stdout.splitlines():
        match = re.match(r"ref:\s+refs/heads/(?P<branch>\S+)\s+HEAD", line)
        if match:
            default_branch = match.group("branch")
            break
    return {
        "repo_url": repo_url,
        "has_refs": has_refs,
        "default_branch": default_branch,
        "ls_remote_returncode": refs.returncode,
        "ls_remote_stderr": refs.stderr.strip(),
    }


def log_dirs() -> list[Path]:
    if not TRAE_LOG_ROOT.exists():
        return []
    return sorted(path for path in TRAE_LOG_ROOT.iterdir() if path.name[:2] == "20" and path.is_dir())


def renderer_logs() -> list[Path]:
    files: list[Path] = []
    for root in log_dirs():
        files.extend(sorted(root.glob("window*/renderer.log")))
    return files


def baseline_for_logs(files: Iterable[Path]) -> dict[Path, int]:
    return {path: sum(1 for _ in path.open("r", encoding="utf-8", errors="replace")) for path in files if path.exists()}


def iter_new_lines(path: Path, start_line: int) -> Iterable[tuple[int, str]]:
    if not path.exists():
        return
    with path.open("r", encoding="utf-8", errors="replace") as handle:
        for idx, line in enumerate(handle, start=1):
            if idx > start_line:
                yield idx, line.rstrip("\n")


def extract_event_from_line(path: Path, line_no: int, line: str) -> SessionEvent | None:
    event = SessionEvent(log_file=str(path), line_no=line_no, line_text=line, timestamp=parse_timestamp(line))

    connect_match = re.search(r"ConnectSessionID.*sessionId:\s*(?P<session_id>[a-zA-Z0-9-]+)", line)
    if connect_match:
        event.connect_session_id = connect_match.group("session_id")

    for pattern in SESSION_PATTERNS:
        match = pattern.search(line)
        if match:
            event.session_id = match.groupdict().get("session_id")
            event.message_id = match.groupdict().get("message_id")
            break

    trace_match = TRACE_PATTERN.search(line)
    if trace_match:
        event.trace_id = trace_match.group("trace_id")

    chat_model = CHAT_MODEL_PATTERN.search(line)
    if chat_model:
        event.chat_model = chat_model.group("chat_model")

    agent_id = AGENT_ID_PATTERN.search(line)
    if agent_id:
        event.agent_id = agent_id.group("agent_id")

    agent_type = AGENT_TYPE_PATTERN.search(line)
    if agent_type:
        event.agent_type = agent_type.group("agent_type")

    agent_name = AGENT_NAME_PATTERN.search(line)
    if agent_name:
        event.agent_name = agent_name.group("agent_name")

    for pattern in ERROR_PATTERNS:
        match = pattern.search(line)
        if match:
            event.error = match.groupdict().get("error") or match.group(0)
            break

    event.profile_match = (
        event.chat_model == REQUIRED_MODEL and
        event.agent_id == REQUIRED_AGENT_ID and
        event.agent_type == REQUIRED_AGENT_TYPE and
        event.agent_name == REQUIRED_AGENT_NAME
    )

    if any([
        event.session_id,
        event.message_id,
        event.trace_id,
        event.connect_session_id,
        event.chat_model,
        event.agent_id,
        event.error,
    ]):
        return event
    return None


def scan_events_since(baseline: dict[Path, int]) -> list[SessionEvent]:
    current_files = renderer_logs()
    events: list[SessionEvent] = []
    for path in current_files:
        start_line = baseline.get(path, 0)
        for line_no, line in iter_new_lines(path, start_line):
            event = extract_event_from_line(path, line_no, line)
            if event:
                events.append(event)
    return events


def event_priority(event: SessionEvent) -> tuple[int, str, int]:
    score = 0
    if event.profile_match:
        score += 10
    if event.session_id:
        score += 4
    if event.message_id:
        score += 2
    if event.trace_id:
        score += 1
    if event.error:
        score -= 3
    return score, event.timestamp or "", event.line_no or 0


def choose_best_event(events: list[SessionEvent]) -> SessionEvent | None:
    if not events:
        return None
    return sorted(events, key=event_priority, reverse=True)[0]


def choose_best_session_event(events: list[SessionEvent]) -> SessionEvent | None:
    matching = [event for event in events if event.session_id and event.profile_match]
    if matching:
        return choose_best_event(matching)
    fallback = [event for event in events if event.session_id]
    if fallback:
        return choose_best_event(fallback)
    return None


def choose_best_error_event(events: list[SessionEvent]) -> SessionEvent | None:
    candidates = [event for event in events if event.error]
    if not candidates:
        return None
    preferred = [event for event in candidates if event.error and "workbench.action.chat.open" in event.error]
    return sorted(preferred or candidates, key=event_priority, reverse=True)[0]


def parse_since(value: str) -> dt.datetime:
    if value.isdigit():
        return dt.datetime.fromtimestamp(int(value), tz=dt.timezone.utc).astimezone()
    parsed = dt.datetime.fromisoformat(value)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=dt.datetime.now().astimezone().tzinfo)
    return parsed


def extract_recent_events(since: dt.datetime) -> list[SessionEvent]:
    events: list[SessionEvent] = []
    for path in renderer_logs():
        with path.open("r", encoding="utf-8", errors="replace") as handle:
            for idx, line in enumerate(handle, start=1):
                ts = parse_timestamp(line)
                if not ts:
                    continue
                try:
                    line_dt = dt.datetime.fromisoformat(ts)
                except ValueError:
                    continue
                if line_dt < since:
                    continue
                event = extract_event_from_line(path, idx, line.rstrip("\n"))
                if event:
                    events.append(event)
    return events


def wait_for_session_event(since: dt.datetime, timeout: int) -> tuple[list[SessionEvent], SessionEvent | None]:
    deadline = time.time() + timeout
    latest_events: list[SessionEvent] = []
    best_event: SessionEvent | None = None
    while time.time() < deadline:
        latest_events = extract_recent_events(since)
        best_event = choose_best_session_event(latest_events)
        if best_event and best_event.profile_match:
            break
        time.sleep(1)
    return latest_events, best_event


def load_sheet(path: Path):
    workbook = load_workbook(path)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    header_map = {str(value): idx for idx, value in enumerate(headers, start=1) if value}
    return workbook, sheet, header_map


def find_row(sheet, header_map: dict[str, int], uid: str | None, round_name: str | None, row_number: int | None) -> int:
    if row_number:
        return row_number

    uid_col = header_map.get("UID")
    round_col = header_map.get("轮次")
    session_col = header_map.get("Trae Session ID")

    for row_idx in range(2, sheet.max_row + 1):
        uid_match = True if uid is None else str(sheet.cell(row_idx, uid_col).value) == str(uid)
        round_match = True if round_name is None else str(sheet.cell(row_idx, round_col).value) == round_name
        session_empty = session_col is None or sheet.cell(row_idx, session_col).value in (None, "")
        if uid_match and round_match and session_empty:
            return row_idx

    for row_idx in range(2, sheet.max_row + 1):
        uid_match = True if uid is None else str(sheet.cell(row_idx, uid_col).value) == str(uid)
        round_match = True if round_name is None else str(sheet.cell(row_idx, round_col).value) == round_name
        if uid_match and round_match:
            return row_idx

    raise SystemExit("No matching row found. Pass --row if you want to override row selection.")


def update_workbook(workbook_path: Path, uid: str | None, round_name: str | None, row_number: int | None, values: dict[str, str]) -> dict[str, object]:
    workbook, sheet, header_map = load_sheet(workbook_path)
    row_idx = find_row(sheet, header_map, uid, round_name, row_number)

    for field, value in values.items():
        if field not in header_map:
            raise SystemExit(f"Unknown column: {field}")
        sheet.cell(row_idx, header_map[field]).value = value

    workbook.save(workbook_path)
    workbook.close()
    return {"workbook": str(workbook_path), "row": row_idx, "updated": values}


def update_workbook_existing_fields(
    workbook_path: Path,
    uid: str | None,
    round_name: str | None,
    row_number: int | None,
    values: dict[str, str],
) -> dict[str, object]:
    workbook, sheet, header_map = load_sheet(workbook_path)
    row_idx = find_row(sheet, header_map, uid, round_name, row_number)

    applied: dict[str, str] = {}
    skipped: dict[str, str] = {}
    for field, value in values.items():
        col_idx = header_map.get(field)
        if not col_idx:
            skipped[field] = value
            continue
        sheet.cell(row_idx, col_idx).value = value
        applied[field] = value

    workbook.save(workbook_path)
    workbook.close()
    return {
        "workbook": str(workbook_path),
        "row": row_idx,
        "updated": applied,
        "skipped_missing_fields": skipped,
    }


def build_first_round_session_fields(session_id: str) -> dict[str, str]:
    return {
        "第一轮SessionID": session_id,
        "第一轮Session ID": session_id,
    }


def all_renderer_logs() -> list[Path]:
    files: list[Path] = []
    if not TRAE_LOG_ROOT.exists():
        return files
    for root in log_dirs():
        files.extend(root.glob("window*/renderer*.log"))
    return sorted(
        [path for path in files if path.is_file()],
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )


def extract_session_process_trace(session_id: str, max_lines: int = 80) -> dict[str, object]:
    if not session_id:
        return {"session_id": session_id, "trace_lines": [], "source_files": []}

    matched_lines: list[str] = []
    source_files: list[str] = []
    patterns = [
        f'"session_id":"{session_id}"',
        f'"sessionId":"{session_id}"',
        f'sessionId":"{session_id}"',
        f"sessionId: {session_id}",
        f'"session_id":"{session_id}"',
        f'"session_id": "{session_id}"',
    ]

    for path in all_renderer_logs():
        try:
            with path.open("r", encoding="utf-8", errors="replace") as handle:
                for raw_line in handle:
                    line = raw_line.rstrip("\n")
                    if any(pattern in line for pattern in patterns):
                        matched_lines.append(line)
                        if str(path) not in source_files:
                            source_files.append(str(path))
                        if len(matched_lines) >= max_lines:
                            break
        except OSError:
            continue
        if len(matched_lines) >= max_lines:
            break

    return {
        "session_id": session_id,
        "trace_lines": matched_lines,
        "source_files": source_files,
    }


def infer_satisfaction_from_suggestion(text: str) -> str | None:
    match = re.search(r"建议满意度[:：]\s*(满意|不满意|无法判断)", text)
    if not match:
        return None
    value = match.group(1)
    return value if value in SATISFACTION_STATUSES else None


def build_codex_review_prompt(session_id: str, trace_lines: list[str]) -> str:
    trace_block = "\n".join(trace_lines)
    return textwrap.dedent(
        f"""\
        你在审查一个 Trae 会话的本地过程轨迹，请仅根据这些轨迹判断这轮“产物及过程是否满意”是否存在明显风险。

        规则：
        - 只能基于给定轨迹，不要臆测未提供的信息。
        - 如果轨迹不足以判断，就写“无法判断”。
        - 重点看过程是否混乱、是否缺少验证、是否只在做机械操作、是否暴露出明显返工风险。
        - 输出必须简洁，严格使用以下格式：

        建议满意度：满意/不满意/无法判断
        判断依据：
        - ...
        - ...
        建议：
        - ...
        - ...

        Session ID: {session_id}

        轨迹：
        {trace_block}
        """
    ).strip() + "\n"


def review_session_trace_with_codex(session_id: str) -> dict[str, object]:
    trace_payload = extract_session_process_trace(session_id)
    trace_lines = [str(line) for line in trace_payload.get("trace_lines") or []]
    if not trace_lines:
        return {
            "session_id": session_id,
            "ok": False,
            "trace_found": False,
            "message": "未在本地 Trae renderer 日志中找到这个 session 的过程轨迹，已跳过自动建议。",
            "trace_payload": trace_payload,
        }

    prompt = build_codex_review_prompt(session_id, trace_lines)
    with tempfile.NamedTemporaryFile("w", encoding="utf-8", delete=False, suffix=".txt") as output_file:
        output_path = Path(output_file.name)

    try:
        try:
            result = run_command(
                [
                    "codex",
                    "exec",
                    "-m",
                    "gpt-5.4-mini",
                    "--skip-git-repo-check",
                    "--color",
                    "never",
                    "-C",
                    str(PROJECT_ROOT),
                    "-o",
                    str(output_path),
                    "-",
                ],
                cwd=PROJECT_ROOT,
                timeout=180,
                input_text=prompt,
            )
        except subprocess.TimeoutExpired:
            return {
                "session_id": session_id,
                "ok": False,
                "trace_found": True,
                "message": "Codex CLI 检查超时，已跳过自动建议，你可以继续手动判断是否满意。",
                "trace_payload": trace_payload,
                "suggested_satisfaction": None,
            }
        except OSError as exc:
            return {
                "session_id": session_id,
                "ok": False,
                "trace_found": True,
                "message": f"Codex CLI 调用失败：{exc}",
                "trace_payload": trace_payload,
                "suggested_satisfaction": None,
            }
        output_text = output_path.read_text(encoding="utf-8").strip() if output_path.exists() else ""
    finally:
        output_path.unlink(missing_ok=True)

    ok = result.returncode == 0 and bool(output_text)
    message = output_text if output_text else (result.stderr.strip() or result.stdout.strip() or "Codex 未返回可用建议。")
    return {
        "session_id": session_id,
        "ok": ok,
        "trace_found": True,
        "message": message,
        "suggested_satisfaction": infer_satisfaction_from_suggestion(message),
        "trace_payload": trace_payload,
        "codex_returncode": result.returncode,
    }


def seed_project_workbook(template_path: Path, output_path: Path, project_id: str) -> dict[str, object]:
    workbook = load_workbook(template_path)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    header_map = {str(value): idx for idx, value in enumerate(headers, start=1) if value}

    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)

    for offset, round_name in enumerate(ROUND_NAMES, start=2):
        row_values = {
            "UID": project_id,
            "轮次": round_name,
            "状态": "待提交",
            "提交任务": "提交任务",
        }
        for field, value in row_values.items():
            col_idx = header_map.get(field)
            if col_idx:
                sheet.cell(offset, col_idx).value = value

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()
    return {
        "template": str(template_path),
        "workbook": str(output_path),
        "project_id": project_id,
        "rounds_seeded": ROUND_NAMES,
    }


def cmd_list_rows(args: argparse.Namespace) -> int:
    workbook, sheet, header_map = load_sheet(Path(args.workbook))
    fields = ["UID", "轮次", "Trae Session ID", "状态", "User Prompt"]
    rows: list[dict[str, object]] = []
    for row_idx in range(2, sheet.max_row + 1):
        row_data = {"row": row_idx}
        for field in fields:
            col = header_map.get(field)
            row_data[field] = sheet.cell(row_idx, col).value if col else None
        rows.append(row_data)
    print(json.dumps(rows, ensure_ascii=False, indent=2, default=str))
    workbook.close()
    return 0


def cmd_update_row(args: argparse.Namespace) -> int:
    workbook_path = Path(args.workbook)
    values: dict[str, str] = {}
    for item in args.set_values:
        if "=" not in item:
            raise SystemExit(f"Invalid --set value: {item!r}. Expected Field=Value.")
        field, value = item.split("=", 1)
        values[field.strip()] = value
    result = update_workbook(workbook_path, args.uid, args.round_name, args.row, values)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


def clone_repo_if_requested(repo_dir: Path, repo_url: str) -> dict[str, object]:
    if repo_dir.exists() and (repo_dir / ".git").exists():
        return {"repo_dir": str(repo_dir), "clone_performed": False, "status": "already-cloned"}
    repo_dir.parent.mkdir(parents=True, exist_ok=True)
    result = run_command(["git", "clone", repo_url, str(repo_dir)], timeout=60)
    return {
        "repo_dir": str(repo_dir),
        "clone_performed": True,
        "returncode": result.returncode,
        "stdout": result.stdout,
        "stderr": result.stderr,
    }


def repo_web_url(repo_url: str) -> str:
    if repo_url.startswith("git@github.com:"):
        repo_path = repo_url[len("git@github.com:") :]
        if repo_path.endswith(".git"):
            repo_path = repo_path[:-4]
        return f"https://github.com/{repo_path}"
    return repo_url[:-4] if repo_url.endswith(".git") else repo_url


def load_json(path: Path) -> dict[str, object]:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: dict[str, object]) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def load_round_context(round_json_path: Path) -> tuple[dict[str, object], dict[str, object], Path]:
    round_payload = load_json(round_json_path)
    idea_dir = round_json_path.resolve().parents[1]
    metadata_path = idea_dir / "metadata.json"
    metadata = load_json(metadata_path) if metadata_path.exists() else {}
    return round_payload, metadata, idea_dir


def git(repo_dir: Path, *args: str, timeout: int = 30) -> subprocess.CompletedProcess[str]:
    return run_command(["git", *args], cwd=repo_dir, timeout=timeout)


def git_head_sha(repo_dir: Path, rev: str = "HEAD") -> str | None:
    result = git(repo_dir, "rev-parse", "--verify", rev)
    value = result.stdout.strip()
    return value or None


def git_remote_branch_exists(repo_dir: Path, branch: str, remote: str = "origin") -> bool:
    return git(repo_dir, "ls-remote", "--exit-code", "--heads", remote, branch).returncode == 0


def parse_github_repo(repo_url: str) -> tuple[str, str]:
    match = re.search(r"github\.com[:/](?P<owner>[^/]+)/(?P<repo>[^/.]+?)(?:\.git)?$", repo_url)
    if not match:
        raise SystemExit(f"Unsupported GitHub repo URL: {repo_url}")
    return match.group("owner"), match.group("repo")


def get_github_credentials() -> dict[str, str]:
    result = run_command(
        ["git", "credential-osxkeychain", "get"],
        input_text="protocol=https\nhost=github.com\n\n",
        timeout=20,
    )
    if result.returncode != 0:
        raise SystemExit("Failed to read GitHub credentials from osxkeychain.")

    pairs: dict[str, str] = {}
    for line in result.stdout.splitlines():
        if "=" not in line:
            continue
        key, value = line.split("=", 1)
        pairs[key] = value

    if not pairs.get("username") or not pairs.get("password"):
        raise SystemExit("GitHub credentials were not found in osxkeychain.")
    return {"username": pairs["username"], "password": pairs["password"]}


def github_api_json(method: str, api_url: str, username: str, password: str, payload: dict[str, object] | None = None) -> tuple[int, dict[str, object]]:
    data = None
    headers = {
        "Accept": "application/vnd.github+json",
        "User-Agent": "solo-coder-flow",
    }
    if payload is not None:
        data = json.dumps(payload).encode("utf-8")
        headers["Content-Type"] = "application/json"

    token = base64.b64encode(f"{username}:{password}".encode("utf-8")).decode("ascii")
    headers["Authorization"] = f"Basic {token}"
    request = urlrequest.Request(api_url, data=data, headers=headers, method=method.upper())

    try:
        with urlrequest.urlopen(request, timeout=30) as response:
            body = response.read().decode("utf-8")
            return response.status, json.loads(body) if body else {}
    except urlerror.HTTPError as exc:
        body = exc.read().decode("utf-8")
        parsed = json.loads(body) if body else {}
        return exc.code, parsed


def git_create_empty_root_commit(repo_dir: Path, message: str) -> str:
    empty_tree = run_command(["git", "mktree"], cwd=repo_dir, timeout=20, input_text="").stdout
    tree_sha = empty_tree.strip()
    if not tree_sha:
        raise SystemExit("Failed to create an empty git tree for bootstrap.")

    commit_result = run_command(
        ["git", "commit-tree", tree_sha, "-m", message],
        cwd=repo_dir,
        timeout=20,
    )
    commit_sha = commit_result.stdout.strip()
    if commit_result.returncode != 0 or not commit_sha:
        raise SystemExit(f"Failed to create bootstrap commit: {commit_result.stderr.strip() or commit_result.stdout.strip()}")
    return commit_sha


def git_update_branch_ref(repo_dir: Path, branch: str, commit_sha: str) -> None:
    result = git(repo_dir, "update-ref", f"refs/heads/{branch}", commit_sha)
    if result.returncode != 0:
        raise SystemExit(f"Failed to update branch ref {branch}: {result.stderr.strip() or result.stdout.strip()}")


def ensure_publish_base(repo_dir: Path, parent_branch: str, round_branch: str, round_name: str) -> dict[str, object]:
    actions: list[dict[str, object]] = []

    if not git_has_commits(repo_dir):
        root_commit = git_create_empty_root_commit(repo_dir, f"[{round_name}] bootstrap parent branch")
        git_update_branch_ref(repo_dir, parent_branch, root_commit)
        actions.append({"action": "bootstrap-parent-root", "branch": parent_branch, "commit": root_commit})
        if not git_branch_exists(repo_dir, round_branch):
            git_update_branch_ref(repo_dir, round_branch, root_commit)
            actions.append({"action": "bootstrap-round-from-parent", "branch": round_branch, "commit": root_commit})
        return {"actions": actions, "parent_commit": root_commit}

    if not git_branch_exists(repo_dir, parent_branch):
        head_sha = git_head_sha(repo_dir)
        if not head_sha:
            raise SystemExit(f"Parent branch {parent_branch} is missing and HEAD is unavailable.")
        create_parent = git(repo_dir, "branch", parent_branch, head_sha)
        actions.append({
            "action": "create-parent-from-head",
            "branch": parent_branch,
            "returncode": create_parent.returncode,
            "stdout": create_parent.stdout,
            "stderr": create_parent.stderr,
            "commit": head_sha,
        })
        if create_parent.returncode != 0:
            raise SystemExit(f"Failed to create parent branch {parent_branch}: {create_parent.stderr.strip() or create_parent.stdout.strip()}")

    if not git_branch_exists(repo_dir, round_branch):
        create_round = git(repo_dir, "branch", round_branch, parent_branch)
        actions.append({
            "action": "create-round-from-parent",
            "branch": round_branch,
            "returncode": create_round.returncode,
            "stdout": create_round.stdout,
            "stderr": create_round.stderr,
        })
        if create_round.returncode != 0:
            raise SystemExit(f"Failed to create round branch {round_branch}: {create_round.stderr.strip() or create_round.stdout.strip()}")

    parent_commit = git_head_sha(repo_dir, parent_branch)
    return {"actions": actions, "parent_commit": parent_commit}


def render_pr_body(round_payload: dict[str, object], metadata: dict[str, object]) -> str:
    review_result = dict(round_payload.get("review_result") or {})
    prompt = str(round_payload.get("last_prompt") or round_payload.get("prompt") or "").strip()
    lines = [
        "## Summary",
        "",
        f"- Idea ID: `{round_payload.get('idea_id') or metadata.get('idea_id') or ''}`",
        f"- Round: `{round_payload.get('round_name') or ''}`",
        f"- Task Type: `{round_payload.get('task_type') or ''}`",
        f"- Business Domain: `{round_payload.get('business_domain') or ''}`",
        f"- Modification Scope: `{round_payload.get('modification_scope') or ''}`",
        "",
        "## User Prompt",
        "",
        "```text",
        prompt,
        "```",
        "",
        "## Trae Session",
        "",
        f"- Session ID: `{round_payload.get('session_id') or '未记录'}`",
        f"- Model: `{round_payload.get('model') or round_payload.get('required_model') or REQUIRED_MODEL}`",
        f"- Agent: `{round_payload.get('agent') or round_payload.get('required_agent') or REQUIRED_AGENT_NAME}`",
        "",
        "## Review",
        "",
        f"- Task Completed: `{review_result.get('task_completed') or '未填写'}`",
        f"- Satisfied: `{review_result.get('satisfied') or '未填写'}`",
        f"- Dissatisfaction Notes: `{review_result.get('reason') or '无'}`",
    ]
    return "\n".join(lines).strip() + "\n"


def create_or_find_pull_request(repo_url: str, base_branch: str, head_branch: str, title: str, body: str) -> dict[str, object]:
    owner, repo = parse_github_repo(repo_url)
    credentials = get_github_credentials()
    head_query = urlparse.quote(f"{owner}:{head_branch}", safe="")
    base_query = urlparse.quote(base_branch, safe="")
    list_url = f"https://api.github.com/repos/{owner}/{repo}/pulls?state=open&head={head_query}&base={base_query}"
    list_status, list_payload = github_api_json("GET", list_url, credentials["username"], credentials["password"])
    if list_status == 200 and isinstance(list_payload, list) and list_payload:
        pr = list_payload[0]
        return {
            "created": False,
            "number": pr.get("number"),
            "url": pr.get("html_url"),
            "api_status": list_status,
            "state": pr.get("state"),
        }

    create_url = f"https://api.github.com/repos/{owner}/{repo}/pulls"
    payload = {
        "title": title,
        "head": head_branch,
        "base": base_branch,
        "body": body,
    }
    create_status, create_payload = github_api_json("POST", create_url, credentials["username"], credentials["password"], payload)
    if create_status not in {200, 201}:
        raise SystemExit(f"Failed to create pull request: HTTP {create_status} {json.dumps(create_payload, ensure_ascii=False)}")

    return {
        "created": True,
        "number": create_payload.get("number"),
        "url": create_payload.get("html_url"),
        "api_status": create_status,
        "state": create_payload.get("state"),
    }

def ensure_repo_initialized(repo_dir: Path, repo_url: str, remote_state: dict[str, object]) -> dict[str, object]:
    repo_dir.mkdir(parents=True, exist_ok=True)
    actions: list[dict[str, object]] = []
    repo_has_files = any(repo_dir.iterdir())

    if remote_state.get("has_refs"):
        if not (repo_dir / ".git").exists() and not repo_has_files:
            clone_result = clone_repo_if_requested(repo_dir, repo_url)
            actions.append({"step": "clone", **clone_result})
        elif not (repo_dir / ".git").exists():
            init_result = git(repo_dir, "init", "-b", "main")
            actions.append({
                "step": "git-init-existing-dir",
                "returncode": init_result.returncode,
                "stdout": init_result.stdout,
                "stderr": init_result.stderr,
            })
        else:
            actions.append({"step": "clone", "status": "skipped-existing-repo"})
    elif not (repo_dir / ".git").exists():
        init_result = git(repo_dir, "init", "-b", "main")
        actions.append({
            "step": "git-init",
            "returncode": init_result.returncode,
            "stdout": init_result.stdout,
            "stderr": init_result.stderr,
        })

    if not (repo_dir / ".git").exists():
        return {"repo_dir": str(repo_dir), "actions": actions, "ok": False}

    current_remote = git(repo_dir, "remote", "get-url", "origin")
    if current_remote.returncode != 0:
        add_remote = git(repo_dir, "remote", "add", "origin", repo_url)
        actions.append({
            "step": "remote-add-origin",
            "returncode": add_remote.returncode,
            "stdout": add_remote.stdout,
            "stderr": add_remote.stderr,
        })
    elif current_remote.stdout.strip() != repo_url:
        set_remote = git(repo_dir, "remote", "set-url", "origin", repo_url)
        actions.append({
            "step": "remote-set-url",
            "returncode": set_remote.returncode,
            "stdout": set_remote.stdout,
            "stderr": set_remote.stderr,
        })
    else:
        actions.append({"step": "remote-origin", "status": "already-correct", "url": repo_url})

    if remote_state.get("has_refs"):
        fetch_result = git(repo_dir, "fetch", "origin")
        actions.append({
            "step": "fetch-origin",
            "returncode": fetch_result.returncode,
            "stdout": fetch_result.stdout,
            "stderr": fetch_result.stderr,
        })

    return {"repo_dir": str(repo_dir), "actions": actions, "ok": True}


def git_has_commits(repo_dir: Path) -> bool:
    return git(repo_dir, "rev-parse", "--verify", "HEAD").returncode == 0


def git_branch_exists(repo_dir: Path, branch: str) -> bool:
    return git(repo_dir, "rev-parse", "--verify", f"refs/heads/{branch}").returncode == 0


def git_current_branch(repo_dir: Path) -> str | None:
    result = git(repo_dir, "branch", "--show-current")
    branch = result.stdout.strip()
    return branch or None


def checkout_branch(repo_dir: Path, branch: str, start_point: str | None = None, orphan: bool = False) -> dict[str, object]:
    if git_branch_exists(repo_dir, branch):
        result = git(repo_dir, "checkout", branch)
        return {"action": "checkout-existing", "branch": branch, "returncode": result.returncode, "stdout": result.stdout, "stderr": result.stderr}
    if orphan:
        result = git(repo_dir, "checkout", "--orphan", branch)
        return {"action": "checkout-orphan", "branch": branch, "returncode": result.returncode, "stdout": result.stdout, "stderr": result.stderr}
    if start_point:
        result = git(repo_dir, "checkout", "-b", branch, start_point)
    else:
        result = git(repo_dir, "checkout", "-b", branch)
    return {"action": "checkout-create", "branch": branch, "start_point": start_point, "returncode": result.returncode, "stdout": result.stdout, "stderr": result.stderr}


def choose_seed_ref(repo_dir: Path, remote_state: dict[str, object]) -> str | None:
    if remote_state.get("has_refs"):
        default_branch = remote_state.get("default_branch") or "main"
        return f"origin/{default_branch}"
    if git_has_commits(repo_dir):
        current = git_current_branch(repo_dir)
        if current:
            return current
        return "HEAD"
    return None


def prepare_round_repo(round_json_path: Path, repo_dir_override: str | None = None) -> dict[str, object]:
    round_payload, metadata, idea_dir = load_round_context(round_json_path)
    repo_url = str(metadata.get("repo_url") or DEFAULT_REPO_URL)
    repo_dir = Path(repo_dir_override).resolve() if repo_dir_override else idea_dir / "repo"
    remote_state = detect_remote_repo_state(repo_url)
    repo_setup = ensure_repo_initialized(repo_dir, repo_url, remote_state)
    if not repo_setup.get("ok"):
        return {"ok": False, "repo_setup": repo_setup}

    folder_name = str(metadata.get("folder_name") or build_folder_name(str(round_payload.get("idea_id") or "idea"), str(round_payload.get("prompt") or "idea")))
    seed_ref = choose_seed_ref(repo_dir, remote_state)
    base_branch = str(round_payload.get("base_branch_name") or build_base_branch_name(str(round_payload.get("idea_id") or "idea")))
    parent_branch = str(round_payload.get("parent_branch_name") or base_branch)
    round_branch = str(round_payload.get("branch_name") or build_branch_name(str(round_payload.get("idea_id") or "idea"), str(round_payload.get("round_name") or "第一轮")))

    branch_actions = []
    if parent_branch == base_branch:
        if git_branch_exists(repo_dir, base_branch):
            branch_actions.append(checkout_branch(repo_dir, base_branch))
        else:
            branch_actions.append(checkout_branch(repo_dir, base_branch, start_point=seed_ref, orphan=seed_ref is None))
    else:
        current_branch = git_current_branch(repo_dir)
        if git_branch_exists(repo_dir, parent_branch):
            branch_actions.append(checkout_branch(repo_dir, parent_branch))
        elif current_branch == parent_branch:
            branch_actions.append({"action": "reuse-current-unborn", "branch": parent_branch, "returncode": 0, "stdout": "", "stderr": ""})
        else:
            parent_start = current_branch if git_has_commits(repo_dir) and current_branch else None
            branch_actions.append(checkout_branch(repo_dir, parent_branch, start_point=parent_start, orphan=parent_start is None))

    if git_branch_exists(repo_dir, round_branch):
        branch_actions.append(checkout_branch(repo_dir, round_branch))
    else:
        start_point = parent_branch if git_has_commits(repo_dir) and git_branch_exists(repo_dir, parent_branch) else None
        branch_actions.append(checkout_branch(repo_dir, round_branch, start_point=start_point, orphan=False))

    work_folder = repo_dir / folder_name
    work_folder.mkdir(parents=True, exist_ok=True)

    compare_url = f"{repo_web_url(repo_url)}/compare/{parent_branch}...{round_branch}?expand=1"
    round_payload["repo_dir"] = str(repo_dir)
    round_payload["work_folder"] = str(work_folder)
    round_payload["compare_url"] = compare_url
    round_payload["repo_prepared_at"] = dt.datetime.now().astimezone().isoformat()
    round_payload["base_branch_name"] = base_branch
    round_payload["parent_branch_name"] = parent_branch
    round_payload["branch_name"] = round_branch
    round_payload["repo_setup"] = repo_setup
    round_payload["branch_actions"] = branch_actions
    round_payload["repo_remote_state"] = remote_state
    write_json(round_json_path, round_payload)

    if metadata:
        repo_plan = dict(metadata.get("repo_plan") or {})
        repo_plan["base_branch"] = base_branch
        repo_plan["parent_branch"] = parent_branch
        repo_plan["round_branch"] = round_branch
        repo_plan["compare_url"] = compare_url
        repo_plan["repo_dir"] = str(repo_dir)
        metadata["repo_plan"] = repo_plan
        write_json(idea_dir / "metadata.json", metadata)

    branch_errors = [action for action in branch_actions if int(action.get("returncode", 0)) != 0]
    return {
        "ok": not branch_errors,
        "repo_dir": str(repo_dir),
        "work_folder": str(work_folder),
        "base_branch": base_branch,
        "parent_branch": parent_branch,
        "round_branch": round_branch,
        "compare_url": compare_url,
        "repo_setup": repo_setup,
        "branch_actions": branch_actions,
        "remote_state": remote_state,
        "branch_errors": branch_errors,
    }


def summarize_status_paths(status_lines: list[str]) -> list[str]:
    paths: list[str] = []
    for raw in status_lines:
        line = raw.rstrip()
        if not line:
            continue
        path = line[3:]
        if " -> " in path:
            path = path.split(" -> ", 1)[1]
        paths.append(path)
    return paths


def infer_scope_from_files(paths: list[str]) -> str:
    unique_paths = sorted(set(paths))
    if len(unique_paths) <= 1:
        return "单文件"
    top_levels = {path.split("/", 1)[0] for path in unique_paths}
    frontend_tokens = {"web", "app", "client", "frontend", "ui", "pages", "components"}
    backend_tokens = {"server", "api", "backend", "db", "database", "prisma", "supabase"}
    if top_levels & frontend_tokens and top_levels & backend_tokens:
        return "跨系统多模块"
    if len(top_levels) == 1:
        return "模块内多文件"
    if len(top_levels) >= 3:
        return "跨系统多模块"
    return "跨模块多文件"


def collect_review_context(repo_dir: Path, base_branch: str | None, round_branch: str | None) -> dict[str, object]:
    status_result = git(repo_dir, "status", "--porcelain")
    status_lines = [line for line in status_result.stdout.splitlines() if line.strip()]
    working_paths = summarize_status_paths(status_lines)

    diff_name_args = ["diff", "--name-only"]
    diff_stat_args = ["diff", "--stat"]
    if base_branch and git_branch_exists(repo_dir, base_branch) and git_has_commits(repo_dir):
        diff_name_args = ["diff", "--name-only", f"{base_branch}...HEAD"]
        diff_stat_args = ["diff", "--stat", f"{base_branch}...HEAD"]

    diff_names = git(repo_dir, *diff_name_args)
    diff_stat = git(repo_dir, *diff_stat_args)
    branch = git_current_branch(repo_dir)

    changed_files = sorted(set(working_paths + [line for line in diff_names.stdout.splitlines() if line.strip()]))
    inferred_scope = infer_scope_from_files(changed_files)
    top_levels = sorted({path.split("/", 1)[0] for path in changed_files})

    package_hints = []
    for marker in ["package.json", "pnpm-lock.yaml", "yarn.lock", "requirements.txt", "pyproject.toml", "Cargo.toml", "go.mod"]:
        if (repo_dir / marker).exists():
            package_hints.append(marker)

    return {
        "repo_dir": str(repo_dir),
        "current_branch": branch,
        "base_branch": base_branch,
        "round_branch": round_branch,
        "status_lines": status_lines,
        "changed_files": changed_files,
        "changed_file_count": len(changed_files),
        "top_level_paths": top_levels,
        "diff_stat": diff_stat.stdout.strip(),
        "inferred_modification_scope": inferred_scope,
        "package_hints": package_hints,
        "run_checks_note": "未自动执行生成项目里的测试或构建脚本，避免误运行不受控命令；如需自动跑检查，需要你在那一轮明确允许。",
    }


def cmd_prepare_repo(args: argparse.Namespace) -> int:
    result = prepare_round_repo(Path(args.round_json), args.repo_dir)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


def cmd_review_round(args: argparse.Namespace) -> int:
    round_payload, metadata, idea_dir = load_round_context(Path(args.round_json))
    repo_dir = Path(args.repo_dir).resolve() if args.repo_dir else Path(str(round_payload.get("repo_dir") or idea_dir / "repo"))
    if not (repo_dir / ".git").exists():
        raise SystemExit("Repo is not prepared yet. Run prepare-repo first.")

    review = collect_review_context(
        repo_dir=repo_dir,
        base_branch=str(round_payload.get("parent_branch_name") or round_payload.get("base_branch_name") or ""),
        round_branch=str(round_payload.get("branch_name") or ""),
    )
    round_payload["review_context"] = review
    round_payload["review_context_updated_at"] = dt.datetime.now().astimezone().isoformat()
    write_json(Path(args.round_json), round_payload)

    review_path = idea_dir / "rounds" / f"{round_payload.get('round_name')}-review-context.json"
    write_json(review_path, review)

    payload = {
        "round_json": args.round_json,
        "review_context_path": str(review_path),
        "review": review,
        "required_model": REQUIRED_MODEL,
        "required_agent": REQUIRED_AGENT_NAME,
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0


def next_round_name(current_round: str) -> str:
    ordered = ["第一轮", "第二轮", "第三轮", "第四轮", "第五轮"]
    try:
        idx = ordered.index(current_round)
    except ValueError:
        return "第二轮"
    return ordered[min(idx + 1, len(ordered) - 1)]


def build_reason_text(product_reasons: list[str] | None, process_reasons: list[str] | None) -> str:
    parts: list[str] = []
    clean_product = [item.strip() for item in (product_reasons or []) if item and item.strip()]
    clean_process = [item.strip() for item in (process_reasons or []) if item and item.strip()]
    if clean_product:
        parts.append("产物不满意：" + "；".join(clean_product))
    if clean_process:
        parts.append("过程不满意：" + "；".join(clean_process))
    return "\n".join(parts)


def build_followup_prompt(folder_name: str, issues: list[str], previous_round: str, next_round: str) -> str:
    issue_lines = "\n".join(f"- {issue}" for issue in issues)
    return (
        f"继续上一轮任务，不要推倒重来，仍然必须使用 {REQUIRED_MODEL} 模型和 {REQUIRED_AGENT_NAME} 智能体。\n\n"
        f"当前是 {next_round}，请基于上一轮已经生成的代码继续修改，只允许在 `{folder_name}/` 及其相关必要配置内调整。\n\n"
        f"这轮优先解决以下问题：\n{issue_lines}\n\n"
        "额外要求：\n"
        "- 不要新开无关需求\n"
        "- 不要把项目重写成另一套方案\n"
        "- 先修复问题，再补齐缺失项\n"
        "- 如果上一轮已经有可用部分，尽量保留\n"
        "- 最后明确说明本轮具体修了什么、还有什么没完成\n\n"
        "这是上一轮的延续，所以验收仍然围绕上一轮核心目标展开。\n"
    )


def cmd_make_next_prompt(args: argparse.Namespace) -> int:
    round_json_path = Path(args.round_json)
    round_payload, metadata, idea_dir = load_round_context(round_json_path)
    issues = [issue.strip() for issue in (args.issue or []) if issue.strip()]
    if args.issues_file:
        issues.extend([line.strip() for line in Path(args.issues_file).read_text(encoding="utf-8").splitlines() if line.strip()])
    if not issues:
        raise SystemExit("Provide at least one --issue or --issues-file.")

    current_round = str(round_payload.get("round_name") or "第一轮")
    next_round = args.next_round_name or next_round_name(current_round)
    folder_name = str(metadata.get("folder_name") or build_folder_name(str(round_payload.get("idea_id") or "idea"), str(round_payload.get("prompt") or "idea")))
    prompt_text = build_followup_prompt(folder_name, issues, current_round, next_round)

    prompt_path = idea_dir / "prompts" / f"{next_round}-优化提示词.md"

    next_round_json_path = idea_dir / "rounds" / f"{next_round}.json"
    next_payload = dict(round_payload)
    next_payload["round_name"] = next_round
    next_payload["prompt"] = prompt_text
    next_payload["prompt_path"] = str(prompt_path)
    next_payload["branch_name"] = build_branch_name(str(round_payload.get("idea_id") or "idea"), next_round)
    next_payload["parent_branch_name"] = str(round_payload.get("branch_name") or build_base_branch_name(str(round_payload.get("idea_id") or "idea")))
    next_payload["pr_title"] = build_pr_title(
        str(round_payload.get("idea_id") or "idea"),
        next_round,
        str(metadata.get("idea_short_title") or metadata.get("folder_name") or folder_name),
    )
    next_payload["previous_round_name"] = current_round
    next_payload["carry_over_issues"] = issues
    next_payload["session_id"] = None
    next_payload["message_id"] = None
    next_payload["trace_id"] = None
    next_payload["created_at"] = dt.datetime.now().astimezone().isoformat()
    next_payload.pop("review_context", None)
    next_payload.pop("review_result", None)
    next_payload.pop("review_context_updated_at", None)
    next_payload.pop("attachment", None)
    next_payload.pop("packaged_at", None)
    next_payload.pop("last_prompt", None)
    next_payload.pop("updated_at", None)

    repo_url = str(metadata.get("repo_url") or DEFAULT_REPO_URL)
    parent_branch = str(next_payload.get("parent_branch_name") or next_payload.get("base_branch_name") or build_base_branch_name(str(round_payload.get("idea_id") or "idea")))
    next_payload["compare_url"] = f"{repo_web_url(repo_url)}/compare/{parent_branch}...{next_payload['branch_name']}?expand=1"

    if not next_payload.get("task_type") or next_payload.get("task_type") == "0-1代码生成":
        next_payload["task_type"] = "Bug修复"

    payload = {
        "next_round_name": next_round,
        "prompt_path": str(prompt_path),
        "round_json_path": str(next_round_json_path),
        "branch_name": next_payload["branch_name"],
        "issues": issues,
        "prompt_preview": prompt_text,
    }

    if args.print_only:
        print(json.dumps(payload, ensure_ascii=False, indent=2))
        return 0

    if not args.yes:
        print("下一轮提示词预览：")
        print(prompt_text)
        if not prompt_yes_no("确认保存这份下一轮提示词并继续吗？", default=False):
            print(json.dumps({**payload, "saved": False}, ensure_ascii=False, indent=2))
            return 0

    prompt_path.write_text(prompt_text, encoding="utf-8")
    write_json(next_round_json_path, next_payload)
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0


def cmd_record_review(args: argparse.Namespace) -> int:
    round_json_path = Path(args.round_json)
    round_payload, metadata, _ = load_round_context(round_json_path)
    if args.reason and (args.product_reason or args.process_reason):
        raise SystemExit("Pass either --reason or --product-reason/--process-reason, not both.")

    reason = args.reason or build_reason_text(args.product_reason, args.process_reason)
    if args.satisfied == "不满意" and not reason:
        raise SystemExit("When satisfied=不满意, pass --reason or at least one --product-reason/--process-reason.")

    review_info = {
        "task_completed": args.task_completed,
        "satisfied": args.satisfied,
        "reason": reason,
        "recorded_at": dt.datetime.now().astimezone().isoformat(),
    }
    round_payload["review_result"] = review_info
    if args.github_url:
        round_payload["compare_url"] = args.github_url
    write_json(round_json_path, round_payload)

    if not args.workbook:
        print(json.dumps({"round_json": str(round_json_path), "review_result": review_info}, ensure_ascii=False, indent=2))
        return 0

    github_url = args.github_url or str(
        round_payload.get("github_url")
        or round_payload.get("pr_url")
        or round_payload.get("compare_url")
        or ""
    )
    folder_name = str(metadata.get("folder_name") or "")
    branch_name = str(round_payload.get("branch_name") or "")
    branch_folder = args.branch_folder or " | ".join(part for part in [branch_name, folder_name] if part)
    values = {
        "任务是否完成": args.task_completed,
        "产物及过程是否满意": args.satisfied,
        "不满意原因": reason or "",
        "github地址": github_url,
        "分支/文件夹": branch_folder,
    }
    if args.attachment:
        values["附件"] = args.attachment
    if args.status:
        values["状态"] = args.status

    workbook_update = update_workbook(Path(args.workbook), args.uid, args.round_name, args.row, values)
    payload = {
        "round_json": str(round_json_path),
        "review_result": review_info,
        "workbook_update": workbook_update,
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0


def cmd_publish_round(args: argparse.Namespace) -> int:
    round_json_path = Path(args.round_json)
    round_payload, metadata, idea_dir = load_round_context(round_json_path)
    repo_dir = Path(args.repo_dir).resolve() if args.repo_dir else Path(str(round_payload.get("repo_dir") or idea_dir / "repo"))
    if not (repo_dir / ".git").exists():
        raise SystemExit("Repo is not prepared yet. Run prepare-repo first.")

    repo_url = str(metadata.get("repo_url") or DEFAULT_REPO_URL)
    round_name = str(round_payload.get("round_name") or "第一轮")
    idea_id = str(round_payload.get("idea_id") or metadata.get("idea_id") or "idea")
    folder_name = str(metadata.get("folder_name") or build_folder_name(idea_id, str(round_payload.get("prompt") or idea_id)))
    parent_branch = str(round_payload.get("parent_branch_name") or round_payload.get("base_branch_name") or build_base_branch_name(idea_id))
    round_branch = str(round_payload.get("branch_name") or build_branch_name(idea_id, round_name))

    bootstrap = ensure_publish_base(repo_dir, parent_branch, round_branch, round_name)
    current_branch = git_current_branch(repo_dir)
    checkout_result = None
    if current_branch != round_branch:
        checkout_result = checkout_branch(repo_dir, round_branch)
        if int(checkout_result.get("returncode", 0)) != 0:
            raise SystemExit(f"Failed to checkout round branch {round_branch}: {checkout_result.get('stderr') or checkout_result.get('stdout')}")

    status_before = git(repo_dir, "status", "--porcelain")
    status_before_lines = [line for line in status_before.stdout.splitlines() if line.strip()]

    add_result = git(repo_dir, "add", "-A")
    if add_result.returncode != 0:
        raise SystemExit(f"git add failed: {add_result.stderr.strip() or add_result.stdout.strip()}")

    commit_message = args.commit_message or str(
        round_payload.get("pr_title")
        or build_pr_title(idea_id, round_name, str(metadata.get("idea_short_title") or folder_name))
    )
    commit_result = git(repo_dir, "commit", "-m", commit_message, timeout=60)
    commit_output = (commit_result.stdout or "") + (commit_result.stderr or "")
    if commit_result.returncode == 0:
        committed = True
    elif "nothing to commit" in commit_output and git_has_commits(repo_dir):
        committed = False
    else:
        raise SystemExit(f"git commit failed: {commit_output.strip()}")

    head_sha = git_head_sha(repo_dir)
    if not head_sha:
        raise SystemExit("HEAD is still unavailable after commit.")

    skip_git_push = SKIP_GIT_PUSH and not args.push
    parent_push_args = ["push", "origin", f"{parent_branch}:refs/heads/{parent_branch}"]
    round_push_args = ["push", "-u", "origin", f"{round_branch}:refs/heads/{round_branch}"]
    if args.dry_run:
        parent_push_args.insert(1, "--dry-run")
        round_push_args.insert(1, "--dry-run")

    if skip_git_push:
        parent_push = subprocess.CompletedProcess(parent_push_args, 0, stdout="skipped by default\n", stderr="")
        round_push = subprocess.CompletedProcess(round_push_args, 0, stdout="skipped by default\n", stderr="")
    else:
        parent_push = git(repo_dir, *parent_push_args, timeout=120)
        if parent_push.returncode != 0:
            raise SystemExit(f"Failed to push parent branch {parent_branch}: {parent_push.stderr.strip() or parent_push.stdout.strip()}")

        round_push = git(repo_dir, *round_push_args, timeout=120)
        if round_push.returncode != 0:
            raise SystemExit(f"Failed to push round branch {round_branch}: {round_push.stderr.strip() or round_push.stdout.strip()}")

    compare_url = f"{repo_web_url(repo_url)}/compare/{parent_branch}...{round_branch}?expand=1"
    pr_body = Path(args.pr_body_file).read_text(encoding="utf-8") if args.pr_body_file else render_pr_body(round_payload, metadata)
    pr_result: dict[str, object] | None = None
    if not args.dry_run and not args.skip_pr and not skip_git_push:
        pr_result = create_or_find_pull_request(
            repo_url=repo_url,
            base_branch=parent_branch,
            head_branch=round_branch,
            title=str(round_payload.get("pr_title") or commit_message),
            body=pr_body,
        )

    github_url = str((pr_result or {}).get("url") or compare_url)
    round_payload["parent_branch_name"] = parent_branch
    round_payload["branch_name"] = round_branch
    round_payload["compare_url"] = compare_url
    round_payload["github_url"] = github_url
    round_payload["commit_message"] = commit_message
    round_payload["commit_sha"] = head_sha
    round_payload["published_at"] = dt.datetime.now().astimezone().isoformat()
    round_payload["pr_url"] = (pr_result or {}).get("url")
    round_payload["pr_number"] = (pr_result or {}).get("number")
    round_payload["publish_result"] = {
        "bootstrap": bootstrap,
        "checkout": checkout_result,
        "status_before": status_before_lines,
        "committed": committed,
        "parent_push": {
            "returncode": parent_push.returncode,
            "stdout": parent_push.stdout,
            "stderr": parent_push.stderr,
        },
        "round_push": {
            "returncode": round_push.returncode,
            "stdout": round_push.stdout,
            "stderr": round_push.stderr,
        },
        "dry_run": args.dry_run,
        "push_skipped": skip_git_push,
        "skip_pr": args.skip_pr,
        "pr_result": pr_result,
    }
    write_json(round_json_path, round_payload)

    workbook_update = None
    if args.workbook:
        branch_folder = args.branch_folder or " | ".join(part for part in [round_branch, folder_name] if part)
        values = {
            "github地址": github_url,
            "分支/文件夹": branch_folder,
        }
        if args.status:
            values["状态"] = args.status
        workbook_update = update_workbook(Path(args.workbook), args.uid, args.round_name, args.row, values)

    payload = {
        "round_json": str(round_json_path),
        "repo_dir": str(repo_dir),
        "commit_message": commit_message,
        "commit_sha": head_sha,
        "parent_branch": parent_branch,
        "round_branch": round_branch,
        "compare_url": compare_url,
        "github_url": github_url,
        "dry_run": args.dry_run,
        "skip_pr": args.skip_pr,
        "pr_result": pr_result,
        "workbook_update": workbook_update,
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0


def create_idea_scaffold(
    idea_id: str,
    idea: str,
    round_name: str,
    ideas_root: str,
    repo_url: str,
    workbook: str,
    uid: str | None,
    clone_repo: bool,
    folder_name_override: str | None = None,
) -> dict[str, object]:
    ideas_root_path = Path(ideas_root).resolve()
    folder_name = folder_name_override or build_folder_name(idea_id, idea)
    idea_dir = ideas_root_path / folder_name
    idea_dir.mkdir(parents=True, exist_ok=True)

    prompts_dir = idea_dir / "prompts"
    rounds_dir = idea_dir / "rounds"
    pr_dir = idea_dir / "pr"
    artifacts_dir = idea_dir / "artifacts"
    repo_dir = idea_dir / "repo"
    for path in [prompts_dir, rounds_dir, pr_dir, artifacts_dir]:
        path.mkdir(parents=True, exist_ok=True)

    task_type = infer_task_type(idea)
    business_domain = infer_business_domain(idea)
    modification_scope = infer_modification_scope(task_type)
    optimized_prompt = build_optimized_prompt(idea_id, round_name, folder_name, idea)
    pr_title = build_pr_title(idea_id, round_name, extract_short_title(idea))
    repo_state = detect_remote_repo_state(repo_url)

    prompt_raw_path = prompts_dir / f"{round_name}-提示词.md"
    prompt_optimized_path = prompts_dir / f"{round_name}-优化建议.md"
    idea_path = idea_dir / "idea.md"
    pr_path = pr_dir / f"{round_name}.md"
    round_json_path = rounds_dir / f"{round_name}.json"
    metadata_path = idea_dir / "metadata.json"

    prompt_raw_path.write_text(idea.strip() + "\n", encoding="utf-8")
    prompt_optimized_path.write_text(optimized_prompt, encoding="utf-8")
    idea_path.write_text(
        textwrap.dedent(
            f"""\
            # {idea_id}

            - 轮次：{round_name}
            - 任务类型建议：{task_type}
            - 业务领域建议：{business_domain}
            - 修改范围建议：{modification_scope}

            ## 原始点子

            {idea}
            """
        ),
        encoding="utf-8",
    )
    pr_path.write_text(
        build_pr_template(
            idea_id,
            round_name,
            optimized_prompt,
            task_type,
            business_domain,
            modification_scope,
        ),
        encoding="utf-8",
    )

    metadata = {
        "idea_id": idea_id,
        "idea_short_title": extract_short_title(idea),
        "folder_name": folder_name,
        "idea_dir": str(idea_dir),
        "repo_url": repo_url,
        "required_model": REQUIRED_MODEL,
        "required_agent": REQUIRED_AGENT_NAME,
        "required_agent_id": REQUIRED_AGENT_ID,
        "round_name": round_name,
        "suggested_labels": {
            "任务类型": task_type,
            "业务领域": business_domain,
            "修改范围": modification_scope,
        },
        "repo_plan": {
            "base_branch": build_base_branch_name(idea_id),
            "round_branch": build_branch_name(idea_id, round_name),
            "pr_title": pr_title,
            "remote_repo_state": repo_state,
        },
        "workbook_hint": {
            "path": workbook,
            "uid": uid,
            "round_name": round_name,
        },
        "paths": {
            "idea": str(idea_path),
            "prompt_raw": str(prompt_raw_path),
            "prompt_optimized": str(prompt_optimized_path),
            "pr": str(pr_path),
            "round_json": str(round_json_path),
        },
        "created_at": dt.datetime.now().astimezone().isoformat(),
    }

    round_record = {
        "idea_id": idea_id,
        "round_name": round_name,
        "required_model": REQUIRED_MODEL,
        "required_agent": REQUIRED_AGENT_NAME,
        "prompt_path": str(prompt_raw_path),
        "prompt": idea.strip() + "\n",
        "optimized_prompt_path": str(prompt_optimized_path),
        "optimized_prompt": optimized_prompt,
        "branch_name": build_branch_name(idea_id, round_name),
        "base_branch_name": build_base_branch_name(idea_id),
        "parent_branch_name": build_base_branch_name(idea_id),
        "pr_title": pr_title,
        "session_id": None,
        "message_id": None,
        "trace_id": None,
        "task_type": task_type,
        "business_domain": business_domain,
        "modification_scope": modification_scope,
    }

    metadata_path.write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    round_json_path.write_text(json.dumps(round_record, ensure_ascii=False, indent=2), encoding="utf-8")

    clone_info = None
    if clone_repo:
        clone_info = clone_repo_if_requested(repo_dir, repo_url)

    return {
        "idea_dir": str(idea_dir),
        "folder_name": folder_name,
        "prompt_optimized_path": str(prompt_optimized_path),
        "pr_template_path": str(pr_path),
        "round_record_path": str(round_json_path),
        "metadata_path": str(metadata_path),
        "required_model": REQUIRED_MODEL,
        "required_agent": REQUIRED_AGENT_NAME,
        "repo_plan": metadata["repo_plan"],
        "clone_info": clone_info,
        "prompt_raw_path": str(prompt_raw_path),
    }


def cmd_init_idea(args: argparse.Namespace) -> int:
    result = create_idea_scaffold(
        idea_id=args.idea_id,
        idea=args.idea,
        round_name=args.round_name,
        ideas_root=args.ideas_root,
        repo_url=args.repo_url,
        workbook=args.workbook,
        uid=args.uid,
        clone_repo=args.clone_repo,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


def prompt_text(label: str, default: str | None = None, allow_empty: bool = False) -> str:
    while True:
        suffix = f" [{default}]" if default else ""
        value = input(f"{label}{suffix}\n> ").strip()
        if value:
            return value
        if default is not None:
            return default
        if allow_empty:
            return ""
        print("该项不能为空，请重新输入。")


def prompt_multiline(label: str) -> str:
    print(f"{label}\n> 第一行开始输入，空行结束")
    lines: list[str] = []
    while True:
        line = input()
        if not line.strip():
            if lines:
                return "\n".join(lines).strip()
            print("提示词不能为空，请继续输入。")
            continue
        lines.append(line)


def prompt_multiline_optional(label: str) -> str:
    print(f"{label}\n> 第一行开始输入，空行结束；直接回车可跳过")
    lines: list[str] = []
    while True:
        line = input()
        if not line.strip():
            return "\n".join(lines).strip()
        lines.append(line)


def prompt_reason_detail(label: str) -> str:
    if not prompt_yes_no(label, default=False):
        return ""
    return prompt_multiline_optional("请填写原因：")


def prompt_yes_no(label: str, default: bool = False) -> bool:
    suffix = "[Y/n]" if default else "[y/N]"
    answer = input(f"{label} {suffix}\n> ").strip().lower()
    if not answer:
        return default
    return answer in {"y", "yes"}


def prompt_choice(label: str, options: list[str], default: str | None = None) -> str:
    while True:
        print(label)
        for idx, option in enumerate(options, start=1):
            marker = " (默认)" if option == default else ""
            print(f"{idx}. {option}{marker}")
        raw = input("> ").strip()
        if not raw and default:
            return default
        if raw.isdigit():
            choice_idx = int(raw)
            if 1 <= choice_idx <= len(options):
                return options[choice_idx - 1]
        if raw in options:
            return raw
        print("输入无效，请输入序号或枚举值本身。")


def shell_arg(value: str) -> str:
    return shlex.quote(value)


def project_workspace_dir(projects_root: str, uid: str) -> Path:
    return Path(projects_root).resolve() / uid


def project_state_path(workspace_dir: Path) -> Path:
    return workspace_dir / "workflow-state.json"


def append_local_git_excludes(repo_dir: Path, patterns: list[str]) -> None:
    exclude_path = repo_dir / ".git" / "info" / "exclude"
    existing = exclude_path.read_text(encoding="utf-8") if exclude_path.exists() else ""
    lines = existing.splitlines()
    updated = False
    for pattern in patterns:
        if pattern not in lines:
            lines.append(pattern)
            updated = True
    if updated:
        exclude_path.write_text("\n".join(lines).strip() + "\n", encoding="utf-8")


def checkout_or_create_branch(repo_dir: Path, branch: str, remote_state: dict[str, object]) -> dict[str, object]:
    if git_branch_exists(repo_dir, branch):
        result = git(repo_dir, "checkout", branch)
        return {"action": "checkout-existing", "branch": branch, "returncode": result.returncode, "stdout": result.stdout, "stderr": result.stderr}
    if git_remote_branch_exists(repo_dir, branch):
        result = git(repo_dir, "checkout", "-b", branch, f"origin/{branch}")
        return {"action": "checkout-remote", "branch": branch, "returncode": result.returncode, "stdout": result.stdout, "stderr": result.stderr}
    seed_ref = choose_seed_ref(repo_dir, remote_state)
    return checkout_branch(repo_dir, branch, start_point=seed_ref, orphan=seed_ref is None)


def build_project_github_url(repo_url: str, branch_name: str, uid: str) -> str:
    return f"{repo_web_url(repo_url)}/tree/{urlparse.quote(branch_name, safe='')}/{uid}"


def ensure_uid_workspace_repo(
    *,
    uid: str,
    workspace_dir: Path,
    repo_url: str,
    branch_name: str,
) -> dict[str, object]:
    remote_state = detect_remote_repo_state(repo_url)
    repo_setup = ensure_repo_initialized(workspace_dir, repo_url, remote_state)
    if not repo_setup.get("ok"):
        raise SystemExit(f"Failed to initialize workspace repo: {json.dumps(repo_setup, ensure_ascii=False)}")

    branch_action = checkout_or_create_branch(workspace_dir, branch_name, remote_state)
    if int(branch_action.get("returncode", 0)) != 0:
        raise SystemExit(f"Failed to checkout branch {branch_name}: {branch_action.get('stderr') or branch_action.get('stdout')}")

    append_local_git_excludes(
        workspace_dir,
        [
            f"{uid}.xlsx",
            "workflow-state.json",
            f"{uid}.zip",
            ".DS_Store",
        ],
    )

    code_dir = workspace_dir / uid
    code_dir.mkdir(parents=True, exist_ok=True)
    placeholder_path = code_dir / "README.md"
    placeholder_created = False
    if not any(code_dir.iterdir()):
        placeholder_path.write_text(
            textwrap.dedent(
                f"""\
                # {uid}

                This folder is reserved for the Solo Coder project `{uid}`.
                Use Trae to build the real project here.
                """
            ),
            encoding="utf-8",
        )
        placeholder_created = True

    branch_tree_url = build_project_github_url(repo_url, branch_name, uid)
    return {
        "repo_url": repo_url,
        "remote_state": remote_state,
        "repo_setup": repo_setup,
        "branch_action": branch_action,
        "branch_name": branch_name,
        "code_dir": str(code_dir),
        "placeholder_created": placeholder_created,
        "branch_tree_url": branch_tree_url,
    }


def finalize_uid_project_publish(
    *,
    repo_dir: Path,
    repo_url: str,
    branch_name: str,
    uid: str,
    rounds: list[dict[str, object]],
    commit_message: str | None = None,
    push: bool | None = None,
) -> dict[str, object]:
    add_result = git(repo_dir, "add", "-A")
    if add_result.returncode != 0:
        raise SystemExit(f"git add failed: {add_result.stderr.strip() or add_result.stdout.strip()}")

    session_ids = [str(item.get("session_id") or "").strip() for item in rounds if str(item.get("session_id") or "").strip()]
    unique_sessions = list(dict.fromkeys(session_ids))
    session_suffix = ", ".join(unique_sessions)
    message = commit_message or f"[{uid}] sessions: {session_suffix}" if session_suffix else f"[{uid}] final publish"
    commit_result = git(repo_dir, "commit", "-m", message, timeout=60)
    commit_output = (commit_result.stdout or "") + (commit_result.stderr or "")
    if commit_result.returncode == 0:
        committed = True
    elif "nothing to commit" in commit_output and git_has_commits(repo_dir):
        committed = False
    else:
        raise SystemExit(f"git commit failed: {commit_output.strip()}")

    push_enabled = (not SKIP_GIT_PUSH) if push is None else push
    push_args = ["push", "-u", "origin", f"{branch_name}:refs/heads/{branch_name}"]
    if push_enabled:
        push_result = git(repo_dir, *push_args, timeout=120)
        if push_result.returncode != 0:
            raise SystemExit(f"Failed to push branch {branch_name}: {push_result.stderr.strip() or push_result.stdout.strip()}")
    else:
        push_result = subprocess.CompletedProcess(push_args, 0, stdout="skipped by default\n", stderr="")

    head_sha = git_head_sha(repo_dir)
    if not head_sha:
        raise SystemExit("HEAD is unavailable after commit.")

    return {
        "branch_name": branch_name,
        "commit_message": message,
        "commit_sha": head_sha,
        "committed": committed,
        "push": {
            "returncode": push_result.returncode,
            "stdout": push_result.stdout,
            "stderr": push_result.stderr,
            "skipped": not push_enabled,
        },
        "github_url": build_project_github_url(repo_url, branch_name, uid),
        "commit_url": f"{repo_web_url(repo_url)}/commit/{head_sha}",
        "branch_tree_url": build_project_github_url(repo_url, branch_name, uid),
    }


def open_trae_for_workspace(workspace_dir: Path, new_window: bool = True) -> dict[str, object]:
    cmd = [str(TRAE_CLI)]
    if new_window:
        cmd.append("-n")
    else:
        cmd.append("-r")
    cmd.append(str(workspace_dir))
    result = run_command(cmd, timeout=30)
    return {
        "command": cmd,
        "returncode": result.returncode,
        "stdout": result.stdout,
        "stderr": result.stderr,
        "workspace_dir": str(workspace_dir),
    }


def warn_if_too_long(field_name: str, value: str, limit: int = 220) -> None:
    if len(value) > limit:
        print(f"提醒：`{field_name}` 当前 {len(value)} 字，已超过 {limit} 字，提交前建议再精简。")


def initialize_uid_workflow(
    *,
    uid: str,
    business_domain: str,
    projects_root: str,
    workbook_template: str,
    repo_url: str,
    branch_name: str,
) -> dict[str, object]:
    workspace_dir = project_workspace_dir(projects_root, uid)
    workspace_dir.mkdir(parents=True, exist_ok=True)

    repo_info = ensure_uid_workspace_repo(
        uid=uid,
        workspace_dir=workspace_dir,
        repo_url=repo_url,
        branch_name=branch_name,
    )

    workbook_path = workspace_dir / f"{uid}.xlsx"
    if not workbook_path.exists():
        seed_project_workbook(Path(workbook_template), workbook_path, uid)

    state = {
        "uid": uid,
        "business_domain": business_domain,
        "workspace_dir": str(workspace_dir),
        "workbook_path": str(workbook_path),
        "repo_url": repo_url,
        "branch_name": branch_name,
        "code_dir": repo_info["code_dir"],
        "project_github_url": build_project_github_url(repo_url, branch_name, uid),
        "current_round_index": 1,
        "completed": False,
        "rounds": [],
        "created_at": dt.datetime.now().astimezone().isoformat(),
        "updated_at": dt.datetime.now().astimezone().isoformat(),
    }
    write_json(project_state_path(workspace_dir), state)
    return state


def load_or_initialize_uid_workflow(
    *,
    uid: str,
    projects_root: str,
    workbook_template: str,
    repo_url: str,
    branch_name: str,
) -> dict[str, object]:
    workspace_dir = project_workspace_dir(projects_root, uid)
    state_path = project_state_path(workspace_dir)
    if state_path.exists():
        return load_json(state_path)

    business_domain = prompt_choice("请告诉我你的业务领域枚举值：", BUSINESS_DOMAINS)
    return initialize_uid_workflow(
        uid=uid,
        business_domain=business_domain,
        projects_root=projects_root,
        workbook_template=workbook_template,
        repo_url=repo_url,
        branch_name=branch_name,
    )


def record_uid_round_interactive(state: dict[str, object], round_name: str) -> dict[str, object]:
    print("")
    print(f"请在 Trae 里完成 {round_name}，完成后回来填写 sessionId。")
    session_id = prompt_text(f"请填写{round_name}的 Trae Session ID")
    print("正在从 Trae 获取这轮的过程轨迹，并调用 Codex CLI（gpt-5.4-mini）检查，请稍等...")
    codex_review = review_session_trace_with_codex(session_id)
    print("自动建议：")
    print(codex_review["message"])
    task_type = prompt_choice(f"请选择{round_name}的任务类型：", TASK_TYPES, default="0-1代码生成")
    task_completed = prompt_choice(f"请选择{round_name}的任务是否完成：", TASK_COMPLETION_STATUSES)

    suggested_satisfaction = codex_review.get("suggested_satisfaction")
    satisfied = prompt_choice(
        f"请选择{round_name}的产物及过程是否满意：",
        SATISFACTION_STATUSES,
        default=str(suggested_satisfaction) if suggested_satisfaction in SATISFACTION_STATUSES else None,
    )
    product_reason = ""
    process_reason = ""
    if satisfied == "不满意":
        product_reason = prompt_reason_detail("是否有产物不满意原因？")
        process_reason = prompt_reason_detail("是否有过程不满意原因？")

    reason_parts: list[str] = []
    if product_reason:
        reason_parts.append(f"产物不满意：因为{product_reason}")
    if process_reason:
        reason_parts.append(f"过程不满意：因为{process_reason}")
    reason = "；".join(reason_parts)

    warn_if_too_long("不满意原因", reason)

    return {
        "round_name": round_name,
        "session_id": session_id,
        "codex_review": codex_review,
        "task_type": task_type,
        "task_completed": task_completed,
        "satisfied": satisfied,
        "product_reason": product_reason,
        "process_reason": process_reason,
        "reason": reason,
    }


def save_uid_workflow_state(workspace_dir: Path, state: dict[str, object]) -> None:
    state["updated_at"] = dt.datetime.now().astimezone().isoformat()
    write_json(project_state_path(workspace_dir), state)


def default_uid_branch_name(repo_url: str) -> str:
    return "pro"


def cmd_uid_wizard(args: argparse.Namespace) -> int:
    uid = args.uid or prompt_text("请给出项目UID:")
    branch_name = args.branch_name or default_uid_branch_name(args.repo_url)
    state = load_or_initialize_uid_workflow(
        uid=uid,
        projects_root=args.projects_root,
        workbook_template=args.workbook,
        repo_url=args.repo_url,
        branch_name=branch_name,
    )

    workspace_dir = Path(str(state["workspace_dir"]))
    workbook_path = Path(str(state["workbook_path"]))
    branch_name = str(state["branch_name"])
    repo_url = str(state["repo_url"])
    if not args.branch_name and branch_name != default_uid_branch_name(repo_url):
        state["branch_name"] = default_uid_branch_name(repo_url)
        state["project_github_url"] = build_project_github_url(repo_url, str(state["branch_name"]), uid)
        branch_name = str(state["branch_name"])
        save_uid_workflow_state(workspace_dir, state)

    repo_info = ensure_uid_workspace_repo(
        uid=uid,
        workspace_dir=workspace_dir,
        repo_url=repo_url,
        branch_name=branch_name,
    )
    state["code_dir"] = repo_info["code_dir"]
    state["project_github_url"] = build_project_github_url(repo_url, branch_name, uid)
    save_uid_workflow_state(workspace_dir, state)

    project_github_url = str(state["project_github_url"])
    default_modification_scope = "模块内多文件"

    print(f"项目目录：{workspace_dir}")
    print(f"代码目录：{state['code_dir']}")
    print(f"项目数据表：{workbook_path}")
    print(f"业务领域：{state['business_domain']}")
    print(f"Git 分支：{branch_name}")
    print(f"Git 目录地址：{project_github_url}")
    print("请自行确保题目不是过于简单的单文件代码理解任务。")

    if bool(state.get("completed")):
        print("这个 UID 已经标记为完成；如需继续，请手动调整 workflow-state.json。")
        return 0

    if not args.no_open_trae:
        trae_target_dir = Path(str(state["code_dir"]))
        trae_open = open_trae_for_workspace(trae_target_dir, new_window=not args.reuse_window)
        if trae_open["returncode"] != 0:
            raise SystemExit(f"Failed to open Trae workspace: {trae_open['stderr'] or trae_open['stdout']}")
        print(f"已使用代码目录打开 Trae：{trae_target_dir}")

    start_index = int(state.get("current_round_index") or 1)
    if start_index > len(ROUND_NAMES):
        print("这个 UID 已达到 5 轮上限；如需重开，请使用新的 UID。")
        return 0

    for idx in range(start_index, len(ROUND_NAMES) + 1):
        round_name = ROUND_NAMES[idx - 1]
        round_data = record_uid_round_interactive(state, round_name)

        first_round_session_id = str(state.get("first_round_session_id") or "")
        if idx == 1 or not first_round_session_id:
            first_round_session_id = str(round_data["session_id"])
            state["first_round_session_id"] = first_round_session_id

        status = "已完成" if round_data["task_completed"] == "完成了任务" else "进行中"

        workbook_values = {
            "Trae Session ID": str(round_data["session_id"]),
            "任务类型": str(round_data["task_type"]),
            "业务领域": str(state["business_domain"]),
            "修改范围": default_modification_scope,
            "任务是否完成": str(round_data["task_completed"]),
            "产物及过程是否满意": str(round_data["satisfied"]),
            "不满意原因": str(round_data["reason"] or ""),
            "github地址": project_github_url,
            "分支/文件夹": uid,
            "状态": status,
        }
        workbook_values.update(build_first_round_session_fields(first_round_session_id))
        workbook_update = update_workbook_existing_fields(workbook_path, uid, round_name, None, workbook_values)

        round_record = {
            **round_data,
            "business_domain": state["business_domain"],
            "first_round_session_id": first_round_session_id,
            "modification_scope": default_modification_scope,
            "github_url": project_github_url,
            "branch_name": branch_name,
            "status": status,
            "recorded_at": dt.datetime.now().astimezone().isoformat(),
            "workbook_update": workbook_update,
        }
        rounds = list(state.get("rounds") or [])
        rounds.append(round_record)
        state["rounds"] = rounds
        state["current_round_index"] = idx + 1
        state["completed"] = round_data["task_completed"] == "完成了任务"
        save_uid_workflow_state(workspace_dir, state)

        print(f"{round_name} 已记录。")
        print(f"GitHub 地址会统一回填为：{project_github_url}")
        if state["completed"] or idx >= len(ROUND_NAMES):
            publish_result = finalize_uid_project_publish(
                repo_dir=workspace_dir,
                repo_url=repo_url,
                branch_name=branch_name,
                uid=uid,
                rounds=list(state.get("rounds") or []),
            )
            for recorded_round in list(state.get("rounds") or []):
                update_workbook_existing_fields(
                    workbook_path,
                    uid,
                    str(recorded_round.get("round_name") or ""),
                    None,
                    {
                        **build_first_round_session_fields(str(state.get("first_round_session_id") or "")),
                        "github地址": project_github_url,
                        "分支/文件夹": uid,
                    },
                )
                recorded_round["commit_sha"] = publish_result["commit_sha"]
                recorded_round["github_url"] = project_github_url
            state["final_publish"] = publish_result
            save_uid_workflow_state(workspace_dir, state)
            print(f"最终 Git 提交：{publish_result['commit_sha']}")
            print(f"GitHub 地址：{project_github_url}")
            print("流程已完成。")
            return 0

        print(f"{round_name} 标记为未完成，请继续在 Trae 中推进，下一轮将记录为 {ROUND_NAMES[idx]}。")

    return 0


def cmd_start(args: argparse.Namespace) -> int:
    idea_id = args.idea_id or prompt_text("请输入项目ID")
    idea = args.idea or prompt_multiline("请输入你打算做的项目提示词")
    uid = args.uid or idea_id
    round_name = args.round_name or "第一轮"

    project_dir = Path(args.ideas_root).resolve() / idea_id
    preview_dir = project_dir
    if preview_dir.exists() and any(preview_dir.iterdir()):
        raise SystemExit(f"目标目录已存在，请换一个项目ID或点子：{preview_dir}")

    project_workbook_path = project_dir / f"{idea_id}.xlsx"
    workbook_seed = seed_project_workbook(Path(args.workbook), project_workbook_path, idea_id)

    init_result = create_idea_scaffold(
        idea_id=idea_id,
        idea=idea,
        round_name=round_name,
        ideas_root=args.ideas_root,
        repo_url=args.repo_url,
        workbook=str(project_workbook_path),
        uid=uid,
        clone_repo=args.clone_repo,
        folder_name_override=idea_id,
    )
    prepare_result = prepare_round_repo(Path(str(init_result["round_record_path"])))
    round_json = str(init_result["round_record_path"])
    prompt_path = str(init_result["prompt_raw_path"])
    round_payload = load_json(Path(round_json))
    print("已完成初始化。")
    print(f"项目目录：{init_result['idea_dir']}")
    print(f"第一轮提示词：{prompt_path}")
    print(f"项目数据表：{project_workbook_path}")
    print(f"轮次记录：{round_json}")
    print(f"工作目录：{prepare_result['work_folder']}")
    print(f"当前分支：{prepare_result['round_branch']}")
    if args.no_send:
        print("已跳过自动发送。")
        send_parts = [
            "python3 ~/solo.py send-prompt-gui",
            f"--prompt-file {shell_arg(prompt_path)}",
            f"--workbook {shell_arg(str(project_workbook_path))}",
            f"--round-name {shell_arg(round_name)}",
            f"--round-json {shell_arg(round_json)}",
            f"--task-type {shell_arg(str(round_payload.get('task_type') or ''))}",
            f"--business-domain {shell_arg(str(round_payload.get('business_domain') or ''))}",
            f"--modification-scope {shell_arg(str(round_payload.get('modification_scope') or ''))}",
        ]
        if uid:
            send_parts.insert(3, f"--uid {shell_arg(uid)}")
        print("你可以手动发送到 Trae：")
        print(" ".join(send_parts))
        return 0

    print("正在发送第一轮提示词到 Trae，并等待会话结果...")
    payload = run_send_prompt_gui(
        prompt=None,
        prompt_file=prompt_path,
        timeout=args.send_timeout,
        reuse_task=False,
        workbook=str(project_workbook_path),
        uid=uid,
        round_name=round_name,
        row=None,
        round_json=round_json,
        task_type=str(round_payload.get("task_type") or ""),
        business_domain=str(round_payload.get("business_domain") or ""),
        modification_scope=str(round_payload.get("modification_scope") or ""),
    )
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    best_event = payload.get("best_session_event")
    return 0 if isinstance(best_event, dict) and best_event.get("profile_match") else 1


def send_prompt_gui(prompt: str, new_task: bool) -> subprocess.CompletedProcess[str]:
    subprocess.run(["pbcopy"], input=prompt, text=True, check=True)
    script_lines = [
        f'tell application "{TRAE_APP}" to activate',
        "delay 0.8",
        'tell application "System Events"',
    ]
    if new_task:
        script_lines.extend([
            'keystroke "n" using {control down, command down}',
            "delay 1.2",
        ])
    script_lines.extend([
        'keystroke "v" using {command down}',
        "delay 0.4",
        "key code 36",
        "end tell",
    ])
    script = "\n".join(script_lines)
    return run_command(["osascript", "-e", script], timeout=30)


def run_send_prompt_gui(
    *,
    prompt: str | None,
    prompt_file: str | None,
    timeout: int,
    reuse_task: bool,
    workbook: str | None,
    uid: str | None,
    round_name: str | None,
    row: int | None,
    round_json: str | None,
    task_type: str | None,
    business_domain: str | None,
    modification_scope: str | None,
) -> dict[str, object]:
    prompt_value = prompt
    if prompt_file:
        prompt_value = Path(prompt_file).read_text(encoding="utf-8")
    if not prompt_value:
        raise SystemExit("Prompt is empty.")

    since = dt.datetime.now().astimezone()
    gui_result = send_prompt_gui(prompt_value, new_task=not reuse_task)
    events, best_event = wait_for_session_event(since, timeout)
    payload: dict[str, object] = {
        "started_at": since.isoformat(),
        "required_model": REQUIRED_MODEL,
        "required_agent": REQUIRED_AGENT_NAME,
        "gui_send_returncode": gui_result.returncode,
        "gui_send_stdout": gui_result.stdout,
        "gui_send_stderr": gui_result.stderr,
        "best_session_event": maybe_asdict(best_event),
        "best_error_event": maybe_asdict(choose_best_error_event(events)),
        "required_profile_ok": bool(best_event and best_event.profile_match),
        "event_count": len(events),
    }

    if workbook and best_event and best_event.profile_match:
        values = {
            "Trae Session ID": best_event.session_id or "",
            "User Prompt": prompt_value.strip(),
        }
        if task_type:
            values["任务类型"] = task_type
        if business_domain:
            values["业务领域"] = business_domain
        if modification_scope:
            values["修改范围"] = modification_scope
        payload["workbook_update"] = update_workbook(Path(workbook), uid, round_name, row, values)

    if round_json and best_event and best_event.profile_match:
        round_path = Path(round_json)
        round_payload = json.loads(round_path.read_text(encoding="utf-8"))
        round_payload["session_id"] = best_event.session_id
        round_payload["message_id"] = best_event.message_id
        round_payload["trace_id"] = best_event.trace_id
        round_payload["last_prompt"] = prompt_value.strip()
        round_payload["model"] = REQUIRED_MODEL
        round_payload["agent"] = REQUIRED_AGENT_NAME
        round_payload["updated_at"] = dt.datetime.now().astimezone().isoformat()
        round_path.write_text(json.dumps(round_payload, ensure_ascii=False, indent=2), encoding="utf-8")
        payload["round_json_update"] = str(round_path)

    return payload


def cmd_send_prompt_gui(args: argparse.Namespace) -> int:
    payload = run_send_prompt_gui(
        prompt=args.prompt,
        prompt_file=args.prompt_file,
        timeout=args.timeout,
        reuse_task=args.reuse_task,
        workbook=args.workbook,
        uid=args.uid,
        round_name=args.round_name,
        row=args.row,
        round_json=args.round_json,
        task_type=args.task_type,
        business_domain=args.business_domain,
        modification_scope=args.modification_scope,
    )
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    best_event = payload.get("best_session_event")
    return 0 if isinstance(best_event, dict) and best_event.get("profile_match") else 1


def cmd_package_idea(args: argparse.Namespace) -> int:
    round_payload, metadata, idea_dir = load_round_context(Path(args.round_json))
    folder_name = str(metadata.get("folder_name") or build_folder_name(str(round_payload.get("idea_id") or "idea"), str(round_payload.get("prompt") or "idea")))
    source_dir = Path(str(round_payload.get("work_folder") or idea_dir / folder_name))
    if not source_dir.exists():
        raise SystemExit(f"Source directory does not exist: {source_dir}")

    output_dir = Path(args.output_dir).resolve() if args.output_dir else idea_dir / "artifacts"
    output_dir.mkdir(parents=True, exist_ok=True)
    archive_base = output_dir / folder_name
    archive_path = shutil.make_archive(str(archive_base), "zip", root_dir=str(source_dir.parent), base_dir=source_dir.name)

    round_payload["attachment"] = archive_path
    round_payload["packaged_at"] = dt.datetime.now().astimezone().isoformat()
    write_json(Path(args.round_json), round_payload)

    payload = {
        "round_json": args.round_json,
        "source_dir": str(source_dir),
        "archive_path": archive_path,
        "folder_name": folder_name,
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Solo Coder helpers for Trae and the workbook.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    start = subparsers.add_parser("start", help="Interactive start flow: ask for project id and prompt, then initialize the idea and repo.")
    start.add_argument("--idea-id")
    start.add_argument("--idea")
    start.add_argument("--uid")
    start.add_argument("--round-name")
    start.add_argument("--ideas-root", default=str(DEFAULT_START_ROOT))
    start.add_argument("--repo-url", default=DEFAULT_REPO_URL)
    start.add_argument("--workbook", default=str(DEFAULT_TEMPLATE_WORKBOOK))
    start.add_argument("--clone-repo", action="store_true")
    start.add_argument("--no-send", action="store_true", help="Only initialize locally; do not auto-send the first-round prompt to Trae.")
    start.add_argument("--send-timeout", type=int, default=60, help="Seconds to wait for the first-round Trae session result.")

    uid_wizard = subparsers.add_parser("uid-wizard", help="Interactive UID workflow: initialize a UID project, open Trae, and record up to five rounds.")
    uid_wizard.add_argument("--uid")
    uid_wizard.add_argument("--projects-root", default=str(DEFAULT_START_ROOT))
    uid_wizard.add_argument("--repo-url", default=DEFAULT_REPO_URL)
    uid_wizard.add_argument("--workbook", default=str(DEFAULT_TEMPLATE_WORKBOOK))
    uid_wizard.add_argument("--branch-name")
    uid_wizard.add_argument("--no-open-trae", action="store_true")
    uid_wizard.add_argument("--reuse-window", action="store_true")

    init_idea = subparsers.add_parser("init-idea", help="Create one idea folder with prompt, PR, and round skeleton files.")
    init_idea.add_argument("--idea-id", required=True)
    init_idea.add_argument("--idea", required=True)
    init_idea.add_argument("--round-name", default="第一轮")
    init_idea.add_argument("--ideas-root", default=str(DEFAULT_IDEAS_ROOT))
    init_idea.add_argument("--repo-url", default=DEFAULT_REPO_URL)
    init_idea.add_argument("--workbook", default="【中强智联】Solo Coder_数据表_表格.xlsx")
    init_idea.add_argument("--uid")
    init_idea.add_argument("--clone-repo", action="store_true")

    prepare_repo = subparsers.add_parser("prepare-repo", help="Prepare the local repo copy, base branch, round branch, and compare URL.")
    prepare_repo.add_argument("--round-json", required=True)
    prepare_repo.add_argument("--repo-dir")

    probe = subparsers.add_parser("probe-chat", help="Diagnostic-only CLI probe. The real workflow should prefer send-prompt-gui.")
    probe.add_argument("--prompt", required=True)
    probe.add_argument("--mode", default="ask", choices=["ask", "edit", "agent"])
    probe.add_argument("--cwd", default=".")
    probe.add_argument("--timeout", type=int, default=20)
    probe.add_argument("--reuse-window", action="store_true")
    probe.add_argument("--new-window", action="store_true")

    gui_send = subparsers.add_parser("send-prompt-gui", help="Activate Trae, optionally open a new task, paste a prompt, send it, and capture the session id.")
    gui_send.add_argument("--prompt")
    gui_send.add_argument("--prompt-file")
    gui_send.add_argument("--timeout", type=int, default=30)
    gui_send.add_argument("--reuse-task", action="store_true")
    gui_send.add_argument("--workbook")
    gui_send.add_argument("--uid")
    gui_send.add_argument("--round-name")
    gui_send.add_argument("--row", type=int)
    gui_send.add_argument("--round-json")
    gui_send.add_argument("--task-type")
    gui_send.add_argument("--business-domain")
    gui_send.add_argument("--modification-scope")

    review_round = subparsers.add_parser("review-round", help="Collect git review context for the current round without running project scripts.")
    review_round.add_argument("--round-json", required=True)
    review_round.add_argument("--repo-dir")

    next_prompt = subparsers.add_parser("make-next-prompt", help="Generate the next-round prompt and round json from issue bullets.")
    next_prompt.add_argument("--round-json", required=True)
    next_prompt.add_argument("--issue", action="append")
    next_prompt.add_argument("--issues-file")
    next_prompt.add_argument("--next-round-name")
    next_prompt.add_argument("--yes", action="store_true", help="Skip the interactive confirmation and save directly.")
    next_prompt.add_argument("--print-only", action="store_true", help="Only print the next prompt preview; do not save files.")

    record_review = subparsers.add_parser("record-review", help="Record review verdicts into round json and optionally backfill the workbook.")
    record_review.add_argument("--round-json", required=True)
    record_review.add_argument("--task-completed", required=True, choices=["未完成任务", "完成了任务"])
    record_review.add_argument("--satisfied", required=True, choices=["满意", "不满意"])
    record_review.add_argument("--reason")
    record_review.add_argument("--product-reason", action="append", default=[], help="可多次传入。常用值：" + "、".join(PRODUCT_REASON_TEMPLATES))
    record_review.add_argument("--process-reason", action="append", default=[], help="可多次传入。常用值：" + "、".join(PROCESS_REASON_TEMPLATES))
    record_review.add_argument("--workbook")
    record_review.add_argument("--uid")
    record_review.add_argument("--round-name")
    record_review.add_argument("--row", type=int)
    record_review.add_argument("--github-url")
    record_review.add_argument("--branch-folder")
    record_review.add_argument("--attachment")
    record_review.add_argument("--status")

    publish_round = subparsers.add_parser("publish-round", help="Commit one round; git push is skipped by default.")
    publish_round.add_argument("--round-json", required=True)
    publish_round.add_argument("--repo-dir")
    publish_round.add_argument("--commit-message")
    publish_round.add_argument("--pr-body-file")
    publish_round.add_argument("--dry-run", action="store_true")
    publish_round.add_argument("--push", action="store_true", help="Actually push git branches instead of the default skip.")
    publish_round.add_argument("--skip-pr", action="store_true")
    publish_round.add_argument("--workbook")
    publish_round.add_argument("--uid")
    publish_round.add_argument("--round-name")
    publish_round.add_argument("--row", type=int)
    publish_round.add_argument("--branch-folder")
    publish_round.add_argument("--status")

    extract = subparsers.add_parser("extract-session", help="Extract session-related events from renderer logs.")
    extract.add_argument("--since", required=True, help="ISO time or unix epoch seconds.")

    package_idea = subparsers.add_parser("package-idea", help="Zip the generated code folder using the same folder name.")
    package_idea.add_argument("--round-json", required=True)
    package_idea.add_argument("--output-dir")

    list_rows = subparsers.add_parser("list-rows", help="List workbook rows that matter for round recording.")
    list_rows.add_argument("--workbook", required=True)

    update = subparsers.add_parser("update-row", help="Update workbook cells by row, UID, and round.")
    update.add_argument("--workbook", required=True)
    update.add_argument("--row", type=int)
    update.add_argument("--uid")
    update.add_argument("--round-name")
    update.add_argument("--set", dest="set_values", action="append", default=[], required=True)

    return parser


def run_probe_chat(prompt: str, mode: str, cwd: Path, reuse_window: bool, new_window: bool, timeout: int) -> dict[str, object]:
    baseline = baseline_for_logs(renderer_logs())
    cmd = [str(TRAE_CLI), "chat", "-m", mode]
    if reuse_window:
        cmd.append("-r")
    if new_window:
        cmd.append("-n")
    cmd.append(prompt)

    result: dict[str, object] = {
        "command": cmd,
        "cwd": str(cwd),
        "started_at": dt.datetime.now().astimezone().isoformat(),
        "workflow_warning": f"CLI probe cannot guarantee {REQUIRED_AGENT_NAME} + {REQUIRED_MODEL}; use send-prompt-gui for the real workflow.",
    }

    try:
        completed = subprocess.run(
            cmd,
            cwd=str(cwd),
            capture_output=True,
            text=True,
            timeout=20,
            check=False,
        )
        result["cli_returncode"] = completed.returncode
        result["cli_stdout"] = completed.stdout
        result["cli_stderr"] = completed.stderr
    except subprocess.TimeoutExpired as exc:
        result["cli_timeout"] = True
        result["cli_stdout"] = exc.stdout
        result["cli_stderr"] = exc.stderr

    deadline = time.time() + timeout
    seen_events: list[SessionEvent] = []
    while time.time() < deadline:
        events = scan_events_since(baseline)
        if events:
            seen_events = events
            best = choose_best_event(events)
            if best and (best.session_id or best.error):
                break
        time.sleep(1)

    latest_dirs = log_dirs()
    if latest_dirs:
        result["latest_log_dir"] = str(latest_dirs[-1])

    result["events"] = [asdict(event) for event in seen_events]
    result["best_event"] = maybe_asdict(choose_best_event(seen_events))
    result["best_session_event"] = maybe_asdict(choose_best_session_event(seen_events))
    result["best_error_event"] = maybe_asdict(choose_best_error_event(seen_events))
    result["finished_at"] = dt.datetime.now().astimezone().isoformat()
    return result


def main(argv: list[str] | None = None) -> int:
    argv = list(argv) if argv is not None else sys.argv[1:]
    known_commands = {
        "start",
        "uid-wizard",
        "init-idea",
        "prepare-repo",
        "probe-chat",
        "send-prompt-gui",
        "review-round",
        "make-next-prompt",
        "record-review",
        "publish-round",
        "extract-session",
        "package-idea",
        "list-rows",
        "update-row",
    }
    if not argv:
        argv = ["uid-wizard"]
    elif argv[0] not in known_commands:
        argv = ["uid-wizard", *argv]

    parser = build_parser()
    args = parser.parse_args(argv)

    if args.command == "start":
        return cmd_start(args)

    if args.command == "uid-wizard":
        return cmd_uid_wizard(args)

    if args.command == "init-idea":
        return cmd_init_idea(args)

    if args.command == "prepare-repo":
        return cmd_prepare_repo(args)

    if args.command == "probe-chat":
        result = run_probe_chat(
            prompt=args.prompt,
            mode=args.mode,
            cwd=Path(args.cwd).resolve(),
            reuse_window=args.reuse_window,
            new_window=args.new_window,
            timeout=args.timeout,
        )
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
        return 0

    if args.command == "send-prompt-gui":
        if not args.prompt and not args.prompt_file:
            raise SystemExit("Pass --prompt or --prompt-file.")
        return cmd_send_prompt_gui(args)

    if args.command == "review-round":
        return cmd_review_round(args)

    if args.command == "make-next-prompt":
        return cmd_make_next_prompt(args)

    if args.command == "record-review":
        return cmd_record_review(args)

    if args.command == "publish-round":
        return cmd_publish_round(args)

    if args.command == "extract-session":
        since = parse_since(args.since)
        events = extract_recent_events(since)
        payload = {
            "since": since.isoformat(),
            "required_model": REQUIRED_MODEL,
            "required_agent": REQUIRED_AGENT_NAME,
            "events": [asdict(event) for event in events],
            "best_event": maybe_asdict(choose_best_event(events)),
            "best_session_event": maybe_asdict(choose_best_session_event(events)),
            "best_error_event": maybe_asdict(choose_best_error_event(events)),
        }
        print(json.dumps(payload, ensure_ascii=False, indent=2))
        return 0

    if args.command == "package-idea":
        return cmd_package_idea(args)

    if args.command == "list-rows":
        return cmd_list_rows(args)

    if args.command == "update-row":
        return cmd_update_row(args)

    parser.error(f"Unsupported command: {args.command}")
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
