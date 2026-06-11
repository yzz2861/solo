#!/usr/bin/env python3
from __future__ import annotations

import argparse
import fcntl
import json
import os
import re
import subprocess
import sys
import time
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urlparse

from manual_lib import (
    ROOT,
    SCRIPT_AUTO_DIR,
    RUN_LOG_ROOT,
    RUNS_DIR,
    default_max_rounds,
    load_runner,
    load_state,
    next_round_name,
    now,
    round_name_to_number,
    update_started_entry,
)


RESULT_DIR = ROOT / "result"
ROUND_SKIP_STATUSES = {"second_started", "more_started", "captured_waiting_qc", "qc_done"}
MORE_LOCK_PATH = RUN_LOG_ROOT / "script_auto" / ".manual_second.lock"
MORE_LOCK_TIMEOUT_SECONDS = 120
LEGACY_MORE_LOCK_PATHS = (
    SCRIPT_AUTO_DIR / ".manual_more.lock",
    SCRIPT_AUTO_DIR / ".manual_second.lock",
)
TRUTHY_ENV_VALUES = {"1", "true", "yes", "y", "on", "open", "enabled"}
FALSY_ENV_VALUES = {"0", "false", "no", "n", "off", "closed", "disabled"}


@dataclass
class MoreRoundCandidate:
    idea_id: str
    workbook: Path
    row: dict[str, str]
    reason: str
    product_reason: str
    process_reason: str
    source_round: str
    target_round: str
    target_round_number: int


@dataclass
class MoreRoundSkip:
    idea_id: str
    workbook: Path
    reason: str
    source_round: str
    target_round: str
    skip_reason: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="批量启动上一轮未完成且不满意项目的下一轮 Trae 修复。")
    parser.add_argument("count", type=int, help="本次要启动多少个符合条件的项目。")
    parser.add_argument("--idea-id", action="append", default=[], help="只处理指定项目；可重复传入，用于重试某一个项目。")
    parser.add_argument("--no-cli", action="store_true", help="只打开 Trae，不调用 send-prompt-gui。")
    parser.add_argument("--dry-run", action="store_true", help="只列出会启动的项目，不打开 Trae。")
    parser.add_argument("--retry-started", action="store_true", help="包含已经发起过下一轮但还没写入结果表的项目。")
    parser.add_argument("--include-open", "--force-open", dest="include_open", action="store_true", help="包含已经在 Trae 打开的项目；默认跳过，避免重复续跑。")
    parser.add_argument("--accept-pending", action="store_true", help="发送前如果轻量接受未命中，再尝试截图/视觉兜底接受待审查文件。")
    parser.add_argument("--no-accept-pending", action="store_true", help="发送前不处理 Trae 待审查文件。")
    parser.add_argument("--max-rounds", type=int, default=default_max_rounds(), help="允许自动续跑到第几轮；默认 5，可用 SOLO_MAX_ROUNDS 调整。")
    return parser.parse_args()


def cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def env_switch_enabled(name: str, *, default: bool) -> bool:
    text = os.environ.get(name)
    if text is None:
        return default
    normalized = text.strip().lower()
    if normalized in TRUTHY_ENV_VALUES:
        return True
    if normalized in FALSY_ENV_VALUES:
        return False
    return default


def more_min_id_floor() -> int | None:
    if not env_switch_enabled("SOLO_MORE_MIN_ID_ENABLED", default=False):
        return None
    text = os.environ.get("SOLO_MORE_MIN_ID", "").strip()
    if not text:
        return None
    try:
        floor = int(text)
    except ValueError as exc:
        raise SystemExit(f"SOLO_MORE_MIN_ID 必须是整数，当前值: {text}") from exc
    if floor <= 0:
        raise SystemExit(f"SOLO_MORE_MIN_ID 必须大于 0，当前值: {text}")
    return floor


def idea_id_number(idea_id: str) -> int | None:
    match = re.search(r"(\d+)$", idea_id.strip())
    return int(match.group(1)) if match else None


def filter_candidates_by_min_id(candidates: list[MoreRoundCandidate], min_id: int | None) -> tuple[list[MoreRoundCandidate], list[MoreRoundCandidate]]:
    if min_id is None:
        return candidates, []
    kept: list[MoreRoundCandidate] = []
    skipped: list[MoreRoundCandidate] = []
    for candidate in candidates:
        number = idea_id_number(candidate.idea_id)
        if number is not None and number >= min_id:
            kept.append(candidate)
        else:
            skipped.append(candidate)
    return kept, skipped


@contextmanager
def more_script_lock():
    paths = [MORE_LOCK_PATH]
    paths.extend(path for path in LEGACY_MORE_LOCK_PATHS if path.exists() and path != MORE_LOCK_PATH)
    handles = []
    started = time.time()
    try:
        while True:
            try:
                for path in paths:
                    path.parent.mkdir(parents=True, exist_ok=True)
                    handle = path.open("a+", encoding="utf-8")
                    fcntl.flock(handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
                    handles.append(handle)
                break
            except BlockingIOError:
                for handle in reversed(handles):
                    fcntl.flock(handle.fileno(), fcntl.LOCK_UN)
                    handle.close()
                handles = []
                if time.time() - started >= MORE_LOCK_TIMEOUT_SECONDS:
                    raise SystemExit(f"等待 more.sh 批次锁超时 {MORE_LOCK_TIMEOUT_SECONDS}s: {MORE_LOCK_PATH}")
                time.sleep(1)
        yield
    finally:
        for handle in reversed(handles):
            fcntl.flock(handle.fileno(), fcntl.LOCK_UN)
            handle.close()


def read_workbook_rows(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise SystemExit(f"读取 xlsx 需要 openpyxl: {exc}") from exc

    wb = load_workbook(path, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        headers = [cell_text(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
        rows: list[dict[str, str]] = []
        for row_idx in range(2, ws.max_row + 1):
            values = [cell_text(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)]
            if not any(values):
                continue
            rows.append({header: values[idx] for idx, header in enumerate(headers) if header})
        return rows
    finally:
        wb.close()


def started_round_pairs() -> set[tuple[str, str]]:
    started: set[tuple[str, str]] = set()
    entries = load_state().get("entries", [])
    if not isinstance(entries, list):
        return started
    for item in entries:
        if not isinstance(item, dict):
            continue
        idea_id = cell_text(item.get("idea_id"))
        status = cell_text(item.get("status"))
        round_name = cell_text(item.get("round_name"))
        if idea_id and round_name and status in ROUND_SKIP_STATUSES:
            started.add((idea_id, round_name))
    return started


def _append_reason_segment(parts: list[str], value: str) -> None:
    text = value.strip()
    if text:
        parts.append(text)


def split_unsatisfied_reasons(reason: str) -> tuple[str, str]:
    text = reason.strip()
    if not text:
        return "", ""

    marker_pattern = re.compile(r"(产物及过程不满意|产物不满意|过程不满意|不满意原因)[：:]")
    matches = list(marker_pattern.finditer(text))
    if not matches:
        return text, ""

    product_parts: list[str] = []
    process_parts: list[str] = []
    prefix = text[: matches[0].start()].strip()
    if prefix:
        product_parts.append(prefix)

    for index, match in enumerate(matches):
        label = match.group(1)
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        value = text[start:end].strip()
        if label == "过程不满意":
            _append_reason_segment(process_parts, value)
        else:
            _append_reason_segment(product_parts, value)

    return "\n".join(product_parts).strip(), "\n".join(process_parts).strip()


def filter_product_reason_candidates(candidates: list[MoreRoundCandidate]) -> tuple[list[MoreRoundCandidate], list[MoreRoundSkip]]:
    kept: list[MoreRoundCandidate] = []
    skipped: list[MoreRoundSkip] = []
    for candidate in candidates:
        if candidate.product_reason:
            kept.append(candidate)
            continue
        skipped.append(
            MoreRoundSkip(
                idea_id=candidate.idea_id,
                workbook=candidate.workbook,
                reason=candidate.reason,
                source_round=candidate.source_round,
                target_round=candidate.target_round,
                skip_reason="仅有过程不满意原因，没有可用于下一轮修改的产物问题",
            )
        )
    return kept, skipped


def find_more_round_candidates(skip_rounds: set[tuple[str, str]] | None = None, *, max_rounds: int) -> list[MoreRoundCandidate]:
    skip_rounds = skip_rounds or set()
    candidates: list[MoreRoundCandidate] = []
    seen_idea_ids: set[str] = set()
    for workbook in sorted(RESULT_DIR.glob("*.xlsx"), key=lambda p: p.name):
        if workbook.name.startswith("~$"):
            continue
        if workbook.stem in seen_idea_ids:
            continue
        rows = read_workbook_rows(workbook)
        if not rows:
            continue
        row = rows[-1]
        if cell_text(row.get("任务是否完成")) != "未完成任务":
            continue
        if cell_text(row.get("产物及过程是否满意")) != "不满意":
            continue
        round_name = cell_text(row.get("轮次")) or "第一轮"
        round_number = round_name_to_number(round_name)
        if round_number >= max_rounds:
            continue
        reason = cell_text(row.get("不满意原因"))
        if not reason:
            continue
        product_reason, process_reason = split_unsatisfied_reasons(reason)
        target_round = next_round_name(round_name)
        if (workbook.stem, target_round) in skip_rounds:
            continue
        candidates.append(
            MoreRoundCandidate(
                idea_id=workbook.stem,
                workbook=workbook,
                row=row,
                reason=reason,
                product_reason=product_reason,
                process_reason=process_reason,
                source_round=round_name,
                target_round=target_round,
                target_round_number=round_number + 1,
            )
        )
        seen_idea_ids.add(workbook.stem)
    return candidates


def file_uri_path(value: Any) -> Path | None:
    text = cell_text(value)
    if not text:
        return None
    if text.startswith("file://"):
        parsed = urlparse(text)
        return Path(unquote(parsed.path))
    return Path(text)


def idea_id_from_workspace_path(value: Any) -> str:
    path = file_uri_path(value)
    if path is None:
        return ""
    try:
        workspace_root = (ROOT / "workspaces").resolve()
        resolved = path.expanduser().resolve()
        if not resolved.is_relative_to(workspace_root):
            return ""
    except Exception:
        pass
    name = path.name.strip()
    return name if re.fullmatch(r"[A-Za-z]{1,8}\d{3,}", name) else ""


def trae_open_workspace_ids_from_storage(runner: Any) -> tuple[set[str], list[str]]:
    try:
        if not runner.trae_main_pid():
            return set(), []
    except Exception:
        return set(), []
    storage_path = Path.home() / "Library" / "Application Support" / "Trae CN" / "User" / "globalStorage" / "storage.json"
    try:
        storage = json.loads(storage_path.read_text(encoding="utf-8"))
    except Exception:
        return set(), []
    windows_state = storage.get("windowsState") if isinstance(storage, dict) else {}
    if not isinstance(windows_state, dict):
        return set(), []
    window_items: list[dict[str, Any]] = []
    last_active = windows_state.get("lastActiveWindow")
    if isinstance(last_active, dict):
        window_items.append(last_active)
    opened = windows_state.get("openedWindows")
    if isinstance(opened, list):
        window_items.extend(item for item in opened if isinstance(item, dict))
    open_ids: set[str] = set()
    details: list[str] = []
    for item in window_items:
        idea_id = idea_id_from_workspace_path(item.get("folder"))
        if idea_id:
            open_ids.add(idea_id)
            details.append(idea_id)
    return open_ids, details


def open_idea_ids_from_trae(runner: Any) -> tuple[set[str], list[str]]:
    try:
        titles = runner.trae_window_titles()
    except Exception:
        titles = []
    open_ids: set[str] = set()
    for title in titles:
        for match in re.finditer(r"(?<![A-Za-z0-9_])([A-Za-z]{1,8}\d{3,})(?![A-Za-z0-9_])", title):
            open_ids.add(match.group(1))
    storage_ids, storage_details = trae_open_workspace_ids_from_storage(runner)
    open_ids.update(storage_ids)
    details = [f"title:{title}" for title in titles]
    details.extend(f"workspace:{idea_id}" for idea_id in storage_details)
    return open_ids, details


def filter_open_candidates(candidates: list[MoreRoundCandidate], open_idea_ids: set[str]) -> tuple[list[MoreRoundCandidate], list[MoreRoundCandidate]]:
    if not open_idea_ids:
        return candidates, []
    kept: list[MoreRoundCandidate] = []
    skipped: list[MoreRoundCandidate] = []
    for candidate in candidates:
        if candidate.idea_id in open_idea_ids:
            skipped.append(candidate)
        else:
            kept.append(candidate)
    return kept, skipped


def normalize_for_copy_check(text: str) -> str:
    return re.sub(r"\s+", "", text.strip())


def build_llm_prompt_summary_request(candidate: MoreRoundCandidate) -> str:
    title = cell_text(candidate.row.get("title"))
    business_domain = cell_text(candidate.row.get("业务领域"))
    return f"""请把上一轮的“产物不满意”原因改写成下一轮要发给 Trae 的修改提示词。

项目 ID: {candidate.idea_id}
项目标题: {title or "未填写"}
业务领域: {business_domain or "未填写"}
上一轮: {candidate.source_round}
本轮: {candidate.target_round}

只允许参考下面这些产物问题，不要加入过程不满意、质检过程、验证过程、总结质量等内容：
{candidate.product_reason}

写作要求：
- 不要直接复制上面的原文，要像人在交代返工任务一样自然。
- 不要写“作为 AI”“根据质检报告”“以下是”“请注意”等模板化开头。
- 如果问题很多，只挑 2-3 个最关键的产物问题。
- 说清楚这轮要改什么，保留已有功能。
- 最后自然提醒完成后要保证项目可安装、可运行、可验证。
- 只输出最终提示词正文，不要解释。"""


def generate_more_round_prompt(runner: Any, candidate: MoreRoundCandidate, run_dir: Path, workspace: Path) -> tuple[str, dict[str, Any]]:
    request_text = build_llm_prompt_summary_request(candidate)
    request_path = run_dir / "more_round_llm_request.txt"
    output_path = run_dir / "more_round_llm_prompt.txt"
    stdout_path = run_dir / "more_round_llm_stdout.log"
    stderr_path = run_dir / "more_round_llm_stderr.log"
    request_path.write_text(request_text, encoding="utf-8")
    cmd = [
        "codex",
        "exec",
        "--model",
        str(getattr(runner, "CODEX_QA_MODEL", "gpt-5.5")),
        "--config",
        f'model_reasoning_effort="{getattr(runner, "CODEX_QA_REASONING_EFFORT", "high")}"',
        "--cd",
        str(workspace),
        "--skip-git-repo-check",
        "--sandbox",
        "read-only",
        "--output-last-message",
        str(output_path),
        request_text,
    ]
    proc = subprocess.run(cmd, cwd=str(workspace), text=True, capture_output=True)
    stdout_path.write_text(proc.stdout or "", encoding="utf-8")
    stderr_path.write_text(proc.stderr or "", encoding="utf-8")
    prompt = output_path.read_text(encoding="utf-8", errors="ignore").strip() if output_path.exists() else ""
    result = {
        "returncode": proc.returncode,
        "request_path": str(request_path),
        "output_path": str(output_path),
        "stdout_path": str(stdout_path),
        "stderr_path": str(stderr_path),
    }
    if proc.returncode != 0:
        raise RuntimeError(f"LLM 生成下一轮提示词失败，退出码 {proc.returncode}，stderr: {(proc.stderr or '').strip()[:500]}")
    if not prompt:
        raise RuntimeError(f"LLM 未生成下一轮提示词: {output_path}")
    if normalize_for_copy_check(prompt) == normalize_for_copy_check(candidate.product_reason):
        raise RuntimeError("LLM 输出与上一轮产物不满意原因基本相同，拒绝直接 copy 作为下一轮提示词")
    normalized_process_reason = normalize_for_copy_check(candidate.process_reason)
    if len(normalized_process_reason) >= 20 and normalized_process_reason in normalize_for_copy_check(prompt):
        raise RuntimeError("LLM 输出包含过程不满意原因，拒绝作为下一轮提示词")
    return prompt, result


def main() -> int:
    args = parse_args()
    if args.count <= 0:
        raise SystemExit("count 必须大于 0，例如: script/more.sh 4")
    if args.max_rounds < 2:
        raise SystemExit("--max-rounds 至少为 2。")

    with more_script_lock():
        runner = None if args.include_open else load_runner()
        open_idea_ids: set[str] = set()
        open_titles: list[str] = []
        if runner is not None:
            open_idea_ids, open_titles = open_idea_ids_from_trae(runner)
        skip_rounds = set() if args.retry_started else started_round_pairs()
        candidates = find_more_round_candidates(skip_rounds, max_rounds=args.max_rounds)
        if args.idea_id:
            allowed_idea_ids = {cell_text(idea_id) for idea_id in args.idea_id if cell_text(idea_id)}
            candidates = [candidate for candidate in candidates if candidate.idea_id in allowed_idea_ids]
        min_id = more_min_id_floor()
        candidates, skipped_min_id = filter_candidates_by_min_id(candidates, min_id)
        candidates, skipped_process_only = filter_product_reason_candidates(candidates)
        candidates, skipped_open = filter_open_candidates(candidates, open_idea_ids)
        selected = candidates[: args.count]
        if not selected:
            print(
                f"没有找到满足条件的项目：结果表最后一轮必须同时是 任务是否完成=未完成任务、产物及过程是否满意=不满意，"
                f"并且不满意原因非空；当前最多续跑到第 {args.max_rounds} 轮。"
            )
            if min_id is not None:
                print(f"当前 more.sh ID 下限开关已打开，只取数字部分 >= {min_id} 的项目。")
            if skipped_min_id:
                print("已跳过低于 ID 下限的项目: " + ", ".join(item.idea_id for item in skipped_min_id))
            if skipped_process_only:
                print("已跳过仅过程不满意的项目: " + ", ".join(item.idea_id for item in skipped_process_only))
            if skipped_open:
                print("已跳过 Trae 已打开的项目: " + ", ".join(item.idea_id for item in skipped_open))
            return 0

        if args.dry_run:
            if min_id is not None:
                print(f"当前 more.sh ID 下限开关已打开，只取数字部分 >= {min_id} 的项目。")
            if skipped_min_id:
                print("已跳过低于 ID 下限的项目: " + ", ".join(item.idea_id for item in skipped_min_id))
            if skipped_process_only:
                print("已跳过仅过程不满意的项目:")
                for item in skipped_process_only:
                    print(f"- {item.idea_id}: {item.source_round} -> {item.target_round}: {item.skip_reason}")
            if skipped_open:
                print("已跳过 Trae 已打开的项目: " + ", ".join(item.idea_id for item in skipped_open))
                if open_titles:
                    print("当前 Trae 窗口: " + " | ".join(open_titles))
            print("将启动这些下一轮项目:")
            for item in selected:
                print(f"- {item.idea_id}: {item.source_round} -> {item.target_round}: {item.product_reason[:120]}")
            return 0

        if runner is None:
            runner = load_runner()
        if skipped_process_only:
            print("已跳过仅过程不满意的项目: " + ", ".join(item.idea_id for item in skipped_process_only))
        if skipped_open:
            print("已跳过 Trae 已打开的项目: " + ", ".join(item.idea_id for item in skipped_open))
        failures: list[str] = []
        started: list[MoreRoundCandidate] = []
        for offset, candidate in enumerate(selected, start=1):
            started_ok, failure = start_more_round_candidate(args, runner, offset, selected, candidate)
            if started_ok:
                started.append(candidate)
            if failure:
                failures.append(failure)

    if failures:
        print("失败项目:", "; ".join(failures), file=sys.stderr)
        return 1
    if started:
        print("本次续跑启动完成: " + ", ".join(f"{item.idea_id}({item.target_round})" for item in started))
    else:
        print("本次没有启动新的续跑项目。")
    return 0


def start_more_round_candidate(args: argparse.Namespace, runner: Any, offset: int, selected: list[MoreRoundCandidate], candidate: MoreRoundCandidate) -> tuple[bool, str | None]:
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    run_dir = RUNS_DIR / f"{timestamp}-{candidate.idea_id}-round{candidate.target_round_number}"
    workspace = ROOT / "workspaces" / candidate.idea_id
    run_dir.mkdir(parents=True, exist_ok=True)
    if not workspace.exists():
        print(f"[{offset}/{len(selected)}] 跳过 {candidate.idea_id}: 工作区不存在 {workspace}", file=sys.stderr)
        return False, f"{candidate.idea_id}: 工作区不存在 {workspace}"

    trace = runner.Trace(run_dir)
    (run_dir / "more_round_reason.txt").write_text(candidate.reason + "\n", encoding="utf-8")
    (run_dir / "more_round_product_reason.txt").write_text(candidate.product_reason + "\n", encoding="utf-8")
    (run_dir / "more_round_process_reason.txt").write_text(candidate.process_reason + "\n", encoding="utf-8")
    try:
        prompt, llm_result = generate_more_round_prompt(runner, candidate, run_dir, workspace)
        trace.write("manual_more_llm_prompt_generated", **llm_result, prompt_chars=len(prompt), product_reason_chars=len(candidate.product_reason))
    except Exception as exc:
        message = f"LLM 生成下一轮提示词失败，已跳过 {candidate.idea_id}: {exc}"
        trace.write("manual_more_llm_prompt_failed", idea_id=candidate.idea_id, target_round=candidate.target_round, error=str(exc))
        update_started_entry(
            candidate.idea_id,
            status="more_skipped_llm_failed",
            round_name=candidate.target_round,
            source_round=candidate.source_round,
            run_dir=str(run_dir),
            workspace=str(workspace),
            prompt_source=f"result:{candidate.source_round}_llm_product_unsatisfied_summary",
            source_workbook=str(candidate.workbook),
            error=str(exc),
            more_started_at=now(),
        )
        print(f"[{offset}/{len(selected)}] {message}", file=sys.stderr)
        return False, None

    (run_dir / "prompt.txt").write_text(prompt, encoding="utf-8")
    (run_dir / "idea.json").write_text(
        json.dumps(
            {
                "idea_id": candidate.idea_id,
                "title": candidate.row.get("title", ""),
                "fields": {
                    "业务领域": candidate.row.get("业务领域", ""),
                    "修改范围": candidate.row.get("修改范围", ""),
                },
                "prompt_source": f"result:{candidate.source_round}_llm_product_unsatisfied_summary",
                "round_prompt_source": f"result:{candidate.source_round}_llm_product_unsatisfied_summary",
                "source_workbook": str(candidate.workbook),
                "source_round": candidate.source_round,
                "source_unsatisfied_reason_path": str(run_dir / "more_round_reason.txt"),
                "source_product_reason_path": str(run_dir / "more_round_product_reason.txt"),
                "source_process_reason_path": str(run_dir / "more_round_process_reason.txt"),
                "llm_prompt_path": str(run_dir / "more_round_llm_prompt.txt"),
                "workspace": str(workspace),
                "manual": True,
                "round_name": candidate.target_round,
                "round_number": candidate.target_round_number,
                "max_rounds": args.max_rounds,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    trace.write(
        "manual_more_selected",
        idea_id=candidate.idea_id,
        source_workbook=str(candidate.workbook),
        source_round=candidate.source_round,
        target_round=candidate.target_round,
        reason_chars=len(candidate.reason),
        product_reason_chars=len(candidate.product_reason),
        process_reason_chars=len(candidate.process_reason),
        workspace=str(workspace),
    )

    try:
        update_started_entry(
            candidate.idea_id,
            status="more_starting",
            round_name=candidate.target_round,
            source_round=candidate.source_round,
            run_dir=str(run_dir),
            workspace=str(workspace),
            prompt_source=f"result:{candidate.source_round}_llm_product_unsatisfied_summary",
            source_workbook=str(candidate.workbook),
            more_started_at=now(),
        )
        with runner.trae_window_lock(trace, "manual_more_prompt_send"):
            trace.write("manual_more_reuse_existing_task", idea_id=candidate.idea_id, target_round=candidate.target_round)
            runner.open_trae_project(trace, workspace)
            if not args.no_cli:
                accepted = False
                if args.accept_pending and not args.no_accept_pending:
                    accepted = runner.accept_visible_pending_review_if_present(trace, run_dir, workspace, candidate.idea_id)
                    trace.write("manual_more_pending_review_check", accepted=accepted, aggressive_fallback=True)
                    if not accepted:
                        runner.accept_all_changes_if_present(trace, run_dir, workspace, candidate.idea_id)
                else:
                    trace.write("manual_more_pending_review_check_skipped", reason="visual_accept_requires_explicit_accept_pending")
                payload = runner.send_prompt_with_solo_helper(trace, prompt, workspace, reuse_task=True)
                runner.save_trae_session_metadata(trace, payload)
            else:
                trace.write("trae_project_opened_without_cli", cwd=str(workspace))
            time.sleep(4)
        update_started_entry(
            candidate.idea_id,
            status="more_started",
            round_name=candidate.target_round,
            source_round=candidate.source_round,
            run_dir=str(run_dir),
            workspace=str(workspace),
            prompt_source=f"result:{candidate.source_round}_llm_product_unsatisfied_summary",
            source_workbook=str(candidate.workbook),
            more_started_at=now(),
        )
        trace.write("manual_more_prompt_handoff_ready", idea_id=candidate.idea_id, run_dir=str(run_dir), target_round=candidate.target_round)
        print(f"[{offset}/{len(selected)}] 已启动 {candidate.target_round} {candidate.idea_id}: {workspace}")
    except Exception as exc:
        trace.write("manual_more_start_failed", idea_id=candidate.idea_id, target_round=candidate.target_round, error=str(exc))
        update_started_entry(
            candidate.idea_id,
            status="more_start_failed",
            round_name=candidate.target_round,
            source_round=candidate.source_round,
            run_dir=str(run_dir),
            workspace=str(workspace),
            prompt_source=f"result:{candidate.source_round}_llm_product_unsatisfied_summary",
            source_workbook=str(candidate.workbook),
            error=str(exc),
            more_started_at=now(),
        )
        print(f"[{offset}/{len(selected)}] {candidate.target_round} 启动失败 {candidate.idea_id}: {exc}", file=sys.stderr)
        return False, f"{candidate.idea_id}: {exc}"
    return True, None


if __name__ == "__main__":
    raise SystemExit(main())
