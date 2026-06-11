#!/usr/bin/env python3
from __future__ import annotations

import csv
import fcntl
import hashlib
import importlib.util
import json
import os
import re
import shutil
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any


SCRIPT_DIR = Path(__file__).resolve().parent
ROOT = Path(os.environ.get("SOLO_ROOT", SCRIPT_DIR.parent)).expanduser().resolve()
SCRIPT_AUTO_DIR = ROOT / "script_auto"
RUN_LOG_ROOT = Path(os.environ.get("SOLO_RUN_LOG_ROOT", ROOT / "run_log")).expanduser().resolve()
SCRIPT_RUN_LOG_DIR = RUN_LOG_ROOT / "script"
STATE_PATH = SCRIPT_RUN_LOG_DIR / "manual_state.json"
RUNS_DIR = SCRIPT_RUN_LOG_DIR / "runs"
LOG_DIR = SCRIPT_RUN_LOG_DIR / "logs"
TABLE_CSV_PATH = SCRIPT_RUN_LOG_DIR / "manual_runs.csv"
TABLE_XLSX_PATH = SCRIPT_RUN_LOG_DIR / "manual_runs.xlsx"
SCREENSHOT_DIR = ROOT / "result" / "截图"


def _unique_runtime_target(path: Path) -> Path:
    if not path.exists():
        return path
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return path.with_name(f"{path.name}.legacy-{stamp}")


def _merge_json_dict_file(src: Path, dst: Path) -> bool:
    try:
        src_data = json.loads(src.read_text(encoding="utf-8"))
        dst_data = json.loads(dst.read_text(encoding="utf-8")) if dst.exists() else {}
    except Exception:
        return False
    if not isinstance(src_data, dict) or not isinstance(dst_data, dict):
        return False
    merged = dict(dst_data)
    for key, value in src_data.items():
        if key not in merged:
            merged[key] = value
            continue
        if key == "entries" and isinstance(value, list) and isinstance(merged.get(key), list):
            seen = {json.dumps(item, ensure_ascii=False, sort_keys=True) for item in merged[key] if isinstance(item, dict)}
            for item in value:
                marker = json.dumps(item, ensure_ascii=False, sort_keys=True) if isinstance(item, dict) else str(item)
                if marker not in seen:
                    merged[key].append(item)
                    seen.add(marker)
    dst.parent.mkdir(parents=True, exist_ok=True)
    dst.write_text(json.dumps(merged, ensure_ascii=False, indent=2), encoding="utf-8")
    return True


def _replace_runtime_marker(text: str, marker: str, replacement: Path) -> str:
    idx = text.find(marker)
    if idx < 0:
        return text
    suffix = text[idx + len(marker):].lstrip("/")
    return str(replacement / suffix) if suffix else str(replacement)


def _normalize_manual_runtime_paths(value: Any) -> Any:
    if isinstance(value, dict):
        return {key: _normalize_manual_runtime_paths(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_normalize_manual_runtime_paths(item) for item in value]
    if not isinstance(value, str):
        return value
    replacements = [
        ("/script/runs/", RUNS_DIR),
        ("/script/logs/", LOG_DIR),
        ("/script/backups/", SCRIPT_RUN_LOG_DIR / "backups"),
        ("/script/migration_backups/", SCRIPT_RUN_LOG_DIR / "migration_backups"),
    ]
    result = value
    for marker, replacement in replacements:
        result = _replace_runtime_marker(result, marker, replacement)
    return result


def _normalize_json_file(path: Path) -> None:
    if not path.exists():
        return
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return
    normalized = _normalize_manual_runtime_paths(data)
    if normalized != data:
        path.write_text(json.dumps(normalized, ensure_ascii=False, indent=2), encoding="utf-8")


def _legacy_lock_is_held(path: Path) -> bool:
    try:
        with path.open("a+", encoding="utf-8") as handle:
            try:
                fcntl.flock(handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
            except BlockingIOError:
                return True
            finally:
                try:
                    fcntl.flock(handle.fileno(), fcntl.LOCK_UN)
                except OSError:
                    pass
    except OSError:
        return False
    return False


def _move_legacy_runtime_path(src: Path, dst: Path, *, merge_json: bool = False, lock_file: bool = False) -> None:
    if not src.exists():
        return
    dst.parent.mkdir(parents=True, exist_ok=True)
    if src.is_dir():
        dst.mkdir(parents=True, exist_ok=True)
        for child in src.iterdir():
            shutil.move(str(child), str(_unique_runtime_target(dst / child.name)))
        try:
            src.rmdir()
        except OSError:
            pass
        return
    if not dst.exists():
        shutil.move(str(src), str(dst))
        return
    if lock_file and _legacy_lock_is_held(src):
        return
    if merge_json and _merge_json_dict_file(src, dst):
        archive = RUN_LOG_ROOT / "legacy_imports" / "script"
        archive.mkdir(parents=True, exist_ok=True)
        shutil.move(str(src), str(_unique_runtime_target(archive / src.name)))
        return
    archive = RUN_LOG_ROOT / "legacy_imports" / "script"
    archive.mkdir(parents=True, exist_ok=True)
    shutil.move(str(src), str(_unique_runtime_target(archive / src.name)))


def migrate_legacy_manual_runtime() -> None:
    legacy_pairs = [
        (SCRIPT_DIR / "manual_state.json", STATE_PATH, True),
        (SCRIPT_DIR / "manual_runs.csv", TABLE_CSV_PATH, False),
        (SCRIPT_DIR / "manual_runs.xlsx", TABLE_XLSX_PATH, False),
        (SCRIPT_DIR / "runs", RUNS_DIR, False),
        (SCRIPT_DIR / "logs", LOG_DIR, False),
        (SCRIPT_DIR / "backups", SCRIPT_RUN_LOG_DIR / "backups", False),
        (SCRIPT_DIR / "migration_backups", SCRIPT_RUN_LOG_DIR / "migration_backups", False),
        (SCRIPT_AUTO_DIR / ".manual_more.lock", RUN_LOG_ROOT / "script_auto" / ".manual_second.lock", False, True),
        (SCRIPT_AUTO_DIR / ".manual_second.lock", RUN_LOG_ROOT / "script_auto" / ".manual_second.lock", False, True),
    ]
    for item in legacy_pairs:
        src, dst, merge_json, *rest = item
        _move_legacy_runtime_path(src, dst, merge_json=merge_json, lock_file=bool(rest and rest[0]))
    _normalize_json_file(STATE_PATH)


migrate_legacy_manual_runtime()
ROUND_DIGITS = {
    0: "零",
    1: "一",
    2: "二",
    3: "三",
    4: "四",
    5: "五",
    6: "六",
    7: "七",
    8: "八",
    9: "九",
    10: "十",
}


def positive_int(value: Any, default: int) -> int:
    try:
        parsed = int(str(value).strip())
    except Exception:
        return default
    return parsed if parsed > 0 else default


def default_max_rounds() -> int:
    return positive_int(os.environ.get("SOLO_MAX_ROUNDS"), 5)


def round_number_to_name(number: int) -> str:
    if number <= 0:
        number = 1
    if number <= 10:
        return f"第{ROUND_DIGITS[number]}轮"
    if number < 20:
        return f"第十{ROUND_DIGITS[number - 10]}轮"
    if number == 20:
        return "第二十轮"
    tens, ones = divmod(number, 10)
    ten_text = ROUND_DIGITS.get(tens, str(tens))
    one_text = ROUND_DIGITS.get(ones, str(ones))
    return f"第{ten_text}十{one_text if ones else ''}轮"


def round_name_to_number(round_name: Any) -> int:
    text = str(round_name or "").strip()
    if not text:
        return 1
    match = re.fullmatch(r"第(\d+)轮", text)
    if match:
        return positive_int(match.group(1), 1)
    names = {round_number_to_name(idx): idx for idx in range(1, 31)}
    return names.get(text, 1)


def next_round_name(round_name: Any) -> str:
    return round_number_to_name(round_name_to_number(round_name) + 1)


def result_screenshot_path(idea_id: str, round_name: Any | None = None) -> Path:
    if round_name:
        return SCREENSHOT_DIR / f"{idea_id}-round{round_name_to_number(round_name)}.png"
    return SCREENSHOT_DIR / f"{idea_id}.png"


def load_runner() -> Any:
    runner_path = SCRIPT_AUTO_DIR / "trae_idea_runner.py"
    if not runner_path.exists():
        raise SystemExit(f"找不到 script_auto runner: {runner_path}")
    os.environ.setdefault("SOLO_ROOT", str(ROOT))
    if str(SCRIPT_AUTO_DIR) not in sys.path:
        sys.path.insert(0, str(SCRIPT_AUTO_DIR))
    spec = importlib.util.spec_from_file_location("manual_script_auto_runner", runner_path)
    if spec is None or spec.loader is None:
        raise SystemExit(f"无法加载 runner: {runner_path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules["manual_script_auto_runner"] = module
    spec.loader.exec_module(module)
    return module


def now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def read_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return default


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def write_run_marker(run_dir: Path, filename: str, data: dict[str, Any]) -> Path:
    payload = dict(data)
    payload.setdefault("marked_at", now())
    path = run_dir / filename
    write_json(path, payload)
    return path


def load_state() -> dict[str, Any]:
    data = read_json(STATE_PATH, {})
    return data if isinstance(data, dict) else {}


def save_state(state: dict[str, Any]) -> None:
    write_json(STATE_PATH, state)


def idea_history_hash() -> str:
    path = ROOT / "idea_history.md"
    if not path.exists():
        return ""
    return hashlib.sha1(path.read_bytes()).hexdigest()


def update_started_entry(idea_id: str, **fields: Any) -> None:
    state = load_state()
    entries = state.setdefault("entries", [])
    if not isinstance(entries, list):
        entries = []
        state["entries"] = entries
    for item in entries:
        if isinstance(item, dict) and item.get("idea_id") == idea_id:
            item.update(fields)
            item["updated_at"] = now()
            break
    else:
        item = {"idea_id": idea_id, **fields, "updated_at": now()}
        entries.append(item)
    save_state(state)


def latest_manual_run_dir(idea_id: str) -> Path | None:
    entries = load_state().get("entries", [])
    if isinstance(entries, list):
        for item in reversed(entries):
            if isinstance(item, dict) and item.get("idea_id") == idea_id:
                run_dir = Path(str(item.get("run_dir") or ""))
                if run_dir.exists():
                    return run_dir
    if RUNS_DIR.exists():
        matches = sorted(
            (
                p
                for p in RUNS_DIR.glob(f"*-{idea_id}*")
                if p.is_dir()
                and not (p / ".dry_run").exists()
                and re.search(rf"-{re.escape(idea_id)}(?:-round\d+)?$", p.name)
            ),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        if matches:
            return matches[0]
    return None


def idea_id_from_run_dir(run_dir: Path) -> str:
    match = re.search(r"-(zy\d+)(?:-round\d+)?$", run_dir.name)
    return match.group(1) if match else ""


def auto_finish_candidate_entries() -> list[dict[str, Any]]:
    entries = load_state().get("entries", [])
    if not isinstance(entries, list):
        return []

    candidates: list[dict[str, Any]] = []
    queue_statuses = {"completed_window_closed", "service_exception_window_closed"}
    queue_sources = {"completion", "service_exception"}
    blocked_finish_statuses = {"running", "captured_waiting_qc", "done", "selected"}
    for item in entries:
        if not isinstance(item, dict):
            continue
        status = str(item.get("status") or "")
        finish_status = str(item.get("finish_status") or "")
        retry_manual_required = finish_status == "auto_manual_required"
        if status not in queue_statuses:
            continue
        if finish_status in blocked_finish_statuses or item.get("finish_done_at"):
            continue
        if item.get("finish_started_at") and not retry_manual_required:
            continue
        if item.get("go_close_source") not in queue_sources:
            continue
        if item.get("go_closed") is False or item.get("go_ready_for_finish") is False:
            continue
        run_dir = Path(str(item.get("run_dir") or ""))
        if not run_dir.exists():
            continue
        if (run_dir / "manual_final_row.json").exists():
            continue
        marker = read_json(run_dir / "finish_marker.json", {})
        if isinstance(marker, dict) and marker.get("finish_status") == "done":
            continue
        candidates.append(item)

    candidates.sort(key=lambda item: str(item.get("closed_at") or item.get("updated_at") or item.get("started_at") or ""))
    return candidates


def next_auto_finish_candidate() -> dict[str, Any] | None:
    candidates = auto_finish_candidate_entries()
    return candidates[0] if candidates else None


def auto_finish_pending_count() -> int:
    return len(auto_finish_candidate_entries())


def sessionish(text: str) -> bool:
    value = text.strip()
    if not value:
        return False
    if re.search(r":[0-9a-f]{16,}_[0-9a-f]{16,}\.", value):
        return True
    if re.search(r"\b(?:session[_ -]?id|sessionId|会话)\b", value, re.I) and re.search(r"[0-9a-f]{16,}", value, re.I):
        return True
    return bool(re.fullmatch(r"[0-9a-f]{16,}", value, re.I))


def wait_for_clipboard_text(
    runner: Any,
    timeout_seconds: int,
    trace: Any,
    *,
    initial_value: str = "",
    min_chars: int = 1,
    event_name: str = "manual_clipboard_text_detected",
) -> str:
    deadline = None if timeout_seconds <= 0 else time.time() + timeout_seconds
    initial = initial_value.strip()
    while True:
        value = runner.get_clipboard().strip()
        if value and value != initial and len(value) >= min_chars:
            trace.write(event_name, chars=len(value))
            return value
        if deadline is not None and time.time() >= deadline:
            raise TimeoutError(f"等待剪贴板内容超时: {timeout_seconds}s")
        time.sleep(1)


def wait_for_session_clipboard(runner: Any, timeout_seconds: int, trace: Any) -> str:
    deadline = None if timeout_seconds <= 0 else time.time() + timeout_seconds
    while True:
        value = runner.get_clipboard().strip()
        if sessionish(value):
            trace.write("manual_session_clipboard_detected", chars=len(value))
            return value
        if deadline is not None and time.time() >= deadline:
            raise TimeoutError(f"等待剪贴板 sessionId 超时: {timeout_seconds}s")
        time.sleep(1)


def compact_business_domain(value: Any) -> str:
    text = str(value or "Web前端")
    return re.sub(r"\s+", "", text) or "Web前端"


def manual_table_value(header: str, value: Any) -> str:
    if header == "轮次":
        return table_text(value or "第一轮")
    if header == "业务领域":
        return table_text(compact_business_domain(value))
    return table_text(value)


def normalize_manual_row(row: dict[str, Any]) -> dict[str, Any]:
    updated = dict(row)
    if not updated.get("截图"):
        updated["截图"] = updated.get("screenshot_path", "")
    return updated


MANUAL_HEADERS = [
    "session_id",
    "idea_id",
    "title",
    "轮次",
    "提示词",
    "任务类型",
    "业务领域",
    "修改范围",
    "任务是否完成",
    "产物及过程是否满意",
    "不满意原因",
    "远端Github地址",
    "github地址",
    "commit id",
    "分支文件夹",
    "截图",
    "日志轨迹",
    "流程状态",
    "流程日志",
    "workspace",
    "run_dir",
    "prompt_path",
    "screenshot_path",
    "trajectory_path",
    "full_transcript_path",
    "qa_status",
    "qa_report_path",
    "qa_summary",
    "git_status",
    "git_commit",
    "git_remote",
    "git_branch",
    "updated_at",
]


def table_text(value: Any, limit: int = 30000) -> str:
    text = str(value or "")
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", text)
    return text if len(text) <= limit else text[:limit] + "\n...[truncated]..."


def upsert_manual_table(row: dict[str, Any]) -> None:
    rows: list[dict[str, str]] = []
    if TABLE_CSV_PATH.exists():
        with TABLE_CSV_PATH.open("r", encoding="utf-8", newline="") as f:
            rows = list(csv.DictReader(f))

    keyed: dict[str, dict[str, str]] = {}
    for item in rows:
        idea_id = item.get("idea_id")
        if idea_id:
            item = normalize_manual_row(item)
            round_name = str(item.get("轮次") or "第一轮").strip() or "第一轮"
            keyed[f"{idea_id}::{round_name}"] = {h: manual_table_value(h, item.get(h, "")) for h in MANUAL_HEADERS}
    row = normalize_manual_row(row)
    row_round_name = str(row.get("轮次") or "第一轮").strip() or "第一轮"
    keyed[f"{row['idea_id']}::{row_round_name}"] = {h: manual_table_value(h, row.get(h, "")) for h in MANUAL_HEADERS}

    TABLE_CSV_PATH.parent.mkdir(parents=True, exist_ok=True)
    with TABLE_CSV_PATH.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=MANUAL_HEADERS)
        writer.writeheader()
        for item in keyed.values():
            writer.writerow(item)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "manual_runs"
    ws.append(MANUAL_HEADERS)
    for item in keyed.values():
        ws.append([item.get(h, "") for h in MANUAL_HEADERS])
    fill = PatternFill("solid", fgColor="D9EAF7")
    font = Font(bold=True, color="1F2937")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, header in enumerate(MANUAL_HEADERS, 1):
        width = min(60, max(12, len(header) + 2))
        for row_values in keyed.values():
            width = min(60, max(width, min(60, len(str(row_values.get(header, ""))) + 2)))
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = width
    for cells in ws.iter_rows(min_row=2):
        for cell in cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(TABLE_XLSX_PATH)


def read_text_if_exists(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore") if path.exists() else ""
