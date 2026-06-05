#!/usr/bin/env python3
"""
Upload Solo result workbook rows into the configured Feishu Base table.

Usage:
  /Users/lzy/pro/solo/update.py /Users/lzy/pro/solo/result/zy10270.xlsx
  /Users/lzy/pro/solo/update.py /Users/lzy/pro/solo/result/zy10270.xlsx /Users/lzy/pro/solo/result/zy10271.xlsx

The script upserts by "Trae Session ID". The workbook's "截图" column is ignored
because this Base table no longer needs screenshot attachments.
"""

from __future__ import annotations

import argparse
import json
import os
import subprocess
import sys
import time
from pathlib import Path
from typing import Any


try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    bundled_python = Path.home() / ".cache/codex-runtimes/codex-primary-runtime/dependencies/python/bin/python3"
    if bundled_python.exists() and Path(sys.executable).resolve() != bundled_python.resolve():
        os.execv(str(bundled_python), [str(bundled_python), *sys.argv])
    raise


DEFAULT_BASE_TOKEN = "NVLabI2piaNiVHsn6GYchYlmnRt"
DEFAULT_TABLE_ID = "tblONK81uEWGM2jF"
DEFAULT_VIEW_ID = "vewKdKoVia"

KEY_FIELD = "Trae Session ID"
ATTACHMENT_FIELD = "截图"

HEADER_MAP = {
    "session_id": "Trae Session ID",
    "Trae Session ID": "Trae Session ID",
    "轮次": "轮次",
    "提示词": "User Prompt",
    "User Prompt": "User Prompt",
    "任务类型": "任务类型",
    "业务领域": "业务领域",
    "修改范围": "修改范围",
    "任务是否完成": "任务是否完成",
    "产物及过程是否满意": "产物及过程是否满意",
    "不满意原因": "不满意原因",
    "远端Github地址": "github地址",
    "远端 GitHub 地址": "github地址",
    "github地址": "github地址",
    "Github地址": "github地址",
    "GitHub地址": "github地址",
    "分支文件夹": "分支/文件夹",
    "分支/文件夹": "分支/文件夹",
    "截图": "截图",
    "日志轨迹": "日志轨迹",
}

READ_ONLY_TYPES = {
    "formula",
    "lookup",
    "created_time",
    "modified_time",
    "created_user",
    "modified_user",
    "auto_number",
    "not_support",
}


class UploadError(RuntimeError):
    pass


def auth_hint_for_lark_error(data: dict[str, Any]) -> str:
    error = data.get("error") if isinstance(data, dict) else None
    if not isinstance(error, dict):
        return ""
    message = str(error.get("message") or "")
    if "need_user_authorization" not in message:
        return ""
    return (
        "\nHint: lark-cli user authorization is missing or stale. "
        "Run `lark-cli auth login --domain base` and retry this update command."
    )


def run_lark(args: list[str], check: bool = True, cwd: Path | None = None) -> dict[str, Any]:
    proc = subprocess.run(["lark-cli", *args], text=True, capture_output=True, cwd=str(cwd) if cwd else None)
    output = (proc.stdout or "").strip() or (proc.stderr or "").strip()
    try:
        data = json.loads(output)
    except ValueError as exc:
        raise UploadError(f"lark-cli {' '.join(args)} returned non-json: {output[:800]}") from exc
    if check and proc.returncode != 0:
        raise UploadError(f"lark-cli {' '.join(args)} failed: {data}{auth_hint_for_lark_error(data)}")
    return data


def list_fields(base_token: str, table_id: str) -> list[dict[str, Any]]:
    data = run_lark(
        [
            "base",
            "+field-list",
            "--as",
            "user",
            "--base-token",
            base_token,
            "--table-id",
            table_id,
            "--offset",
            "0",
            "--limit",
            "200",
        ]
    )
    if not data.get("ok"):
        raise UploadError(f"field-list failed: {data}")
    payload = data.get("data", {})
    return payload.get("items") or payload.get("fields") or []


def field_name(field: dict[str, Any]) -> str:
    return str(field.get("field_name") or field.get("name") or "")


def field_id(field: dict[str, Any]) -> str:
    return str(field.get("field_id") or field.get("id") or "")


def normalize_record_rows(payload: dict[str, Any], field_id_to_name: dict[str, str]) -> list[dict[str, Any]]:
    if isinstance(payload.get("items"), list):
        return payload["items"]
    rows = payload.get("data") or []
    field_ids = payload.get("field_id_list") or []
    record_ids = payload.get("record_id_list") or []
    records: list[dict[str, Any]] = []
    for index, row in enumerate(rows):
        fields: dict[str, Any] = {}
        for col_index, value in enumerate(row):
            if col_index >= len(field_ids):
                continue
            name = field_id_to_name.get(field_ids[col_index], field_ids[col_index])
            fields[name] = value
        records.append(
            {
                "record_id": record_ids[index] if index < len(record_ids) else "",
                "fields": fields,
            }
        )
    return records


def list_records(base_token: str, table_id: str, field_id_to_name: dict[str, str]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    offset = 0
    limit = 200
    while True:
        data = run_lark(
            [
                "base",
                "+record-list",
                "--as",
                "user",
                "--base-token",
                base_token,
                "--table-id",
                table_id,
                "--format",
                "json",
                "--offset",
                str(offset),
                "--limit",
                str(limit),
            ]
        )
        if not data.get("ok"):
            raise UploadError(f"record-list failed: {data}")
        payload = data.get("data", {})
        records.extend(normalize_record_rows(payload, field_id_to_name))
        if not payload.get("has_more"):
            break
        offset += limit
    return records


def value_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, float, bool)):
        return str(value).strip()
    if isinstance(value, list):
        return ",".join(value_to_text(item) for item in value if value_to_text(item))
    if isinstance(value, dict):
        return value_to_text(value.get("text") or value.get("name") or value.get("file_token") or "")
    return str(value).strip()


def is_blank_record(record: dict[str, Any], writable_names: set[str]) -> bool:
    fields = record.get("fields", {})
    return not any(value_to_text(fields.get(name)) for name in writable_names)


def read_workbook_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    headers = [value_to_text(cell.value) for cell in sheet[1]]
    rows: list[dict[str, Any]] = []
    for row_cells in sheet.iter_rows(min_row=2, values_only=True):
        raw = {headers[index]: row_cells[index] for index in range(min(len(headers), len(row_cells))) if headers[index]}
        if not any(value_to_text(value) for value in raw.values()):
            continue
        rows.append(raw)
    return rows


def coerce_cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def build_fields(row: dict[str, Any], writable_field_names: set[str]) -> dict[str, Any]:
    fields: dict[str, Any] = {}
    for excel_header, value in row.items():
        base_field = HEADER_MAP.get(excel_header, excel_header)
        if base_field == ATTACHMENT_FIELD or excel_header == "截图":
            continue
        if base_field not in writable_field_names:
            continue
        clean_value = coerce_cell_value(value)
        if clean_value is not None:
            fields[base_field] = clean_value
    return fields


def record_upsert(base_token: str, table_id: str, fields: dict[str, Any], record_id: str | None = None) -> str:
    args = [
        "base",
        "+record-upsert",
        "--as",
        "user",
        "--base-token",
        base_token,
        "--table-id",
        table_id,
        "--json",
        json.dumps(fields, ensure_ascii=False),
    ]
    if record_id:
        args.extend(["--record-id", record_id])
    data = run_lark(args)
    if not data.get("ok"):
        raise UploadError(f"record-upsert failed: {data}")
    return (
        data.get("data", {}).get("record", {}).get("record_id")
        or data.get("data", {}).get("record_id")
        or record_id
        or ""
    )


def load_table_context(args: argparse.Namespace) -> dict[str, Any]:
    fields = list_fields(args.base_token, args.table_id)
    field_by_name = {field_name(field): field for field in fields if field_name(field)}
    field_id_to_name = {field_id(field): field_name(field) for field in fields if field_id(field) and field_name(field)}
    writable_names = {
        name
        for name, field in field_by_name.items()
        if name and field.get("type") not in READ_ONLY_TYPES
    }
    existing_records = list_records(args.base_token, args.table_id, field_id_to_name)
    by_key = {
        value_to_text(record.get("fields", {}).get(KEY_FIELD)): record
        for record in existing_records
        if value_to_text(record.get("fields", {}).get(KEY_FIELD))
    }
    blank_records = [record for record in existing_records if is_blank_record(record, writable_names)]
    return {
        "field_id_to_name": field_id_to_name,
        "writable_names": writable_names,
        "by_key": by_key,
        "blank_records": blank_records,
    }


def upload_workbook(args: argparse.Namespace, table_context: dict[str, Any] | None = None) -> dict[str, Any]:
    workbook_path = Path(args.workbook).expanduser().resolve()
    if not workbook_path.exists():
        raise UploadError(f"workbook not found: {workbook_path}")

    context = table_context or load_table_context(args)
    field_id_to_name = context["field_id_to_name"]
    writable_names = context["writable_names"]
    by_key = context["by_key"]
    blank_records = context["blank_records"]

    workbook_rows = read_workbook_rows(workbook_path)
    result = {
        "workbook": str(workbook_path),
        "rows": len(workbook_rows),
        "created": 0,
        "updated": 0,
        "reused_blank": 0,
        "attachments_uploaded": 0,
        "attachments_skipped": 0,
        "attachments_missing": [],
        "skipped": [],
        "records": [],
    }

    for index, row in enumerate(workbook_rows, start=2):
        fields_to_write = build_fields(row, writable_names)
        key = value_to_text(fields_to_write.get(KEY_FIELD) or row.get("session_id") or row.get(KEY_FIELD))
        if not key:
            result["skipped"].append({"row": index, "reason": f"missing {KEY_FIELD}/session_id"})
            continue
        fields_to_write[KEY_FIELD] = key

        if args.dry_run:
            result["records"].append(
                {
                    "row": index,
                    "action": "dry-run",
                    "key": key,
                    "fields": sorted(fields_to_write),
                }
            )
            continue

        action = "created"
        record = by_key.get(key)
        record_id = ""
        if record:
            record_id = record["record_id"]
            record_upsert(args.base_token, args.table_id, fields_to_write, record_id=record_id)
            result["updated"] += 1
            action = "updated"
        elif blank_records:
            record = blank_records.pop(0)
            record_id = record["record_id"]
            record_upsert(args.base_token, args.table_id, fields_to_write, record_id=record_id)
            result["updated"] += 1
            result["reused_blank"] += 1
            action = "reused_blank"
        else:
            record_id = record_upsert(args.base_token, args.table_id, fields_to_write)
            if not record_id:
                refreshed = list_records(args.base_token, args.table_id, field_id_to_name)
                context["by_key"] = {
                    value_to_text(candidate.get("fields", {}).get(KEY_FIELD)): candidate
                    for candidate in refreshed
                    if value_to_text(candidate.get("fields", {}).get(KEY_FIELD))
                }
                by_key = context["by_key"]
                for candidate in refreshed:
                    if value_to_text(candidate.get("fields", {}).get(KEY_FIELD)) == key:
                        record_id = candidate.get("record_id", "")
                        break
            if not record_id:
                raise UploadError(f"created row for key but could not resolve record_id: {key}")
            result["created"] += 1

        by_key[key] = {"record_id": record_id, "fields": fields_to_write}

        result["records"].append({"row": index, "action": action, "record_id": record_id, "key": key})
        time.sleep(args.delay)

    return result


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Upload a Solo result xlsx into Feishu Base.")
    parser.add_argument("workbooks", nargs="+", help="Path(s) to result xlsx, for example /Users/lzy/pro/solo/result/zy10270.xlsx")
    parser.add_argument("--base-token", default=DEFAULT_BASE_TOKEN)
    parser.add_argument("--table-id", default=DEFAULT_TABLE_ID)
    parser.add_argument("--view-id", default=DEFAULT_VIEW_ID)
    parser.add_argument("--dry-run", action="store_true", help="Parse rows without writing to Feishu.")
    parser.add_argument("--always-upload-attachment", action="store_true", help=argparse.SUPPRESS)
    parser.add_argument("--delay", type=float, default=0.2, help="Delay between row uploads.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    reports: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []
    totals = {
        "files": len(args.workbooks),
        "rows": 0,
        "created": 0,
        "updated": 0,
        "reused_blank": 0,
        "attachments_uploaded": 0,
        "attachments_skipped": 0,
        "attachments_missing": 0,
        "skipped": 0,
        "errors": 0,
    }
    try:
        table_context = load_table_context(args)
    except UploadError as exc:
        print(f"ERROR load table context: {exc}", file=sys.stderr)
        return 1
    for workbook in args.workbooks:
        args.workbook = workbook
        try:
            report = upload_workbook(args, table_context)
            reports.append(report)
            for key in ("rows", "created", "updated", "reused_blank", "attachments_uploaded", "attachments_skipped"):
                totals[key] += int(report.get(key, 0))
            totals["attachments_missing"] += len(report.get("attachments_missing", []))
            totals["skipped"] += len(report.get("skipped", []))
        except UploadError as exc:
            errors.append({"workbook": workbook, "error": str(exc)})
            totals["errors"] += 1
            print(f"ERROR {workbook}: {exc}", file=sys.stderr)
            if not args.dry_run:
                continue
    output = {"summary": totals, "reports": reports, "errors": errors}
    print(json.dumps(output, ensure_ascii=False, indent=2))
    return 1 if errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
