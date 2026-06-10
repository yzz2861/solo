from sqlalchemy.orm import Session
from datetime import datetime
from typing import Optional
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from models import ConstructionApplication, SkylightPlan, TemporaryBlock, ConflictRecord


def generate_conflict_report_csv(db: Session, line_name: Optional[str] = None) -> bytes:
    conflicts = db.query(ConflictRecord).order_by(
        ConflictRecord.severity.desc(),
        ConflictRecord.overlap_start
    )
    if line_name:
        conflicts = conflicts.filter(ConflictRecord.line_name == line_name)
    conflicts = conflicts.all()

    output = io.StringIO()
    writer = csv.writer(output)

    header = [
        "冲突ID", "冲突类型", "严重程度", "线路名称", "区段",
        "起始公里", "结束公里",
        "A项类型", "A项编号", "A项负责人", "A项开始时间", "A项结束时间",
        "B项类型", "B项编号", "B项负责人", "B项开始时间", "B项结束时间",
        "时间重叠开始", "时间重叠结束", "公里重叠开始", "公里重叠结束",
        "状态", "处理意见", "检测时间", "处理时间"
    ]
    writer.writerow(header)

    for c in conflicts:
        row = [
            c.id, c.conflict_type, c.severity, c.line_name, c.section,
            c.start_km, c.end_km,
            c.item_a_type, c.item_a_no, c.item_a_person,
            c.item_a_start.strftime("%Y-%m-%d %H:%M") if c.item_a_start else "",
            c.item_a_end.strftime("%Y-%m-%d %H:%M") if c.item_a_end else "",
            c.item_b_type, c.item_b_no, c.item_b_person,
            c.item_b_start.strftime("%Y-%m-%d %H:%M") if c.item_b_start else "",
            c.item_b_end.strftime("%Y-%m-%d %H:%M") if c.item_b_end else "",
            c.overlap_start.strftime("%Y-%m-%d %H:%M") if c.overlap_start else "",
            c.overlap_end.strftime("%Y-%m-%d %H:%M") if c.overlap_end else "",
            c.overlap_km_start if c.overlap_km_start else "",
            c.overlap_km_end if c.overlap_km_end else "",
            c.status, c.handle_opinion or "",
            c.detected_at.strftime("%Y-%m-%d %H:%M") if c.detected_at else "",
            c.handled_at.strftime("%Y-%m-%d %H:%M") if c.handled_at else "",
        ]
        writer.writerow(row)

    return output.getvalue().encode("utf-8-sig")


def generate_schedule_report_excel(db: Session, line_name: Optional[str] = None) -> bytes:
    output = io.BytesIO()
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def _style_header(ws, row_idx, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

    def _style_data(ws, start_row, end_row, col_count):
        for row in range(start_row, end_row + 1):
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws1 = wb.active
    ws1.title = "施工申请"
    headers1 = ["申请编号", "线路名称", "区段", "起始公里", "结束公里", "施工类型", "施工内容",
               "开始时间", "结束时间", "施工负责人", "联系电话", "状态", "复核意见", "批次号"]
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)
    _style_header(ws1, 1, len(headers1))

    apps = db.query(ConstructionApplication).order_by(ConstructionApplication.start_time)
    if line_name:
        apps = apps.filter(ConstructionApplication.line_name == line_name)
    apps = apps.all()

    for i, app in enumerate(apps, 2):
        ws1.cell(row=i, column=1, value=app.application_no)
        ws1.cell(row=i, column=2, value=app.line_name)
        ws1.cell(row=i, column=3, value=app.section)
        ws1.cell(row=i, column=4, value=app.start_km)
        ws1.cell(row=i, column=5, value=app.end_km)
        ws1.cell(row=i, column=6, value=app.construction_type or "")
        ws1.cell(row=i, column=7, value=app.construction_content or "")
        ws1.cell(row=i, column=8, value=app.start_time.strftime("%Y-%m-%d %H:%M") if app.start_time else "")
        ws1.cell(row=i, column=9, value=app.end_time.strftime("%Y-%m-%d %H:%M") if app.end_time else "")
        ws1.cell(row=i, column=10, value=app.responsible_person)
        ws1.cell(row=i, column=11, value=app.phone or "")
        ws1.cell(row=i, column=12, value=app.status)
        ws1.cell(row=i, column=13, value=app.review_opinion or "")
        ws1.cell(row=i, column=14, value=app.batch_no or "")
    _style_data(ws1, 2, len(apps) + 1, len(headers1))
    for col_idx in range(1, len(headers1) + 1):
        ws1.column_dimensions[chr(64 + col_idx)].width = 15

    ws2 = wb.create_sheet("天窗计划")
    headers2 = ["计划编号", "线路名称", "区段", "起始公里", "结束公里", "天窗类型", "施工内容",
               "开始时间", "结束时间", "施工负责人", "联系电话", "状态", "复核意见", "关联申请ID", "批次号"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    _style_header(ws2, 1, len(headers2))

    plans = db.query(SkylightPlan).order_by(SkylightPlan.start_time)
    if line_name:
        plans = plans.filter(SkylightPlan.line_name == line_name)
    plans = plans.all()

    for i, plan in enumerate(plans, 2):
        ws2.cell(row=i, column=1, value=plan.plan_no)
        ws2.cell(row=i, column=2, value=plan.line_name)
        ws2.cell(row=i, column=3, value=plan.section)
        ws2.cell(row=i, column=4, value=plan.start_km)
        ws2.cell(row=i, column=5, value=plan.end_km)
        ws2.cell(row=i, column=6, value=plan.skylight_type or "")
        ws2.cell(row=i, column=7, value=plan.construction_content or "")
        ws2.cell(row=i, column=8, value=plan.start_time.strftime("%Y-%m-%d %H:%M") if plan.start_time else "")
        ws2.cell(row=i, column=9, value=plan.end_time.strftime("%Y-%m-%d %H:%M") if plan.end_time else "")
        ws2.cell(row=i, column=10, value=plan.responsible_person)
        ws2.cell(row=i, column=11, value=plan.phone or "")
        ws2.cell(row=i, column=12, value=plan.status)
        ws2.cell(row=i, column=13, value=plan.review_opinion or "")
        ws2.cell(row=i, column=14, value=plan.application_id or "")
        ws2.cell(row=i, column=15, value=plan.batch_no or "")
    _style_data(ws2, 2, len(plans) + 1, len(headers2))
    for col_idx in range(1, len(headers2) + 1):
        ws2.column_dimensions[chr(64 + col_idx)].width = 15

    ws3 = wb.create_sheet("临时封锁")
    headers3 = ["封锁单编号", "线路名称", "区段", "起始公里", "结束公里", "封锁原因",
               "开始时间", "结束时间", "施工负责人", "联系电话", "状态", "复核意见", "关联计划ID", "批次号"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
    _style_header(ws3, 1, len(headers3))

    blocks = db.query(TemporaryBlock).order_by(TemporaryBlock.start_time)
    if line_name:
        blocks = blocks.filter(TemporaryBlock.line_name == line_name)
    blocks = blocks.all()

    for i, block in enumerate(blocks, 2):
        ws3.cell(row=i, column=1, value=block.block_no)
        ws3.cell(row=i, column=2, value=block.line_name)
        ws3.cell(row=i, column=3, value=block.section)
        ws3.cell(row=i, column=4, value=block.start_km)
        ws3.cell(row=i, column=5, value=block.end_km)
        ws3.cell(row=i, column=6, value=block.block_reason or "")
        ws3.cell(row=i, column=7, value=block.start_time.strftime("%Y-%m-%d %H:%M") if block.start_time else "")
        ws3.cell(row=i, column=8, value=block.end_time.strftime("%Y-%m-%d %H:%M") if block.end_time else "")
        ws3.cell(row=i, column=9, value=block.responsible_person)
        ws3.cell(row=i, column=10, value=block.phone or "")
        ws3.cell(row=i, column=11, value=block.status)
        ws3.cell(row=i, column=12, value=block.review_opinion or "")
        ws3.cell(row=i, column=13, value=block.plan_id or "")
        ws3.cell(row=i, column=14, value=block.batch_no or "")
    _style_data(ws3, 2, len(blocks) + 1, len(headers3))
    for col_idx in range(1, len(headers3) + 1):
        ws3.column_dimensions[chr(64 + col_idx)].width = 15

    ws4 = wb.create_sheet("冲突汇总")
    headers4 = ["冲突ID", "冲突类型", "严重程度", "线路名称", "区段",
                "起始公里", "结束公里",
                "A项类型", "A项编号", "A项负责人", "A项开始", "A项结束",
                "B项类型", "B项编号", "B项负责人", "B项开始", "B项结束",
                "重叠开始", "重叠结束", "重叠公里起", "重叠公里止",
                "状态", "处理意见", "检测时间", "处理时间"]
    for col, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=col, value=h)
    _style_header(ws4, 1, len(headers4))

    conflicts = db.query(ConflictRecord).order_by(
        ConflictRecord.severity.desc(),
        ConflictRecord.overlap_start
    )
    if line_name:
        conflicts = conflicts.filter(ConflictRecord.line_name == line_name)
    conflicts = conflicts.all()

    severity_colors = {"严重": "FFC7CE", "中等": "FFEB9C", "轻微": "C6EFCE"}
    severity_font_colors = {"严重": "9C0006", "中等": "9C5700", "轻微": "006100"}

    for i, c in enumerate(conflicts, 2):
        ws4.cell(row=i, column=1, value=c.id)
        ws4.cell(row=i, column=2, value=c.conflict_type)
        sev_cell = ws4.cell(row=i, column=3, value=c.severity)
        sev_color = severity_colors.get(c.severity, "FFFFFF")
        sev_font_color = severity_font_colors.get(c.severity, "000000")
        sev_cell.fill = PatternFill(start_color=sev_color, end_color=sev_color, fill_type="solid")
        sev_cell.font = Font(bold=True, color=sev_font_color)
        ws4.cell(row=i, column=4, value=c.line_name)
        ws4.cell(row=i, column=5, value=c.section)
        ws4.cell(row=i, column=6, value=c.start_km)
        ws4.cell(row=i, column=7, value=c.end_km)
        ws4.cell(row=i, column=8, value=c.item_a_type or "")
        ws4.cell(row=i, column=9, value=c.item_a_no or "")
        ws4.cell(row=i, column=10, value=c.item_a_person or "")
        ws4.cell(row=i, column=11, value=c.item_a_start.strftime("%m-%d %H:%M") if c.item_a_start else "")
        ws4.cell(row=i, column=12, value=c.item_a_end.strftime("%m-%d %H:%M") if c.item_a_end else "")
        ws4.cell(row=i, column=13, value=c.item_b_type or "")
        ws4.cell(row=i, column=14, value=c.item_b_no or "")
        ws4.cell(row=i, column=15, value=c.item_b_person or "")
        ws4.cell(row=i, column=16, value=c.item_b_start.strftime("%m-%d %H:%M") if c.item_b_start else "")
        ws4.cell(row=i, column=17, value=c.item_b_end.strftime("%m-%d %H:%M") if c.item_b_end else "")
        ws4.cell(row=i, column=18, value=c.overlap_start.strftime("%m-%d %H:%M") if c.overlap_start else "")
        ws4.cell(row=i, column=19, value=c.overlap_end.strftime("%m-%d %H:%M") if c.overlap_end else "")
        ws4.cell(row=i, column=20, value=c.overlap_km_start if c.overlap_km_start else "")
        ws4.cell(row=i, column=21, value=c.overlap_km_end if c.overlap_km_end else "")
        ws4.cell(row=i, column=22, value=c.status)
        ws4.cell(row=i, column=23, value=c.handle_opinion or "")
        ws4.cell(row=i, column=24, value=c.detected_at.strftime("%Y-%m-%d %H:%M") if c.detected_at else "")
        ws4.cell(row=i, column=25, value=c.handled_at.strftime("%Y-%m-%d %H:%M") if c.handled_at else "")
    _style_data(ws4, 2, len(conflicts) + 1, len(headers4))
    for col_idx in range(1, len(headers4) + 1):
        ws4.column_dimensions[chr(64 + col_idx)].width = 13

    ws5 = wb.create_sheet("统计概览")
    ws5.cell(row=1, column=1, value="天窗排程冲突报告统计概览")
    ws5.cell(row=1, column=1).font = Font(bold=True, size=16)
    ws5.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    ws5.cell(row=3, column=1, value="统计项")
    ws5.cell(row=3, column=2, value="数量")
    _style_header(ws5, 3, 2)

    total_apps = len(apps)
    total_plans = len(plans)
    total_blocks = len(blocks)
    total_conflicts = len(conflicts)
    pending_conflicts = sum(1 for c in conflicts if c.status == "待处理")
    resolved_conflicts = sum(1 for c in conflicts if c.status == "已处理")

    stats = [
        ("施工申请总数", total_apps),
        ("天窗计划总数", total_plans),
        ("临时封锁单数", total_blocks),
        ("冲突总数", total_conflicts),
        ("待处理冲突", pending_conflicts),
        ("已处理冲突", resolved_conflicts),
    ]
    for i, (label, value) in enumerate(stats, 4):
        ws5.cell(row=i, column=1, value=label)
        ws5.cell(row=i, column=2, value=value)
    _style_data(ws5, 4, len(stats) + 3, 2)
    ws5.column_dimensions['A'].width = 20
    ws5.column_dimensions['B'].width = 15

    ws5.cell(row=len(stats) + 5, column=1, value="按严重程度分布")
    ws5.cell(row=len(stats) + 5, column=1).font = Font(bold=True, size=12)

    ws5.cell(row=len(stats) + 6, column=1, value="严重程度")
    ws5.cell(row=len(stats) + 6, column=2, value="数量")
    _style_header(ws5, len(stats) + 6, 2)

    sev_stats = {}
    for c in conflicts:
        sev_stats[c.severity] = sev_stats.get(c.severity, 0) + 1
    sev_order = ["严重", "中等", "轻微"]
    for i, sev in enumerate(sev_order, len(stats) + 7):
        ws5.cell(row=i, column=1, value=sev)
        ws5.cell(row=i, column=2, value=sev_stats.get(sev, 0))
    _style_data(ws5, len(stats) + 7, len(stats) + 9, 2)

    ws5.cell(row=len(stats) + 11, column=1, value="按冲突类型分布")
    ws5.cell(row=len(stats) + 11, column=1).font = Font(bold=True, size=12)

    ws5.cell(row=len(stats) + 12, column=1, value="冲突类型")
    ws5.cell(row=len(stats) + 12, column=2, value="数量")
    _style_header(ws5, len(stats) + 12, 2)

    type_stats = {}
    for c in conflicts:
        type_stats[c.conflict_type] = type_stats.get(c.conflict_type, 0) + 1
    for i, (ctype, count) in enumerate(sorted(type_stats.items()), len(stats) + 13):
        ws5.cell(row=i, column=1, value=ctype)
        ws5.cell(row=i, column=2, value=count)
    _style_data(ws5, len(stats) + 13, len(stats) + 12 + len(type_stats), 2)

    ws5.cell(row=len(stats) + 15 + len(type_stats), column=1, value="报告生成时间")
    ws5.cell(row=len(stats) + 15 + len(type_stats), column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    wb.save(output)
    return output.getvalue()
