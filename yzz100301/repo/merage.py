#!/usr/bin/env python3
import argparse
from copy import copy
from datetime import datetime
from pathlib import Path
import sys

from openpyxl import load_workbook


def nonempty(values):
    return any(value is not None for value in values)


def used_header(ws):
    max_col = 0
    for col in range(1, ws.max_column + 1):
        if ws.cell(1, col).value is not None:
            max_col = col
    return [ws.cell(1, col).value for col in range(1, max_col + 1)]


def copy_cell(src_cell, dst_cell):
    dst_cell.value = src_cell.value
    if src_cell.has_style:
        dst_cell._style = copy(src_cell._style)
    dst_cell.number_format = src_cell.number_format
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.protection = copy(src_cell.protection)
    if src_cell.hyperlink:
        dst_cell._hyperlink = copy(src_cell.hyperlink)
    if src_cell.comment:
        dst_cell.comment = copy(src_cell.comment)


def default_output_path(first_input):
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return first_input.parent / f"merged_{stamp}.xlsx"


def merge_workbooks(paths, output_path):
    first = paths[0]
    out_wb = load_workbook(first)
    out_ws = out_wb.active
    header = used_header(out_ws)
    if not header:
        raise ValueError(f"{first} has no header row")

    if out_ws.max_row > 1:
        out_ws.delete_rows(2, out_ws.max_row - 1)

    write_row = 2
    counts = {}

    for path in paths:
        src_wb = load_workbook(path, data_only=False)
        src_ws = src_wb.active
        src_header = used_header(src_ws)
        if src_header != header:
            raise ValueError(
                f"header mismatch in {path}\n"
                f"expected: {header}\n"
                f"actual:   {src_header}"
            )

        copied = 0
        for row_idx in range(2, src_ws.max_row + 1):
            row_values = [
                src_ws.cell(row_idx, col).value for col in range(1, len(header) + 1)
            ]
            if not nonempty(row_values):
                continue

            out_ws.row_dimensions[write_row].height = src_ws.row_dimensions[
                row_idx
            ].height
            for col_idx in range(1, len(header) + 1):
                copy_cell(src_ws.cell(row_idx, col_idx), out_ws.cell(write_row, col_idx))
            write_row += 1
            copied += 1
        counts[path.name] = copied

    out_ws.freeze_panes = "A2"
    out_ws.auto_filter.ref = out_ws.dimensions
    output_path.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_path)
    return counts, write_row - 2


def parse_args(argv):
    parser = argparse.ArgumentParser(
        description="Merge xlsx files into one workbook and keep only one header row."
    )
    parser.add_argument("files", nargs="+", help="xlsx files to merge, in order")
    parser.add_argument("-o", "--output", help="output xlsx path")
    return parser.parse_args(argv)


def main(argv):
    args = parse_args(argv)
    paths = [Path(file).expanduser().resolve() for file in args.files]
    missing = [str(path) for path in paths if not path.exists()]
    if missing:
        print("Missing files:", file=sys.stderr)
        for path in missing:
            print(f"  {path}", file=sys.stderr)
        return 2

    output_path = (
        Path(args.output).expanduser().resolve()
        if args.output
        else default_output_path(paths[0]).resolve()
    )
    if output_path in paths:
        print("Output path must not be one of the input files.", file=sys.stderr)
        return 2

    try:
        counts, total_rows = merge_workbooks(paths, output_path)
    except Exception as exc:
        print(f"Merge failed: {exc}", file=sys.stderr)
        return 1

    print(f"Output: {output_path}")
    print(f"Data rows: {total_rows}")
    for name, count in counts.items():
        print(f"{name}: {count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
